#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Contrôle de conformité du corpus tarif HA76.

Audit AUTONOME : relit les .md produits sans réutiliser aucune fonction du
générateur, et les confronte à la note de cadrage (forme) et à l'Excel / au PDF
(fidélité). Aucun import de generateur_tarif_HA76.
"""
import openpyxl
import re
import os
import glob
from collections import defaultdict, Counter
from openpyxl.utils import column_index_from_string as CI

XLSX = "/mnt/user-data/uploads/HA76_-infos-tarifs.xlsx"
TXT_PDF = "/tmp/ha76.txt"
OUTDIR = "/mnt/user-data/outputs"
PLAFOND = 200
PREFIXE = "HA76 Porte d'entrée aluminium — "
RE_SOURCE = re.compile(
    r"^\*Source : Tarif—HA76—HT—23-06-2026\.pdf, pages? (\d+|\d+ à \d+) — "
    r"information (originale|complémentaire) — SC(\d{4})\*$")

OK, KO, WARN = [], [], []


def ok(m):
    OK.append(m)


def ko(m):
    KO.append(m)


def warn(m):
    WARN.append(m)


def norm_num(s):
    return re.sub(r"[\s\u202f\u00a0]", "", s)


# ------------------------------------------------------------------ relecture md
def parse(path):
    raw = open(path, encoding="utf-8").read()
    m = re.match(r"^---\n(.*?)\n---\n", raw, re.S)
    fm = m.group(1) if m else ""
    corps = raw[m.end():] if m else raw
    chunks = []
    for bloc in re.split(r"\n(?=## )", corps):
        bloc = bloc.strip()
        if not bloc.startswith("## "):
            continue
        lignes = bloc.split("\n")
        titre = lignes[0][3:].strip()
        src = lignes[1].strip() if len(lignes) > 1 else ""
        texte = " ".join(l.strip() for l in lignes[2:]).strip()
        chunks.append({"titre": titre, "src": src, "corps": texte, "bloc": bloc})
    return fm, chunks


FICHIERS = {}
for p in sorted(glob.glob(f"{OUTDIR}/Tarif_HA76_*.md")):
    nom = os.path.basename(p)[len("Tarif_HA76_"):-3]
    FICHIERS[nom] = parse(p)

# ------------------------------------------------------------------ Excel
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Feuil1"]


def cell(r, letter):
    v = ws.cell(r, CI(letter)).value
    return "" if v is None else str(v).strip()


lignes_modele = [r for r in range(2, ws.max_row + 1) if cell(r, "C")]
groupes = defaultdict(list)
for r in lignes_modele:
    groupes[cell(r, "C")].append(r)
MODELES = list(groupes)

TTC_GELE = set()
for m, rs in groupes.items():
    if len({cell(r, "N") for r in rs}) > 1:
        TTC_GELE.add(m)
    else:
        try:
            if float(cell(rs[0], "N")) < float(cell(rs[0], "L")):
                TTC_GELE.add(m)
        except ValueError:
            pass

# ================================================================== 1. décomptes
attendus = {
    "METHODE": None,
    "PRIX_MODELES": len(MODELES) + sum(1 for m, rs in groupes.items() if cell(rs[0], "M")),
    "CARACTERISTIQUES": len(MODELES),
    "OPTIONS_MODELES": None,
    "COMPAT_EQUIPEMENTS": None,
    "CATALOGUE_OPTIONS": None,
    "FAISABILITES": None,
    "TRANSVERSES": None,
}
print("=== 1. Décomptes ===")
total = 0
for nom, (fm, ch) in FICHIERS.items():
    total += len(ch)
    att = attendus.get(nom)
    if att is None:
        print(f"  {nom:22s} {len(ch):4d}")
    elif att == len(ch):
        print(f"  {nom:22s} {len(ch):4d}  attendu {att}  OK")
        ok(f"décompte {nom}")
    else:
        print(f"  {nom:22s} {len(ch):4d}  attendu {att}  ECHEC")
        ko(f"décompte {nom} : {len(ch)} au lieu de {att}")
print(f"  {'TOTAL':22s} {total:4d}")

# ================================================================== 2. forme
print("\n=== 2. Forme des chunks ===")
err_plafond = err_src = err_prefixe = err_puce = 0
titres = Counter()
for nom, (fm, ch) in FICHIERS.items():
    for c in ch:
        if len(re.findall(r"\S+", c["bloc"])) > PLAFOND:
            err_plafond += 1
            ko(f"plafond dépassé : {c['titre'][:70]}")
        if not RE_SOURCE.match(c["src"]):
            err_src += 1
            ko(f"ligne de source non conforme : {c['src'][:80]}")
        if not c["titre"].startswith(PREFIXE):
            err_prefixe += 1
            ko(f"préfixe de titre absent : {c['titre'][:70]}")
        # une puce est un marqueur en tête de ligne, pas un tiret dans un libellé
        if any(re.match(r"^\s*[-*•]\s", l) for l in c["bloc"].split("\n")):
            err_puce += 1
            ko(f"puce détectée : {c['titre'][:70]}")
        titres[c["titre"]] += 1
for lib, n in [("plafond 200 mots", err_plafond), ("ligne de source", err_src),
               ("préfixe de titre", err_prefixe), ("prose sans puces", err_puce)]:
    print(f"  {lib:24s} {'OK' if n == 0 else f'ECHEC ({n})'}")
    (ok if n == 0 else lambda x: None)(lib)
dbl = [t for t, n in titres.items() if n > 1]
print(f"  {'unicité des titres':24s} {'OK' if not dbl else f'ECHEC ({len(dbl)})'}")
if dbl:
    for t in dbl[:5]:
        ko(f"titre dupliqué : {t[:80]}")
else:
    ok("unicité des titres")

# ================================================================== 3. SC
print("\n=== 3. Continuité SC ===")
for nom, (fm, ch) in FICHIERS.items():
    scs = [int(RE_SOURCE.match(c["src"]).group(3)) for c in ch if RE_SOURCE.match(c["src"])]
    attendu = list(range(2, 2 + len(ch)))
    if scs == attendu:
        print(f"  {nom:22s} SC0002 → SC{scs[-1]:04d}  OK")
        ok(f"SC {nom}")
    else:
        print(f"  {nom:22s} ECHEC")
        ko(f"continuité SC rompue dans {nom}")

# ================================================================== 4. front matter
print("\n=== 4. Front matter ===")
REQUIS = ["document_source", "type_document", "gamme_code", "gamme_nom", "collection",
          "materiau", "version_doc", "date_validite", "remplace", "perimetre",
          "audiences", "glossaire_ref"]
manque = 0
for nom, (fm, ch) in FICHIERS.items():
    abs_ = [k for k in REQUIS if not re.search(rf"^{k}:", fm, re.M)]
    if abs_:
        manque += 1
        ko(f"front matter incomplet dans {nom} : {abs_}")
print(f"  {'complétude':24s} {'OK' if manque == 0 else f'ECHEC ({manque})'}")
(ok if manque == 0 else lambda x: None)("front matter")

# ================================================================== 5. anti-fantôme
print("\n=== 5. Anti-fantôme ===")
prix = FICHIERS["PRIX_MODELES"][1]
c1 = [c for c in prix if "Tarif " in c["titre"] and "1 vantail" in c["titre"]]
c2 = [c for c in prix if "2 vantaux" in c["titre"]]
sans2 = {m for m, rs in groupes.items() if not cell(rs[0], "M")}
fantome = [c for c in c2 if any(re.search(rf"Tarif {re.escape(m)} 2 vantaux", c["titre"]) for m in sans2)]
print(f"  chunks 1 vantail       {len(c1):4d}  modèles {len(MODELES)}  "
      f"{'OK' if len(c1) == len(MODELES) else 'ECHEC'}")
print(f"  chunks 2 vantaux       {len(c2):4d}  modèles tarifés 2V {len(MODELES)-len(sans2)}  "
      f"{'OK' if len(c2) == len(MODELES)-len(sans2) else 'ECHEC'}")
print(f"  chunks fantômes        {len(fantome):4d}  {'OK' if not fantome else 'ECHEC'}")
for cond, lib in [(len(c1) == len(MODELES), "un chunk 1V par modèle"),
                  (len(c2) == len(MODELES) - len(sans2), "aucun 2V manquant"),
                  (not fantome, "aucun chunk fantôme")]:
    (ok if cond else ko)(lib)

# dédoublonnage : le modèle est sur plusieurs lignes Excel, jamais sur deux chunks
dup = [m for m in MODELES
       if sum(1 for c in c1 if re.search(rf"Tarif {re.escape(m)} 1 vantail", c["titre"])) > 1]
print(f"  doublons de modèle     {len(dup):4d}  {'OK' if not dup else 'ECHEC'}")
(ok if not dup else ko)("dédoublonnage des emplacements d'affichage")

# ================================================================== 6. fidélité prix
print("\n=== 6. Fidélité numérique des prix modèles ===")
err = 0
verifies = 0
for c in prix:
    m = re.search(r"— Tarif (.+?) (1 vantail|2 vantaux) \(", c["titre"])
    if not m:
        ko(f"titre de prix illisible : {c['titre'][:70]}")
        continue
    mod, conf = m.group(1), m.group(2)
    if mod not in groupes:
        ko(f"modèle inconnu de l'Excel : {mod}")
        continue
    r0 = groupes[mod][0]
    col_ht, col_ttc = ("L", "N") if conf == "1 vantail" else ("M", "O")
    att_ht = int(round(float(cell(r0, col_ht))))
    trouve = re.findall(r"([\d\u202f  ]+) € HT", c["corps"])
    if len(trouve) != 1 or int(norm_num(trouve[0])) != att_ht:
        err += 1
        ko(f"prix HT infidèle : {mod} {conf} → {trouve} vs {att_ht}")
    else:
        verifies += 1
    ttc_trouve = re.findall(r"([\d\u202f  ]+) € TTC", c["corps"])
    if mod in TTC_GELE and conf == "1 vantail":
        if ttc_trouve:
            err += 1
            ko(f"TTC servi alors qu'il est gelé : {mod}")
    else:
        att_ttc = cell(r0, col_ttc)
        if att_ttc:
            att_ttc = int(round(float(att_ttc)))
            if len(ttc_trouve) != 1 or int(norm_num(ttc_trouve[0])) != att_ttc:
                err += 1
                ko(f"prix TTC infidèle : {mod} {conf} → {ttc_trouve} vs {att_ttc}")
            else:
                verifies += 1
print(f"  montants confrontés à l'Excel : {verifies}   écarts : {err}   "
      f"{'OK' if err == 0 else 'ECHEC'}")
(ok if err == 0 else lambda x: None)("fidélité prix modèles")

# ================================================================== 7. fidélité options
print("\n=== 7. Fidélité numérique des options par modèle ===")
FAM = [("AE", "AF", "AG"), ("AS", "AX", "AY"), ("BB", "BD", "BE"),
       ("BG", "BI", "BJ"), ("BO", "BR", "BS"), ("BW", "BW", "BX")]
# la colonne de libellé du panneau analogue est vide dans l'Excel ; le libellé
# est repris de l'en-tête de bloc de la page modèle du PDF (page 95 vérifiée).
LIB_PDF = {"BW": "Panneau analogue pour fixe latéral"}
ref_opt = {}
for m, rs in groupes.items():
    for r in rs:
        for lib_c, ht_c, ttc_c in FAM:
            if lib_c in LIB_PDF:
                lib = LIB_PDF[lib_c] if cell(r, lib_c) else ""
            else:
                lib = re.sub(r"\s+", " ", cell(r, lib_c).replace("\n", " ")).strip().rstrip(".")
            if not lib:
                continue
            ref_opt[(m, lib.lower())] = (cell(r, ht_c), cell(r, ttc_c))
err = verif = 0
introuvables = 0
for c in FICHIERS["OPTIONS_MODELES"][1]:
    mt = re.search(r"— Option (.+?) sur (.+?) \(ligne", c["titre"])
    if not mt:
        ko(f"titre d'option illisible : {c['titre'][:70]}")
        continue
    lib, mod = mt.group(1).rstrip("…").strip(), mt.group(2)
    base = re.sub(r"^vitrage ornemental du .*$", "vitrage ornemental", lib.lower())
    cands = [v for (mm, ll), v in ref_opt.items()
             if mm == mod and ll.startswith(base[:40])]
    if not cands:
        introuvables += 1
        ko(f"option absente de l'Excel : {lib[:40]} / {mod}")
        continue
    att_ht, att_ttc = cands[0]
    tr_ht = re.findall(r"([\d\u202f  ]+) € HT", c["corps"])
    if float(att_ht) == 0:
        if tr_ht or "sans plus-value" not in c["corps"]:
            err += 1
            ko(f"plus-value nulle mal rendue : {lib[:40]} / {mod}")
        else:
            verif += 1
    else:
        if len(tr_ht) != 1 or int(norm_num(tr_ht[0])) != int(round(float(att_ht))):
            err += 1
            ko(f"plus-value HT infidèle : {lib[:40]} / {mod} → {tr_ht} vs {att_ht}")
        else:
            verif += 1
        if att_ttc:
            tr_ttc = re.findall(r"([\d\u202f  ]+) € TTC", c["corps"])
            if len(tr_ttc) != 1 or int(norm_num(tr_ttc[0])) != int(round(float(att_ttc))):
                err += 1
                ko(f"plus-value TTC infidèle : {lib[:40]} / {mod}")
            else:
                verif += 1
print(f"  montants confrontés : {verif}   écarts : {err}   introuvables : {introuvables}   "
      f"{'OK' if err == 0 and introuvables == 0 else 'ECHEC'}")
(ok if err == 0 and introuvables == 0 else lambda x: None)("fidélité options")

# ================================================================== 8. fidélité Ud
print("\n=== 8. Fidélité des Ud et des limites dimensionnelles ===")
err = verif = 0
for c in FICHIERS["CARACTERISTIQUES"][1]:
    mt = re.search(r"— Caractéristiques (.+?) \(ligne", c["titre"])
    mod = mt.group(1)
    rs = groupes[mod]
    src_ud = next((cell(r, "T") for r in rs if cell(r, "T")), "")
    val = re.search(r"([\d,]+)\s*W/m²\.K", src_ud.replace("m2", "m²"))
    tr = re.search(r"performance thermique est de ([\d,]+) W/m²\.K", c["corps"])
    if val and tr:
        if val.group(1) != tr.group(1):
            err += 1
            ko(f"Ud infidèle : {mod} → {tr.group(1)} vs {val.group(1)}")
        else:
            verif += 1
    elif val or tr:
        err += 1
        ko(f"Ud présent d'un seul côté : {mod}")
    # enveloppe dimensionnelle
    lm = [float(cell(rs[0], x)) for x in ("AK", "AO") if cell(rs[0], x)]
    lM = [float(cell(rs[0], x)) for x in ("AM", "AQ") if cell(rs[0], x)]
    d = re.search(r"de ([\d\u202f ]+) à ([\d\u202f ]+) mm de largeur", c["corps"])
    if lm and lM and d:
        if int(norm_num(d.group(1))) != int(min(lm)) or int(norm_num(d.group(2))) != int(max(lM)):
            err += 1
            ko(f"enveloppe dimensionnelle infidèle : {mod}")
        else:
            verif += 1
print(f"  valeurs confrontées : {verif}   écarts : {err}   {'OK' if err == 0 else 'ECHEC'}")
(ok if err == 0 else lambda x: None)("fidélité Ud et dimensions")

# ================================================================== 9. catalogue
print("\n=== 9. Fidélité du catalogue d'options ===")
montants_xl = Counter()
for r in range(2, ws.max_row + 1):
    if cell(r, "CF") and not cell(r, "C") and cell(r, "CK"):
        montants_xl[int(round(float(cell(r, "CK"))))] += 1
err = verif = 0
for c in FICHIERS["CATALOGUE_OPTIONS"][1]:
    tr = re.findall(r"([\d\u202f  ]+) € HT", c["corps"])
    if not tr:
        continue
    v = int(norm_num(tr[0]))
    if montants_xl[v] == 0:
        err += 1
        ko(f"montant catalogue absent de l'Excel : {v} € — {c['titre'][:60]}")
    else:
        verif += 1
print(f"  montants confrontés : {verif}   écarts : {err}   {'OK' if err == 0 else 'ECHEC'}")
(ok if err == 0 else lambda x: None)("fidélité catalogue")

# ================================================================== 10. règle transverses
print("\n=== 10. Règle transverses et faisabilités : aucun montant ===")
err = 0
for nom in ("TRANSVERSES", "FAISABILITES", "METHODE"):
    for c in FICHIERS[nom][1]:
        if "€" in c["bloc"]:
            err += 1
            ko(f"montant présent dans {nom} : {c['titre'][:60]}")
print(f"  occurrences du symbole € : {err}   {'OK' if err == 0 else 'ECHEC'}")
(ok if err == 0 else lambda x: None)("règle transverses")

# ================================================================== 11. grille gelée
print("\n=== 11. Chapitre gelé : grille des fixes ===")
grille = set()
for r in range(2, ws.max_row + 1):
    if cell(r, "CF") == "Fixes" and cell(r, "CM"):
        for col in [f"C{c}" for c in "NOPQRSTUVWXYZ"] + ["DA", "DB", "DC", "DD", "DE",
                                                          "DF", "DG", "DH", "DI", "DJ", "DK"]:
            try:
                v = cell(r, col)
            except ValueError:
                continue
            if v:
                grille.add(int(round(float(v))))
fuite = 0
for nom, (fm, ch) in FICHIERS.items():
    for c in ch:
        for v in re.findall(r"([\d\u202f  ]+) € HT", c["corps"]):
            if int(norm_num(v)) in grille and nom in ("TRANSVERSES", "FAISABILITES"):
                fuite += 1
print(f"  prix de grille servis : {fuite}   {'OK' if fuite == 0 else 'ECHEC'}")
(ok if fuite == 0 else lambda x: None)("grille des fixes non servie")

# ================================================================== 12. discrimination
print("\n=== 12. Discrimination HA76 / HAM76 ===")
mentions = []
for nom, (fm, ch) in FICHIERS.items():
    for c in ch:
        if "HAM76" in c["bloc"] and "distinction avec la gamme HAM76" not in c["titre"]:
            mentions.append((nom, c["titre"][:60]))
sans_code = [c["titre"] for nom, (fm, ch) in FICHIERS.items() for c in ch
             if not c["titre"].startswith("HA76 ")]
sans_gamme = 0
for nom in ("PRIX_MODELES", "OPTIONS_MODELES", "CARACTERISTIQUES"):
    for c in FICHIERS[nom][1]:
        if "gamme HA76" not in c["corps"]:
            sans_gamme += 1
print(f"  mentions HAM76 hors chunk dédié : {len(mentions)}   {'OK' if not mentions else 'ECHEC'}")
print(f"  titres sans code de gamme       : {len(sans_code)}   {'OK' if not sans_code else 'ECHEC'}")
print(f"  corps sans mention 'gamme HA76' : {sans_gamme}   {'OK' if sans_gamme == 0 else 'ECHEC'}")
for cond, lib in [(not mentions, "pas de contamination HAM76"),
                  (not sans_code, "code de gamme en tête de titre"),
                  (sans_gamme == 0, "gamme nommée dans le corps")]:
    (ok if cond else lambda x: None)(lib)

# ================================================================== 13. liant
print("\n=== 13. Liant inter-fichiers ===")
def couples(nom, rx):
    d = {}
    for c in FICHIERS[nom][1]:
        m = re.search(rx, c["titre"])
        if m:
            d[m.group(1)] = m.group(2)
    return d
p = couples("PRIX_MODELES", r"— Tarif (.+?) (?:1 vantail|2 vantaux) (\(ligne .+?\))")
k = couples("CARACTERISTIQUES", r"— Caractéristiques (.+?) (\(ligne .+?\))")
o = couples("OPTIONS_MODELES", r"— Option .+? sur (.+?) (\(ligne .+?\))")
err = [m for m in p if m in k and p[m] != k[m]] + [m for m in p if m in o and p[m] != o[m]]
manquants = [m for m in p if m not in k]
print(f"  qualifiants divergents : {len(err)}   modèles sans caractéristiques : {len(manquants)}   "
      f"{'OK' if not err and not manquants else 'ECHEC'}")
(ok if not err and not manquants else lambda x: None)("liant inter-fichiers")

# ================================================================== 14. croisement PDF
print("\n=== 14. Croisement PDF par échantillon ===")
pages = open(TXT_PDF, encoding="utf-8", errors="replace").read().split("\f")
ech, trouves = 0, 0
for c in prix[:400:17]:
    mt = re.search(r"— Tarif (.+?) (1 vantail|2 vantaux) \(", c["titre"])
    mod = mt.group(1)
    pg = int(RE_SOURCE.match(c["src"]).group(1))
    ech += 1
    txt = pages[pg - 1]
    val = re.findall(r"([\d\u202f  ]+) € HT", c["corps"])
    brut = norm_num(val[0]) if val else ""
    if mod in txt and (brut in norm_num(txt) or not brut):
        trouves += 1
    else:
        warn(f"échantillon PDF non retrouvé : {mod} page {pg} valeur {brut}")
print(f"  échantillons : {ech}   concordants : {trouves}")
(ok if trouves == ech else lambda x: None)("croisement PDF par échantillon")

# ================================================================== bilan
print("\n" + "=" * 60)
print(f"BILAN : {len(OK)} contrôles réussis, {len(KO)} échecs, {len(WARN)} avertissements")
if KO:
    print("\nÉchecs :")
    for m in KO[:40]:
        print("  ✗", m)
    if len(KO) > 40:
        print(f"  … et {len(KO)-40} autres")
if WARN:
    print("\nAvertissements :")
    for m in WARN[:20]:
        print("  !", m)
