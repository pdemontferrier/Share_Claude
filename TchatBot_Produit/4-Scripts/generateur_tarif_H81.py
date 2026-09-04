#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Générateur de chunks Markdown pour le tarif H81 (gamme pilote).
Produit 5 fichiers conformes à note_cadrage_migration_tarif_v1.md :
  1. prix                     (règle 1-3, un chunk par référence × configuration vantaux)
  2. options                  (règle 4, un chunk par couple option × modèle chiffré)
  3. caractéristiques         (règle 5, un chunk par modèle, enveloppe dimensionnelle)
  4. compatibilité équipements(règle 6, un chunk par équipement)
  5. pages transverses        (règle 7, un chunk par bloc, orientation sans montant)

Transversal : SC continue par fichier depuis SC0002, ligne de source normée,
contrôle plafond 200 mots, journal des colonnes non mappées.
"""
import openpyxl
import re
import unicodedata
from collections import OrderedDict, defaultdict

GAMME = "H81"
DESIGNATION = "Porte d'entrée PVC"
PREFIXE = f"{GAMME} {DESIGNATION}"
XLSX = "/mnt/user-data/uploads/H81-modèles_porte.xlsx"
PDF_SOURCE = "Tarif—H81—HT—08-04-2026.pdf"   # nom affiché (em-dashes) dans la ligne de source
OUTDIR = "/mnt/user-data/outputs"
PLAFOND = 200

# ---------------------------------------------------------------- utilitaires
def col(ws, r, letter):
    return ws[f"{letter}{r}"].value

def clean(v):
    """Nettoie une valeur cellule -> str strippée, '' si vide."""
    if v is None:
        return ""
    return str(v).strip()

def fmt_euro(v):
    """2614 -> '2 614'. Espace insécable fine comme séparateur de milliers."""
    if v in (None, ""):
        return None
    try:
        n = int(round(float(v)))
    except (ValueError, TypeError):
        return None
    s = f"{n:,}".replace(",", "\u202f")   # 2 614
    return s

def norm_ud(v):
    """Normalise l'Ud et extrait la valeur seule.
    'Ud porte vitrée : 1,2W/m2.K' -> '1,2 W/m².K' (sans le libellé de tête)."""
    s = clean(v)
    if not s:
        return ""
    s = s.replace("m2", "m²")
    # isoler la partie après ':' si présente (retire 'Ud porte vitrée :')
    if ":" in s:
        s = s.split(":", 1)[1]
    s = re.sub(r"(\d)\s*W", r"\1 W", s)     # 1,2W -> 1,2 W
    s = re.sub(r"\s+", " ", s).strip()
    return s

def count_words(*parts):
    """Compte les mots (\\S+) sur titre + source + corps."""
    txt = " ".join(parts)
    return len(re.findall(r"\S+", txt))

def lower_keep_acronyms(s):
    """Passe un libellé en minuscule pour lecture en phrase, mais préserve les
    tokens tout-majuscule (acronymes: RAL) et les codes alphanumériques (44/6, I45)."""
    def fix(tok):
        core = tok.strip(".,;:()")
        if not core:
            return tok
        # préserver si tout majuscule (RAL), ou contient un chiffre (44/6, 7016, I45)
        if core.isupper() or any(ch.isdigit() for ch in core):
            return tok
        return tok.lower()
    return " ".join(fix(t) for t in s.split())

def sc_id(n):
    return f"SC{n:04d}"

def source_line(page, sc, nature="originale"):
    return f"*Source : {PDF_SOURCE}, page {page} — information {nature} — {sc}*"

# ---------------------------------------------------------------- chargement
def load_rows():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["modèles portes"]
    return ws

def group_by_model(ws):
    """
    Groupe les lignes par modèle. Une référence = un modèle (qui implique sa ligne).
    Retourne OrderedDict {modele: {'ligne':.., 'collection':.., 'page':.., 'rows':[r,..],
                                    'prix':{...}, 'ud':.., 'vitrage':.., 'dimvit':.., 'dims':{...}}}.
    La 1ère ligne du modèle porte le tronc commun (prix, page, caractéristiques).
    """
    models = OrderedDict()
    for r in range(2, ws.max_row + 1):
        m = clean(col(ws, r, "A"))
        if not m:
            continue
        if m not in models:
            models[m] = {
                "ligne": clean(col(ws, r, "D")).replace("Ligne ", ""),
                "collection": clean(col(ws, r, "E")),
                "gamme_desc": clean(col(ws, r, "C")),
                "page": clean(col(ws, r, "F")),
                "rows": [],
                "prix": {
                    "ht1": col(ws, r, "G"), "ht2": col(ws, r, "H"),
                    "ttc1": col(ws, r, "I"), "ttc2": col(ws, r, "J"),
                },
                "ud": norm_ud(col(ws, r, "O")),
                "vitrage": clean(col(ws, r, "R")).replace("\n", " "),
                "dimvit": clean(col(ws, r, "P")),
                "dims": {p: [col(ws, r, c) for c in cols]
                         for p, cols in {
                             "5103": ["AF","AG","AH","AI"],
                             "5107": ["AJ","AK","AL","AM"],
                             "5114": ["AN","AO","AP","AQ"],
                             "5120": ["AR","AS","AT","AU"],
                         }.items()},
            }
        models[m]["rows"].append(r)
    return models

# ---------------------------------------------------------------- alertes plafond
ALERTS = []
def emit(title, source, body, sc):
    """Assemble un chunk et vérifie le plafond. Retourne le texte du chunk."""
    n = count_words(title, source, body)
    if n > PLAFOND:
        ALERTS.append(f"PLAFOND DÉPASSÉ ({n} mots) : {title}")
    return f"## {title}\n{source}\n\n{body}\n"

def yaml_front(type_doc, nb_chunks):
    return (
        "---\n"
        f"document_source: Tarif_H81_HT_08-04-2026.pdf\n"
        f"type_document: tarif\n"
        f"sous_type: {type_doc}\n"
        f"gamme_code: H81\n"
        f'gamme_nom: "Porte d\'entrée PVC"\n'
        'collection: "TRYBA PVC"\n'
        f"materiau: PVC\n"
        f"version_doc: \"2026.04\"\n"
        f"date_validite: 2026-04-08\n"
        f"nb_chunks: {nb_chunks}\n"
        "audiences: [ADV, commercial]\n"
        "---\n\n"
    )

CONFIGS = [("ht1", "ttc1", "un vantail", "1 vantail"),
           ("ht2", "ttc2", "deux vantaux égaux", "2 vantaux")]

# ---------------------------------------------------------------- 1. PRIX
def gen_prix(models):
    chunks, sc = [], 2
    for m, d in models.items():
        pref = f"{PREFIXE} — Tarif {m}"
        mention = f"(ligne {d['ligne']}, collection {d['collection']})"
        for htk, ttck, phrase, vlabel in CONFIGS:
            ht, ttc = fmt_euro(d["prix"][htk]), fmt_euro(d["prix"][ttck])
            if ht is None:          # anti-fantôme : configuration non tarifée
                continue
            title = f"{pref} {vlabel} {mention}"
            body = (f"En {phrase}, le modèle {m} de la porte d'entrée PVC H81, "
                    f"ligne {d['ligne']}, collection {d['collection']}, est proposé au "
                    f"tarif de {ht} € HT, soit {ttc} € TTC. Ce tarif s'entend hors "
                    f"éco-participation.")
            chunks.append(emit(title, source_line(d["page"], sc_id(sc)), body, sc_id(sc)))
            sc += 1
    return chunks

# ---------------------------------------------------------------- 2. OPTIONS
def gen_options(ws, models):
    """Un chunk par couple option × modèle CHIFFRÉ (colonnes AW/BA/BC)."""
    # index modèle par 1ère ligne pour retrouver ligne/collection/page
    row2model = {}
    for m, d in models.items():
        for r in d["rows"]:
            row2model[r] = m
    chunks, sc = [], 2
    for r in range(2, ws.max_row + 1):
        m = row2model.get(r)
        if not m:
            continue
        opt = clean(col(ws, r, "AW"))
        raw_pvht = col(ws, r, "BA")
        # anti-fantôme : une plus-value doit être un montant strictement positif.
        # 0, vide ou non numérique = option non chiffrée (renvoi transverse) -> exclue.
        try:
            if raw_pvht in (None, "") or float(raw_pvht) <= 0:
                continue
        except (ValueError, TypeError):
            continue
        if not opt:
            continue
        pvht = fmt_euro(raw_pvht)
        pvttc = fmt_euro(col(ws, r, "BC"))
        d = models[m]
        opt_lc = lower_keep_acronyms(opt)
        desc = clean(col(ws, r, "AY")).replace("\n", " ")
        desc_short = (desc.split(".")[0] + ".") if desc else ""
        title = (f"{PREFIXE} — Option {opt_lc} sur {m} "
                 f"(ligne {d['ligne']}, collection {d['collection']})")
        body = (f"Sur le modèle {m} de la porte d'entrée PVC H81, ligne {d['ligne']}, "
                f"collection {d['collection']}, l'option {opt_lc} est proposée en "
                f"plus-value au tarif de {pvht} € HT, soit {pvttc} € TTC. ")
        if desc_short:
            body += desc_short + " "
        body += "Cette plus-value s'entend hors éco-participation."
        chunks.append(emit(title, source_line(d["page"], sc_id(sc)), body, sc_id(sc)))
        sc += 1
    return chunks

# ---------------------------------------------------------------- 3. CARACTÉRISTIQUES
def enveloppe(dims):
    """min des Mini L/H, max des Maxi L/H sur les 4 profils. Retourne (Lmin,Lmax,Hmin,Hmax) ou None."""
    minL = minH = maxL = maxH = None
    for p, vals in dims.items():
        miniL, miniH, maxiL, maxiH = vals
        for v, key in [(miniL,"minL"),(miniH,"minH"),(maxiL,"maxL"),(maxiH,"maxH")]:
            if v in (None, ""):
                continue
            v = float(v)
            if key == "minL": minL = v if minL is None else min(minL, v)
            if key == "minH": minH = v if minH is None else min(minH, v)
            if key == "maxL": maxL = v if maxL is None else max(maxL, v)
            if key == "maxH": maxH = v if maxH is None else max(maxH, v)
    if None in (minL, minH, maxL, maxH):
        return None
    return int(minL), int(maxL), int(minH), int(maxH)

def gen_caracteristiques(models):
    chunks, sc = [], 2
    for m, d in models.items():
        title = (f"{PREFIXE} — Caractéristiques {m} "
                 f"(ligne {d['ligne']}, collection {d['collection']})")
        # vitrage : la cellule peut contenir plusieurs phrases ; on prend la 1ère
        # (description du vitrage) et on met la dimension dans une phrase séparée.
        vit_full = d["vitrage"].strip()
        vit_desc = vit_full.split(".")[0].strip() if vit_full else ""
        vit_extra = ".".join(vit_full.split(".")[1:]).strip().rstrip(".") if vit_full and "." in vit_full else ""
        body = (f"Le modèle {m} de la porte d'entrée PVC H81, ligne {d['ligne']}, "
                f"collection {d['collection']}, reçoit en vitrage de base "
                f"{vit_desc[0].lower() + vit_desc[1:] if vit_desc else 'un vitrage standard'}")
        if d["dimvit"]:
            body += f", de dimensions {d['dimvit']}"
        body += ". "
        if vit_extra:
            body += vit_extra + ". "
        if d["ud"]:
            body += f"Sa performance thermique est un coefficient Ud de {d['ud']}. "
        env = enveloppe(d["dims"])
        if env:
            Lmin, Lmax, Hmin, Hmax = env
            body += (f"Selon le profil de dormant retenu (5103, 5107, 5114 ou 5120), il se "
                     f"fabrique dans une plage de largeur comprise entre {Lmin} et {Lmax} mm "
                     f"et de hauteur comprise entre {Hmin} et {Hmax} mm, les limites exactes "
                     f"dépendant du profil choisi.")
        chunks.append(emit(title, source_line(d["page"], sc_id(sc)), body.strip(), sc_id(sc)))
        sc += 1
    return chunks

# ---------------------------------------------------------------- 4. COMPATIBILITÉ ÉQUIPEMENTS
EQUIP_PAGE = "14"   # page transversale équipements (à confirmer)
def gen_compat(ws, models):
    equip_oui = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        eq = clean(col(ws, r, "T")); fa = clean(col(ws, r, "S")).lower(); m = clean(col(ws, r, "A"))
        if eq and fa == "oui" and m:
            equip_oui[eq].append(m)
    chunks, sc = [], 2
    recurrents = ["Judas optique", "Heurtoir", "Passe-lettres", "Chatière"]
    # article défini par équipement (genre/élision)
    articles = {"Judas optique": "le judas optique", "Heurtoir": "le heurtoir",
                "Passe-lettres": "le passe-lettres", "Chatière": "la chatière"}
    for eq in recurrents:
        modeles = equip_oui.get(eq, [])
        if not modeles:
            continue
        eq_lc = eq.lower()
        eq_art = articles.get(eq, eq_lc)
        title = f"{PREFIXE} — Compatibilité de l'équipement {eq_lc} par modèle"
        liste = ", ".join(modeles[:-1]) + (" et " + modeles[-1] if len(modeles) > 1 else modeles[0])
        body = (f"Sur la porte d'entrée PVC H81, {eq_art} peut équiper "
                f"les modèles suivants : {liste}. Sur les autres modèles de la gamme, "
                f"{eq_art} n'est pas disponible. Cet équipement n'est pas chiffré dans le tarif.")
        chunks.append(emit(title, source_line(EQUIP_PAGE, sc_id(sc)), body, sc_id(sc)))
        sc += 1
    return chunks

# ---------------------------------------------------------------- 5. PAGES TRANSVERSES
def gen_transverses():
    blocs = [
        ("20", "teintes (plaxage et offre couleurs)",
         "La porte d'entrée PVC H81 peut recevoir un plaxage de teinte, dont la faisabilité "
         "dépend du modèle. Les teintes sont réparties en groupes tarifaires : un groupe sans "
         "plus-value, et des groupes avec plus-value exprimée en pourcentage du prix. L'offre de "
         "couleurs disponibles et le détail des plus-values figurent à la page « Offre couleurs » "
         "du tarif. Le montant applicable dépend de la teinte et du groupe retenus ; il doit être "
         "lu directement sur cette page du tarif."),
        ("21", "laquage RAL",
         "La porte d'entrée PVC H81 peut être laquée dans une large gamme de teintes RAL, avec une "
         "faisabilité et des exceptions selon le modèle et les accessoires. Le laquage des "
         "accessoires fait l'objet d'une plus-value au mètre linéaire. La liste des teintes RAL "
         "disponibles et les plus-values associées figurent à la page « Laquage » du tarif. Le "
         "montant applicable dépend de la teinte et de la configuration ; il doit être lu "
         "directement sur cette page du tarif."),
        ("23", "plus-values vitrages",
         "La porte d'entrée PVC H81 en version entièrement vitrée peut recevoir différents vitrages "
         "en plus-value : vitrages thermiques, phoniques, de sécurité et ornementaux. Chaque vitrage "
         "est défini par ses performances et une plus-value au mètre carré de surface vitrée. La "
         "liste des vitrages et leurs plus-values figurent à la page « Plus-value vitrages » du "
         "tarif. Le montant applicable dépend du vitrage choisi et de la surface ; il doit être lu "
         "directement sur cette page."),
    ]
    chunks, sc = [], 2
    for page, label, body in blocs:
        title = f"{PREFIXE} — Existence et localisation des tarifs de {label}"
        chunks.append(emit(title, source_line(page, sc_id(sc)), body, sc_id(sc)))
        sc += 1
    return chunks

# ---------------------------------------------------------------- journal colonnes non mappées
MAPPED = set("A B C D E F G H I J O P R S T AW AY BA BC AF AG AH AI AJ AK AL AM AN AO AP AQ AR AS AT AU AE".split())
def journal_non_mappees(ws):
    lignes = []
    for c in range(1, ws.max_column + 1):
        letter = openpyxl.utils.get_column_letter(c)
        h = ws.cell(row=1, column=c).value
        nonempty = sum(1 for r in range(2, ws.max_row+1) if ws.cell(row=r,column=c).value not in (None,""))
        if nonempty and letter not in MAPPED:
            lignes.append(f"  {letter} ({h}) : {nonempty} lignes remplies — NON MAPPÉE")
    return lignes

# ---------------------------------------------------------------- écriture
def write_file(fname, type_doc, chunks):
    path = f"{OUTDIR}/{fname}"
    with open(path, "w", encoding="utf-8") as f:
        f.write(yaml_front(type_doc, len(chunks)))
        f.write("\n".join(chunks))
    return path, len(chunks)

if __name__ == "__main__":
    ws = load_rows()
    models = group_by_model(ws)
    print(f"Modèles groupés : {len(models)}\n")

    results = []
    results.append(write_file("Tarif_H81_PRIX.md", "prix", gen_prix(models)))
    results.append(write_file("Tarif_H81_OPTIONS.md", "options", gen_options(ws, models)))
    results.append(write_file("Tarif_H81_CARACTERISTIQUES.md", "caracteristiques", gen_caracteristiques(models)))
    results.append(write_file("Tarif_H81_COMPAT_EQUIPEMENTS.md", "compatibilite_equipements", gen_compat(ws, models)))
    results.append(write_file("Tarif_H81_PAGES_TRANSVERSES.md", "pages_transverses", gen_transverses()))

    total = 0
    print("=== Fichiers générés ===")
    for path, n in results:
        print(f"  {path.split('/')[-1]:38s} : {n} chunks")
        total += n
    print(f"  {'TOTAL':38s} : {total} chunks")

    print("\n=== Alertes plafond ===")
    print("  Aucune." if not ALERTS else "\n".join("  " + a for a in ALERTS))

    print("\n=== Journal : colonnes remplies NON mappées ===")
    for l in journal_non_mappees(ws):
        print(l)
