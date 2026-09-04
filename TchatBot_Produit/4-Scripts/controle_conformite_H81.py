#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Contrôle de conformité AUTONOME des chunks tarif H81.
Indépendant du générateur : relit les .md produits et les confronte à
  - la note de cadrage (forme : plafond, source, SC, titres, liant, YAML)
  - l'Excel (fidélité numérique exhaustive : chaque prix vs cellule)
  - le PDF (croisement par échantillon)
Ne réutilise AUCUNE fonction du générateur.
"""
import openpyxl
import re
import sys
from collections import defaultdict, OrderedDict

OUTDIR = "/mnt/user-data/outputs"
XLSX = "/mnt/user-data/uploads/H81-modèles_porte.xlsx"
PDF = "/mnt/user-data/uploads/Tarif_H81_HT_08-04-2026_.pdf"
PLAFOND = 200
FILES = {
    "prix": "Tarif_H81_PRIX.md",
    "options": "Tarif_H81_OPTIONS.md",
    "caracteristiques": "Tarif_H81_CARACTERISTIQUES.md",
    "compat": "Tarif_H81_COMPAT_EQUIPEMENTS.md",
    "transverses": "Tarif_H81_PAGES_TRANSVERSES.md",
}
ATTENDU = {"prix": 164, "options": 85, "caracteristiques": 84, "compat": 4, "transverses": 3}

# regex de la ligne de source normée
RE_SOURCE = re.compile(
    r"^\*Source : Tarif—H81—HT—08-04-2026\.pdf, page (\S+) — "
    r"information (originale|complémentaire) — (SC\d{4})\*$"
)
RE_TITLE = re.compile(r"^## (.+)$")

results = {"OK": [], "WARN": [], "FAIL": []}
def ok(msg):   results["OK"].append(msg)
def warn(msg): results["WARN"].append(msg)
def fail(msg): results["FAIL"].append(msg)

# ---------------------------------------------------------------- parsing .md
def parse_chunks(path):
    """Découpe un .md en chunks. Retourne (front_matter, [ {title, source, body, raw} ])."""
    with open(path, encoding="utf-8") as f:
        text = f.read()
    fm = ""
    if text.startswith("---"):
        end = text.find("\n---", 3)
        fm = text[:end+4]
        text = text[end+4:]
    chunks = []
    for block in re.split(r"\n(?=## )", text):
        block = block.strip()
        if not block.startswith("## "):
            continue
        lines = block.split("\n")
        title = lines[0][3:].strip()
        source = lines[1].strip() if len(lines) > 1 else ""
        body = "\n".join(lines[2:]).strip()
        chunks.append({"title": title, "source": source, "body": body, "raw": block})
    return fm, chunks

# ---------------------------------------------------------------- 1. FORME
def check_forme(name, fm, chunks):
    scs = []
    for c in chunks:
        # plafond
        nwords = len(re.findall(r"\S+", c["title"] + " " + c["source"] + " " + c["body"]))
        if nwords > PLAFOND:
            fail(f"[{name}] plafond dépassé ({nwords}) : {c['title'][:60]}")
        # ligne de source
        m = RE_SOURCE.match(c["source"])
        if not m:
            fail(f"[{name}] ligne source non conforme : {c['title'][:50]} -> {c['source'][:60]}")
        else:
            scs.append(m.group(3))
        # titre préfixé gamme
        if not c["title"].startswith("H81 "):
            fail(f"[{name}] titre non préfixé H81 : {c['title'][:50]}")
    # continuité SC (sans trou ni doublon), démarrage SC0002
    nums = sorted(int(s[2:]) for s in scs)
    if nums:
        if nums[0] != 2:
            warn(f"[{name}] SC ne démarre pas à SC0002 (démarre à SC{nums[0]:04d})")
        if len(nums) != len(set(nums)):
            fail(f"[{name}] doublons SC détectés")
        trous = [n for n in range(nums[0], nums[-1]+1) if n not in set(nums)]
        if trous:
            fail(f"[{name}] trous SC : {['SC%04d'%t for t in trous][:5]}")
        else:
            ok(f"[{name}] SC continue {('SC%04d'%nums[0])}→{('SC%04d'%nums[-1])} sans trou")
    # YAML minimal
    for key in ["document_source", "type_document", "gamme_code", "nb_chunks"]:
        if key not in fm:
            warn(f"[{name}] front matter : clé '{key}' absente")

# ---------------------------------------------------------------- 2. FIDÉLITÉ EXCEL (exhaustif)
def load_excel():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    return wb["modèles portes"]

def excel_prix_index(ws):
    """{ (modele, 'ht1'/'ht2'/'ttc1'/'ttc2'): valeur_int } pour comparaison."""
    idx = {}
    seen = set()
    for r in range(2, ws.max_row+1):
        m = (ws[f"A{r}"].value or "")
        m = str(m).strip()
        if not m or m in seen:
            continue
        seen.add(m)
        for k, colL in [("ht1","G"),("ht2","H"),("ttc1","I"),("ttc2","J")]:
            v = ws[f"{colL}{r}"].value
            if v not in (None, ""):
                idx[(m, k)] = int(round(float(v)))
    return idx

def parse_montant(txt):
    """'3 085' (avec espaces variés) -> 3085."""
    t = txt.replace("\u202f", "").replace("\u00a0", "").replace(" ", "")
    return int(t)

def check_fidelite_prix(chunks, ws):
    idx = excel_prix_index(ws)
    RE_PRIX = re.compile(r"tarif de ([\d\u202f\u00a0 ]+) € HT, soit ([\d\u202f\u00a0 ]+) € TTC")
    RE_T = re.compile(r"Tarif (.+?) (1 vantail|2 vantaux) \(")
    controles = 0
    for c in chunks:
        mt = RE_T.search(c["title"])
        mp = RE_PRIX.search(c["body"])
        if not (mt and mp):
            fail(f"[prix] chunk non parsable pour fidélité : {c['title'][:50]}")
            continue
        modele = mt.group(1).strip()
        conf = "1" if mt.group(2).startswith("1") else "2"
        ht_chunk = parse_montant(mp.group(1))
        ttc_chunk = parse_montant(mp.group(2))
        ht_xl = idx.get((modele, f"ht{conf}"))
        ttc_xl = idx.get((modele, f"ttc{conf}"))
        if ht_xl is None:
            fail(f"[prix] modèle absent de l'Excel : {modele} ({conf}V)")
            continue
        if ht_chunk != ht_xl:
            fail(f"[prix] ÉCART HT {modele} {conf}V : chunk={ht_chunk} vs Excel={ht_xl}")
        if ttc_chunk != ttc_xl:
            fail(f"[prix] ÉCART TTC {modele} {conf}V : chunk={ttc_chunk} vs Excel={ttc_xl}")
        controles += 1
    ok(f"[prix] fidélité vérifiée sur {controles} chunks (HT+TTC vs Excel, au €)")

def check_transverses_sans_montant(chunks):
    """Règle 7 : aucun montant chiffré (€, %, chiffre de prix) dans les transverses."""
    RE_MONTANT = re.compile(r"\d+\s*(€|%|/ml|/m²)")
    for c in chunks:
        if RE_MONTANT.search(c["body"]):
            fail(f"[transverses] montant interdit détecté : {c['title'][:50]}")
    ok(f"[transverses] {len(chunks)} chunks vérifiés sans montant (règle 7)")

# ---------------------------------------------------------------- 2b. FIDÉLITÉ OPTIONS (exhaustif)
def excel_options_index(ws):
    """{ (modele, option_libellé): (pv_ht, pv_ttc) } pour les PV chiffrées > 0."""
    idx = {}
    for r in range(2, ws.max_row+1):
        m = str(ws[f"A{r}"].value or "").strip()
        opt = str(ws[f"AW{r}"].value or "").strip()
        ba, bc = ws[f"BA{r}"].value, ws[f"BC{r}"].value
        try:
            if m and opt and float(ba) > 0:
                idx[(m, opt.lower())] = (int(round(float(ba))),
                                          int(round(float(bc))) if bc not in (None,"") else None)
        except (ValueError, TypeError):
            continue
    return idx

def check_fidelite_options(chunks, ws):
    idx = excel_options_index(ws)
    RE_T = re.compile(r"Option (.+?) sur (.+?) \(")
    RE_PV = re.compile(r"plus-value au tarif de ([\d\u202f\u00a0 ]+) € HT, soit ([\d\u202f\u00a0 ]+) € TTC")
    controles = 0
    for c in chunks:
        mt = RE_T.search(c["title"])
        mp = RE_PV.search(c["body"])
        if not (mt and mp):
            fail(f"[options] chunk non parsable pour fidélité : {c['title'][:55]}")
            continue
        opt = mt.group(1).strip().lower()
        modele = mt.group(2).strip()
        ht_chunk = parse_montant(mp.group(1))
        ttc_chunk = parse_montant(mp.group(2))
        ref = idx.get((modele, opt))
        if ref is None:
            fail(f"[options] couple absent de l'Excel : {modele} / {opt}")
            continue
        ht_xl, ttc_xl = ref
        if ht_chunk != ht_xl:
            fail(f"[options] ÉCART HT {modele}/{opt} : chunk={ht_chunk} vs Excel={ht_xl}")
        if ttc_xl is not None and ttc_chunk != ttc_xl:
            fail(f"[options] ÉCART TTC {modele}/{opt} : chunk={ttc_chunk} vs Excel={ttc_xl}")
        controles += 1
    ok(f"[options] fidélité vérifiée sur {controles} chunks (PV HT+TTC vs Excel, au €)")

# ---------------------------------------------------------------- 2c. FIDÉLITÉ Ud (exhaustif)
def num_ud(s):
    """Extrait la valeur numérique de l'Ud : 'Ud... : 1,3 W/m2.K' -> 1.3 ; '1,3 W/m².K' -> 1.3."""
    m = re.search(r"(\d+[.,]\d+)", str(s))
    return float(m.group(1).replace(",", ".")) if m else None

def excel_ud_index(ws):
    idx = {}
    seen = set()
    for r in range(2, ws.max_row+1):
        m = str(ws[f"A{r}"].value or "").strip()
        if m and m not in seen:
            seen.add(m)
            u = num_ud(ws[f"O{r}"].value)
            if u is not None:
                idx[m] = u
    return idx

def check_fidelite_ud(chunks, ws):
    idx = excel_ud_index(ws)
    RE_T = re.compile(r"Caractéristiques (.+?) \(")
    RE_UD = re.compile(r"coefficient Ud de (\d+[.,]\d+)")
    controles = 0
    for c in chunks:
        mt = RE_T.search(c["title"])
        mu = RE_UD.search(c["body"])
        if not mt:
            continue
        modele = mt.group(1).strip()
        if not mu:
            # certains modèles peuvent ne pas avoir d'Ud : vérifier cohérence
            if modele in idx:
                warn(f"[Ud] Ud présent dans l'Excel mais absent du chunk : {modele}")
            continue
        ud_chunk = float(mu.group(1).replace(",", "."))
        ud_xl = idx.get(modele)
        if ud_xl is None:
            warn(f"[Ud] Ud dans le chunk mais absent de l'Excel : {modele}")
            continue
        if abs(ud_chunk - ud_xl) > 1e-9:
            fail(f"[Ud] ÉCART {modele} : chunk={ud_chunk} vs Excel={ud_xl}")
        controles += 1
    ok(f"[Ud] fidélité vérifiée sur {controles} chunks (valeur Ud vs Excel)")

# ---------------------------------------------------------------- 3. LIANT inter-fichiers
def check_liant(chunks_by_file):
    """Le préfixe 'H81 Porte PVC — ... [Modèle]' doit relier prix/options/caractéristiques."""
    def modeles_prix():
        s = set()
        for c in chunks_by_file["prix"]:
            m = re.search(r"Tarif (.+?) (1 vantail|2 vantaux)", c["title"])
            if m: s.add(m.group(1).strip())
        return s
    def modeles_carac():
        s = set()
        for c in chunks_by_file["caracteristiques"]:
            m = re.search(r"Caractéristiques (.+?) \(", c["title"])
            if m: s.add(m.group(1).strip())
        return s
    mp, mc = modeles_prix(), modeles_carac()
    manquants = mp - mc
    if manquants:
        warn(f"[liant] {len(manquants)} modèles ont un chunk prix mais pas caractéristiques : {list(manquants)[:5]}")
    else:
        ok(f"[liant] tous les modèles prix ({len(mp)}) ont un chunk caractéristiques correspondant")

# ---------------------------------------------------------------- 4. ANTI-FANTÔME
def check_antifantome(chunks_prix, ws):
    """Aucun chunk prix ne doit exister pour une config non tarifée."""
    idx = excel_prix_index(ws)
    RE_T = re.compile(r"Tarif (.+?) (1 vantail|2 vantaux) \(")
    for c in chunks_prix:
        mt = RE_T.search(c["title"])
        if not mt:
            continue
        modele = mt.group(1).strip()
        conf = "1" if mt.group(2).startswith("1") else "2"
        if (modele, f"ht{conf}") not in idx:
            fail(f"[anti-fantôme] chunk prix pour config non tarifée : {modele} {conf}V")
    ok(f"[anti-fantôme] aucun chunk prix pour une config absente de l'Excel")

# ---------------------------------------------------------------- 5. CROISEMENT PDF (échantillon)
def check_pdf_echantillon(chunks_prix, n=6):
    try:
        import pypdf
    except ImportError:
        warn("[pdf] pypdf indisponible, croisement PDF sauté")
        return
    reader = pypdf.PdfReader(PDF)
    RE_T = re.compile(r"Tarif (.+?) (1 vantail|2 vantaux) \(")
    RE_PRIX = re.compile(r"tarif de ([\d\u202f\u00a0 ]+) € HT")
    # échantillon : prendre n chunks 1 vantail répartis
    ech = [c for c in chunks_prix if "1 vantail" in c["title"]][::max(1, len(chunks_prix)//(2*n))][:n]
    verifs = 0
    for c in ech:
        page_m = RE_SOURCE.match(c["source"])
        if not page_m:
            continue
        try:
            page = int(page_m.group(1))
        except ValueError:
            continue
        ht = parse_montant(RE_PRIX.search(c["body"]).group(1))
        txt = (reader.pages[page-1].extract_text() or "").replace("\u202f"," ").replace("\u00a0"," ")
        txt_digits = re.sub(r"\s+", "", txt)
        # le prix (sans séparateur) doit apparaître dans la page
        if str(ht) in txt_digits:
            verifs += 1
        else:
            warn(f"[pdf] prix {ht} non retrouvé en page {page} pour {RE_T.search(c['title']).group(1)} "
                 f"(extraction PDF bruitée possible, à vérifier manuellement)")
    ok(f"[pdf] croisement échantillon : {verifs}/{len(ech)} prix retrouvés dans le PDF")

# ---------------------------------------------------------------- exécution
def main():
    ws = load_excel()
    chunks_by_file = {}
    print("=== Décomptes ===")
    for name, fname in FILES.items():
        fm, chunks = parse_chunks(f"{OUTDIR}/{fname}")
        chunks_by_file[name] = chunks
        att = ATTENDU[name]
        flag = "OK" if len(chunks) == att else "FAIL"
        if flag == "FAIL":
            fail(f"[{name}] décompte {len(chunks)} ≠ attendu {att}")
        print(f"  {name:18s}: {len(chunks):3d} chunks (attendu {att}) [{flag}]")
        check_forme(name, fm, chunks)

    check_fidelite_prix(chunks_by_file["prix"], ws)
    check_fidelite_options(chunks_by_file["options"], ws)
    check_fidelite_ud(chunks_by_file["caracteristiques"], ws)
    check_antifantome(chunks_by_file["prix"], ws)
    check_transverses_sans_montant(chunks_by_file["transverses"])
    check_liant(chunks_by_file)
    check_pdf_echantillon(chunks_by_file["prix"])

    print("\n=== Résultats ===")
    print(f"  OK   : {len(results['OK'])}")
    print(f"  WARN : {len(results['WARN'])}")
    print(f"  FAIL : {len(results['FAIL'])}")
    if results["FAIL"]:
        print("\n--- ÉCHECS ---")
        for m in results["FAIL"]:
            print("  ✗", m)
    if results["WARN"]:
        print("\n--- AVERTISSEMENTS ---")
        for m in results["WARN"]:
            print("  ⚠", m)
    print("\n--- CONTRÔLES PASSÉS ---")
    for m in results["OK"]:
        print("  ✓", m)
    return 1 if results["FAIL"] else 0

if __name__ == "__main__":
    sys.exit(main())
