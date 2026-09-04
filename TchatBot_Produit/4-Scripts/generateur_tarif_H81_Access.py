#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Générateur de chunks Markdown pour le tarif H81 Access.

Étape 4 de la migration décrite par `note_cadrage_migration_tarif_H81_Access_v1.md`,
dont il applique les règles 1 à 7 héritées de H81 et les adaptations A1 à A7.

Adaptations par rapport à `generateur_tarif_H81.py` :
  - remapping intégral des colonnes (le classeur H81 Access n'a pas la structure
    de H81 : modèle en F, prix en I/J/K/L, plages en AA–AP, page absente) ;
  - gestion des deux blocs juxtaposés dans la feuille unique `Feuil1` :
    « fiche porte » (lignes 2 à 45) et « chapitres tarifaires » (lignes 47 à 257) ;
  - table des pages en dur, la colonne page n'existant pas ;
  - exclusion du chapitre `Fixes` du traitement forfaitaire (anti-fantôme :
    ses 24 lignes portent un 0 structurel dans les colonnes HT et TTC) ;
  - traitement dimensionnel de la grille des fixes et du meneau battant,
    en cotes exactes, la sémantique d'intervalle n'étant pas déclarée par le tarif ;
  - mapping des unités de facturation repris du PDF, absentes de l'Excel ;
  - formulation A2 du couple HT/TTC, sans dérivation suggérée ;
  - équipements en formulation négative (constante négative sur les 44 lignes) ;
  - journal exhaustif : colonnes non mappées, lignes exclues et motif,
    corrections de libellé, postes repris du PDF, unités non établies.

Huit fichiers produits :
  METHODE, PRIX_PORTES, PRIX_FIXES, OPTIONS, PLUS_VALUES_PROPORTIONNELLES,
  CARACTERISTIQUES, FAISABILITES, TRANSVERSES.
"""
import json
import os
import re
import unicodedata

import openpyxl

# --------------------------------------------------------------------------
# Constantes de gamme et de document
# --------------------------------------------------------------------------

_CANDIDATS = ["/mnt/user-data/uploads/H81_Access-modèles_porte.xlsx",
              "/home/claude/h81a/H81_Access-modèles_porte.xlsx"]
XLSX = next(p for p in _CANDIDATS if os.path.exists(p))
OUT = os.environ.get("H81A_OUT", "/mnt/user-data/outputs")

PLAFOND = 200

GAMME = "H81 Access"
GAMME_NOM = "Porte de service PVC"
COLLECTION = "TRYBA PVC"
MATERIAU = "PVC"
PREFIXE = "H81 Access Porte de service PVC"
# Auto-portance : la mention figure aussi dans le corps, le sens de contamination
# probable étant H81 -> H81 Access, `H81` étant un préfixe strict de `H81 Access`.
PRODUIT = "la porte de service PVC H81 Access"

DOC_AFFICHE = "Tarif—H81—Access—HT—08-04-2026.pdf"
DOC_AFFICHE_TTC = "Tarif—H81—Access—TTC—08-04-2026.pdf"
DOC_YAML = "Tarif_H81_Access_HT_08-04-2026.pdf"
DOC_YAML_TTC = "Tarif_H81_Access_TTC_08-04-2026.pdf"
VERSION_DOC = "2026.04"
DATE_VALIDITE = "2026-04-08"
AUDIENCES = "[ADV, commercial]"

JOURNAL = {
    "colonnes_non_mappees": [],
    "lignes_exclues": [],
    "corrections_libelle": [],
    "postes_repris_du_pdf": [],
    "unites_non_etablies": [],
    "arbitrages_signales": [],
}

# --------------------------------------------------------------------------
# Table des pages, établie en dur contre le PDF (la colonne page est absente
# de l'Excel ; le sommaire est faux à partir de la page 11 et le tableau
# alphabétique de la page 6 l'est pour Persane : seules les pages font foi).
# --------------------------------------------------------------------------

PAGE_MODELE = {
    "Porte vitrée": 21, "Porte panneau plein": 22, "Cypris": 23, "Dahpnis": 24,
    "Madrane": 25, "Persane": 26, "Santéria": 27, "T1L": 28, "T2L": 29,
    "Melbourne": 30, "Vienne": 31,
}

PAGE_CHAPITRE = {
    "Pack EVO": 11,
    "PV vitrages": 16,
    "Fixes": 17,
    "Vitrages pour fixes": 18,
    "Croisillons": 19,
    "Remplissage": 20,
    "PV vitrage panneaux": 32,
    "Garnitures - béquilles et poignées": 33,
    "Options et accessoires": 34,
    "Élargisseurs": 35,
    "Profilés complémentaires": 36,
    "Tapées de doublage": 37,
    "Accouplements statiques": 38,
    "Seuils": 39,
    "Fabrications spéciales : cintres": 41,
    "Exemple de calculs": 42,
}

PAGE_LIMITES = 7
PAGE_DESCRIPTIF_LIGNES = 8
PAGE_TYPOLOGIE = 9
PAGE_FERRAGE = 12
PAGE_PAUMELLES = 13
PAGE_COULEURS = 14
PAGE_ACCESSOIRES_COULEUR = 15
PAGE_CINTRES_FAISABILITE = 40
PAGE_COTES_FABRICATION = 43
PAGE_LARGEUR_PASSAGE = 44
PAGE_EVOLUTIONS = 45

# Chapitres écartés du traitement forfaitaire.
CHAPITRES_EXCLUS = {
    # Traité dimensionnellement dans PRIX_FIXES ; ses 24 lignes de grille portent
    # un 0 structurel en HT et TTC qu'un test naïf servirait comme prix.
    "Fixes",
    # Exclusion de périmètre assumée (note §7) : aucun prix qui ne figure ailleurs,
    # deux erreurs arithmétiques, et une forme — l'addition — que la règle 3 interdit.
    "Exemple de calculs",
}

# --------------------------------------------------------------------------
# Corrections de libellé, tranchées en faveur du PDF qui se déclare document de
# référence à sa page 2. Aucune ne touche un montant. Toutes sont journalisées.
# Clé : (ligne Excel, colonne) -> (valeur corrigée, motif, statut)
# statut « note » : arbitrage déjà consigné au §5 de la note de cadrage.
# statut « nouveau » : relevé à l'étape 4, à confirmer par le service Produits.
# --------------------------------------------------------------------------

CORRECTIONS = {
    (49, "BA"): ("GMECA", "référence de gâche mécanique, page 11", "note"),
    (60, "AZ"): ("44/2-20G-Isol'3 4", "code vitrage de sécurité Evo, page 16", "note"),
    (181, "BB"): ("PP28 / PP32", "codes des panneaux standards, page 20", "note"),
    (182, "BB"): ("PP28 / PP32", "codes des panneaux standards, page 20", "note"),
    (186, "BB"): ("PR28 / PR32", "PR33 n'existe dans aucune source, page 20", "nouveau"),
    (189, "BA"): ("Moulures Chêne d'Or (en sus de la PV du panneau support)",
                  "la mention « et Blanc veiné » est absente du PDF, page 20", "nouveau"),
    (230, "BB"): ("IH3-76", "le profilé d'accouplement est unique, page 38", "nouveau"),
    (231, "BB"): ("I7Cx2 / 3312x2 / I7-20x2",
                  "références des profilés d'accouplement, page 38", "nouveau"),
    (225, "BB"): ("4300 à 4304", "la plage des tapées est identique en blanc et en "
                                 "plaxage, page 37", "nouveau"),
    (227, "BB"): ("4305 à 4308", "la plage des tapées est identique en blanc et en "
                                 "plaxage, page 37", "nouveau"),
    (242, "BA"): ("arc de cercle 1 vantail + 1 fixe", "coquille de saisie, page 41", "nouveau"),
    (254, "BA"): ("PV ferrage Evo", "coquille de saisie, page 42", "nouveau"),
    (190, "BA"): ("Panneau bicolore : moulures blanches face int. et Chêne d'Or "
                  "face ext.",
                  "la mention « ou Blanc veiné » est absente du PDF, page 20", "nouveau"),
    (229, "BB"): ("IH3-76", "casse de la référence du profilé, page 38", "nouveau"),
    (232, "BB"): ("EVLE84 / EVP1EVP2", "références d'accouplement d'angle, page 38",
                  "nouveau"),
    (206, "BA"): ("Ferme-porte TS 3000", "libellé du PDF, page 34", "nouveau"),
    (207, "BA"): ("Ferme-porte TS 5000", "libellé du PDF, page 34", "nouveau"),
    (208, "BA"): ("Stop-porte, références SP-W et SP-BR", "libellé du PDF, page 34",
                  "nouveau"),
    (241, "BA"): ("arc de cercle 1 vantail", "abréviation développée, page 41", "nouveau"),
    (243, "BA"): ("arc de cercle 2 vantaux", "abréviation développée, page 41", "nouveau"),
    (244, "BA"): ("arc de cercle 2 vantaux + 1 fixe", "abréviation développée, page 41",
                  "nouveau"),
    (245, "BA"): ("arc de cercle 2 vantaux + 2 fixes", "abréviation développée, page 41",
                  "nouveau"),
    (183, "BB"): ("PP28/38dB / PP32/38dB", "séparateur de liste de codes, page 20",
                  "nouveau"),
    (184, "BB"): ("PP28/38dB / PP32/38dB", "séparateur de liste de codes, page 20",
                  "nouveau"),
    (185, "BB"): ("PR28 / PR32", "séparateur de liste de codes, page 20", "nouveau"),
}

# Corrections de libellé applicables par égalité de chaîne, tous chapitres.
CORRECTIONS_GLOBALES = [
    ("Elargisseurs", "Élargisseurs", "orthographe, page 35"),
    ("Côtes de référence tarif", "Cotes de référence tarif", "orthographe, page 37"),
    ("Longeur meneau battant mm", "Longueur meneau battant (mm)", "orthographe, page 17"),
    ("Elégance", "Élégance", "orthographe de la famille, page 4"),
    ("Decor 1F", "Décor 1 face", "abréviation développée, page 20"),
    ("Décor 1F", "Décor 1 face", "abréviation développée, page 20"),
    ("PA sans volet roulant", "Profilé d'accouplement droit sans volet roulant",
     "abréviation développée, page 38"),
    ("profilés de finition et d'accouplement",
     "Profilé de finition ou d'accouplement", "libellé du PDF, page 36"),
    ("Couvre-joints", "Couvre-joint intérieur ou extérieur", "libellé du PDF, page 36"),
    ("Compensateur de feuillure", "Compensateur de feuillure pour dormant rénovation",
     "libellé du PDF, page 36"),
    ("Elargissers sans armatures", "Élargisseur sans armature", "orthographe, page 35"),
    ("Elargissers avec armatures", "Élargisseur avec armature", "orthographe, page 35"),
]

# Noms de chapitre tels qu'affichés dans les chunks.
CHAPITRE_AFFICHE = {
    "Pack EVO": "Pack Evo",
    "PV vitrages": "plus-values vitrages",
    "Vitrages pour fixes": "vitrages pour fixes",
    "Croisillons": "croisillons",
    "Remplissage": "remplissage, panneaux de soubassement",
    "PV vitrage panneaux": "plus-values vitrages pour panneaux",
    "Garnitures - béquilles et poignées": "garnitures, béquilles et poignées de tirage",
    "Options et accessoires": "options et accessoires",
    "Élargisseurs": "élargisseurs",
    "Profilés complémentaires": "profilés complémentaires",
    "Tapées de doublage": "tapées de doublage",
    "Accouplements statiques": "accouplements statiques",
    "Seuils": "seuils",
    "Fabrications spéciales : cintres": "fabrications spéciales, cintres",
}

# Famille de panneau de soubassement, absente de l'Excel, reprise des en-têtes de
# section du PDF (page 20) et rattachée par le montant, jamais par l'ordre des
# lignes : c'est la règle C4 de CA76, importée.
FAMILLE_PANNEAU = {
    ("39", "34"): "panneau standard",
    ("180", "156"): "panneau standard",
    ("205", "178"): "panneau phonique 38 dB",
    ("436", "377"): "panneau phonique 38 dB",
    ("167", "144"): "panneau lisse renforcé",
    ("391", "142"): "panneau lisse renforcé",
    ("224", "194"): "panneau rainuré renforcé",
}

# Configuration de garniture, reprise de la page 33 et rattachée par le montant :
# l'Excel ne nomme que la garniture extérieure, l'intérieure étant toujours BDEL.
GARNITURE_EXT = {"BDEL", "BPEL", "BDSL", "BPSL"}

# --------------------------------------------------------------------------
# Unités de facturation. Elles n'existent pas dans l'Excel : les colonnes
# s'appellent `HT` et `TTC` sans qualifieur. Elles sont relevées dans le PDF.
# Règle A3 (importée de C4) : tout poste chiffré déclare son unité ; un poste
# dont l'unité ne peut être établie n'est pas généré et part au journal.
# --------------------------------------------------------------------------

# (chapitre, tableau) -> (unité courte, glose de facturation)
UNITES = {
    ("Pack EVO", "PV"): ("forfait", "Cette plus-value est forfaitaire, par châssis."),
    ("Pack EVO", "Option spécifique"): (
        "forfait", "Ce tarif est forfaitaire, à la pièce."),
    ("Pack EVO", "Bequilles et poignées"): (
        "forfait", "Ce tarif est forfaitaire, par garniture."),
    ("PV vitrages", None): (
        "au mètre carré",
        "Cette plus-value s'applique au mètre carré de surface vitrée du châssis ; "
        "le calcul du total revient à l'ADV."),
    ("Vitrages pour fixes", None): (
        "au mètre carré",
        "Cette plus-value s'applique au mètre carré de surface vitrée du fixe ; "
        "le calcul du total revient à l'ADV."),
    ("Croisillons", "Méthode de calcul"): (
        "au champ",
        "Ce tarif s'entend par champ, un champ étant une surface de vitrage "
        "délimitée par les croisillons, le dormant ou l'ouvrant ; le comptage "
        "des champs et le calcul du total reviennent à l'ADV."),
    ("Remplissage", "Panneaux de soubassement"): (
        "au mètre carré",
        "Cette plus-value s'applique au mètre carré de panneau ; le calcul du "
        "total revient à l'ADV."),
    ("Remplissage", "Moulures 2 faces"): (
        "à la face",
        "Cette plus-value s'applique par face, en sus de la plus-value du "
        "panneau support."),
    ("PV vitrage panneaux", "Option pour vitrage pa PVC"): (
        "forfait",
        "Cette plus-value est forfaitaire et tient compte du vitrage de "
        "sécurité 44/2 en Evo."),
    ("Garnitures - béquilles et poignées", "Monocolore"): (
        "forfait", "Ce tarif est forfaitaire, par garniture."),
    ("Options et accessoires", "opt de vantail et de sécurité"): (
        "forfait", "Ce tarif est forfaitaire, à la pièce."),
    ("Élargisseurs", None): (
        "au mètre linéaire",
        "Ce tarif s'entend au mètre linéaire posé sur châssis ; le calcul du "
        "total revient à l'ADV."),
    ("Profilés complémentaires", "Profilé de finition ou d'accouplement"): (
        "au mètre linéaire",
        "Ce tarif s'entend au mètre linéaire posé sur châssis ; le calcul du "
        "total revient à l'ADV."),
    ("Profilés complémentaires",
     "Compensateur de feuillure pour dormant rénovation"): (
        "au châssis", "Ce tarif s'entend par châssis, posé sur châssis."),
    ("Profilés complémentaires", "Couvre-joint intérieur ou extérieur"): (
        "au mètre linéaire",
        "Ce tarif s'entend au mètre linéaire posé sur châssis ; le calcul du "
        "total revient à l'ADV."),
    ("Tapées de doublage", "Cotes de référence tarif"): (
        "au mètre linéaire",
        "Ce tarif s'entend au mètre linéaire posé sur châssis, la longueur "
        "facturée étant la largeur plus deux fois la hauteur de fabrication du "
        "châssis ; le calcul revient à l'ADV."),
    ("Accouplements statiques", None): (
        "au mètre linéaire",
        "Ce tarif s'entend au mètre linéaire posé sur châssis ; le calcul du "
        "total revient à l'ADV."),
    ("Fabrications spéciales : cintres", "Principe de tarification"): (
        "au châssis",
        "Cette plus-value s'ajoute au prix du châssis rectangulaire englobant le "
        "châssis de forme spéciale, pour une unité, vitrage non compris."),
    ("Fabrications spéciales : cintres", "Meneau / Traverse"): (
        "à la fixation", "Ce tarif s'entend par fixation."),
}

# Surcharges par désignation, là où l'unité varie à l'intérieur d'un tableau
# (chapitre Seuils : deux postes à la pièce, trois au mètre linéaire — page 39).
UNITES_PAR_DESIGNATION = {
    ("Seuils", "AS10100_RA1"): ("à la pièce", "Ce tarif s'entend à la pièce."),
    ("Seuils", "AS10100_RA2"): ("à la pièce", "Ce tarif s'entend à la pièce."),
    ("Seuils", "APE-70"): ("au mètre linéaire",
                           "Ce tarif s'entend au mètre linéaire ; le calcul du "
                           "total revient à l'ADV."),
    ("Seuils", "PE725"): ("au mètre linéaire",
                          "Ce tarif s'entend au mètre linéaire ; le calcul du "
                          "total revient à l'ADV."),
    ("Seuils", "KP484RCY"): ("au mètre linéaire",
                             "Ce tarif s'entend au mètre linéaire ; le calcul du "
                             "total revient à l'ADV."),
}

# --------------------------------------------------------------------------
# Utilitaires
# --------------------------------------------------------------------------


def mots(txt):
    return len(re.findall(r"\S+", txt))


_APOS = 0


def txt(v):
    """Nettoyage d'une cellule. L'apostrophe typographique U+2019, qui coexiste
    avec l'apostrophe droite dans les deux sources sans distinction de sens, est
    normalisée sur l'apostrophe droite : sans quoi deux graphies du même code
    vitrage cohabiteraient dans le corpus et casseraient la recherche."""
    global _APOS
    if v is None:
        return ""
    s = re.sub(r"\s+", " ", str(v)).strip()
    if "\u2019" in s:
        _APOS += 1
        s = s.replace("\u2019", "'")
    return s


def eur(v):
    """Montant en euros, entier, sans séparateur de milliers, comme le tarif."""
    if v in (None, ""):
        return None
    try:
        return str(int(round(float(v))))
    except (TypeError, ValueError):
        return None


def enumere(items):
    items = [str(i) for i in items]
    if len(items) == 1:
        return items[0]
    return ", ".join(items[:-1]) + " et " + items[-1]


def bas(s):
    """Passe un libellé en minuscule pour lecture en phrase, en préservant les
    acronymes tout-majuscule et les codes alphanumériques."""
    def fix(tok):
        core = tok.strip(".,;:()«»\"'")
        if not core:
            return tok
        if core.isupper() or any(ch.isdigit() for ch in core):
            return tok
        return tok.lower()
    return " ".join(fix(t) for t in s.split())


class Fichier:
    """Accumule des chunks, numérote les SC depuis SC0002 (SC0001 est réservé par
    le moteur Wikit), contrôle le plafond de 200 mots marqueur compris."""

    def __init__(self, nom, sous_type):
        self.nom = nom
        self.sous_type = sous_type
        self.chunks = []
        self.sc = 2

    def ajoute(self, titre, page, corps):
        corps = re.sub(r"\s+", " ", corps).strip()
        src = "*Source : %s, page %d — information originale — SC%04d*" % (
            DOC_AFFICHE, page, self.sc)
        bloc = "## %s\n%s\n\n%s\n" % (titre, src, corps)
        n = mots("## " + titre + " " + src + " " + corps)
        if n > PLAFOND:
            raise ValueError("Plafond dépassé (%d mots) : %s" % (n, titre))
        self.chunks.append({"sc": self.sc, "titre": titre, "page": page,
                            "corps": corps, "bloc": bloc, "mots": n})
        self.sc += 1

    def ecrit(self):
        fm = [
            "---",
            "document_source: %s" % DOC_YAML,
            # Le TTC provient de l'édition TTC, de pagination identique. Champ
            # ajouté au format figé au §8 de la note : arbitrage A8 signalé.
            "document_source_ttc: %s" % DOC_YAML_TTC,
            "type_document: tarif",
            "sous_type: %s" % self.sous_type,
            "gamme_code: %s" % GAMME,
            'gamme_nom: "%s"' % GAMME_NOM,
            'collection: "%s"' % COLLECTION,
            "materiau: %s" % MATERIAU,
            'version_doc: "%s"' % VERSION_DOC,
            "date_validite: %s" % DATE_VALIDITE,
            "nb_chunks: %d" % len(self.chunks),
            "audiences: %s" % AUDIENCES,
            "---",
            "",
        ]
        os.makedirs(OUT, exist_ok=True)
        path = os.path.join(OUT, self.nom)
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(fm) + "\n" + "\n".join(c["bloc"] for c in self.chunks))
        return path


def empaquete(elements, entete_mots, plafond=PLAFOND, sep=1):
    """Découpe gloutonne d'une liste d'éléments textuels sous le plafond de mots.
    La coupure est pilotée par le comptage des mots, jamais par une constante."""
    tranches, courante, poids = [], [], entete_mots
    for i, t in enumerate(elements):
        m = mots(t) + (sep if courante else 0)
        if courante and poids + m > plafond:
            tranches.append(courante)
            courante, poids = [], entete_mots
        courante.append(i)
        poids += m
    if courante:
        tranches.append(courante)
    return tranches


# --------------------------------------------------------------------------
# Lecture du classeur : deux blocs juxtaposés dans la feuille unique
# --------------------------------------------------------------------------

COL = {c: i for i, c in enumerate(
    ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O",
     "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"], start=1)}
for i in range(27, 106):
    COL[openpyxl.utils.get_column_letter(i)] = i

LIGNE_PORTES = (2, 45)
LIGNE_CHAPITRES = (47, 257)


def cel(ws, r, lettre):
    return ws.cell(r, COL[lettre]).value


_GLOBALES_VUES = {}


def corrige_global(valeur, r, lettre):
    """Les corrections applicables par égalité de chaîne portent sur des libellés
    répétés ; le journal les agrège au lieu de les répéter ligne à ligne."""
    for avant, apres, motif in CORRECTIONS_GLOBALES:
        if valeur == avant:
            cle = (avant, apres)
            if cle not in _GLOBALES_VUES:
                entree = {"portee": "toutes occurrences", "avant": avant,
                          "apres": apres, "motif": motif, "statut": "nouveau",
                          "occurrences": 0, "colonnes": []}
                _GLOBALES_VUES[cle] = entree
                JOURNAL["corrections_libelle"].append(entree)
            entree = _GLOBALES_VUES[cle]
            entree["occurrences"] += 1
            if lettre not in entree["colonnes"]:
                entree["colonnes"].append(lettre)
            return apres
    return valeur


def corrige(valeur, r, lettre):
    if (r, lettre) in CORRECTIONS:
        apres, motif, statut = CORRECTIONS[(r, lettre)]
        JOURNAL["corrections_libelle"].append(
            {"ligne": r, "colonne": lettre, "avant": valeur, "apres": apres,
             "motif": motif, "statut": statut})
        return apres
    return corrige_global(valeur, r, lettre)


def norm_ud(v):
    """'Ud porte pleine :  1,3 W/m2.K' -> ('porte pleine', '1,3 W/m².K').
    L'anomalie de saisie du double espace sur T2L est normalisée ici."""
    s = txt(v).replace("m2", "m²")
    if not s:
        return "", ""
    libelle, valeur = "", s
    if ":" in s:
        gauche, valeur = s.split(":", 1)
        libelle = gauche.replace("Ud", "").strip()
    valeur = re.sub(r"(\d)\s*W", r"\1 W", valeur).strip()
    return libelle, valeur


def lit_portes(ws):
    """Bloc « fiche porte », colonnes A à AX, lignes 2 à 45 : onze modèles à
    raison de quatre lignes par modèle, une par équipement."""
    modeles = []
    index = {}
    for r in range(LIGNE_PORTES[0], LIGNE_PORTES[1] + 1):
        nom = txt(cel(ws, r, "F"))
        if not nom:
            continue
        if nom not in index:
            ligne = txt(cel(ws, r, "D")).replace("Ligne ", "")
            famille = corrige_global(txt(cel(ws, r, "E")), r, "E")
            ud_lib, ud_val = norm_ud(cel(ws, r, "O"))
            d = {
                "nom": nom, "ligne": ligne, "famille": famille,
                "page": PAGE_MODELE.get(nom),
                "lignes": [],
                "ht1": cel(ws, r, "I"), "ht2": cel(ws, r, "J"),
                "ttc1": cel(ws, r, "K"), "ttc2": cel(ws, r, "L"),
                "ud_libelle": ud_lib, "ud": ud_val,
                "dim_mini_panneau": txt(cel(ws, r, "P")),
                "modele_base": txt(cel(ws, r, "R")),
                "evo_ht": cel(ws, r, "V"), "evo_ttc": cel(ws, r, "W"),
                "teintes": [], "equipements": [],
                "dims": {},
                "opt_titre": txt(cel(ws, r, "AQ")),
                "opt_libelle": txt(cel(ws, r, "AR")),
                "opt_desc": txt(cel(ws, r, "AT")),
                "opt_ht": cel(ws, r, "AU"), "opt_ttc": cel(ws, r, "AW"),
                "opt_renvoi": txt(cel(ws, r, "AX")),
            }
            cols = ["AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH",
                    "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP"]
            for k, prof in enumerate(["5103", "5107", "5114", "5120"]):
                bloc = [cel(ws, r, c) for c in cols[4 * k:4 * k + 4]]
                d["dims"][prof] = bloc      # Mini L, Mini H, Maxi L, Maxi H
            index[nom] = d
            modeles.append(d)
        d = index[nom]
        d["lignes"].append(r)
        eq = txt(cel(ws, r, "T"))
        fa = txt(cel(ws, r, "S")).lower()
        if eq:
            d["equipements"].append((eq, fa))
        te = txt(cel(ws, r, "X"))
        if te and te not in d["teintes"]:
            d["teintes"].append(te)
    return modeles


def lit_chapitres(ws):
    """Bloc « chapitres tarifaires », colonnes AY à DA, lignes 47 à 257."""
    lignes = []
    for r in range(LIGNE_CHAPITRES[0], LIGNE_CHAPITRES[1] + 1):
        chap = corrige(txt(cel(ws, r, "AY")), r, "AY")
        if not chap:
            continue
        lignes.append({
            "r": r,
            "chapitre": chap,
            "tableau": corrige(txt(cel(ws, r, "AZ")), r, "AZ"),
            "designation": corrige(txt(cel(ws, r, "BA")), r, "BA"),
            "details": corrige(txt(cel(ws, r, "BB")), r, "BB"),
            "ht": cel(ws, r, "BC"),
            "ttc": cel(ws, r, "BD"),
            "hauteur": cel(ws, r, "BE"),
            "grille_ht": [cel(ws, r, openpyxl.utils.get_column_letter(c))
                          for c in range(58, 82)],
            "grille_ttc": [cel(ws, r, openpyxl.utils.get_column_letter(c))
                           for c in range(82, 106)],
        })
    return lignes


MAPPEES = set("D E F I J K L O P R S T V W X AA AB AC AD AE AF AG AH AI AJ AK AL "
              "AM AN AO AP AQ AR AT AU AW AX AY AZ BA BB BC BD BE".split())
MAPPEES |= {openpyxl.utils.get_column_letter(c) for c in range(58, 106)}


def journal_colonnes(ws):
    for c in range(1, ws.max_column + 1):
        lettre = openpyxl.utils.get_column_letter(c)
        entete = txt(ws.cell(1, c).value)
        remplies = sum(1 for r in range(2, ws.max_row + 1)
                       if ws.cell(r, c).value not in (None, ""))
        if remplies and lettre not in MAPPEES:
            JOURNAL["colonnes_non_mappees"].append(
                {"colonne": lettre, "entete": entete, "lignes_remplies": remplies})


# --------------------------------------------------------------------------
# Grille des fixes et meneau battant : extraction dimensionnelle
# --------------------------------------------------------------------------

LARGEURS = list(range(300, 2700, 100))


def extrait_grille(lignes):
    """Renvoie [(hauteur, [(largeur, ht, ttc), ...]), ...] pour les 24 hauteurs.
    Les colonnes HT/TTC de ces lignes portent un 0 structurel : elles ne sont
    jamais lues comme un prix (anti-fantôme)."""
    grille = []
    for L in lignes:
        if L["chapitre"] != "Fixes" or txt(L["tableau"]).lower().startswith("meneau"):
            continue
        h = L["hauteur"]
        if h in (None, ""):
            continue
        cellules = []
        for j, larg in enumerate(LARGEURS):
            ht, ttc = eur(L["grille_ht"][j]), eur(L["grille_ttc"][j])
            if ht is None or ttc is None:
                JOURNAL["lignes_exclues"].append(
                    {"ligne": L["r"], "objet": "fixe H%s x L%d" % (h, larg),
                     "motif": "cellule non renseignée en HT ou en TTC"})
                continue
            cellules.append((larg, ht, ttc))
        grille.append((int(h), cellules))
    return grille


def extrait_meneau(lignes):
    paliers = []
    for L in lignes:
        if L["chapitre"] != "Fixes" or not txt(L["tableau"]).lower().startswith("meneau"):
            continue
        ht, ttc = eur(L["ht"]), eur(L["ttc"])
        if L["hauteur"] in (None, "") or ht is None or ttc is None:
            JOURNAL["lignes_exclues"].append(
                {"ligne": L["r"], "objet": "meneau battant",
                 "motif": "longueur ou montant non renseigné"})
            continue
        paliers.append((int(L["hauteur"]), ht, ttc))
    return paliers


# --------------------------------------------------------------------------
# F1 — Méthode
# --------------------------------------------------------------------------

def construit_methode():
    f = Fichier("Tarif_H81_Access_METHODE.md", "methode")

    f.ajoute(
        PREFIXE + " — Lecture de la table de tarif des fixes",
        PAGE_CHAPITRE["Fixes"],
        "Sur %s, la table de tarif des fixes latéraux vitrés et des impostes "
        "vitrées se lit en lecture directe, en fonction de la dimension largeur "
        "par hauteur du fixe, pour un châssis blanc teinté masse. Le prix lu "
        "inclut un double vitrage Isol'3 4-20G-4 de coefficient Ug égal à "
        "1,1 W/m².K. Le minimum de réalisation des fixes est de 280 mm, selon le "
        "dormant. Les cotes tarifées vont de 300 à 2 600 mm au pas de 100 mm sur "
        "les deux axes. Le tarif n'énonce pas ce que couvre une colonne : il ne "
        "donne aucune règle de lecture pour une cote intermédiaire. Les chunks de "
        "prix énoncent donc les cotes exactes du tarif, et aucune interpolation "
        "ne doit être faite entre deux cotes." % PRODUIT)

    f.ajoute(
        PREFIXE + " — Comptage des croisillons au champ",
        PAGE_CHAPITRE["Croisillons"],
        "Sur %s, qu'il s'agisse de croisillons incorporés ou rapportés, le prix "
        "des croisillons se calcule au nombre de champs, et non au mètre "
        "linéaire. Un champ est une surface de vitrage délimitée par les "
        "croisillons, par le dormant ou par l'ouvrant. Deux types de champs sont "
        "distingués : les champs à croisillons avec jonction en T ou en croix, et "
        "les champs à croisillons sans jonction, dits filants, dont les tarifs "
        "diffèrent. Le comptage des champs et la multiplication par le prix "
        "unitaire reviennent à l'ADV." % PRODUIT)

    f.ajoute(
        PREFIXE + " — Cote de référence du prix des tapées de doublage",
        PAGE_CHAPITRE["Tapées de doublage"],
        "Sur %s, les dimensions de fabrication du châssis font office de cote de "
        "référence pour le calcul du prix des tapées de doublage. La longueur "
        "facturée est la largeur, plus deux fois la hauteur. Les tapées posées "
        "sur les traverses horizontales sont filantes ; celles posées sur les "
        "montants viennent buter contre les tapées horizontales. La face "
        "extérieure des tapées n'est plaxée qu'à mi-hauteur. Le calcul de la "
        "longueur et du total revient à l'ADV." % PRODUIT)

    f.ajoute(
        PREFIXE + " — Calcul de la longueur du meneau battant selon le dormant",
        PAGE_CHAPITRE["Fixes"],
        "Sur %s, la longueur du meneau battant se déduit de la largeur hors tout "
        "du châssis, diminuée d'une valeur propre au profil de dormant : 253 mm "
        "pour le dormant 5103 L69, 283 mm pour le 5107 L84, 319 mm pour le 5114 "
        "LZ102 et 335 mm pour le 5120 L110. La longueur obtenue détermine ensuite "
        "le palier de tarif à lire. Le calcul revient à l'ADV ; les valeurs de "
        "tarif ne sont jamais interpolées entre deux paliers." % PRODUIT)

    f.ajoute(
        PREFIXE + " — Formules de cintrage en arc de cercle",
        PAGE_CINTRES_FAISABILITE,
        "Sur %s, le tarif inscrit deux formules pour les châssis en arc de "
        "cercle. Le rayon R se détermine en fonction de la largeur L et de la "
        "flèche f par la formule R égale la somme du carré de L divisé par quatre "
        "f et de f, le tout multiplié par 0,50. La flèche f se détermine en "
        "fonction du rayon R et de la largeur L par la formule f égale R moins la "
        "racine carrée de la différence entre le carré de R et le quart du carré "
        "de L. Une fermeture correcte en partie haute impose que la flèche f "
        "reste inférieure à 0,4 fois la largeur L. L'exécution de ces formules "
        "revient à l'ADV." % PRODUIT)

    f.ajoute(
        PREFIXE + " — Principe de tarification des fabrications spéciales cintrées",
        PAGE_CHAPITRE["Fabrications spéciales : cintres"],
        "Sur %s, le prix d'un châssis de forme spéciale se construit à partir du "
        "prix du châssis rectangulaire englobant ce châssis, auquel s'ajoutent "
        "les plus-values de cintrage, applicables pour une unité. Les prix des "
        "compléments et accessoires s'ajoutent au prix du châssis nu, et les "
        "dimensions des compléments sont incluses dans les cotes de fabrication L "
        "et H. La faisabilité du châssis doit être vérifiée avant tout chiffrage, "
        "de même que la faisabilité de la découpe du panneau, esthétique et "
        "dimensionnelle, d'après le relevé de cotes." % PRODUIT)

    f.ajoute(
        PREFIXE + " — Lecture des montants HT et TTC du tarif",
        2,
        "Sur %s, chaque poste du tarif porte deux montants distincts, un montant "
        "hors taxes et un montant toutes taxes comprises, tous deux publiés par le "
        "tarif. Le montant TTC ne se déduit pas du montant HT : le rapport entre "
        "les deux varie d'un poste à l'autre. Les deux montants doivent donc être "
        "lus tels qu'ils figurent au tarif, jamais recalculés, jamais dérivés l'un "
        "de l'autre. Tous les prix s'entendent hors éco-participation. Le logiciel "
        "Syscon évoluant en permanence, la version PDF du tarif reste le seul "
        "document de référence." % PRODUIT)

    f.ajoute(
        PREFIXE + " — Vocabulaire des paumelles",
        PAGE_PAUMELLES,
        "Sur %s, l'organe de rotation du vantail est la paumelle, de référence "
        "PPE-4 : les termes gond et charnière ne sont pas employés par le tarif "
        "et n'en sont pas des synonymes. Les paumelles sont en aluminium, d'une "
        "largeur de 99 mm et d'une longueur de 112 mm, fixées directement dans le "
        "dormant et dans l'ouvrant, et disponibles en blanc, titane et noir. Le "
        "tarif parle d'engondage et de dégondage du vantail ; il n'emploie pas le "
        "terme anti-dégondage, qui relève d'une autre catégorie." % PRODUIT)

    f.ajoute(
        PREFIXE + " — Vocabulaire du ferrage et du verrouillage",
        PAGE_FERRAGE,
        "Sur %s, l'organe de verrouillage est une serrure à galets à relevage, en "
        "ferrage cinq points manuel de série ou en ferrage six points manuel avec "
        "le Pack Evo. Le terme crémone ne s'applique pas à cette gamme et n'est "
        "pas employé par le tarif. Le ferrage six points comporte deux crochets "
        "massifs, un pêne dormant, un pêne demi-tour et deux pênes manuels. Le "
        "ferrage automatique n'est pas disponible." % PRODUIT)

    f.ajoute(
        PREFIXE + " — Portée des prix servis et interdiction de calculer",
        2,
        "Sur %s, les prix servis sont ceux du tarif, transcrits littéralement. "
        "Aucun total, aucune somme, aucune multiplication par une surface, une "
        "longueur, un nombre de champs ou un pourcentage ne doit être effectuée : "
        "ces opérations reviennent à l'ADV. Aucun prix n'est interpolé entre deux "
        "cotes ni entre deux paliers. Toute construction qui ne peut être traitée "
        "à l'aide du tarif n'est pas réalisable." % PRODUIT)

    return f


# --------------------------------------------------------------------------
# F2 — Prix des portes
# --------------------------------------------------------------------------

CONFIGS = [("ht1", "ttc1", "1 vantail", "en un vantail"),
           ("ht2", "ttc2", "2 vantaux", "en deux vantaux égaux")]


def construit_prix_portes(modeles):
    f = Fichier("Tarif_H81_Access_PRIX_PORTES.md", "prix_portes")
    for d in modeles:
        for kht, kttc, etiquette, phrase in CONFIGS:
            ht, ttc = eur(d[kht]), eur(d[kttc])
            if ht is None or ttc is None:
                # Anti-fantôme : une configuration non tarifée ne produit pas de chunk.
                JOURNAL["lignes_exclues"].append(
                    {"modele": d["nom"], "objet": "prix %s" % etiquette,
                     "motif": "montant HT ou TTC non renseigné"})
                continue
            titre = "%s — Tarif %s %s (ligne %s, famille %s)" % (
                PREFIXE, d["nom"], etiquette, d["ligne"], d["famille"])
            corps = (
                "Le modèle %s de la porte de service PVC H81 Access, ligne %s, "
                "famille %s, est proposé %s au tarif de %s € HT ; le tarif TTC "
                "correspondant est de %s €. Ce prix est celui du modèle de base, "
                "sans option, et ne varie pas avec la dimension du châssis. Il "
                "s'entend hors éco-participation." % (
                    d["nom"], d["ligne"], d["famille"], phrase, ht, ttc))
            f.ajoute(titre, d["page"], corps)
    return f


# --------------------------------------------------------------------------
# F3 — Prix des fixes (dimensionnel) et du meneau battant
# --------------------------------------------------------------------------

def construit_prix_fixes(lignes):
    f = Fichier("Tarif_H81_Access_PRIX_FIXES.md", "prix_fixes")
    page = PAGE_CHAPITRE["Fixes"]

    for hauteur, cellules in extrait_grille(lignes):
        elements = ["%d mm, %s € HT, %s € TTC" % (l, ht, ttc)
                    for l, ht, ttc in cellules]
        entete = (
            "Sur la porte de service PVC H81 Access, le fixe latéral vitré et "
            "l'imposte vitrée d'une hauteur de %d mm sont tarifés en châssis "
            "blanc teinté masse, double vitrage Isol'3 4-20G-4 inclus, selon "
            "la largeur exacte : " % hauteur)
        gabarit_titre = "%s — Tarif fixe latéral vitré et imposte vitrée, hauteur %d mm, largeurs de 0000 à 0000 mm" % (PREFIXE, hauteur)
        cloture = ". Ces prix se lisent directement, sans interpolation."
        cout = mots(gabarit_titre) + 14 + mots(entete) + mots(cloture)
        for tranche in empaquete(elements, cout):
            l0 = cellules[tranche[0]][0]
            l1 = cellules[tranche[-1]][0]
            titre = ("%s — Tarif fixe latéral vitré et imposte vitrée, "
                     "hauteur %d mm, largeurs de %d à %d mm" % (
                         PREFIXE, hauteur, l0, l1))
            corps = entete + " ; ".join(elements[i] for i in tranche) + cloture
            f.ajoute(titre, page, corps)

    paliers = extrait_meneau(lignes)
    elements = ["%d mm, %s € HT, %s € TTC" % (lg, ht, ttc) for lg, ht, ttc in paliers]
    entete = (
        "Sur la porte de service PVC H81 Access, le meneau battant est tarifé par "
        "palier de longueur, en valeur blanc teinté masse, la longueur se "
        "déduisant de la largeur hors tout selon le profil de dormant. Les "
        "paliers tarifés sont : ")
    gabarit_titre = "%s — Tarif meneau battant, longueurs de 0000 à 0000 mm" % PREFIXE
    cloture = ". Ces prix se lisent directement, sans interpolation."
    cout = mots(gabarit_titre) + 14 + mots(entete) + mots(cloture)
    for tranche in empaquete(elements, cout):
        l0 = paliers[tranche[0]][0]
        l1 = paliers[tranche[-1]][0]
        titre = "%s — Tarif meneau battant, longueurs de %d à %d mm" % (PREFIXE, l0, l1)
        corps = entete + " ; ".join(elements[i] for i in tranche) + cloture
        f.ajoute(titre, page, corps)

    return f


# --------------------------------------------------------------------------
# F4 — Options et plus-values forfaitaires
# --------------------------------------------------------------------------

def resout_unite(L):
    cle_des = (L["chapitre"], L["designation"])
    if cle_des in UNITES_PAR_DESIGNATION:
        return UNITES_PAR_DESIGNATION[cle_des]
    for cle in ((L["chapitre"], L["tableau"]), (L["chapitre"], None)):
        if cle in UNITES:
            return UNITES[cle]
    return None


def libelle_poste(L):
    """Libellé auto-porteur d'un poste de gamme. Les discriminants absents de
    l'Excel sont repris du PDF et rattachés par le montant, jamais par l'ordre
    des lignes (règle C4 de CA76, importée par A3)."""
    ch, tb, de, dt = L["chapitre"], L["tableau"], L["designation"], L["details"]
    ht, ttc = eur(L["ht"]), eur(L["ttc"])

    systeme = {"standard": "en système standard",
               "evo": "en système Evo",
               "standard+evo": "en système standard et en système Evo"}

    if ch == "PV vitrages":
        return "plus-value vitrage %s pour modèle à composition libre %s" % (
            tb, systeme.get(de, ""))

    if ch == "Vitrages pour fixes":
        return "plus-value vitrage %s pour fixe %s" % (tb, systeme.get(de, ""))

    if ch == "Croisillons":
        jonction = ("croisillon filant" if dt.lower() == "filant"
                    else "croisillon en T ou en croix")
        profil = "" if dt.lower() == "filant" else ", profil %s" % dt
        return "%s, %s%s" % (jonction, bas(de), profil)

    if ch == "Remplissage" and tb == "Panneaux de soubassement":
        famille = FAMILLE_PANNEAU.get((ht, ttc))
        if famille is None:
            JOURNAL["unites_non_etablies"].append(
                {"ligne": L["r"], "chapitre": ch,
                 "motif": "famille de panneau non rattachable par le montant"})
            return None
        teinte = bas(de) if de.lower() != "panneaux rainurés renforcés" else "blanc"
        codes = ", codes %s" % dt if dt else ""
        return "%s en %s%s" % (famille, teinte, codes)

    if ch == "Remplissage" and tb == "Moulures 2 faces":
        return bas(de.split("(")[0].strip().rstrip(","))

    if ch == "PV vitrage panneaux":
        nom = bas(de)
        prefixe = "" if nom.startswith("vitrage") else "vitrage "
        return "%s%s pour panneau PVC" % (prefixe, nom)

    if ch.startswith("Garnitures"):
        if de not in GARNITURE_EXT:
            JOURNAL["unites_non_etablies"].append(
                {"ligne": L["r"], "chapitre": ch,
                 "motif": "configuration de garniture non identifiée"})
            return None
        return ("garniture monocolore, béquille intérieure BDEL et garniture "
                "extérieure %s" % de)

    if ch == "Options et accessoires":
        return bas(de)

    if ch == "Pack EVO":
        if tb == "Option spécifique":
            nom = {"GELEC": "gâche électrique 12 volts",
                   "GMECA": "gâche mécanique"}.get(de, bas(de))
            return "%s, référence %s, disponible avec le Pack Evo uniquement" % (nom, de)
        if tb == "Bequilles et poignées":
            return ("béquille et poignée de tirage sécurité monocolore, "
                    "références BDSL et BPSL, disponibles avec le Pack Evo uniquement")

    if ch in ("Élargisseurs", "Profilés complémentaires", "Accouplements statiques"):
        # La colonne Détails porte tantôt un nombre de faces plaxées, tantôt une
        # référence de profilé : la catégorie n'est jamais annoncée à tort.
        if not dt:
            detail = ""
        elif re.fullmatch(r"\d+ faces?", dt):
            detail = " en %s" % dt
        else:
            detail = ", référence %s" % dt
        return "%s, %s%s" % (bas(tb), bas(de), detail)

    if ch == "Tapées de doublage":
        return "tapée de doublage, %s, références %s" % (bas(de), dt)

    if ch == "Seuils":
        return "seuil ou profilé sous seuil, référence %s" % de

    if ch == "Fabrications spéciales : cintres":
        if tb == "Meneau / Traverse":
            return "fixation spéciale de meneau ou de traverse sur partie cintrée"
        return "cintrage en %s" % bas(de)

    JOURNAL["unites_non_etablies"].append(
        {"ligne": L["r"], "chapitre": ch, "tableau": tb, "designation": de,
         "motif": "aucun patron de libellé pour ce chapitre"})
    return None


def construit_options(modeles, lignes):
    f = Fichier("Tarif_H81_Access_OPTIONS.md", "options")

    # --- A3 : maille option × modèle, pour les deux options portées par le modèle
    for d in modeles:
        ht, ttc = eur(d["evo_ht"]), eur(d["evo_ttc"])
        if ht is None or ttc is None:
            JOURNAL["lignes_exclues"].append(
                {"modele": d["nom"], "objet": "Pack Evo",
                 "motif": "montant HT ou TTC non renseigné"})
        else:
            titre = "%s — Option Pack Evo sur %s (ligne %s, famille %s)" % (
                PREFIXE, d["nom"], d["ligne"], d["famille"])
            corps = (
                "Sur le modèle %s de la porte de service PVC H81 Access, ligne %s, "
                "famille %s, le Pack Evo est proposé en plus-value au tarif de "
                "%s € HT ; le tarif TTC correspondant est de %s €. Cette "
                "plus-value est forfaitaire, par châssis, et s'ajoute au prix du "
                "modèle de base. Le Pack Evo comprend le ferrage six points manuel, "
                "le double vitrage clair de sécurité Isol'3 44/2-20G-4 en standard "
                "et l'accès aux options spécifiques du pack. Ce montant s'entend "
                "hors éco-participation." % (
                    d["nom"], d["ligne"], d["famille"], ht, ttc))
            f.ajoute(titre, d["page"], corps)

    for d in modeles:
        ht, ttc = eur(d["opt_ht"]), eur(d["opt_ttc"])
        libelle = d["opt_libelle"]
        if not libelle:
            if ht is not None:
                # Anti-fantôme : un montant sans libellé d'option n'est pas un poste.
                JOURNAL["lignes_exclues"].append(
                    {"modele": d["nom"], "objet": "option de vitrage ou de panneau",
                     "motif": "montant de structure sans libellé d'option ; "
                              "aucun bloc correspondant sur la page-modèle du PDF"})
            continue
        if d["opt_renvoi"]:
            # Zéro accompagné d'un renvoi transverse : exclu (lecture CA76).
            JOURNAL["lignes_exclues"].append(
                {"modele": d["nom"], "objet": "option %s" % libelle,
                 "motif": "montant nul assorti du renvoi « %s » : renvoi "
                          "transverse, traité en plus-values" % d["opt_renvoi"]})
            continue
        if ht is None or ttc is None:
            JOURNAL["lignes_exclues"].append(
                {"modele": d["nom"], "objet": "option %s" % libelle,
                 "motif": "montant HT ou TTC non renseigné"})
            continue
        titre = "%s — Option %s sur %s (ligne %s, famille %s)" % (
            PREFIXE, bas(libelle), d["nom"], d["ligne"], d["famille"])
        corps = (
            "Sur le modèle %s de la porte de service PVC H81 Access, ligne %s, "
            "famille %s, l'option %s est proposée en plus-value au tarif de "
            "%s € HT ; le tarif TTC correspondant est de %s €. %s Cette "
            "plus-value est forfaitaire et s'ajoute au prix du modèle de base. "
            "Elle s'entend hors éco-participation." % (
                d["nom"], d["ligne"], d["famille"], bas(libelle), ht, ttc,
                d["opt_desc"] + "." if d["opt_desc"] and not d["opt_desc"].endswith(".")
                else d["opt_desc"]))
        f.ajoute(titre, d["page"], corps)

    # --- Postes de gamme, issus du bloc chapitres
    for L in lignes:
        ch = L["chapitre"]
        if ch in CHAPITRES_EXCLUS:
            if ch == "Exemple de calculs":
                JOURNAL["lignes_exclues"].append(
                    {"ligne": L["r"], "objet": "%s / %s" % (ch, L["designation"]),
                     "motif": "exclusion de périmètre assumée : aucun prix "
                              "inédit, deux erreurs arithmétiques, forme additive"})
            continue
        if ch == "Pack EVO" and L["tableau"] == "PV":
            JOURNAL["lignes_exclues"].append(
                {"ligne": L["r"], "objet": "Pack Evo, plus-value de gamme",
                 "motif": "servi à la maille option × modèle sur les onze "
                          "modèles ; le doublon de gamme n'apporte rien"})
            continue

        ht, ttc = eur(L["ht"]), eur(L["ttc"])
        if ht is None or ttc is None:
            JOURNAL["lignes_exclues"].append(
                {"ligne": L["r"], "objet": "%s / %s" % (ch, L["designation"]),
                 "motif": "montant HT ou TTC non renseigné"})
            continue

        unite = resout_unite(L)
        if unite is None:
            JOURNAL["unites_non_etablies"].append(
                {"ligne": L["r"], "chapitre": ch, "tableau": L["tableau"],
                 "designation": L["designation"]})
            continue
        _, glose = unite

        libelle = libelle_poste(L)
        if libelle is None:
            JOURNAL["lignes_exclues"].append(
                {"ligne": L["r"], "objet": "%s / %s" % (ch, L["designation"]),
                 "motif": "libellé auto-porteur non établi ; poste non généré"})
            continue
        libelle = re.sub(r"\s+", " ", libelle).strip().rstrip(",")
        titre = "%s — Option %s (tarif de gamme)" % (PREFIXE, libelle)
        if ht == "0":
            phrase_prix = ("ne donne lieu à aucune plus-value : son tarif est de "
                           "0 € HT et de 0 € TTC")
        else:
            phrase_prix = ("est proposé au tarif de %s € HT ; le tarif TTC "
                           "correspondant est de %s €" % (ht, ttc))
        corps = ("Sur %s, le poste %s, chapitre %s, %s. %s Ce montant s'entend "
                 "hors éco-participation." % (
                     PRODUIT, libelle, CHAPITRE_AFFICHE.get(ch, bas(ch)),
                     phrase_prix, glose))
        f.ajoute(titre, PAGE_CHAPITRE[ch], corps)

    # --- Deux postes à 0 € nu relevés dans le PDF et absents de l'Excel (règle A3)
    postes_pdf = [
        ("ferrage cinq points manuel de série", PAGE_FERRAGE,
         "Sur %s, le ferrage galets cinq points manuel, serrure à galets à "
         "relevage, est le ferrage de série : il ne donne lieu à aucune "
         "plus-value, son tarif étant de 0 € HT. Le ferrage six points manuel "
         "relève du Pack Evo et fait l'objet d'une plus-value distincte. Le "
         "ferrage automatique n'est pas disponible sur cette gamme. Ce poste ne "
         "donne lieu à aucune facturation, ni forfaitaire ni proportionnelle. Ce "
         "montant s'entend hors éco-participation." % PRODUIT),
        ("changement de teinte des paumelles", PAGE_PAUMELLES,
         "Sur %s, le changement de teinte des paumelles ne donne lieu à aucune "
         "plus-value : son tarif est de 0 € HT. Les paumelles aluminium de "
         "référence PPE-4, d'une largeur de 99 mm et d'une longueur de 112 mm, "
         "sont disponibles en trois teintes, blanc, titane et noir. Ce poste ne "
         "donne lieu à aucune facturation, ni forfaitaire ni proportionnelle. Ce "
         "montant s'entend hors éco-participation." % PRODUIT),
    ]
    for libelle, page, corps in postes_pdf:
        JOURNAL["postes_repris_du_pdf"].append(
            {"poste": libelle, "page": page,
             "motif": "montant nul non porté par l'Excel ; zéro nu, donc absence "
                      "de plus-value et information utile (règle A3)"})
        f.ajoute("%s — Option %s (tarif de gamme)" % (PREFIXE, libelle), page, corps)

    return f


# --------------------------------------------------------------------------
# F5 — Plus-values proportionnelles
# --------------------------------------------------------------------------

def construit_plus_values_proportionnelles():
    f = Fichier("Tarif_H81_Access_PLUS_VALUES_PROPORTIONNELLES.md",
                "plus_values_proportionnelles")

    f.ajoute(
        PREFIXE + " — Plus-value de plaxage sur le prix du modèle",
        PAGE_COULEURS,
        "Sur %s, l'offre de couleurs comporte deux groupes. Le groupe 1, blanc "
        "teinté masse en monocolore, est sans plus-value. Le groupe 2, plaxage "
        "une face selon le modèle, intérieur blanc teinté masse et extérieur "
        "Chêne d'Or veiné 48 ou gris anthracite perlé 115, porte une plus-value "
        "de 15 %% du prix. Seules ces teintes sont disponibles. Le taux est "
        "transcrit tel qu'il figure au tarif ; son application au prix du modèle "
        "revient à l'ADV. La faisabilité du plaxage dépend du modèle." % PRODUIT)

    f.ajoute(
        PREFIXE + " — Plus-value de vitrage sur les châssis cintrés en arc de cercle",
        PAGE_CHAPITRE["Fabrications spéciales : cintres"],
        "Sur %s, un châssis cintré en arc de cercle porte, en sus de la "
        "plus-value de cintrage du châssis, une plus-value de vitrage de 100 %% . "
        "Elle s'applique au prix du vitrage, la plus-value de cintrage étant "
        "quant à elle donnée vitrage non compris. Le taux est transcrit tel qu'il "
        "figure au tarif ; son application revient à l'ADV. Aucune autre forme "
        "spéciale n'est réalisable sur cette gamme." % PRODUIT)

    return f


# --------------------------------------------------------------------------
# F6 — Caractéristiques
# --------------------------------------------------------------------------

def enveloppe(dims):
    minL = minH = maxL = maxH = None
    for prof, (miL, miH, maL, maH) in dims.items():
        for v, k in ((miL, "minL"), (miH, "minH"), (maL, "maxL"), (maH, "maxH")):
            if v in (None, ""):
                continue
            v = float(v)
            if k == "minL":
                minL = v if minL is None else min(minL, v)
            elif k == "minH":
                minH = v if minH is None else min(minH, v)
            elif k == "maxL":
                maxL = v if maxL is None else max(maxL, v)
            else:
                maxH = v if maxH is None else max(maxH, v)
    if None in (minL, minH, maxL, maxH):
        return None
    return int(minL), int(maxL), int(minH), int(maxH)


def construit_caracteristiques(modeles):
    f = Fichier("Tarif_H81_Access_CARACTERISTIQUES.md", "caracteristiques")
    for d in modeles:
        titre = "%s — Caractéristiques %s (ligne %s, famille %s)" % (
            PREFIXE, d["nom"], d["ligne"], d["famille"])
        # Le descriptif du modèle de base est cité tel quel, ponctuation comprise :
        # il n'est ni reformulé ni tronqué.
        base = d["modele_base"]
        corps = ("Le modèle %s de la porte de service PVC H81 Access, ligne %s, "
                 "famille %s, a pour modèle de base, dans les termes du tarif : "
                 "« %s » " % (d["nom"], d["ligne"], d["famille"], base))
        if d["ud"]:
            libelle = d["ud_libelle"] or "porte"
            corps += "Le tarif lui attribue un coefficient Ud %s de %s. " % (
                libelle, d["ud"])
        if d["dim_mini_panneau"]:
            corps += "La dimension minimale du panneau est de %s. " % d["dim_mini_panneau"]
        env = enveloppe(d["dims"])
        if env:
            Lmin, Lmax, Hmin, Hmax = env
            corps += ("Selon le profil de dormant retenu, 5103 L69, 5107 L84, "
                      "5114 LZ102 ou 5120 L110, il se fabrique dans une plage de "
                      "largeur comprise entre %d et %d mm et de hauteur comprise "
                      "entre %d et %d mm ; les limites exactes dépendent du profil "
                      "choisi et figurent à la page du modèle." % (
                          Lmin, Lmax, Hmin, Hmax))
        f.ajoute(titre, d["page"], corps)
    return f


# --------------------------------------------------------------------------
# F7 — Faisabilités
# --------------------------------------------------------------------------

# (sujet, complément de nom) pour chaque équipement.
ARTICLES_EQUIP = {
    "Passe-lettres": ("le passe-lettres", "du passe-lettres"),
    "Chatière": ("la chatière", "de la chatière"),
    "Judas optique": ("le judas optique", "du judas optique"),
    "Heurtoir": ("le heurtoir", "du heurtoir"),
}


def construit_faisabilites(modeles):
    f = Fichier("Tarif_H81_Access_FAISABILITES.md", "faisabilites")
    noms = [d["nom"] for d in modeles]

    # A4 — équipements : constante négative sur les 44 lignes, formulation en négatif.
    equipements = []
    for d in modeles:
        for eq, fa in d["equipements"]:
            if eq not in equipements:
                equipements.append(eq)
    for eq in equipements:
        compatibles = [d["nom"] for d in modeles
                       if any(e == eq and fa == "oui" for e, fa in d["equipements"])]
        art, art_de = ARTICLES_EQUIP.get(eq, (bas(eq), "de " + bas(eq)))
        if compatibles:
            corps = ("Sur %s, %s peut équiper les modèles suivants : %s. Sur les "
                     "autres modèles de la gamme, il n'est pas disponible. Cet "
                     "équipement n'est pas chiffré au tarif." % (
                         PRODUIT, art, enumere(compatibles)))
        else:
            corps = ("Sur %s, %s n'est montable sur aucun modèle de la gamme. "
                     "Cela vaut pour les onze modèles du tarif : %s. Aucune "
                     "page-modèle ne déclare cette compatibilité et cet "
                     "équipement n'est pas chiffré au tarif." % (
                         PRODUIT, art, enumere(noms)))
        f.ajoute("%s — Faisabilité %s par modèle" % (PREFIXE, art_de),
                 min(PAGE_MODELE.values()), corps)

    # Teintes par modèle
    plaxables = [d["nom"] for d in modeles if "Plaxage" in d["teintes"]]
    blancs = [d["nom"] for d in modeles if d["nom"] not in plaxables]
    f.ajoute(
        "%s — Faisabilité du plaxage par modèle" % PREFIXE,
        PAGE_COULEURS,
        "Sur %s, le plaxage une face n'est réalisable que sur quatre modèles : "
        "%s. Les sept autres modèles, %s, ne sont disponibles qu'en blanc teinté "
        "masse. Le tarif ne prévoit pas d'autre teinte que celles de l'offre "
        "couleurs, et la plus-value de plaxage figure aux plus-values "
        "proportionnelles." % (PRODUIT, enumere(plaxables), enumere(blancs)))

    # Typologie de châssis
    f.ajoute(
        "%s — Faisabilité des typologies de châssis" % PREFIXE,
        PAGE_TYPOLOGIE,
        "Sur %s, sont réalisables, en version standard comme en version Evo : la "
        "porte à un vantail et à deux vantaux, la partie fixe vitrée attenante ou "
        "accouplée, qu'il s'agisse d'un fixe latéral, d'une imposte ou des deux, "
        "l'ouverture extérieure à un vantail et l'arc de cercle. Ne sont "
        "réalisables ni en standard ni en Evo : l'ouverture extérieure à deux "
        "vantaux, la porte avec volet roulant, le plein cintre, l'anse de panier "
        "et le trapèze." % PRODUIT)

    # Ferrage
    f.ajoute(
        "%s — Faisabilité des ferrages" % PREFIXE,
        PAGE_FERRAGE,
        "Sur %s, le ferrage de série est un ferrage galets cinq points manuel, "
        "serrure à galets à relevage. Le ferrage six points manuel est disponible "
        "avec le Pack Evo. Le ferrage automatique n'est pas disponible sur cette "
        "gamme. Chaque ferrage impose ses propres limites de hauteur hors tout, "
        "distinctes selon le profil de dormant et selon le nombre de vantaux ; "
        "elles figurent aux pages du ferrage standard et du Pack Evo." % PRODUIT)

    # Cintres
    f.ajoute(
        "%s — Faisabilité des formes cintrées et de leurs teintes" % PREFIXE,
        PAGE_CINTRES_FAISABILITE,
        "Sur %s, seul l'arc de cercle est réalisable en forme spéciale. Le plein "
        "cintre, l'anse de panier, le gothique, l'oeil de boeuf et l'oblong ne "
        "sont pas réalisables. En teinte, le blanc et le plaxage Chêne d'Or sont "
        "réalisables en cintre ; le plaxage gris anthracite perlé ne l'est pas, "
        "en raison d'une problématique de brillance dans le cintre. Les rayons "
        "minimaux de cintrage dépendent du profil de dormant ou d'ouvrant et "
        "figurent à la page des fabrications spéciales." % PRODUIT)

    # Limites dimensionnelles
    f.ajoute(
        "%s — Limites dimensionnelles selon le profil de dormant" % PREFIXE,
        PAGE_LIMITES,
        "Sur %s, la largeur minimale est définie par le rayon d'ouverture et la "
        "saillie de la poignée : en dessous, la poignée entre en conflit avec le "
        "dormant. La dimension maximale est définie par le poids maximal accepté "
        "par les paumelles, soit 160 kg. Les valeurs mini et maxi de largeur hors "
        "tout diffèrent selon le profil de dormant, 5103 L69, 5107 L84, 5114 "
        "LZ102 ou 5120 L110, et selon le nombre de vantaux ; elles figurent au "
        "tableau des limites dimensionnelles. Sur un vantail semi-fixe, la "
        "largeur minimale de pareclose est de 206 mm." % PRODUIT)

    # Contradiction produit exposée, non arbitrée
    f.ajoute(
        "%s — Mention contradictoire sur les limites de fabrication" % PREFIXE,
        min(PAGE_MODELE.values()),
        "Sur %s, les onze pages-modèles portent en note de bas de page que les "
        "limites de fabrication indiquées valent pour une menuiserie en aluminium "
        "blanc et une serrure six points. Cette mention contredit deux "
        "caractéristiques déclarées ailleurs par le même tarif : la gamme est en "
        "PVC, et le ferrage de série est un ferrage cinq points, le six points "
        "relevant du Pack Evo. La divergence est signalée telle quelle et n'est "
        "pas arbitrée ; elle doit être levée par le service Produits avant tout "
        "usage des limites en question." % PRODUIT)

    return f


# --------------------------------------------------------------------------
# F8 — Transverses
# --------------------------------------------------------------------------

def construit_transverses():
    f = Fichier("Tarif_H81_Access_TRANSVERSES.md", "transverses")

    blocs = [
        ("descriptif des lignes et des familles", PAGE_DESCRIPTIF_LIGNES,
         "Sur %s, le tarif classe les modèles en quatre ensembles : la ligne "
         "Vitrée, vitrage TRYBA, en double vitrage réalisé en usine ; la ligne "
         "Contemporaine, panneau plein lisse renforcé en système Evo de 32 mm "
         "d'épaisseur ; la ligne Traditionnelle, famille Création, à composition "
         "libre pleine ou vitrée avec meneaux, panneaux lisses ou rainurés "
         "renforcés de 32 mm ; et la ligne Traditionnelle, famille Élégance, à "
         "panneaux renforcés avec moulures thermoformées de 32 mm. Le détail "
         "figure à la page du descriptif des lignes." % PRODUIT),
        ("couleur des accessoires", PAGE_ACCESSOIRES_COULEUR,
         "Sur %s, la couleur des garnitures, du seuil, des paumelles et des "
         "joints dépend de l'aspect extérieur retenu, blanc ou plaxé, l'intérieur "
         "étant blanc. Les garnitures et les paumelles existent en blanc, titane "
         "ou noir ; le seuil en gris ou noir ; la teinte des joints varie selon "
         "que l'extérieur est blanc ou plaxé. Le détail des combinaisons figure à "
         "la page de la couleur des accessoires. Ces choix ne portent aucun "
         "montant." % PRODUIT),
        ("réglage et poids admissible des paumelles", PAGE_PAUMELLES,
         "Sur %s, les paumelles de référence PPE-4 offrent un réglage "
         "tridimensionnel et acceptent un poids de vantail maximal. Les plages de "
         "réglage horizontal, vertical et en compression, ainsi que le poids maxi, "
         "figurent à la page des paumelles. Avec le dormant 5103 L69, un "
         "élargisseur de 10 mm est fortement conseillé en partie haute pour "
         "assurer l'engondage et le dégondage du vantail. Une notice de montage et "
         "de réglage est disponible sur le site Infos Concessionnaires." % PRODUIT),
        ("cotes de fabrication et largeur de passage", PAGE_COTES_FABRICATION,
         "Sur %s, les cotes de fabrication sont données par profil de dormant et "
         "par composition, un vantail, un vantail plus fixe, deux vantaux, un "
         "vantail plus imposte fixe. La largeur de passage libre se déduit de la "
         "largeur de fabrication par une valeur propre au profil de dormant. Ces "
         "abaques figurent aux pages des cotes de fabrication et de la largeur de "
         "passage. Elles ne portent aucun montant." % PRODUIT),
        ("évolutions du tarif", PAGE_EVOLUTIONS,
         "Sur %s, le tarif tient un tableau des évolutions, qui consigne la date "
         "de chaque hausse, son taux et la partie du tarif concernée. La dernière "
         "évolution consignée porte sur l'ensemble du tarif. Ce tableau permet de "
         "vérifier qu'un prix cité correspond bien à la version en vigueur. Il "
         "figure à la page des évolutions du tarif." % PRODUIT),
        ("cylindre de sécurité livré en standard",
         PAGE_CHAPITRE["Garnitures - béquilles et poignées"],
         "Sur %s, la porte est équipée en standard d'un cylindre de sécurité "
         "débrayable, livré avec cinq clés et une carte de propriété. La taille "
         "du cylindre en rosette standard dépend de la configuration de "
         "garnitures retenue. Le détail figure à la page des garnitures. Ce "
         "cylindre ne fait l'objet d'aucune plus-value distincte au tarif." % PRODUIT),
    ]
    for label, page, corps in blocs:
        f.ajoute("%s — Existence et localisation des tarifs de %s" % (PREFIXE, label),
                 page, corps)
    return f


# --------------------------------------------------------------------------
# Point d'entrée
# --------------------------------------------------------------------------

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Feuil1"]

    journal_colonnes(ws)
    modeles = lit_portes(ws)
    lignes = lit_chapitres(ws)

    JOURNAL["arbitrages_signales"].append(
        {"reference": "A8",
         "objet": "champ front matter `document_source_ttc`",
         "detail": "le §8 de la note ne liste que `document_source`, en édition HT, "
                   "alors que chaque chunk sert un montant TTC issu de l'édition "
                   "TTC. Le champ est ajouté sur le précédent FT84 ; à confirmer."})
    for d in modeles:
        vitre = "vitrage" in d["modele_base"].lower() or "vitré" in d["modele_base"].lower()
        if d["ud_libelle"] == "porte vitrée" and not vitre:
            JOURNAL["arbitrages_signales"].append(
                {"reference": "divergence source",
                 "objet": "libellé du coefficient Ud du modèle %s" % d["nom"],
                 "detail": "la page %d porte « Ud porte vitrée » alors que le "
                           "modèle de base est un panneau plein. Le libellé est "
                           "transcrit tel quel et attribué au tarif ; l'écart est "
                           "remonté au service Produits." % d["page"]})
    nouvelles = sum(1 for e in JOURNAL["corrections_libelle"]
                    if e.get("statut") == "nouveau")
    deja = sum(1 for e in JOURNAL["corrections_libelle"]
               if e.get("statut") == "note")
    JOURNAL["arbitrages_signales"].append(
        {"reference": "A9",
         "objet": "corrections de libellé de statut « nouveau »",
         "detail": "%d corrections relevées à l'étape 4 s'ajoutent aux %d déjà "
                   "arbitrées au §5 de la note. Toutes suivent la même règle — le "
                   "PDF se déclare document de référence à sa page 2 — et aucune "
                   "ne touche un montant. Elles restent à confirmer par le "
                   "service Produits." % (nouvelles, deja)})

    fichiers = [
        construit_methode(),
        construit_prix_portes(modeles),
        construit_prix_fixes(lignes),
        construit_options(modeles, lignes),
        construit_plus_values_proportionnelles(),
        construit_caracteristiques(modeles),
        construit_faisabilites(modeles),
        construit_transverses(),
    ]

    os.makedirs(OUT, exist_ok=True)
    total = 0
    print("=== Fichiers générés ===")
    for f in fichiers:
        f.ecrit()
        pmax = max((c["mots"] for c in f.chunks), default=0)
        print("  %-50s %4d chunks   (max %d mots)" % (f.nom, len(f.chunks), pmax))
        total += len(f.chunks)
    print("  %-50s %4d chunks" % ("TOTAL", total))

    with open(os.path.join(OUT, "journal_migration_H81_Access.json"), "w",
              encoding="utf-8") as fh:
        json.dump(JOURNAL, fh, ensure_ascii=False, indent=2)

    print("\n=== Journal ===")
    for cle, valeurs in JOURNAL.items():
        print("  %-26s : %d entrée(s)" % (cle, len(valeurs)))
    if JOURNAL["unites_non_etablies"]:
        print("\n  ATTENTION — postes non générés faute d'unité de facturation :")
        for e in JOURNAL["unites_non_etablies"]:
            print("   ", e)


if __name__ == "__main__":
    main()
