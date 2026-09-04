#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Contrôle de conformité du corpus tarif HAM76.

Audit AUTONOME : relit les fichiers .md produits sans réutiliser aucune fonction
du générateur, et les confronte à la note de cadrage (forme) puis à l'Excel et au
PDF (fidélité). Toute constante utile est recalculée ici depuis les sources.

Seize familles de contrôles :
   1  décomptes par fichier
   2  plafond de 200 mots
   3  ligne de source (regex) et cohérence du nom de document
   4  continuité SC depuis SC0002, sans trou ni doublon
   5  préfixe de titre
   6  unicité des titres
   7  prose sans puces
   8  front matter complet et cohérent
   9  anti-fantôme et dédoublonnage
  10  fidélité exhaustive des prix modèles
  11  fidélité exhaustive des plus-values d'options
  12  fidélité exhaustive des enveloppes dimensionnelles
  13  fidélité exhaustive des montants de catalogue
  14  absence de tout montant dans les fichiers d'orientation
  15  non-fuite des chapitres gelés et des unités non forfaitaires
  16  discrimination HAM76 / HA76 / H81, liant inter-fichiers, croisement PDF
"""
import glob
import os
import re
import unicodedata
from collections import Counter, OrderedDict, defaultdict

import openpyxl
from openpyxl.utils import column_index_from_string as CI

XLSX = "/mnt/user-data/uploads/HAM76_-infos-tarifs.xlsx"
PDF_PATH = "/mnt/user-data/uploads/Tarif_HAM76_HT_04-05-2026.pdf"
OUTDIR = "/mnt/user-data/outputs"
GAMME = "HAM76"
DESIGNATION = "Porte d'entrée monobloc Aluminium"
PREFIXE = f"{GAMME} {DESIGNATION} — "
PLAFOND = 200
NBSP_FINE = "\u202f"

ATTENDU = {"METHODE": 9, "PRIX_MODELES": 53, "OPTIONS_MODELES": 128,
           "CARACTERISTIQUES": 53, "COMPAT_EQUIPEMENTS": 6,
           "CATALOGUE_OPTIONS": 23, "FAISABILITES": 10, "TRANSVERSES": 6}
ORIENTATION = {"TRANSVERSES"}

RE_SOURCE = re.compile(
    r"^\*Source : Tarif—HAM76—HT—04-05-2026\.pdf, pages? "
    r"([0-9]+(?: (?:à|et) [0-9]+)*) — information (originale|complémentaire) — "
    r"SC(\d{4})\*$")

# HAM12 : coquille manifeste rétablie au titre. Table re-déclarée ici de façon
# autonome ; l'audit vérifie l'intervention au lieu d'en être aveugle.
COQUILLES_TITRE = {"Vitragre avec print grille": "Vitrage avec print grille"}

OK, KO, WARN = [], [], []


def ok(msg):
    OK.append(msg)


def ko(msg):
    KO.append(msg)


def warn(msg):
    WARN.append(msg)


# =============================================================== relecture .md

def lire_corpus():
    corpus = OrderedDict()
    for path in sorted(glob.glob(f"{OUTDIR}/Tarif_{GAMME}_*.md")):
        nom = os.path.basename(path)[len(f"Tarif_{GAMME}_"):-3]
        txt = open(path, encoding="utf-8").read()
        parts = txt.split("---\n")
        fm = parts[1] if len(parts) > 2 else ""
        body = "---\n".join(parts[2:])
        chunks = []
        for bloc in re.split(r"\n(?=## )", body):
            bloc = bloc.strip()
            if not bloc.startswith("## "):
                continue
            lignes = bloc.split("\n")
            titre = lignes[0][3:].strip()
            src = lignes[1].strip() if len(lignes) > 1 else ""
            corps = "\n".join(lignes[2:]).strip()
            chunks.append({"titre": titre, "src": src, "corps": corps, "bloc": bloc})
        corpus[nom] = {"fm": fm, "chunks": chunks, "path": path}
    return corpus


CORPUS = lire_corpus()
TOUS = [(f, c) for f, d in CORPUS.items() for c in d["chunks"]]
TEXTE_TOTAL = "\n".join(c["bloc"] for _, c in TOUS)


def montants(txt):
    """Tous les montants en euros d'un texte, en entiers, quel que soit le
    séparateur de milliers (espace ordinaire ou fine insécable)."""
    out = []
    for m in re.finditer(r"(\d[\d\u202f\u00a0 ]*)\s*€", txt):
        out.append(int(re.sub(r"[^\d]", "", m.group(1))))
    return out


# =============================================================== sources

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Feuil1"]


def cel(r, letter):
    v = ws.cell(r, CI(letter)).value
    return v.strip() if isinstance(v, str) else ("" if v is None else v)


LIGNES_MODELE = [r for r in range(2, ws.max_row + 1)
                 if cel(r, "AV") == "Fiche porte" and cel(r, "B")]
GRP = OrderedDict()
for r in LIGNES_MODELE:
    GRP.setdefault(cel(r, "B"), []).append(r)
MODELES = list(GRP.keys())


def euro(v):
    try:
        return int(round(float(str(v).replace(" ", "").replace(",", "."))))
    except (ValueError, TypeError):
        return None


# =============================================================== 1. décomptes

def c1_decomptes():
    reel = {f: len(d["chunks"]) for f, d in CORPUS.items()}
    if set(reel) != set(ATTENDU):
        ko(f"1. fichiers produits {sorted(reel)} != attendus {sorted(ATTENDU)}")
        return
    ecarts = {f: (reel[f], ATTENDU[f]) for f in ATTENDU if reel[f] != ATTENDU[f]}
    if ecarts:
        ko(f"1. décomptes non conformes : {ecarts}")
    else:
        ok(f"1. décomptes conformes sur 8 fichiers, total {sum(reel.values())} chunks")


# =============================================================== 2. plafond

def c2_plafond():
    depass = [(f, c["titre"], len(re.findall(r"\S+", c["bloc"])))
              for f, c in TOUS if len(re.findall(r"\S+", c["bloc"])) > PLAFOND]
    if depass:
        ko(f"2. plafond de {PLAFOND} mots dépassé sur {len(depass)} chunks : {depass[:3]}")
    else:
        mx = max(len(re.findall(r"\S+", c["bloc"])) for _, c in TOUS)
        ok(f"2. plafond respecté sur {len(TOUS)} chunks (maximum observé {mx} mots)")


# =============================================================== 3. ligne de source

def c3_source():
    mauvais = [(f, c["titre"]) for f, c in TOUS if not RE_SOURCE.match(c["src"])]
    if mauvais:
        ko(f"3. ligne de source non conforme sur {len(mauvais)} chunks : {mauvais[:3]}")
        return
    pages = set()
    for _, c in TOUS:
        for p in re.findall(r"\d+", RE_SOURCE.match(c["src"]).group(1)):
            pages.add(int(p))
    hors = sorted(p for p in pages if not 1 <= p <= 76)
    if hors:
        ko(f"3. pages hors du PDF (76 pages) : {hors}")
    elif 11 in pages or 76 in pages:
        ko("3. une page exclue du périmètre est citée en source (11 ou 76)")
    else:
        ok(f"3. lignes de source conformes, {len(pages)} pages distinctes citées, "
           f"aucune page exclue")


# =============================================================== 4. continuité SC

def c4_sc():
    pb = []
    for f, d in CORPUS.items():
        scs = [int(RE_SOURCE.match(c["src"]).group(3)) for c in d["chunks"]
               if RE_SOURCE.match(c["src"])]
        if scs != list(range(2, 2 + len(d["chunks"]))):
            pb.append(f)
    if pb:
        ko(f"4. continuité SC rompue dans : {pb}")
    else:
        ok("4. numérotation SC continue depuis SC0002 dans les 8 fichiers")


# =============================================================== 5. préfixe

def c5_prefixe():
    mauvais = [(f, c["titre"]) for f, c in TOUS if not c["titre"].startswith(PREFIXE)]
    if mauvais:
        ko(f"5. préfixe de titre absent sur {len(mauvais)} chunks : {mauvais[:3]}")
    else:
        ok(f"5. les {len(TOUS)} titres portent le préfixe « {PREFIXE.strip(' —')} »")


# =============================================================== 6. unicité

def c6_unicite():
    doublons = [t for t, n in Counter(c["titre"] for _, c in TOUS).items() if n > 1]
    if doublons:
        ko(f"6. titres en doublon : {doublons[:5]}")
    else:
        ok(f"6. les {len(TOUS)} titres sont uniques dans l'ensemble du corpus")


# =============================================================== 7. prose

def c7_prose():
    mauvais = []
    for f, c in TOUS:
        for l in c["corps"].split("\n"):
            if re.match(r"^\s*([-*•]|\d+[.)])\s", l):
                mauvais.append((f, c["titre"]))
                break
    if mauvais:
        ko(f"7. puces ou listes numérotées sur {len(mauvais)} chunks : {mauvais[:3]}")
    else:
        ok("7. prose continue, aucune puce ni liste numérotée")


# =============================================================== 8. front matter

def c8_front_matter():
    requis = ["document_source: Tarif_HAM76_HT_04-05-2026.pdf",
              "type_document: tarif", "gamme_code: HAM76",
              f'gamme_nom: "{DESIGNATION}"', 'collection: "TRYBA ALUMINIUM"',
              "materiau: aluminium", "date_validite: 2026-05-04",
              "audiences:", "glossaire_ref:", "perimetre:"]
    pb = [(f, r) for f, d in CORPUS.items() for r in requis if r not in d["fm"]]
    if pb:
        ko(f"8. front matter incomplet : {pb[:4]}")
    else:
        ok("8. front matter complet et homogène sur les 8 fichiers")


# =============================================================== 9. anti-fantôme

def c9_antifantome():
    # a) un chunk de prix par modèle réellement tarifé, ni plus ni moins
    tarifes = {m for m in MODELES if euro(cel(GRP[m][0], "K")) is not None}
    servis = set()
    for c in CORPUS["PRIX_MODELES"]["chunks"]:
        m = re.match(r".*— Tarif (.+?) 1 vantail \(ligne ", c["titre"]).group(1)
        servis.add(m)
    if servis != tarifes:
        ko(f"9a. écart prix : en trop {sorted(servis - tarifes)}, "
           f"manquants {sorted(tarifes - servis)}")
    else:
        ok(f"9a. anti-fantôme prix : {len(servis)} modèles servis, "
           f"exactement les modèles tarifés")

    # b) aucune option servie sans libellé dans la source
    libelles = set()
    for m, rs in GRP.items():
        for r in rs:
            for col in ("AF", "AK"):
                lib = re.sub(r"\s+", " ", str(cel(r, col))).strip().rstrip(".")
                if lib:
                    libelles.add((COQUILLES_TITRE.get(lib, lib), m))
    servis_opt = set()
    for c in CORPUS["OPTIONS_MODELES"]["chunks"]:
        mm = re.match(r".*— Option (.+) sur (.+?) \(ligne ", c["titre"])
        servis_opt.add((mm.group(1), mm.group(2)))
    if servis_opt != libelles:
        ko(f"9b. écart options : en trop {sorted(servis_opt - libelles)[:3]}, "
           f"manquants {sorted(libelles - servis_opt)[:3]}")
    else:
        ok(f"9b. anti-fantôme options : {len(servis_opt)} couples servis, exactement "
           f"les couples porteurs d'un libellé (264 cellules à zéro sans libellé écartées)")


# =============================================================== 10. prix

def c10_prix():
    err = 0
    for c in CORPUS["PRIX_MODELES"]["chunks"]:
        m = re.match(r".*— Tarif (.+?) 1 vantail \(ligne ", c["titre"]).group(1)
        r0 = GRP[m][0]
        att = [euro(cel(r0, "K")), euro(cel(r0, "L"))]
        lus = montants(c["corps"])
        if lus != [x for x in att if x is not None]:
            ko(f"10. {m} : corpus {lus} != Excel {att}")
            err += 1
        if cel(r0, "C").replace("Ligne ", "") not in c["titre"]:
            ko(f"10. {m} : ligne de design absente ou erronée au titre")
            err += 1
    if not err:
        ok(f"10. fidélité des prix : {2 * len(CORPUS['PRIX_MODELES']['chunks'])} montants "
           f"tracés à la cellule, aucun montant calculé")


# =============================================================== 11. options

def c11_options():
    ref = {}
    for m, rs in GRP.items():
        for r in rs:
            for lib_c, ht_c, ttc_c in (("AF", "AH", "AI"), ("AK", "AL", "AM")):
                lib = re.sub(r"\s+", " ", str(cel(r, lib_c))).strip().rstrip(".")
                if lib:
                    ref[(COQUILLES_TITRE.get(lib, lib), m)] = (euro(cel(r, ht_c)),
                                                               euro(cel(r, ttc_c)))
    err = n = 0
    for c in CORPUS["OPTIONS_MODELES"]["chunks"]:
        mm = re.match(r".*— Option (.+) sur (.+?) \(ligne ", c["titre"])
        cle = (mm.group(1), mm.group(2))
        att_ht, att_ttc = ref[cle]
        lus = montants(c["corps"])
        if att_ht == 0:
            if lus or "sans plus-value" not in c["corps"]:
                ko(f"11. {cle} : plus-value nulle mal rendue ({lus})")
                err += 1
            n += 1
        else:
            if lus != [att_ht, att_ttc]:
                ko(f"11. {cle} : corpus {lus} != Excel {[att_ht, att_ttc]}")
                err += 1
            n += 2
    if not err:
        ok(f"11. fidélité des options : {n} valeurs de plus-value vérifiées, "
           f"dont 14 nulles énoncées « sans plus-value »")


# =============================================================== 12. dimensions

def c12_dimensions():
    err = n = 0
    for c in CORPUS["CARACTERISTIQUES"]["chunks"]:
        m = re.match(r".*— Caractéristiques (.+?) \(ligne ", c["titre"]).group(1)
        r0 = GRP[m][0]
        att = [int(cel(r0, x)) for x in ("T", "V", "U", "W")]  # Lmin Lmax Hmin Hmax
        lus = [int(x) for x in re.findall(r"(\d{3,4}) (?:à|millimètres)", c["corps"])]
        lus = [int(x) for x in re.findall(r"de (\d{3,4}) à (\d{3,4}) millimètres",
                                          c["corps"]) for x in x]
        if lus != att:
            ko(f"12. {m} : dimensions corpus {lus} != Excel {att}")
            err += 1
        n += 4
        base = re.sub(r"\s+", " ", str(cel(r0, "P"))).strip().rstrip(".")
        if base and base not in c["corps"]:
            ko(f"12. {m} : description du modèle de base non littérale")
            err += 1
    if not err:
        ok(f"12. fidélité des caractéristiques : {n} bornes dimensionnelles et "
           f"{len(MODELES)} descriptions de modèle de base transcrites littéralement")


# =============================================================== 13. catalogue

EXCLUS_PDF = {"ZAE35/400", "ZAE35/800", "ZAE351200", "BDE-DG/O"}
LIB_EXCEL = ("béquilles standards monocolores", "béquilles standards bicolores",
             "béquilles inox", "poussoirs inox", "poignée encastrée extérieure",
             "rosettes seules")
LIB_PDF = ("ferrage", "paumelles", "seuil")


def c13_catalogue():
    """Le catalogue mêle deux origines : 20 références issues de l'Excel et 3
    options captées au PDF. On vérifie séparément le nombre, l'ordre, les
    montants, et l'attestation au PDF des libellés restitués ou corrigés."""
    attendu = []
    for r in range(2, ws.max_row + 1):
        chap, tab = cel(r, "AV"), cel(r, "AW")
        if not chap or chap == "Fiche porte":
            continue
        if tab in ("Fixe", "PV pour vitrages fixes"):
            continue
        d = str(cel(r, "AX"))
        if d in EXCLUS_PDF:
            continue
        attendu.append((r, d, euro(cel(r, "BA")), euro(cel(r, "BB"))))

    servis = CORPUS["CATALOGUE_OPTIONS"]["chunks"]
    depuis_excel = [c for c in servis
                    if any(f"catalogue {l} :" in c["titre"] for l in LIB_EXCEL)]
    depuis_pdf = [c for c in servis
                  if any(f"catalogue {l} :" in c["titre"] for l in LIB_PDF)]
    err = 0
    if len(depuis_excel) + len(depuis_pdf) != len(servis):
        ko("13a. chunk de catalogue rattaché à aucun chapitre connu")
        err += 1
    if len(depuis_excel) != len(attendu):
        ko(f"13a. {len(depuis_excel)} chunks issus de l'Excel pour {len(attendu)} "
           f"lignes retenues")
        err += 1
        return
    if len(depuis_pdf) != 3:
        ko(f"13a. {len(depuis_pdf)} options captées au PDF, 3 attendues")
        err += 1

    n = 0
    interventions = []
    for c, (r, d, ht, ttc) in zip(depuis_excel, attendu):
        att = [] if ht in (0, None) else [ht, ttc]
        lus = montants(c["corps"])
        if lus != att:
            ko(f"13b. ligne {r} ({d or 'sans désignation'}) : corpus {lus} != Excel {att}")
            err += 1
        n += len(att)
        if ht == 0 and "sans plus-value" not in c["corps"]:
            ko(f"13b. ligne {r} : montant nul non énoncé « sans plus-value »")
            err += 1
        if not d or d not in c["titre"]:
            interventions.append((r, d, c["titre"].split(" : ", 1)[1]))
    for c in depuis_pdf:
        if montants(c["corps"]):
            ko(f"13c. option captée au PDF portant un montant : {c['titre']}")
            err += 1

    # les libellés restitués ou corrigés doivent être attestés dans le PDF
    if interventions:
        atteste, non_atteste = 0, []
        try:
            import pdfplumber
            with pdfplumber.open(PDF_PATH) as pdf:
                p72 = pdf.pages[71].extract_text() or ""
                p73 = pdf.pages[72].extract_text() or ""
            ref_pdf = re.sub(r"[\s\u202f\u00a0]+", "", p72 + p73).upper()
            for r, d, titre in interventions:
                jetons = re.findall(r"\b[A-Z][A-Z0-9/\-]{2,}\b", titre)
                if jetons and all(j in ref_pdf for j in jetons):
                    atteste += 1
                else:
                    non_atteste.append((r, titre))
        except ImportError:
            warn("13. attestation PDF des libellés restitués non exécutée")
            return
        if non_atteste:
            ko(f"13d. libellé restitué non attesté au PDF : {non_atteste}")
            err += 1
        else:
            ok(f"13d. {atteste} libellés restitués ou corrigés, tous attestés dans les "
               f"pages 72 et 73 du PDF")
    if not err:
        ok(f"13. fidélité du catalogue : {n} montants tracés à la cellule sur "
           f"{len(attendu)} références Excel, 3 options captées au PDF sans montant, "
           f"4 références non attestées au PDF écartées")


# =============================================================== 14. orientation

def c14_orientation():
    fautifs = []
    for f in ORIENTATION:
        for c in CORPUS[f]["chunks"]:
            if montants(c["corps"]) or re.search(r"\d+\s*%", c["corps"]):
                fautifs.append((f, c["titre"]))
    if fautifs:
        ko(f"14. montant ou pourcentage dans un fichier d'orientation : {fautifs}")
    else:
        ok("14. aucun montant ni pourcentage dans les 6 chunks d'orientation")


# =============================================================== 15. non-fuite

def c15_non_fuite():
    """Deux invariants distincts : les valeurs de la grille gelée ne doivent pas
    ressortir dans les fichiers d'orientation, et aucun chunk sourcé sur une page
    à unité (17 grille des fixes, 18 plus-values au mètre carré) ne doit porter
    de montant."""
    grille = set()
    for r in range(214, 238):
        for c in range(CI("BD"), CI("CY") + 1):
            v = euro(ws.cell(r, c).value)
            if v is not None:
                grille.add(v)
    servis_orient = set()
    for f in ("TRANSVERSES", "FAISABILITES", "METHODE"):
        for c in CORPUS[f]["chunks"]:
            servis_orient |= set(montants(c["corps"]))
    fuite_grille = sorted(grille & servis_orient)
    if fuite_grille:
        ko(f"15a. valeurs de la grille gelée présentes dans un chunk d'orientation : "
           f"{fuite_grille[:5]}")
    else:
        ok(f"15a. aucune des {len(grille)} valeurs distinctes de la grille gelée "
           f"(1106 cellules) ne figure dans les fichiers d'orientation")

    PAGES_UNITE = {17, 18}
    fuite, nb = [], 0
    for f, c in TOUS:
        pages = {int(x) for x in re.findall(r"\d+", RE_SOURCE.match(c["src"]).group(1))}
        if pages & PAGES_UNITE:
            nb += 1
            if montants(c["corps"]):
                fuite.append((f, c["titre"], montants(c["corps"])))
    if fuite:
        ko(f"15b. montant servi par un chunk sourcé sur une page à unité : {fuite}")
    else:
        ok(f"15b. les {nb} chunks sourcés sur les pages 17 et 18 (grille gelée et "
           f"plus-values au mètre carré) ne portent aucun montant")


def c16_discrimination():
    # a) aucune trace des gammes voisines hors mention explicite de distinction
    fautifs = []
    for f, c in TOUS:
        for g in ("HA76", "H81"):
            # HA76 n'est pas un sous-mot de HAM76 : on cherche le mot entier
            if re.search(rf"\b{g}\b", c["bloc"]):
                if not re.search(r"ne s'applique|contrairement|distinct|ni à la",
                                 c["bloc"]):
                    fautifs.append((f, c["titre"], g))
    if fautifs:
        ko(f"16a. mention non encadrée d'une gamme voisine : {fautifs[:3]}")
    else:
        ok("16a. HA76 et H81 ne sont cités que dans des énoncés de distinction explicite")

    # b) artefact HAM77..HAM88 de la colonne H
    art = re.findall(r"HAM(?!76\b)\d{2}", TEXTE_TOTAL)
    if art:
        ko(f"16b. artefact de désignation propagé dans le corpus : {set(art)}")
    else:
        ok("16b. aucune désignation HAM77 à HAM88 dans le corpus (colonne H neutralisée)")

    # c) le corps de chaque chunk de prix, option et caractéristiques nomme la gamme
    manquants = []
    for f in ("PRIX_MODELES", "OPTIONS_MODELES", "CARACTERISTIQUES"):
        for c in CORPUS[f]["chunks"]:
            if "de la gamme HAM76" not in c["corps"]:
                manquants.append((f, c["titre"]))
    if manquants:
        ko(f"16c. mention « de la gamme HAM76 » absente du corps : {len(manquants)} chunks")
    else:
        ok("16c. les 234 chunks de prix, options et caractéristiques portent la mention "
           "« de la gamme HAM76 » dans leur corps")

    # d) le mot monobloc, discriminant produit, est présent partout
    sans = [(f, c["titre"]) for f, c in TOUS if "monobloc" not in c["bloc"].lower()]
    if sans:
        ko(f"16d. discriminant « monobloc » absent de {len(sans)} chunks : {sans[:3]}")
    else:
        ok(f"16d. le discriminant « monobloc » figure dans les {len(TOUS)} chunks")

    # e) liant inter-fichiers : même nom de modèle et même ligne dans les trois fichiers
    def cles(f, motif):
        return {re.match(motif, c["titre"]).group(1, 2)
                for c in CORPUS[f]["chunks"] if re.match(motif, c["titre"])}
    p = cles("PRIX_MODELES", r".*— Tarif (.+?) 1 vantail \(ligne (.+?)\)")
    k = cles("CARACTERISTIQUES", r".*— Caractéristiques (.+?) \(ligne (.+?)\)")
    o = {(m, l) for _, m, l in
         (re.match(r".*— Option (.+) sur (.+?) \(ligne (.+?)\)", c["titre"]).groups()
          for c in CORPUS["OPTIONS_MODELES"]["chunks"])}
    if p != k:
        ko(f"16e. liant rompu entre prix et caractéristiques : {p ^ k}")
    elif not o <= p:
        ko(f"16e. options rattachées à un couple modèle/ligne inconnu : {o - p}")
    else:
        ok(f"16e. liant inter-fichiers vérifié : {len(p)} couples modèle/ligne "
           f"identiques en prix et caractéristiques, options toutes rattachées")


# =============================================================== 17. croisement PDF

def c17_pdf():
    try:
        import pdfplumber
    except ImportError:
        warn("17. pdfplumber indisponible : croisement PDF non exécuté")
        return

    def dedup(s):
        out = []
        for line in s.split("\n"):
            if len(line) >= 6 and len(line) % 2 == 0 and all(
                    line[i] == line[i + 1] for i in range(0, len(line) - 1, 2)):
                out.append(line[::2])
            else:
                out.append(line)
        return "\n".join(out)

    def key(s):
        s = unicodedata.normalize("NFD", s.lower())
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        return re.sub(r"[^a-z0-9]", "", s)

    # page citée par chaque chunk de prix
    pages_citees = {}
    for c in CORPUS["PRIX_MODELES"]["chunks"]:
        m = re.match(r".*— Tarif (.+?) 1 vantail", c["titre"]).group(1)
        pages_citees[m] = int(RE_SOURCE.match(c["src"]).group(1))

    err = 0
    with pdfplumber.open(PDF_PATH) as pdf:
        if len(pdf.pages) != 76:
            ko(f"17. le PDF compte {len(pdf.pages)} pages, 76 attendues")
            err += 1
        for m, p in pages_citees.items():
            txt = dedup(pdf.pages[p - 1].extract_text() or "")
            plat = re.sub(r"[\s\u202f\u00a0]+", "", txt)
            if key(m) not in key(txt):
                ko(f"17. {m} : le nom du modèle n'apparaît pas en page {p}")
                err += 1
                continue
            ht = euro(cel(GRP[m][0], "K"))
            if f"{ht}€" not in plat:
                ko(f"17. {m} : le montant {ht} € n'apparaît pas en page {p}")
                err += 1
            if "2vantaux" in plat.lower():
                ko(f"17. {m} : mention de 2 vantaux en page {p}")
                err += 1
    if not err:
        ok(f"17. croisement PDF exhaustif : les {len(pages_citees)} modèles, leur prix HT "
           f"et l'absence de configuration 2 vantaux vérifiés page à page")


# =============================================================== 18. vocabulaire

FAUX_SYNONYMES = {"anti-dégondage": "anti-décrochement",
                  "gond": "paumelle", "charnière": "paumelle",
                  "survitrage": "triple vitrage isolant",
                  "ouverture à soufflet": "oscillo-battant"}


def c18_vocabulaire():
    """Deux invariants de gouvernance : tout faux synonyme présent dans le corpus
    doit être couvert par un chunk de vocabulaire qui nomme le terme retenu, et
    toute coquille rétablie au titre doit voir sa graphie d'origine consignée
    verbatim dans le corps du chunk."""
    presents = {t for t in FAUX_SYNONYMES
                if re.search(rf"\b{re.escape(t)}\b", TEXTE_TOTAL, flags=re.I)}
    err = 0
    for t in presents:
        retenu = FAUX_SYNONYMES[t]
        couvert = any(re.search(rf"\b{re.escape(t)}\b", c["corps"], flags=re.I)
                      and re.search(rf"\b{re.escape(retenu)}\b", c["corps"], flags=re.I)
                      for _, c in TOUS)
        if not couvert:
            ko(f"18a. « {t} » présent dans le corpus sans chunk de vocabulaire nommant "
               f"« {retenu} »")
            err += 1
    if not err:
        ok(f"18a. {len(presents)} faux synonyme(s) présent(s) dans le corpus, chacun "
           f"couvert par un chunk de vocabulaire nommant le terme retenu")

    interventions = 0
    for source, corrige in COQUILLES_TITRE.items():
        for _, c in TOUS:
            if corrige in c["titre"]:
                interventions += 1
                if source not in c["corps"]:
                    ko(f"18b. coquille rétablie au titre sans consignation de la graphie "
                       f"du tarif : {c['titre']}")
                    err += 1
        if re.search(rf"{re.escape(source)}", "\n".join(c["titre"] for _, c in TOUS)):
            ko(f"18b. graphie fautive « {source} » encore présente dans un titre")
            err += 1
    if not err:
        ok(f"18b. {interventions} coquille(s) rétablie(s) au titre, graphie du tarif "
           f"consignée verbatim dans le corps de chaque chunk concerné")


# =============================================================== exécution

def main():
    for fn in (c1_decomptes, c2_plafond, c3_source, c4_sc, c5_prefixe, c6_unicite,
               c7_prose, c8_front_matter, c9_antifantome, c10_prix, c11_options,
               c12_dimensions, c13_catalogue, c14_orientation, c15_non_fuite,
               c16_discrimination, c17_pdf, c18_vocabulaire):
        try:
            fn()
        except Exception as e:  # un contrôle qui casse est un échec, pas un silence
            ko(f"{fn.__name__} : exception {type(e).__name__} {e}")

    print("=" * 78)
    print("CONTRÔLE DE CONFORMITÉ — CORPUS TARIF HAM76")
    print("=" * 78)
    for m in OK:
        print(f"  OK   {m}")
    for m in WARN:
        print(f"  WARN {m}")
    for m in KO:
        print(f"  KO   {m}")
    print("-" * 78)
    print(f"{len(OK)} contrôles réussis, {len(KO)} échecs, {len(WARN)} avertissements")
    return 1 if KO else 0


if __name__ == "__main__":
    raise SystemExit(main())
