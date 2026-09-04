#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Générateur de chunks Markdown pour le tarif HAM76
(Porte d'entrée monobloc Aluminium, collection TRYBA ALUMINIUM).

Adapté du gabarit HA76, lui-même dérivé de la note souche H81.
Règles normatives HAM1 à HAM11 (note_cadrage_migration_tarif_HAM76.md).

Huit fichiers produits :
  1. METHODE              cadre de lecture du tarif
  2. PRIX_MODELES         un chunk par modèle (configuration unique : 1 vantail)
  3. OPTIONS_MODELES      un chunk par couple option x modèle porteur d'un libellé
  4. CARACTERISTIQUES     un chunk par modèle (modèle de base, enveloppe dimensionnelle)
  5. COMPAT_EQUIPEMENTS   un chunk par équipement et par type de poignée de tirage
  6. CATALOGUE_OPTIONS    un chunk par référence forfaitaire hors modèle
  7. FAISABILITES         faisabilités non tarifaires
  8. TRANSVERSES          orientation sans montant (%, €/m², grille gelée)

Points d'adaptation propres à HAM76, tous fondés sur l'étape 2 :
  - une seule configuration de vantaux (2 vantaux irréalisable, page 10) ;
  - aucun coefficient Ud au tarif ;
  - un seul couple de dormants dans le bloc modèles (AL10101 / AL10108) ;
  - colonne H interdite de lecture (artefact HAM77..HAM88) ;
  - anti-fantôme fondé sur la présence du LIBELLÉ, jamais sur le montant ;
  - compatibilité des équipements uniformément négative ;
  - trois options chiffrées captées au PDF, absentes de l'Excel ;
  - quatre références Excel non attestées au PDF, écartées.

Transversal : SC continue par fichier depuis SC0002, ligne de source normée,
plafond 200 mots, prose sans puces, journal des colonnes non mappées,
fidélité littérale (aucun montant calculé, aucune conversion HT/TTC).
"""
import json
import re
import unicodedata
from collections import OrderedDict, defaultdict

import openpyxl
from openpyxl.utils import column_index_from_string as CI, get_column_letter as CL

XLSX = "/mnt/user-data/uploads/HAM76_-infos-tarifs.xlsx"
PDF_PATH = "/mnt/user-data/uploads/Tarif_HAM76_HT_04-05-2026.pdf"
PDF_SOURCE = "Tarif—HAM76—HT—04-05-2026.pdf"      # nom affiché, em-dashes
PDF_FILENAME = "Tarif_HAM76_HT_04-05-2026.pdf"    # nom réel, underscores
OUTDIR = "/mnt/user-data/outputs"
PAGEMAP_OUT = "/home/claude/ham76/pagemap_HAM76.json"
PLAFOND = 200

GAMME = "HAM76"
DESIGNATION = "Porte d'entrée monobloc Aluminium"
PREFIXE = f"{GAMME} {DESIGNATION}"
NBSP_FINE = "\u202f"

ALERTS = []
JOURNAL = []

# =============================================================== utilitaires


def clean(v):
    return "" if v is None else str(v).strip()


def fmt_euro(v):
    """4069 -> '4 069' (espace fine insécable U+202F). None si non numérique."""
    if v in (None, ""):
        return None
    try:
        n = int(round(float(str(v).replace(" ", "").replace(",", "."))))
    except (ValueError, TypeError):
        return None
    return f"{n:,}".replace(",", NBSP_FINE)


def to_num(v):
    try:
        return float(str(v).replace(" ", "").replace(NBSP_FINE, "").replace(",", "."))
    except (ValueError, TypeError):
        return None


def norm_texte(v):
    """Aplatit un champ multi-lignes de l'Excel en prose, sans toucher au U+202F."""
    s = clean(v).replace("\r", "\n")
    s = re.sub(r"\n+", " ", s)
    s = re.sub(r"[^\S\u202f]+", " ", s).strip()
    return s.rstrip(".").strip()


def count_words(*parts):
    return len(re.findall(r"\S+", " ".join(p for p in parts if p)))


def source_line(page, sc, nature="originale"):
    mot = "pages" if ("à" in str(page) or "et" in str(page)) else "page"
    return f"*Source : {PDF_SOURCE}, {mot} {page} — information {nature} — SC{sc:04d}*"


def dedup_lignes(txt):
    """Le PDF HAM76 porte des couches de texte dupliquées caractère à caractère
    ('PPoorrttee'). Piège équivalent au \\s+ de HA76 : dédoubler avant tout usage."""
    out = []
    for line in txt.split("\n"):
        if len(line) >= 6 and len(line) % 2 == 0 and all(
                line[i] == line[i + 1] for i in range(0, len(line) - 1, 2)):
            out.append(line[::2])
        else:
            out.append(line)
    return "\n".join(out)


def cle(s):
    s = unicodedata.normalize("NFD", s.lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", s)


def enumere(noms):
    """['A','B','C'] -> 'A, B et C'."""
    noms = list(noms)
    if not noms:
        return ""
    if len(noms) == 1:
        return noms[0]
    return ", ".join(noms[:-1]) + " et " + noms[-1]


class Fichier:
    """Accumule des chunks, numérote les SC depuis SC0002, contrôle le plafond."""

    def __init__(self, nom, perimetre):
        self.nom = nom
        self.perimetre = perimetre
        self.chunks = []
        self.sc = 2

    def add(self, titre, corps, page, nature="originale"):
        src = source_line(page, self.sc, nature)
        # \s+ engloberait l'espace fine insécable des montants : on l'exclut
        corps = re.sub(r"[^\S\u202f]+", " ", corps).strip()
        n = count_words(titre, src, corps)
        if n > PLAFOND:
            ALERTS.append(f"[{self.nom}] plafond dépassé ({n} mots) : {titre}")
        self.chunks.append((titre, src, corps, n))
        self.sc += 1

    def write(self):
        fm = [
            "---",
            f"document_source: {PDF_FILENAME}",
            "type_document: tarif",
            f"sous_type: {self.nom.lower()}",
            f"gamme_code: {GAMME}",
            f'gamme_nom: "{DESIGNATION}"',
            'collection: "TRYBA ALUMINIUM"',
            "materiau: aluminium",
            'version_doc: "2026.05"',
            "date_validite: 2026-05-04",
            f"nb_chunks: {len(self.chunks)}",
            "audiences: [ADV, commercial]",
            "---",
            "",
        ]
        body = [f"## {t}\n{s}\n\n{c}\n" for t, s, c, _ in self.chunks]
        path = f"{OUTDIR}/Tarif_{GAMME}_{self.nom}.md"
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(fm) + "\n".join(body))
        return path, len(self.chunks)


# =============================================================== lecture Excel

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Feuil1"]
HDR = {CL(c): clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)}


def cell(r, letter):
    return clean(ws.cell(r, CI(letter)).value)


# --- HAM2 : groupement par modèle. La ligne Excel est un emplacement
#     d'affichage (l'équipement), pas un axe tarifaire.
LIGNES_MODELE = [r for r in range(2, ws.max_row + 1)
                 if cell(r, "AV") == "Fiche porte" and cell(r, "B")]

groupes = OrderedDict()
for r in LIGNES_MODELE:
    groupes.setdefault(cell(r, "B"), []).append(r)

tailles = {len(v) for v in groupes.values()}
if tailles != {4}:
    ALERTS.append(f"[HAM2] groupes de lignes de taille inattendue : {sorted(tailles)}")

# --- HAM9 / colonne H interdite : contrôle de l'artefact HAM77..HAM88
_artefact = {cell(r, "H") for r in LIGNES_MODELE
             if cell(r, "H") and "HAM76" not in cell(r, "H")}
if _artefact:
    JOURNAL.append(
        f"Colonne H (Gamme description) interdite de lecture : {len(_artefact)} valeurs "
        f"aberrantes de type HAM77 à HAM88. La désignation est dérivée de la colonne G.")

# --- constantes de gamme vérifiées, non supposées
_gammes = {cell(r, "G") for r in LIGNES_MODELE}
if _gammes != {GAMME}:
    ALERTS.append(f"[gamme] codes gamme inattendus dans la colonne G : {_gammes}")


def ligne_de(m):
    return cell(groupes[m][0], "C").replace("Ligne ", "").strip()


MODELES = list(groupes.keys())
NB_MODELES = len(MODELES)

# =============================================================== table modèle -> page


def construire_pagemap():
    """Table modèle -> page PDF construite par lecture des en-têtes de page.
    Une page modèle porte 'Ligne X' en première ligne et le nom du modèle en
    seconde. Couverture exigée à 100 %, sinon arrêt."""
    import pdfplumber
    par_cle = {cle(m): m for m in MODELES}
    pages = {}
    with pdfplumber.open(PDF_PATH) as pdf:
        nb_pages = len(pdf.pages)
        for i, p in enumerate(pdf.pages, start=1):
            txt = dedup_lignes(p.extract_text() or "")
            lignes = [l.strip() for l in txt.split("\n") if l.strip()]
            if len(lignes) < 2 or not lignes[0].lower().startswith("ligne "):
                continue
            m = par_cle.get(cle(lignes[1]))
            if m is None:
                continue
            if m in pages:
                ALERTS.append(f"[pagemap] modèle {m} trouvé sur deux pages : "
                              f"{pages[m]} et {i}")
            pages[m] = i
    manquants = [m for m in MODELES if m not in pages]
    if manquants:
        raise SystemExit(f"[pagemap] couverture incomplète, modèles sans page : {manquants}")
    JOURNAL.append(
        f"Table modèle -> page construite par lecture des en-têtes de page : "
        f"{len(pages)}/{NB_MODELES} modèles, pages {min(pages.values())} à "
        f"{max(pages.values())}, sur un PDF de {nb_pages} pages.")
    json.dump(pages, open(PAGEMAP_OUT, "w", encoding="utf-8"),
              ensure_ascii=False, indent=0)
    return pages


PAGES_MODELE = construire_pagemap()

# =============================================================== 1. METHODE

P_TYPOLOGIE, P_ALPHA, P_MENTIONS, P_FIXES, P_VITR_FIXES = 10, 9, 2, 17, 18


def f_methode():
    f = Fichier("METHODE", "règles de lecture du tarif HAM76 et périmètre du corpus")

    f.add(f"{PREFIXE} — Portée du tarif et distinction avec les gammes HA76 et H81",
          "Ce tarif porte exclusivement sur la porte d'entrée monobloc aluminium de la "
          "gamme HAM76, collection TRYBA ALUMINIUM, dans son édition du 4 mai 2026. Il ne "
          "s'applique ni à la porte d'entrée aluminium de la gamme HA76, ni à la porte "
          "d'entrée PVC de la gamme H81, qui sont des gammes distinctes disposant chacune "
          "de son propre tarif, de ses propres modèles et de ses propres prix. Le tarif "
          "HAM76 signale lui-même cette distinction : contrairement aux portes H81 et HA76, "
          "le sens du panneau y est déterminé par le sens d'ouverture de la porte. Les trois "
          "gammes partagent un vocabulaire technique très proche, mais leurs montants ne "
          "sont jamais interchangeables. Aucun prix, aucune plus-value et aucune "
          "caractéristique de ce tarif ne doit être rapporté à HA76 ni à H81.",
          P_TYPOLOGIE)

    f.add(f"{PREFIXE} — Unité tarifaire : le modèle et sa ligne de design",
          f"Le prix d'une porte HAM76 est attaché au modèle individuel. Le tarif comprend "
          f"{NB_MODELES} modèles, répartis en neuf lignes de design : Création, Evasion, "
          "Intemporelle, Epurée, Actuelle, Evolution, Nature, Tradition et Harmonie. Chaque "
          "modèle appartient à une seule ligne. La ligne est une propriété descriptive du "
          "modèle et non un axe de prix : une même ligne porte plusieurs prix et un même "
          "prix couvre plusieurs lignes. Le tarif HAM76 ne comporte aucune notion de "
          "collection ni de référence commerciale au-delà de la ligne.",
          P_ALPHA)

    f.add(f"{PREFIXE} — Le prix d'un modèle ne dépend pas de la dimension",
          "Le tarif d'un modèle de porte HAM76 est forfaitaire : il ne varie ni avec la "
          "largeur ni avec la hauteur commandées. Les limites dimensionnelles figurant au "
          "tarif sont des limites de fabrication et non des paliers de prix. Une demande de "
          "prix portant sur une dimension particulière reçoit donc le tarif du modèle, sous "
          "réserve que la dimension demandée reste dans les limites de fabrication de ce "
          "modèle.",
          P_ALPHA)

    f.add(f"{PREFIXE} — Une seule configuration de vantaux : la porte un vantail",
          "La gamme HAM76 n'existe qu'en porte un vantail. Le tarif déclare la porte deux "
          "vantaux irréalisable, au même titre que les formes spéciales, la porte avec volet "
          "roulant et l'ouverture extérieure. Il n'y a donc pas lieu de demander la "
          "configuration de vantaux avant de restituer un prix : chaque modèle n'a qu'un "
          "seul tarif. Une demande de prix en deux vantaux sur la gamme HAM76 reçoit une "
          "réponse d'indisponibilité et un renvoi à la page 10 du tarif, et jamais un "
          "montant.",
          P_TYPOLOGIE)

    f.add(f"{PREFIXE} — Les prix hors taxes et toutes taxes comprises sont lus, jamais calculés",
          "Les montants hors taxes s'entendent hors éco-participation. Les montants toutes "
          "taxes comprises ne se déduisent pas des montants hors taxes par application d'un "
          "taux : le rapport entre les deux varie d'une ligne de tarif à l'autre, aussi bien "
          "entre les modèles qu'entre les références du catalogue. Chaque montant doit donc "
          "être restitué tel qu'il figure au tarif, et jamais obtenu par calcul, par "
          "conversion ni par interpolation entre deux valeurs. L'édition publiée du tarif "
          "HAM76 ne porte que les montants hors taxes.",
          P_MENTIONS)

    f.add(f"{PREFIXE} — Montants exprimés au mètre carré ou en pourcentage",
          "Certaines plus-values du tarif HAM76 ne sont pas des forfaits. Les plus-values de "
          "vitrages pour fixes sont exprimées au mètre carré de surface vitrée. Les "
          "plus-values de l'offre couleurs sont exprimées en pourcentage du prix. Ces "
          "montants ne figurent pas dans ce corpus, car leur usage suppose un calcul que le "
          "corpus ne doit pas induire. Une demande portant sur ces plus-values est orientée "
          "vers la page du tarif qui les porte.",
          P_VITR_FIXES)

    f.add(f"{PREFIXE} — Ce que ce corpus ne contient pas",
          "Ce corpus ne contient ni la grille de tarif des fixes latéraux vitrés et impostes "
          "vitrées, ni les plus-values exprimées au mètre carré ou en pourcentage. Il ne "
          "contient pas non plus la fiche info produit insérée dans le tarif, qui relève du "
          "corpus technique. Sur ces points, la réponse consiste à renvoyer l'utilisateur à "
          "la page correspondante du tarif. Toute demande dont la réponse n'est pas présente "
          "en toutes lettres dans un chunk doit recevoir une réponse d'absence, jamais une "
          "estimation.",
          P_FIXES)

    f.add(f"{PREFIXE} — Portée des limites de fabrication indiquées par modèle",
          "Les limites de fabrication portées au tarif pour chaque modèle de la gamme HAM76 "
          "sont données pour une menuiserie en aluminium blanc équipée d'une serrure six "
          "points. Les restrictions propres aux autres teintes et aux options de ferrage se "
          "lisent aux chapitres correspondants du tarif et peuvent réduire ces limites. Une "
          "limite citée pour un modèle ne vaut donc pas engagement de faisabilité pour une "
          "teinte ou un ferrage particuliers.",
          P_VITR_FIXES)

    f.add(f"{PREFIXE} — Vocabulaire retenu : paumelle et anti-décrochement",
          "Le terme retenu dans l'ensemble du corpus TRYBA pour désigner le dispositif "
          "empêchant le décrochement du vantail est anti-décrochement. La page 13 du tarif "
          "HAM76 emploie le terme anti-dégondage pour décrire la configuration à trois lames "
          "de la paumelle : les deux termes désignent le même dispositif, et la formulation "
          "du tarif est reprise telle quelle dans le chunk qui la transcrit. Les termes gond "
          "et charnière ne sont pas employés en HAM76 : la pièce qui articule l'ouvrant sur "
          "le dormant est une paumelle. La classification de ce dispositif au regard de la "
          "résistance à l'effraction n'est pas énoncée par le tarif et ne peut donc pas en "
          "être déduite.",
          13, nature="complémentaire")
    return f


# =============================================================== 2. PRIX


def f_prix():
    f = Fichier("PRIX_MODELES", "prix des modèles de porte HAM76")
    for m, rs in groupes.items():
        r0 = rs[0]
        ht = fmt_euro(cell(r0, "K"))
        if ht is None:                                     # anti-fantôme HAM1
            JOURNAL.append(f"[anti-fantôme] modèle {m} sans prix HT : aucun chunk de prix.")
            continue
        ttc = fmt_euro(cell(r0, "L"))
        lg = ligne_de(m)
        titre = f"{PREFIXE} — Tarif {m} 1 vantail (ligne {lg})"
        if ttc:
            corps = (f"En un vantail, seule configuration réalisable de la gamme, le modèle "
                     f"{m} de la gamme HAM76, porte d'entrée monobloc aluminium de la ligne "
                     f"{lg}, est proposé au tarif de {ht} € HT, soit {ttc} € TTC. Le montant "
                     f"hors taxes s'entend hors éco-participation. La porte deux vantaux "
                     f"n'est pas réalisable en HAM76.")
        else:
            corps = (f"En un vantail, seule configuration réalisable de la gamme, le modèle "
                     f"{m} de la gamme HAM76, porte d'entrée monobloc aluminium de la ligne "
                     f"{lg}, est proposé au tarif de {ht} € HT hors éco-participation. Le "
                     f"tarif ne porte pas de montant toutes taxes comprises pour ce modèle.")
        f.add(titre, corps, PAGES_MODELE[m])
    return f


# =============================================================== 3. OPTIONS MODELE

# (colonne libellé, colonne PV HT, colonne PV TTC, famille rédigée)
FAMILLES = [
    ("AF", "AH", "AI", "vitrage ornemental"),
    ("AK", "AL", "AM", "insert de panneau"),
]

# HAM12 — coquille orthographique manifeste dans un libellé d'option porté en titre.
# Le titre est une clé de récupération et non une transcription : l'orthographe est
# rétablie, et la graphie du tarif est consignée dans le corps du chunk, verbatim.
# Ne s'applique ni aux montants, ni aux codes de référence, ni aux descriptions
# transcrites dans le corps d'un chunk.
COQUILLES_TITRE = {"Vitragre avec print grille": "Vitrage avec print grille"}

# La colonne AC (famille « Vitrage ornemental ») est lacunaire sur Indigo alors
# que les trois options existent. La famille est donc portée par la constante de
# FAMILLES, attestée sur chaque page modèle du PDF, et la lacune est consignée.
JOURNAL.append(
    "[options] famille « Vitrage ornemental » absente de la colonne AC pour Indigo : "
    "restituée depuis la page modèle du PDF, et consignée.")


def f_options():
    f = Fichier("OPTIONS_MODELES",
                "options et plus-values rattachées à un modèle de porte HAM76")
    stats = defaultdict(int)
    for m, rs in groupes.items():
        lg = ligne_de(m)
        page = PAGES_MODELE[m]
        for col_lib, col_ht, col_ttc, famille in FAMILLES:
            for r in rs:
                lib = norm_texte(cell(r, col_lib))
                if not lib:
                    # HAM5 anti-fantôme : la présence du libellé commande, jamais
                    # la présence du montant (264 cellules à 0 sans libellé).
                    if cell(r, col_ht) != "":
                        stats["fantome_" + famille] += 1
                    continue
                lib_source = lib
                lib = COQUILLES_TITRE.get(lib, lib)
                pv_ht = fmt_euro(cell(r, col_ht))
                pv_ttc = fmt_euro(cell(r, col_ttc))
                if pv_ht is None:
                    stats["sans_montant_" + famille] += 1
                    JOURNAL.append(f"[anti-fantôme] option « {lib} » sur {m} sans montant : "
                                   f"écartée.")
                    continue
                titre = f"{PREFIXE} — Option {lib} sur {m} (ligne {lg})"
                nulle = to_num(cell(r, col_ht)) == 0
                if nulle:
                    corps = (f"L'option de {famille} {lib} peut équiper le modèle {m} de la "
                             f"gamme HAM76, porte d'entrée monobloc aluminium de la ligne "
                             f"{lg}. Elle est proposée sans plus-value : elle n'entraîne "
                             f"aucun supplément sur le prix du modèle de base.")
                else:
                    corps = (f"L'option de {famille} {lib} peut équiper le modèle {m} de la "
                             f"gamme HAM76, porte d'entrée monobloc aluminium de la ligne "
                             f"{lg}. Elle est facturée en plus-value {pv_ht} € HT, soit "
                             f"{pv_ttc} € TTC, sur le prix du modèle de base.")
                if lib != lib_source:
                    corps += (f" Le tarif orthographie cette option "
                              f"« {lib_source} ».")
                    JOURNAL.append(f"[HAM12] libellé « {lib_source} » rétabli en "
                                   f"« {lib} » au titre sur {m} ; graphie du tarif "
                                   f"consignée dans le corps du chunk.")
                f.add(titre, corps, page)
                stats["retenus_" + famille] += 1
    for k in sorted(stats):
        JOURNAL.append(f"[options] {k} : {stats[k]}")
    return f


# =============================================================== 4. CARACTERISTIQUES

DORMANTS = "AL10101 et AL10108"


def f_caracteristiques():
    f = Fichier("CARACTERISTIQUES",
                "caractéristiques des modèles de porte HAM76 hors prix")
    for m, rs in groupes.items():
        r0 = rs[0]
        lg = ligne_de(m)
        base = norm_texte(cell(r0, "P"))
        lmin, hmin = cell(r0, "T"), cell(r0, "U")
        lmax, hmax = cell(r0, "V"), cell(r0, "W")
        titre = f"{PREFIXE} — Caractéristiques {m} (ligne {lg})"
        corps = (f"Le modèle {m} de la gamme HAM76, porte d'entrée monobloc aluminium de la "
                 f"ligne {lg}, est décrit au tarif par : {base}. Ses limites de fabrication, "
                 f"pour les dormants {DORMANTS}, vont de {lmin} à {lmax} millimètres en "
                 f"largeur et de {hmin} à {hmax} millimètres en hauteur. Ces limites sont "
                 f"données pour une menuiserie en aluminium blanc équipée d'une serrure six "
                 f"points. Le tarif ne porte aucun coefficient thermique par modèle : "
                 f"cette valeur relève de la documentation technique de la gamme.")
        f.add(titre, corps, PAGES_MODELE[m])
    return f


# =============================================================== 5. COMPAT EQUIPEMENTS

EQUIPEMENTS = ["Judas optique", "Heurtoir", "Passe-lettres", "Chatière"]
POIGNEES = ["Poussoir Inox", "Poignées encastrée"]
POIGNEE_LIB = {"Poussoir Inox": "poussoir inox",
               "Poignées encastrée": "poignée encastrée"}


def _plage_modeles():
    ps = sorted(PAGES_MODELE.values())
    return f"{ps[0]} à {ps[-1]}"


def f_compat():
    f = Fichier("COMPAT_EQUIPEMENTS",
                "compatibilité de montage des équipements et des poignées de tirage HAM76")
    plage = _plage_modeles()

    for eq in EQUIPEMENTS:
        compat, total = [], 0
        for m, rs in groupes.items():
            for r in rs:
                if cell(r, "R") == eq:
                    total += 1
                    if "check" in cell(r, "Q").lower():
                        compat.append(m)
        if total != NB_MODELES:
            ALERTS.append(f"[HAM7] équipement {eq} : {total} lignes pour {NB_MODELES} modèles")
        titre = f"{PREFIXE} — Compatibilité de l'équipement {eq.lower()} par modèle"
        if compat:
            corps = (f"Dans la gamme HAM76, porte d'entrée monobloc aluminium, l'équipement "
                     f"{eq.lower()} est montable sur les modèles suivants : {enumere(compat)}. "
                     f"Sur les autres modèles de la gamme, il n'est pas réalisable. Cet "
                     f"équipement n'est pas chiffré au tarif HAM76.")
        else:
            corps = (f"Dans la gamme HAM76, porte d'entrée monobloc aluminium, l'équipement "
                     f"{eq.lower()} n'est montable sur aucun modèle. Les {NB_MODELES} pages "
                     f"modèles du tarif signalent toutes cette compatibilité de montage comme "
                     f"non réalisable. Cet équipement n'est pas chiffré au tarif HAM76. Une "
                     f"demande portant sur un {eq.lower()} en HAM76 reçoit donc une réponse "
                     f"d'indisponibilité, et non un prix.")
        f.add(titre, corps, plage)

    for p in POIGNEES:
        oui, dims, remarques = [], set(), []
        for m, rs in groupes.items():
            for r in rs:
                if cell(r, "AA") != p:
                    continue
                if "check" in cell(r, "X").lower():
                    oui.append(m)
                    if cell(r, "AB"):
                        dims.add(cell(r, "AB"))
                    if cell(r, "Y"):
                        remarques.append(f"{m} ({cell(r, 'Y')})")
        lib = POIGNEE_LIB[p]
        titre = (f"{PREFIXE} — Faisabilité de la poignée de tirage extérieure "
                 f"{lib} par modèle")
        dim = enumere(sorted(dims))
        corps = (f"Dans la gamme HAM76, porte d'entrée monobloc aluminium, la poignée de "
                 f"tirage extérieure de type {lib} est réalisable sur {len(oui)} modèles : "
                 f"{enumere(oui)}. Elle n'est pas réalisable sur les autres modèles de la "
                 f"gamme. La dimension minimale d'ouvrant exigée est de {dim}.")
        if remarques:
            corps += (f" Deux modèles portent une restriction particulière au tarif : "
                      f"{enumere(remarques)}.")
        corps += " Cette faisabilité n'est pas chiffrée sur les pages modèles du tarif."
        f.add(titre, corps, plage)
    return f


# =============================================================== 6. CATALOGUE

# HAM8 + arbitrage A2 : est retenue toute référence attestée dans le PDF publié.
# Écartées : ZAE35/400, ZAE35/800, ZAE351200 et BDE-DG/O, tarifées dans l'Excel
# mais introuvables dans le PDF (divergence de périmètre remontée au service Produits).
CATALOGUE_EXCLUS = {"ZAE35/400", "ZAE35/800", "ZAE351200", "BDE-DG/O"}

CHAP_PAGES = {
    ("Garnitures - bequilles standards", "Monocolore"): 72,
    ("Garnitures - bequilles standards", "Bicolore"): 72,
    ("Garnitures - bequilles standards", "Bequille inox"): 72,
    ("Garnitures - poussoirs et rosettes", "Poussoirs inox"): 73,
    ("Garnitures - poussoirs et rosettes", "Poignée encastrée ext"): 73,
    ("Garnitures - poussoirs et rosettes", "Rosettes seules"): 73,
}

CHAP_LIBELLE = {
    "Monocolore": "béquilles standards monocolores",
    "Bicolore": "béquilles standards bicolores",
    "Bequille inox": "béquilles inox",
    "Poussoirs inox": "poussoirs inox",
    "Poignée encastrée ext": "poignée encastrée extérieure",
    "Rosettes seules": "rosettes seules",
}

# Chapitres écartés du corpus chiffré : montants au m² (HAM9) ou grille gelée (A3).
CHAP_EXCLUS = {"PV pour vitrages fixes": "montants au mètre carré de surface vitrée",
               "Fixe": "grille dimensionnelle gelée, sans règle de lecture publiée"}

# Libellés restitués depuis le PDF, consignés au journal.
LIBELLE_RESTITUE = {
    281: "rosette encastrée inox, niveau de sécurité R20, référence ROC",
}
# Libellé Excel manifestement fautif, tranché en faveur du PDF (page 72) :
# l'Excel porte « BDEL int - BPEL int » à 102 €, alors que le PDF tarife
# BDEL/BPEL à 0 € et BDSL/BPSL à 102 €.
LIBELLE_CORRIGE = {
    258: "BDEL intérieure avec BDEL ou BPEL extérieure",
    259: "BDEL intérieure avec BDSL ou BPSL extérieure",
}

# Descriptifs transcrits de la page 73 du PDF (l'Excel ne porte aucun détail).
POUSSOIRS_DETAIL = {
    "ZAE751/400": "une hauteur de 400 mm, un entraxe de 300 mm et un diamètre de 30 mm",
    "ZAE48": "une hauteur de 405 mm, un entraxe de 300 mm et un diamètre de 30 mm",
    "ZAE751": "une hauteur de 330 mm, un entraxe de 210 mm et un diamètre de 30 mm",
    "PR/500": "une hauteur de 500 mm et une section rectangulaire de 40 sur 20 mm, en inox brossé",
    "ZAE770": "une hauteur de 470 mm, un entraxe de 300 mm et un diamètre de 32 mm",
    "ZAE751/800": "une hauteur de 800 mm, un entraxe de 600 mm et un diamètre de 30 mm",
    "PR/800": "une hauteur de 800 mm et une section rectangulaire de 40 sur 20 mm, en inox brossé",
    "ZAE751/1200": "une hauteur de 1200 mm, un entraxe de 1000 mm et un diamètre de 30 mm",
    "ZAE48/1200": "une hauteur de 1200 mm, un entraxe de 1000 mm et un diamètre de 30 mm",
    "ZAE902": "une hauteur de 940 mm, un entraxe de 800 mm et un diamètre de 30 mm",
    "ZAE751/1800": "une hauteur de 1800 mm, un entraxe de 1600 mm et un diamètre de 30 mm",
    "ZAE48/1800": "une hauteur de 1800 mm, un entraxe de 1600 mm et un diamètre de 30 mm",
}

# Nature grammaticale de la référence, par tableau du catalogue.
CHAP_NATURE = {
    "Monocolore": ("l'ensemble de béquilles", "m"),
    "Bicolore": ("l'ensemble de béquilles", "m"),
    "Bequille inox": ("la béquille inox", "f"),
    "Poussoirs inox": ("le poussoir", "m"),
    "Poignée encastrée ext": ("la poignée encastrée extérieure de référence", "f"),
    "Rosettes seules": ("la", "f"),
}

# Trois options chiffrées présentes au PDF et absentes de l'Excel (arbitrage A2).
CATALOGUE_PDF = [
    ("ferrage", "Ferrage 6 points automatique", 0, None, 12,
     "Le ferrage six points automatique est l'un des deux systèmes de ferrage livrés en "
     "standard sur la gamme HAM76. Il réunit deux crochets massifs, un pêne dormant, un "
     "pêne demi-tour Soft-Lock et deux pênes à déclenchement automatique Soft-Lock. Il est "
     "incompatible avec une gâche mécanique et avec une gâche électrique."),
    ("paumelles", "Changement de teinte des paumelles", 0, None, 12,
     "Les paumelles aluminium tubulaires de la gamme HAM76 comptent trois lames, un "
     "diamètre de 22 mm et une longueur de 185 mm. Elles sont disponibles en trois "
     "teintes : blanc, titane et noir. Le changement de teinte est proposé sans plus-value."),
    ("seuil", "Plinthe automatique, référence PA", 0, None, 13,
     "La plinthe automatique remplace le seuil : fixée sur l'ouvrant, elle est en position "
     "haute porte ouverte et descend en butée contre le sol à la fermeture. Le tarif "
     "signale que les valeurs thermiques sont réduites avec cette solution, que "
     "l'étanchéité n'est pas assurée, et qu'elle n'est disponible qu'en porte un vantail."),
]


def f_catalogue():
    f = Fichier("CATALOGUE_OPTIONS",
                "références d'options tarifées indépendamment du modèle en HAM76")
    exclus = defaultdict(int)
    for r in range(2, ws.max_row + 1):
        chap, tab = cell(r, "AV"), cell(r, "AW")
        if not chap or chap == "Fiche porte":
            continue
        if tab in CHAP_EXCLUS:
            exclus[tab] += 1
            continue
        desig = LIBELLE_CORRIGE.get(r) or LIBELLE_RESTITUE.get(r) or cell(r, "AX")
        if not desig:
            JOURNAL.append(f"[catalogue] ligne {r} sans désignation : écartée.")
            continue
        if cell(r, "AX") in CATALOGUE_EXCLUS:
            exclus["références non attestées au PDF"] += 1
            JOURNAL.append(f"[catalogue] référence {cell(r, 'AX')} tarifée dans l'Excel et "
                           f"absente du PDF publié : écartée, remontée au service Produits.")
            continue
        ht = fmt_euro(cell(r, "BA"))
        if ht is None:
            exclus["sans montant"] += 1
            continue
        ttc = fmt_euro(cell(r, "BB"))
        lib = CHAP_LIBELLE.get(tab, tab.lower())
        page = CHAP_PAGES.get((chap, tab), 72)
        titre = f"{PREFIXE} — Tarif catalogue {lib} : {desig}"
        detail = POUSSOIRS_DETAIL.get(cell(r, "AX"))
        phrase_detail = f" Ce poussoir présente {detail}." if detail else ""
        nature, genre = CHAP_NATURE.get(tab, ("la référence", "f"))
        e = "e" if genre == "f" else ""
        if to_num(cell(r, "BA")) == 0:
            corps = (f"Au catalogue des garnitures de la gamme HAM76, porte d'entrée "
                     f"monobloc aluminium, {nature} {desig} est proposé{e} sans "
                     f"plus-value.{phrase_detail} Cette option est indépendante du modèle de "
                     f"porte commandé.")
        else:
            corps = (f"Au catalogue des garnitures de la gamme HAM76, porte d'entrée "
                     f"monobloc aluminium, {nature} {desig} est tarifé{e} {ht} € HT, soit "
                     f"{ttc} € TTC.{phrase_detail} Ce montant est indépendant du modèle de "
                     f"porte commandé.")
        if tab == "Bicolore":
            corps += (" La béquille intérieure est une BDEL déclinée en version gauche ou "
                      "droite, en blanc, titane ou noir.")
        if tab == "Bequille inox":
            corps += (" La version BDE/Osec est celle équipée d'une rosette de sécurité.")
        if tab == "Poignée encastrée ext":
            corps += (" La rosace encastrée est toujours incluse en cas de poignée encastrée "
                      "extérieure, et cette poignée est toujours noire quelle que soit la "
                      "teinte de l'ouvrant.")
        if tab == "Rosettes seules":
            corps += (" La rosette standard inox de niveau de sécurité R20, référence R-IN, "
                      "est livrée de série sans plus-value.")
        if tab == "Poussoirs inox":
            corps += " La largeur minimale d'ouvrant pour un poussoir est de 800 mm."
        f.add(titre, corps, page)

    for cat, lib, ht_v, ttc_v, page, desc in CATALOGUE_PDF:
        titre = f"{PREFIXE} — Tarif catalogue {cat} : {lib}"
        montant = ("est proposé sans plus-value" if ht_v == 0
                   else f"est tarifé {fmt_euro(ht_v)} € HT")
        corps = (f"Au tarif de la gamme HAM76, porte d'entrée monobloc aluminium, "
                 f"« {lib} » {montant}. {desc} Ce montant est indépendant du modèle de porte "
                 f"commandé. Le tarif publié ne porte pour cette option qu'un montant hors "
                 f"taxes.")
        f.add(titre, corps, page)
        JOURNAL.append(f"[catalogue] option « {lib} » captée au PDF page {page}, "
                       f"absente de l'Excel.")

    for k, n in exclus.items():
        JOURNAL.append(f"[catalogue] écarté : {k} ({n} lignes).")
    return f


# =============================================================== 7. FAISABILITES


def f_faisabilites():
    f = Fichier("FAISABILITES",
                "faisabilités non tarifaires de la gamme HAM76")

    f.add(f"{PREFIXE} — Typologie des châssis réalisables et non réalisables",
          "Dans la gamme HAM76, porte d'entrée monobloc aluminium, sont réalisables la porte "
          "un vantail et la partie fixe vitrée, qu'il s'agisse d'un fixe latéral ou d'une "
          "imposte montés d'usine, ou d'une composition fixe et imposte à monter sur "
          "chantier. Ne sont pas réalisables la porte deux vantaux, les formes spéciales "
          "telles que le plein cintre, l'arc de cercle, l'anse de panier et le trapèze, la "
          "porte avec volet roulant, et l'ouverture extérieure.",
          P_TYPOLOGIE)

    f.add(f"{PREFIXE} — Sens d'ouverture de la porte et sens du panneau",
          "Sur la gamme HAM76, contrairement aux portes H81 et HA76, le sens du panneau est "
          "déterminé par le sens d'ouverture de la porte. Il n'est donc pas possible de "
          "sélectionner un sens A ou un sens B au passage de commande. Les sens réalisables "
          "sont présentés sur chaque page modèle du tarif.",
          P_TYPOLOGIE)

    f.add(f"{PREFIXE} — Limites dimensionnelles générales : deux jeux de valeurs divergents",
          "Le tarif HAM76 porte deux jeux de limites de fabrication en un vantail pour les "
          "mêmes dormants AL10101 et AL10108, et ils divergent. La page 9 annonce des "
          "limites valables pour tous les modèles de la gamme, de 810 mm à 1160 mm en "
          "largeur et de 2075 mm à 2280 mm en hauteur. Les pages modèles portent des limites "
          "de 798 mm à 1248 mm en largeur et de 1892 mm ou 2065 mm à 2445 mm en hauteur "
          "selon le modèle. Ces deux valeurs sont exposées telles quelles : la divergence "
          "est remontée au service Produits et n'est pas arbitrée. Toute demande "
          "dimensionnelle limite doit être confirmée auprès de l'usine.",
          "9 et 19 à 71")

    f.add(f"{PREFIXE} — Limites de réalisation en hauteur hors tout et gâches ponctuelles",
          "Le ferrage de la gamme HAM76 impose des limites de réalisation en hauteur hors "
          "tout propres à chaque dormant. Pour les dormants AL10101 en L68 et AL10108 en "
          "LZ108, la hauteur hors tout va de 1893 mm à 2586 mm en un vantail. Pour le "
          "dormant AL10100 en L70, elle va de 1895 mm à 2588 mm. Une gâche ponctuelle est "
          "requise pour une hauteur hors tout comprise entre 1906 mm et 2115 mm sur les "
          "dormants AL10101 et AL10108, et entre 1908 mm et 2117 mm sur le dormant AL10100.",
          12)

    f.add(f"{PREFIXE} — Paumelles : conception, réglages et poids de vantail admissible",
          "Les paumelles de la gamme HAM76 sont divisées en trois lames. La lame centrale, "
          "vissée dans l'ouvrant, est coincée entre les deux lames fixées dans le dormant, "
          "ce qui apporte un système anti-dégondage directement dans la paumelle. La "
          "référence PPE-5 admet un réglage horizontal de moins 3 mm à plus 3 mm, un réglage "
          "vertical de moins 3 mm à plus 4 mm, un réglage en compression de moins 0,4 mm à "
          "plus 0,4 mm, et un poids de vantail maximal de 160 kg.",
          13)

    f.add(f"{PREFIXE} — Seuil standard et restrictions de la plinthe automatique",
          "Le seuil standard de la gamme HAM76 porte la référence AS10100, mesure 20 mm de "
          "hauteur et existe en teinte grise ou noire. La plinthe automatique, référence PA, "
          "permet de se passer de seuil. Le tarif signale trois restrictions : les valeurs "
          "thermiques sont réduites avec cette solution, l'étanchéité n'est pas assurée, et "
          "la plinthe automatique n'est disponible qu'en porte un vantail.",
          13)

    f.add(f"{PREFIXE} — Couleur des accessoires selon les teintes intérieure et extérieure",
          "Le tarif HAM76 définit les teintes claires comme le blanc pur 9010 lisse "
          "brillant, le blanc signalisation 9016 granité mat, l'ivoire clair 1015 granité "
          "mat et le gris clair 7035 granité mat, et les teintes sombres comme les autres "
          "teintes RAL granité mat du groupe 1. Quelle que soit la combinaison des teintes "
          "intérieure et extérieure, la garniture est proposée en BDE inox ou BDEL blanc, "
          "titane ou noir, le seuil en gris ou noir, les paumelles 3D en blanc, titane ou "
          "noir, la barrette thermique en noir et le joint de frappe battant en gris.",
          16)

    f.add(f"{PREFIXE} — Réalisation des fixes latéraux, des impostes et des meneaux",
          "Dans la gamme HAM76, la dimension minimale de réalisation d'un fixe est de 300 mm "
          "pour tous les dormants, mais un fixe de 300 mm sur 300 mm est techniquement "
          "irréalisable. Le prix d'un fixe inclut un triple vitrage retardateur d'effraction. "
          "Le meneau dormant AL10111, référence MD94, n'est réalisable qu'en configuration en "
          "T : le meneau en croix est irréalisable. La plinthe AL10103J se monte sur une "
          "partie fixe avec seuil, pour éviter un accouplement de châssis.",
          P_FIXES)

    f.add(f"{PREFIXE} — Largeur de passage libre d'une porte un vantail",
          "La largeur de passage libre d'une porte HAM76 un vantail se calcule à partir de "
          "la largeur de fabrication, pour les dormants AL10101 et AL10108 avec l'ouvrant "
          "AL10104. Le tarif énonce le seuil suivant : si la largeur de fabrication est "
          "supérieure à 1009,75 mm, alors le passage libre sera supérieur à 830 mm. La "
          "formule complète et le schéma coté figurent en page 74 du tarif, à laquelle il "
          "convient de se reporter pour une largeur de passage précise.",
          74)

    f.add(f"{PREFIXE} — Configuration unique de la gamme : tous les modèles en un vantail",
          f"Les {NB_MODELES} modèles de la gamme HAM76 sont tarifés dans une configuration "
          f"unique, la porte un vantail. Aucun modèle ne dispose d'un prix en deux vantaux, "
          f"cette configuration étant déclarée irréalisable sur la gamme. Il n'existe donc "
          f"pas de modèle partiellement tarifé : la question du choix d'une configuration ne "
          f"se pose pas en HAM76.",
          P_TYPOLOGIE)
    return f


# =============================================================== 8. TRANSVERSES


def f_transverses():
    f = Fichier("TRANSVERSES",
                "existence et localisation des tarifs transverses HAM76, sans montant")

    f.add(f"{PREFIXE} — Existence et localisation des tarifs de l'offre couleurs, groupes 1 et 2",
          "L'offre couleurs de la gamme HAM76 est organisée en groupes de teintes. Le groupe "
          "1 réunit des teintes monocolores et bicolores en granité mat proposées sans "
          "plus-value. Le groupe 2 réunit des laquages aluminium, des teintes RAL granité "
          "mat supplémentaires, la gamme Futura et des combinaisons bicolores, et donne lieu "
          "à une plus-value exprimée en pourcentage du prix. Ce pourcentage n'est pas "
          "reproduit ici parce que son usage suppose un calcul : il se lit directement en "
          "page 14 du tarif, où figure aussi la liste complète des teintes de chaque groupe.",
          14)

    f.add(f"{PREFIXE} — Existence et localisation des tarifs du groupe sublimation et des RAL granités autres",
          "La gamme HAM76 propose deux offres de teintes complémentaires. Le groupe "
          "sublimation réunit des tons bois exclusifs, en monocolore et en bicolore, et le "
          "tarif signale que la sublimation n'est pas réalisée sur la ligne Tradition. Les "
          "RAL granités autres réunissent une longue liste de teintes monocolores et "
          "bicolores, sous réserve de faisabilité et de demande à l'usine. Ces deux offres "
          "donnent lieu à une plus-value exprimée en pourcentage, non reproduite ici : elle "
          "se lit en page 15 du tarif.",
          15)

    f.add(f"{PREFIXE} — Existence et localisation du tarif des fixes latéraux vitrés et impostes vitrées",
          "La gamme HAM76 dispose d'un tarif des fixes latéraux vitrés et des impostes "
          "vitrées, présenté sous forme de table à double entrée en largeur et en hauteur, "
          "de 300 mm à 2600 mm par pas de 100 mm, pour les teintes du groupe de laquage 1. "
          "Cette table n'est pas reproduite dans ce corpus : le tarif la donne en lecture "
          "directe sans énoncer de règle de lecture pour les cotes intermédiaires, si bien "
          "que toute demande hors cote ronde exigerait un arrondi. Le prix d'un fixe se lit "
          "donc directement en page 17 du tarif.",
          P_FIXES)

    f.add(f"{PREFIXE} — Existence et localisation des plus-values de vitrages pour fixes",
          "La gamme HAM76 propose des plus-values de vitrage sur les fixes, réparties en "
          "vitrages thermiques, phoniques, de sécurité et ornementaux, chacun identifié par "
          "une désignation de composition et un code vitrage, avec ses coefficients de "
          "transmission thermique et de transmission lumineuse. Ces plus-values sont "
          "exprimées au mètre carré de surface vitrée et ne sont donc pas reproduites ici : "
          "leur application suppose un calcul de surface. Elles se lisent en page 18 du "
          "tarif.",
          P_VITR_FIXES)

    f.add(f"{PREFIXE} — Existence et localisation du tarif de l'ensemble béquille intérieure et poussoir extérieur",
          "La gamme HAM76 permet d'associer une béquille intérieure à un poussoir extérieur. "
          "Dans cette configuration, la béquille intérieure est systématiquement une BDEL "
          "blanche, titane ou noire. Le tarif ne porte pas de prix propre à cet ensemble : "
          "il renvoie au prix du poussoir choisi, qui se lit au chapitre des poussoirs en "
          "page 73.",
          72)

    f.add(f"{PREFIXE} — Existence et localisation du journal des évolutions du tarif",
          "Le tarif HAM76 tient un journal de ses évolutions. Trois modifications y sont "
          "consignées pour l'édition du 4 mai 2026 : l'ajout de nouveaux modèles le 15 "
          "décembre 2025, une hausse générale appliquée à l'ensemble du tarif à compter du "
          "1er mai 2026, et la suppression des rainures sur les modèles Aude, Daphné et "
          "Victoire le 4 mai 2026. Le taux de la hausse générale n'est pas reproduit ici "
          "parce que son usage supposerait de l'appliquer à un prix : il se lit en page 75 "
          "du tarif. Les prix servis par ce corpus sont ceux de l'édition du 4 mai 2026 : un "
          "prix antérieur ne peut pas en être reconstitué et doit être lu sur l'édition "
          "correspondante du tarif.",
          75)
    return f


# =============================================================== exécution

def main():
    fichiers = [f_methode(), f_prix(), f_options(), f_caracteristiques(),
                f_compat(), f_catalogue(), f_faisabilites(), f_transverses()]
    total = 0
    print("Fichier                  Chunks")
    print("-" * 34)
    for f in fichiers:
        path, n = f.write()
        total += n
        print(f"{f.nom:<24} {n:>6}")
    print("-" * 34)
    print(f"{'TOTAL':<24} {total:>6}")

    with open(f"{OUTDIR}/journal_generation_HAM76.txt", "w", encoding="utf-8") as fh:
        fh.write("JOURNAL DE GÉNÉRATION — TARIF HAM76\n\n")
        for j in JOURNAL:
            fh.write(f"- {j}\n")
        fh.write("\nALERTES\n")
        if ALERTS:
            for a in ALERTS:
                fh.write(f"- {a}\n")
        else:
            fh.write("- aucune\n")

    print(f"\nAlertes : {len(ALERTS)}")
    for a in ALERTS:
        print("  !", a)


if __name__ == "__main__":
    main()
