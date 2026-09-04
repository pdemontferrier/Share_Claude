#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generateur des chunks Markdown du tarif FT84 (fenetre de toit PVC, collection TRYBA PVC).

Source primaire  : FT84_-infos-tarifs.xlsx (feuille Feuil1, 34 lignes de donnees)
Source controle  : Tarif_FT84_HT_28-05-2026.pdf et Tarif_FT84_TTC_28-05-2026.pdf (24 pages)

Regles appliquees : voir note_cadrage_migration_tarif_FT84_v1.md, regles F1 a F9.

Le generateur ne calcule aucun montant : chaque valeur servie est lue dans une cellule
de l'Excel ou relevee litteralement dans le PDF, et consignee au journal.
"""

import os, re, json
import openpyxl

_CANDIDATS = ["/home/claude/ft84/FT84_-infos-tarifs.xlsx",
              "/mnt/user-data/uploads/FT84_-infos-tarifs.xlsx"]
XLSX = next(p for p in _CANDIDATS if os.path.exists(p))
OUT = os.environ.get("FT84_OUT", "/home/claude/ft84/out")
PLAFOND = 200

GAMME = "FT84"
GAMME_NOM = "Fenêtre de toit PVC"   # libellé de gamme, identique aux titres et à Instructions
DESIGNATION = "Fenetre de toit PVC"
PREFIXE = "FT84 Fenêtre de toit PVC"
DOC_AFFICHE = "Tarif—FT84—HT—28-05-2026.pdf"
DOC_AFFICHE_TTC = "Tarif—FT84—TTC—28-05-2026.pdf"
DOC_YAML = "Tarif_FT84_HT_28-05-2026.pdf"
DOC_YAML_TTC = "Tarif_FT84_TTC_28-05-2026.pdf"
VERSION_DOC = "V.28/05/2026"
DATE_VALIDITE = "2026-05-28"

JOURNAL = {"colonnes_non_mappees": [], "lignes_exclues": [], "postes_geles": [],
           "unites_non_etablies": [], "discriminants_repris_du_pdf": [],
           "divergences_exposees": [], "pages_exclues": []}

# --------------------------------------------------------------------------
# Tables relevees dans le PDF (absentes de l'Excel). Chaque entree porte sa page.
# --------------------------------------------------------------------------

# Bandes de largeur du regime sur mesure, lues dans les en-tetes de la page 10.
LARGEURS_SM = [(460, 524), (525, 590), (591, 790), (791, 990),
               (991, 1080), (1081, 1190), (1191, 1310)]

# Bandes de hauteur du regime sur mesure, lues dans la colonne H de la page 10.
HAUTEURS_SM = [("740", 740, 890), ("891", 891, 1090), ("1091", 1091, 1290),
               ("1291", 1291, 1420), ("1421", 1421, 1510), ("1511", 1511, 1560)]

# Dimensions ponctuelles du regime stock, page 11.
LARGEURS_ST = [495, 605, 725, 887, 1085, 1285]
HAUTEURS_ST = [919, 1119, 1339, 1541]

# Abergement ardoises : nombre de toles par bande de hauteur, page 10 (sur mesure)
# et page 11 (stock). Le montant vient de l'Excel (colonnes HT / TTC scalaires),
# le libelle et le nombre de toles viennent du PDF, rattaches par le montant.
TOLES_SM = {"740": 7, "891": 8, "1091": 9, "1291": 10, "1421": 10, "1511": 11}
TOLES_ST = {919: 8, 1119: 9, 1339: 10, 1541: 11}

# Page du PDF ou figure chaque nature d'information.
PAGES = {"cotes_utiles": 6, "couverture": 6, "pente": 6, "vts_exclusif": 6,
         "lecture": 8, "composition_chassis": 8, "tva": 8,
         "cotes_reference": 9, "sur_mesure": 10, "stock": 11,
         "pv_couleur": 13, "couleurs_vts": 14, "vitrage_description": 16,
         "vitrage_pv": 17, "ventilation": 19, "motorisation": 20,
         "jumeles": 21, "bon_commande": 22, "evolutions": 23}

PRODUITS = {
    "FT84": {
        "titre": "de la fenêtre de toit FT84 seule",
        "intro": "de la fenêtre de toit FT84 seule",
    },
    "VTS": {
        "titre": "du volet de toit solaire TRYBA VTS",
        "intro": "du volet de toit solaire autonome TRYBA VTS",
    },
}

# --------------------------------------------------------------------------
# Utilitaires
# --------------------------------------------------------------------------

def mots(txt):
    return len(re.findall(r"\S+", txt))


def eur(v):
    """Montant en euros, sans separateur de milliers, comme les grilles du tarif."""
    return str(int(v))


def enumere(items):
    """Enumeration francaise : virgules puis « et » devant le dernier terme."""
    items = [str(i) for i in items]
    if len(items) == 1:
        return items[0]
    return ", ".join(items[:-1]) + " et " + items[-1]


def bande(mini, maxi):
    return "de %d à %d mm" % (mini, maxi)


class Fichier:
    """Accumule des chunks, numerote les SC depuis SC0002, controle le plafond."""

    def __init__(self, nom, sous_type, resume):
        self.nom = nom
        self.sous_type = sous_type
        self.resume = resume
        self.chunks = []
        self.sc = 2

    def ajoute(self, titre, page, corps, origine="originale", edition="HT"):
        corps = re.sub(r"\s+", " ", corps).strip()
        doc = DOC_AFFICHE if edition == "HT" else DOC_AFFICHE_TTC
        src = "*Source : %s, page %d — information %s — SC%04d*" % (
            doc, page, origine, self.sc)
        bloc = "## %s\n%s\n\n%s\n" % (titre, src, corps)
        n = mots("## " + titre + " " + src + " " + corps)
        if n > PLAFOND:
            raise ValueError("Plafond depasse (%d mots) : %s" % (n, titre))
        self.chunks.append({"sc": self.sc, "titre": titre, "page": page,
                            "corps": corps, "bloc": bloc, "mots": n})
        self.sc += 1

    def ecrit(self):
        fm = [
            "---",
            "document_source: %s" % DOC_YAML,
            "document_source_ttc: %s" % DOC_YAML_TTC,
            "type_document: tarif",
            "sous_type: %s" % self.sous_type,
            "gamme_code: %s" % GAMME,
            "gamme_nom: \"%s\"" % GAMME_NOM,
            "gammes_couvertes: [FT84]",
            "collection: \"TRYBA PVC\"",
            "materiau: PVC",
            "version_doc: \"%s\"" % VERSION_DOC,
            "date_validite: %s" % DATE_VALIDITE,
            "nb_chunks: %d" % len(self.chunks),
            "audiences: [ADV, commercial]",
            "---",
            "",
        ]
        corps = "\n".join(c["bloc"] for c in self.chunks)
        path = os.path.join(OUT, self.nom)
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(fm) + "\n" + corps)
        return path


def empaquete(elements, entete_mots, plafond=PLAFOND, sep=1):
    """Decoupe glouton d'une liste d'elements textuels sous le plafond de mots.

    Renvoie une liste de tranches (listes d'indices). La coupure est pilotee par
    le comptage des mots, jamais par une constante.
    """
    tranches, courante, poids = [], [], entete_mots
    for i, txt in enumerate(elements):
        m = mots(txt) + (sep if courante else 0)
        if courante and poids + m > plafond:
            tranches.append(courante)
            courante, poids = [], entete_mots
        courante.append(i)
        poids += m
    if courante:
        tranches.append(courante)
    return tranches


# --------------------------------------------------------------------------
# Lecture de l'Excel
# --------------------------------------------------------------------------

def lit_excel():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Feuil1"]
    entetes = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    col_sm_ht, col_sm_ttc, col_st_ht, col_st_ttc = {}, {}, {}, {}
    for idx, h in enumerate(entetes, start=1):
        if not h:
            continue
        h = str(h).strip()
        m = re.fullmatch(r"(HT|TTC) (\d+) à (\d+)", h)
        if m:
            cle = (int(m.group(2)), int(m.group(3)))
            (col_sm_ht if m.group(1) == "HT" else col_sm_ttc)[cle] = idx
            continue
        m = re.fullmatch(r"(HT|TTC) (\d+)", h)
        if m:
            (col_st_ht if m.group(1) == "HT" else col_st_ttc)[int(m.group(2))] = idx

    lignes = []
    for r in range(2, ws.max_row + 1):
        chap = ws.cell(r, 3).value
        haut = ws.cell(r, 11).value
        lignes.append({
            "r": r,
            "cle": ws.cell(r, 2).value,
            "chapitre": (chap or "").strip() if isinstance(chap, str) else chap,
            "tableau": (ws.cell(r, 4).value or "").strip(),
            "designation": (ws.cell(r, 5).value or "").strip(),
            "details": ws.cell(r, 6).value,
            "ht": ws.cell(r, 9).value,
            "ttc": ws.cell(r, 10).value,
            "hauteur": haut,
            "cells": {c: ws.cell(r, c).value for c in range(12, ws.max_column + 1)},
        })

    # Journal des colonnes vides ou constantes : aucune information ecartee en silence.
    for idx, h in enumerate(entetes, start=1):
        vals = [ws.cell(r, idx).value for r in range(2, ws.max_row + 1)]
        remplis = [v for v in vals if v not in (None, "")]
        if not remplis:
            JOURNAL["colonnes_non_mappees"].append(
                {"colonne": h, "motif": "entierement vide"})
        elif len(set(map(str, remplis))) == 1 and idx in (1, 7, 8):
            JOURNAL["colonnes_non_mappees"].append(
                {"colonne": h, "motif": "constante sans information",
                 "valeur": str(remplis[0])})

    return lignes, col_sm_ht, col_sm_ttc, col_st_ht, col_st_ttc


def decoupe_hauteur(txt):
    """'740 - FR84' -> ('740', 'FT84'). Quatre graphies de separateur coexistent
    et la coquille FR84 est corrigee vers FT84 : divergence de libelle, tranchee
    en faveur du PDF qui se declare document de reference (page 2)."""
    if txt is None:
        return None, None
    s = str(txt).strip()
    m = re.match(r"^\s*(\d+)\s*-?\s*(.*)$", s)
    if not m:
        return None, None
    h, prod = m.group(1), m.group(2).strip()
    if prod.upper() in ("FR84", "FT84"):
        if prod.upper() == "FR84":
            JOURNAL["divergences_exposees"].append(
                {"objet": "libelle Excel 'FR84'", "correction": "FT84",
                 "motif": "coquille de saisie, tranchee en faveur du PDF"})
        return h, "FT84"
    if prod.upper() == "VTS":
        return h, "VTS"
    if "val" in prod.lower():
        return h, "VAL_VIT"
    return h, prod


# --------------------------------------------------------------------------
# F1 — Methode, cotes et lecture du tarif
# --------------------------------------------------------------------------

def construit_methode():
    f = Fichier("Tarif_FT84_METHODE.md", "methode_et_cotes",
                "cotes, regimes tarifaires, lecture des grilles, vocabulaire")

    f.ajoute(
        "%s — Cotes de tarif et cotes de fabrication : une seule dimension à relever" % PREFIXE,
        PAGES["cotes_reference"],
        "Le tarif de la fenêtre de toit FT84 distingue par la notation les cotes de "
        "référence tarif, notées L exposant T et H exposant T, des cotes de commande et "
        "de fabrication, notées L et H. Il énonce cependant que les cotes de fabrication "
        "sont égales aux cotes de prix, elles-mêmes égales aux dimensions de l'existant "
        "diminuées du jeu de pose. Une seule dimension est donc à relever pour chiffrer "
        "et pour commander, contrairement à d'autres gammes où la cote de tarif et la "
        "cote de fabrication diffèrent. Toutes les cotes sont exprimées en millimètres.")

    f.ajoute(
        "%s — Prendre les dimensions du tableau intérieur, jamais celles du dormant existant" % PREFIXE,
        PAGES["cotes_reference"],
        "Le tarif de la fenêtre de toit FT84 avertit qu'il faut systématiquement prendre "
        "les dimensions des tableaux intérieurs, et qu'il ne s'agit pas du cadre dormant "
        "existant. C'est le point de vigilance principal du chiffrage de cette gamme : "
        "une cote relevée sur le dormant en place conduit à une dimension fausse, donc à "
        "un prix faux et à un châssis inadapté. Le bon de commande reprend d'ailleurs "
        "cette contrainte en intitulant ses deux colonnes de dimensions tableau intérieur, "
        "largeur et hauteur, en millimètres.")

    f.ajoute(
        "%s — Cotes utiles : passage dormant et clair de vitre intérieur" % PREFIXE,
        PAGES["cotes_utiles"],
        "À partir de la dimension de fabrication de la fenêtre de toit FT84, le tarif "
        "donne deux déductions à appliquer, en largeur comme en hauteur. Le calcul des "
        "dimensions de passage dormant retire deux fois 62 mm. Le calcul des dimensions "
        "de clair de vitre retire deux fois 96 mm. Le croquis associé indique une cote A "
        "de 34 mm et une cote B de 98 mm. Les dimensions de clair de vitre s'entendent "
        "hors joint. Ces déductions sont données par le tarif : leur application revient "
        "à l'ADV, qui les effectue lui-même.")

    f.ajoute(
        "%s — Deux régimes tarifaires distincts : dimensions sur mesure et dimensions stock" % PREFIXE,
        PAGES["sur_mesure"],
        "Le tarif de la fenêtre de toit FT84 comporte deux régimes qui ne se mélangent "
        "pas. Le régime dimensions sur mesure tarife par grille, une cellule couvrant une "
        "bande de largeur et une bande de hauteur, et ses montants sont des prix bruts. "
        "Le régime dimensions stock tarife un nombre limité de dimensions fixes, "
        "identifiées par un numéro de modèle, et ses montants sont des prix nets sur "
        "lesquels aucune remise d'achat ne s'applique. Les deux régimes n'ouvrent pas les "
        "mêmes options ni les mêmes coloris, et le prix d'une même option y diffère.")

    f.ajoute(
        "%s — Lecture des grilles sur mesure : une cellule couvre une bande de dimensions" % PREFIXE,
        PAGES["lecture"],
        "Dans le régime dimensions sur mesure de la fenêtre de toit FT84, les en-têtes de "
        "la grille portent directement l'intervalle couvert, sous la forme de 460 à 590 "
        "millimètres par exemple. Une cellule de prix ne correspond donc pas à une cote "
        "ponctuelle mais à une bande. Une dimension quelconque se lit dans la bande qui la "
        "contient, sans arrondi ni interpolation. La lecture se fait largeur puis hauteur. "
        "Les bandes de largeur tarifées vont de 460 à 1310 mm et les bandes de hauteur de "
        "740 à 1560 mm, chaque bande de hauteur n'ouvrant pas nécessairement toutes les "
        "bandes de largeur.")

    f.ajoute(
        "%s — Correspondance entre le numéro de modèle stock et ses dimensions" % PREFIXE,
        PAGES["stock"],
        "Dans le régime dimensions stock de la fenêtre de toit FT84, chaque dimension "
        "offerte porte un numéro de modèle à deux chiffres. Le premier chiffre désigne la "
        "hauteur : 1 pour 919 mm, 2 pour 1119 mm, 3 pour 1339 mm, 4 pour 1541 mm. Le "
        "second désigne la largeur : 1 pour 495 mm, 2 pour 605 mm, 3 pour 725 mm, 4 pour "
        "887 mm, 5 pour 1085 mm, 6 pour 1285 mm. Le modèle 23 vaut ainsi 725 mm de largeur "
        "sur 1119 mm de hauteur. Toutes les combinaisons ne sont pas offertes : dix-sept "
        "modèles seulement figurent au tarif.")

    f.ajoute(
        "%s — Ce que comprend le prix du châssis nu" % PREFIXE,
        PAGES["composition_chassis"],
        "Les prix des grilles de la fenêtre de toit FT84 sont des valeurs pour châssis nu. "
        "Ils comprennent le vitrage standard, le capotage extérieur en RAL 7043 ou "
        "RAL 8019, l'abergement en RAL 7043 ou RAL 8019, le pare-pluie sur quatre côtés "
        "sous forme d'écran sous-toiture, le pare-vapeur intérieur sur quatre côtés, les "
        "solins hauts et bas, et le sachet d'accessoires de pose comprenant clous et vis. "
        "Ce tarif concerne des fenêtres de toit en PVC blanc intérieur.")

    f.ajoute(
        "%s — Prix bruts en sur mesure, prix nets en stock" % PREFIXE,
        PAGES["stock"],
        "Les montants du régime dimensions sur mesure de la fenêtre de toit FT84 sont "
        "présentés comme des prix bruts. Ceux du régime dimensions stock sont des prix "
        "nets, valables uniquement pour la fenêtre de toit, et le tarif précise qu'aucune "
        "remise d'achat ne sera appliquée sur cette grille de prix. Les deux régimes ne "
        "sont donc pas comparables terme à terme, et un montant de l'un ne se substitue "
        "jamais à un montant de l'autre.")

    f.ajoute(
        "%s — Éditions hors taxes et toutes taxes comprises du tarif" % PREFIXE,
        PAGES["lecture"],
        "Le tarif de la fenêtre de toit FT84 existe en deux éditions de même date, une "
        "édition hors taxes et une édition toutes taxes comprises, de pagination "
        "identique sur leurs vingt-quatre pages. Les montants hors taxes et toutes taxes "
        "comprises servis dans les chunks de ce corpus sont lus chacun dans son édition, "
        "à la même page. Le montant toutes taxes comprises ne se déduit d'aucun "
        "coefficient appliqué au montant hors taxes : il est transcrit, jamais calculé. "
        "Les prix des deux éditions s'entendent hors éco-participation.")

    f.ajoute(
        "%s — Vocabulaire propre au tarif de la fenêtre de toit" % PREFIXE,
        PAGES["composition_chassis"],
        "L'abergement désigne l'ensemble des pièces assurant le raccord d'étanchéité entre "
        "la fenêtre de toit et la couverture. Les solins sont les pièces de raccord hautes "
        "et basses. Le capotage extérieur est composé de tôles en aluminium prélaqué. Le "
        "pare-pluie est l'écran sous-toiture posé sur les quatre côtés. Le sigle VTS "
        "désigne le volet de toit solaire TRYBA, volet roulant autonome. La mention "
        "Val. Vit. désigne une valeur vitrage servant de base de calcul. Le verrouillage "
        "par crochets massifs et les gâches fixées dans les armatures appartiennent au "
        "vocabulaire propre de cette fenêtre.")

    return f


# --------------------------------------------------------------------------
# F2 — Prix des chassis, regime dimensions sur mesure
# --------------------------------------------------------------------------

def construit_prix_sur_mesure(lignes, c_ht, c_ttc):
    f = Fichier("Tarif_FT84_PRIX_SUR_MESURE.md", "prix_chassis_sur_mesure",
                "grilles dimensionnelles du regime sur mesure")

    index = {}
    for L in lignes:
        h, prod = decoupe_hauteur(L["hauteur"])
        if h and prod and "sur mesure" in str(L["chapitre"]).lower():
            index[(h, prod)] = L

    for cle, hmin, hmax in HAUTEURS_SM:
        for prod in ("FT84", "VTS"):
            L = index.get((cle, prod))
            if L is None:
                continue
            valeurs = []
            for (lmin, lmax) in LARGEURS_SM:
                vh = L["cells"].get(c_ht[(lmin, lmax)])
                vt = L["cells"].get(c_ttc[(lmin, lmax)])
                if vh is None or vt is None:
                    continue  # anti-fantome : aucune cellule inventee
                valeurs.append(((lmin, lmax), int(vh), int(vt)))
            if not valeurs:
                continue

            phrases = ["largeur %s, %s € HT et %s € TTC"
                       % (bande(lmin, lmax), eur(vh), eur(vt))
                       for (lmin, lmax), vh, vt in valeurs]

            couverte = {b for b, _, _ in valeurs}
            absentes = [b for b in LARGEURS_SM if b not in couverte]
            clauses = []
            if prod == "VTS" and (460, 524) in absentes:
                clauses.append(
                    "La largeur de 460 à 524 mm porte au tarif la mention Impossible : "
                    "le volet n'est pas réalisable en deçà de 525 mm de largeur.")
                absentes = [b for b in absentes if b != (460, 524)]
            if absentes:
                seuil = min(b[0] for b in absentes)
                clauses.append(
                    "Au-delà de %d mm de largeur, cette bande de hauteur n'est pas "
                    "tarifée." % (seuil - 1))
            clause = " ".join(clauses)

            p = PRODUITS[prod]
            base_titre = ("%s — Tarif sur mesure %s, hauteur %s"
                          % (PREFIXE, p["titre"], bande(hmin, hmax)))
            intro = ("Dans le régime dimensions sur mesure, à une hauteur de "
                     "fabrication %s, le prix brut %s s'établit ainsi selon la largeur "
                     "de fabrication : " % (bande(hmin, hmax), p["intro"]))
            queue = ("Chaque dimension se lit dans la bande qui la contient ; aucun prix "
                     "n'est interpolé.")

            entete = (mots("## " + base_titre) + 10 + mots(intro)
                      + mots(clause) + mots(queue) + 3)
            tranches = empaquete(phrases, entete)

            for t in tranches:
                sel = [valeurs[i] for i in t]
                if len(tranches) > 1:
                    titre = base_titre + ", largeurs de %d à %d mm" % (
                        sel[0][0][0], sel[-1][0][1])
                else:
                    titre = base_titre
                corps = intro + " ; ".join(phrases[i] for i in t) + ". "
                if t is tranches[-1]:
                    corps += clause + " "
                corps += queue
                f.ajoute(titre, PAGES["sur_mesure"], corps)
    return f


# --------------------------------------------------------------------------
# F3 — Prix des chassis, regime dimensions stock
# --------------------------------------------------------------------------

def code_modele(hauteur, largeur):
    return int("%d%d" % (HAUTEURS_ST.index(hauteur) + 1, LARGEURS_ST.index(largeur) + 1))


def libelle_codes(codes):
    """Enumere les codes modele ; ne condense en intervalle que s'ils sont contigus."""
    if len(codes) == 1:
        return "modèle %d" % codes[0]
    contigu = all(b - a == 1 for a, b in zip(codes, codes[1:]))
    if contigu and len(codes) > 2:
        return "modèles %d à %d" % (codes[0], codes[-1])
    return "modèles " + enumere(codes)


def construit_prix_stock(lignes, c_ht, c_ttc):
    f = Fichier("Tarif_FT84_PRIX_STOCK.md", "prix_chassis_stock",
                "postes forfaitaires par code de dimension normalisee")

    index = {}
    for L in lignes:
        h, prod = decoupe_hauteur(L["hauteur"])
        if h and prod and "stock" in str(L["chapitre"]).lower():
            index[(int(h), prod)] = L

    for haut in HAUTEURS_ST:
        for prod in ("FT84", "VTS"):
            L = index.get((haut, prod))
            if L is None:
                continue
            valeurs = []
            for larg in LARGEURS_ST:
                vh = L["cells"].get(c_ht[larg])
                vt = L["cells"].get(c_ttc[larg])
                if vh is None or vt is None:
                    continue  # anti-fantome
                valeurs.append((larg, code_modele(haut, larg), int(vh), int(vt)))
            if not valeurs:
                continue

            JOURNAL["discriminants_repris_du_pdf"].append(
                {"objet": "numeros de modele stock", "page": PAGES["stock"],
                 "hauteur": haut, "produit": prod,
                 "codes": [v[1] for v in valeurs],
                 "rattachement": "par la coordonnee hauteur x largeur, le code figurant "
                                 "dans la meme colonne que le prix ; le rattachement par "
                                 "le montant est impossible, plusieurs modeles portant "
                                 "le meme montant"})

            phrases = ["le modèle %d, largeur %d mm, à %s € HT et %s € TTC"
                       % (code, larg, eur(vh), eur(vt))
                       for larg, code, vh, vt in valeurs]

            couverte = {v[0] for v in valeurs}
            absentes = [l for l in LARGEURS_ST if l not in couverte]
            clauses = []
            # La mention « Impossible » n'est portee au tarif que la ou une dimension
            # stock existe : elle suppose un modele offert a cette largeur. La ou aucun
            # modele n'existe, la case est vide sans mention et la clause serait inventee.
            offert_ft84 = {larg for larg in LARGEURS_ST
                           if index[(haut, "FT84")]["cells"].get(c_ht[larg]) is not None}
            if prod == "VTS" and 495 in absentes and 495 in offert_ft84:
                clauses.append(
                    "La largeur de 495 mm porte au tarif la mention Impossible : le volet "
                    "n'est pas réalisable en deçà de 525 mm de largeur.")
                absentes = [l for l in absentes if l != 495]
            if absentes:
                clauses.append(
                    "%s de %s mm %s offerte%s en dimensions stock à cette hauteur et ne "
                    "porte%s aucun numéro de modèle."
                    % ("La largeur" if len(absentes) == 1 else "Les largeurs",
                       enumere(absentes),
                       "n'est pas" if len(absentes) == 1 else "ne sont pas",
                       "" if len(absentes) == 1 else "s",
                       "" if len(absentes) == 1 else "nt"))
            clause = " ".join(clauses)

            p = PRODUITS[prod]
            base_titre = "%s — Tarif stock %s, hauteur %d mm" % (PREFIXE, p["titre"], haut)
            intro = ("Dans le régime dimensions stock, pour une hauteur de fabrication "
                     "de %d mm, les prix nets %s sont les suivants : "
                     % (haut, p["intro"]))
            queue = ("Ces prix sont nets : aucune remise d'achat n'est appliquée sur "
                     "cette grille.")

            entete = (mots("## " + base_titre) + 16 + mots(intro)
                      + mots(clause) + mots(queue) + 3)
            tranches = empaquete(phrases, entete)

            for t in tranches:
                codes = [valeurs[i][1] for i in t]
                titre = base_titre + ", " + libelle_codes(codes)
                corps = intro + " ; ".join(phrases[i] for i in t) + ". "
                if t is tranches[-1] and clause:
                    corps += clause + " "
                corps += queue
                f.ajoute(titre, PAGES["stock"], corps)
    return f


# --------------------------------------------------------------------------
# F4 — Options et plus-values forfaitaires
# --------------------------------------------------------------------------

def construit_options(lignes):
    f = Fichier("Tarif_FT84_OPTIONS.md", "options_et_plus_values",
                "postes forfaitaires chiffres en euros")

    # -- Abergement ardoises, regime sur mesure. Le montant vient des colonnes
    #    scalaires HT / TTC de l'Excel ; le libelle et le nombre de toles sont
    #    releves page 10 et rattaches par le montant.
    vus = set()
    for cle, hmin, hmax in HAUTEURS_SM:
        L = None
        for x in lignes:
            h, prod = decoupe_hauteur(x["hauteur"])
            if h == cle and prod == "FT84" and "sur mesure" in str(x["chapitre"]).lower():
                L = x
                break
        if L is None:
            continue
        ht, ttc, toles = int(L["ht"]), int(L["ttc"]), TOLES_SM[cle]
        if (ht, ttc) in vus:
            continue
        vus.add((ht, ttc))
        bandes = [b for b in HAUTEURS_SM
                  if TOLES_SM[b[0]] == toles and int(
                      [x for x in lignes
                       if decoupe_hauteur(x["hauteur"]) == (b[0], "FT84")
                       and "sur mesure" in str(x["chapitre"]).lower()][0]["ht"]) == ht]
        libelle_b = enumere([bande(b[1], b[2]) for b in bandes])
        JOURNAL["unites_non_etablies"].append(
            {"poste": "abergement ardoises sur mesure, %d tôles" % toles,
             "page": PAGES["sur_mesure"],
             "motif": "le tarif juxtapose un nombre de tôles droite et gauche et un "
                      "montant, sans dire si le montant couvre l'ensemble ou la tôle"})
        f.ajoute(
            "%s — Abergement ardoises en dimensions sur mesure, hauteur %s"
            % (PREFIXE, libelle_b),
            PAGES["sur_mesure"],
            "Pour une couverture en ardoises, le tarif de la fenêtre de toit FT84 porte "
            "un abergement ardoises spécifique. Pour une hauteur de fabrication %s, il "
            "compte %d tôles à droite et %d tôles à gauche, au prix brut de %s € HT et "
            "%s € TTC. Le tarif ne précise pas si ce montant couvre l'ensemble des tôles "
            "ou la tôle prise isolément : l'unité de facturation n'est pas établie et se "
            "vérifie page %d du tarif."
            % (libelle_b, toles, toles, eur(ht), eur(ttc), PAGES["sur_mesure"]))

    # -- Abergement ardoises, regime stock
    for haut in HAUTEURS_ST:
        L = None
        for x in lignes:
            h, prod = decoupe_hauteur(x["hauteur"])
            if h == str(haut) and prod == "FT84" and "stock" in str(x["chapitre"]).lower():
                L = x
                break
        if L is None:
            continue
        ht, ttc, toles = int(L["ht"]), int(L["ttc"]), TOLES_ST[haut]
        JOURNAL["unites_non_etablies"].append(
            {"poste": "abergement ardoises stock, %d tôles" % toles,
             "page": PAGES["stock"], "motif": "unite de facturation non enoncee"})
        f.ajoute(
            "%s — Abergement ardoises en dimensions stock, hauteur %d mm" % (PREFIXE, haut),
            PAGES["stock"],
            "Pour une couverture en ardoises, le tarif de la fenêtre de toit FT84 porte "
            "un abergement ardoises spécifique au régime dimensions stock. Pour une "
            "hauteur de fabrication de %d mm, il compte %d tôles à droite et %d tôles à "
            "gauche, au prix net de %s € HT et %s € TTC. Le tarif ne précise pas si ce "
            "montant couvre l'ensemble des tôles ou la tôle prise isolément : l'unité de "
            "facturation n'est pas établie et se vérifie page %d du tarif. Le croquis de "
            "la tôle d'abergement pour couverture ardoises, sur la même page, porte les "
            "cotes 330 mm, 138,50 mm et 137,50 mm."
            % (haut, toles, toles, eur(ht), eur(ttc), PAGES["stock"]))

    # -- Postes forfaitaires hors grille, lus directement dans l'Excel
    forfaits = {
        ("Couleurs", "PV couleur", "Dimensions sur-mesure"): (
            "%s — Plus-value du décor intérieur PVC plaxé Sierra en dimensions sur mesure" % PREFIXE,
            PAGES["pv_couleur"],
            "Le décor intérieur standard de la fenêtre de toit FT84 est le PVC blanc. Le "
            "PVC plaxé Sierra est proposé en plus-value. En dimensions sur mesure, cette "
            "plus-value est de {ht} € HT et {ttc} € TTC par fenêtre FT84. Le montant "
            "s'entend pour une fenêtre : le total revient à l'ADV, qui le multiplie par "
            "le nombre de fenêtres concernées. La teinte Sierra s'accompagne d'habillages "
            "intérieurs Sierra, d'une poignée brune et d'une poignée de tirage brune."),
        ("Couleurs", "PV couleur", "Dimensions stock"): (
            "%s — Plus-value du décor intérieur PVC plaxé Sierra en dimensions stock" % PREFIXE,
            PAGES["pv_couleur"],
            "Le décor intérieur standard de la fenêtre de toit FT84 est le PVC blanc. En "
            "dimensions stock, la plus-value pour le PVC plaxé Sierra est de {ht} € HT "
            "net et {ttc} € TTC net par fenêtre FT84. Le montant s'entend pour une "
            "fenêtre : le total revient à l'ADV. Cette plus-value diffère de celle du "
            "régime dimensions sur mesure et ne s'y substitue pas."),
        ("Options", "Ouverture motorisée", "Kit motorisation"): (
            "%s — Prix du kit de motorisation solaire et radio pour ouverture motorisée" % PREFIXE,
            PAGES["motorisation"],
            "Le kit de motorisation solaire et radio pour l'ouverture motorisée de la "
            "fenêtre de toit FT84 est tarifé {ht} € HT et {ttc} € TTC. Le montant "
            "s'entend pour un kit. Le kit comprend un moteur 12 V à fixer sur la traverse "
            "basse du dormant à l'aide de deux platines fournies, une cellule solaire à "
            "fixer sur la traverse basse du dormant, un récepteur, une télécommande et "
            "une batterie. Les pré-perçages de fixation sont faits en usine."),
        ("Options", "Châssis jumelés", "Dimensions sur-mesure"): (
            "%s — Prix de la tôle de jumelage horizontal en dimensions sur mesure" % PREFIXE,
            PAGES["jumeles"],
            "Le jumelage horizontal de deux châssis de fenêtre de toit FT84 requiert une "
            "tôle de jumelage. En dimensions sur mesure, cette tôle est tarifée {ht} € HT "
            "et {ttc} € TTC. Le montant s'entend pour une tôle, donc pour un jumelage : "
            "le total revient à l'ADV. La largeur entre tableaux admise pour un jumelage "
            "horizontal est comprise entre 140 mm au minimum et 460 mm au maximum."),
        ("Options", "Châssis jumelés", "Dimensions stock"): (
            "%s — Prix de la tôle de jumelage horizontal en dimensions stock" % PREFIXE,
            PAGES["jumeles"],
            "Le jumelage horizontal de deux châssis de fenêtre de toit FT84 requiert une "
            "tôle de jumelage. En dimensions stock, cette tôle est tarifée {ht} € HT et "
            "{ttc} € TTC. Le montant s'entend pour une tôle, donc pour un jumelage. Cette "
            "valeur diffère de celle du régime dimensions sur mesure et ne s'y substitue "
            "pas. La largeur entre tableaux admise pour un jumelage horizontal est "
            "comprise entre 140 mm au minimum et 460 mm au maximum."),
    }

    for L in lignes:
        if not L["chapitre"] or "Grille" in L["tableau"]:
            continue
        tab = L["tableau"].replace("motoriséé", "motorisée")
        cle = (L["chapitre"].strip(), tab.strip(), L["designation"].strip())
        if cle not in forfaits:
            JOURNAL["lignes_exclues"].append(
                {"ligne": L["r"], "motif": "poste forfaitaire non cartographie", "cle": str(cle)})
            continue
        titre, page, gabarit = forfaits[cle]
        if L["ht"] is None or L["ttc"] is None:
            JOURNAL["lignes_exclues"].append(
                {"ligne": L["r"], "motif": "montant HT ou TTC absent", "cle": str(cle)})
            continue
        f.ajoute(titre, page,
                 gabarit.format(ht=eur(int(L["ht"])), ttc=eur(int(L["ttc"]))))
        if "motoriséé" in L["tableau"]:
            JOURNAL["divergences_exposees"].append(
                {"objet": "libelle Excel 'Ouverture motoriséé'",
                 "correction": "Ouverture motorisée",
                 "motif": "coquille de saisie, tranchee en faveur du PDF"})
    return f


# --------------------------------------------------------------------------
# F5 — Faisabilites et restrictions (aucun montant)
# --------------------------------------------------------------------------

def construit_faisabilites():
    f = Fichier("Tarif_FT84_FAISABILITES.md", "faisabilites_et_restrictions",
                "restrictions et impossibilites, sans aucun montant")

    f.ajoute("%s — Volet de toit solaire impossible en deçà de 525 mm de largeur" % PREFIXE,
             PAGES["sur_mesure"],
             "Le tarif de la fenêtre de toit FT84 avertit que le volet roulant est "
             "impossible pour toute fenêtre de toit dont la largeur est inférieure à "
             "525 mm. La ligne du volet de toit solaire TRYBA VTS porte pour cette raison "
             "la mention Impossible dans les deux régimes : en dimensions sur mesure sur "
             "la bande de 460 à 524 mm, pour toutes les hauteurs ; en dimensions stock sur "
             "la largeur de 495 mm, aux hauteurs de 919 et 1119 mm, seules hauteurs où "
             "cette largeur est offerte. Cette case n'est pas une gratuité, c'est une "
             "impossibilité produit.")

    f.ajoute("%s — Types de couverture admis et exclus" % PREFIXE, PAGES["couverture"],
             "Le tarif de la fenêtre de toit FT84 déclare la pose réalisable sur trois "
             "types de couverture : la tuile plate, la tuile mécanique et l'ardoise. Il "
             "déclare la pose non réalisable sur tuile canal, dite aussi romane, ainsi que "
             "sur bac acier, sur zinc et sur bardeau bitumineux. Le bon de commande ne "
             "propose d'ailleurs que les trois couvertures admises. Une divergence interne "
             "est signalée sur la tuile romane, dont le même bloc de la page 6 donne une "
             "pente minimale alors que la liste des couvertures la déclare exclue.")

    f.ajoute("%s — Pente de toit minimale et maximale" % PREFIXE, PAGES["pente"],
             "Le tarif de la fenêtre de toit FT84 fixe une pente de toit minimale de 20 "
             "degrés et une pente maximale de 70 degrés. Il ajoute qu'en cas de tuile "
             "romaine, la pente minimale est de 25 degrés, ce qui contredit la liste des "
             "types de couverture de la même page, où la tuile canal romane est déclarée "
             "non réalisable. Les deux énoncés sont rapportés tels quels, sans arbitrage. "
             "Le bon de commande demande la pente en degrés pour chaque position.")

    f.ajoute("%s — Restrictions de l'offre en dimensions stock" % PREFIXE, PAGES["stock"],
             "Le tarif de la fenêtre de toit FT84 limite l'offre en dimensions stock aux "
             "possibilités suivantes : les deux coloris intérieurs de la FT84, le RAL 7043 "
             "pour l'extérieur de la FT84 ainsi que pour le coffre et les coulisses du "
             "volet TRYBA VTS, le RAL 7016 pour le tablier du volet, et le vitrage "
             "standard pour la FT84. Toute autre combinaison relève du régime dimensions "
             "sur mesure.")

    f.ajoute("%s — Dimensions non offertes en dimensions stock" % PREFIXE, PAGES["stock"],
             "Le régime dimensions stock de la fenêtre de toit FT84 n'offre pas toutes les "
             "combinaisons de hauteur et de largeur. À 919 mm de hauteur, la largeur "
             "1085 mm n'est pas offerte. À 1119 mm de hauteur, la largeur 1285 mm n'est "
             "pas offerte. À 1339 mm de hauteur, la largeur 495 mm n'est pas offerte. À "
             "1541 mm de hauteur, seules les largeurs 725 mm et 887 mm sont offertes. Ces "
             "cases vides ne portent aucun numéro de modèle et ne sont pas commandables en "
             "stock ; elles relèvent du régime dimensions sur mesure.")

    f.ajoute("%s — Enveloppe dimensionnelle tarifée en dimensions sur mesure" % PREFIXE,
             PAGES["sur_mesure"],
             "En dimensions sur mesure, le tarif de la fenêtre de toit FT84 couvre les "
             "largeurs de fabrication de 460 à 1310 mm et les hauteurs de fabrication de "
             "740 à 1560 mm. Toutes les bandes de largeur ne sont pas ouvertes à toutes "
             "les hauteurs : au-delà de 1420 mm de hauteur, les largeurs supérieures à "
             "1080 mm ne sont pas tarifées. Une dimension hors de ces plages n'est pas "
             "tarifée ; le tarif énonce que toute construction qui ne peut être traitée à "
             "l'aide du tarif n'est pas réalisable.")

    f.ajoute("%s — Restrictions de l'ouverture motorisée" % PREFIXE, PAGES["motorisation"],
             "Le système d'ouverture motorisée de la fenêtre de toit FT84 permet une "
             "ouverture en projection et s'adresse aux fenêtres de toit difficilement "
             "accessibles, cage d'escalier ou mezzanine par exemple. Le tarif avertit que "
             "ce système ne pourra pas être proposé sur une FT84 déjà posée, ni sur une "
             "autre fenêtre de toit que la TRYBA FT84. La fenêtre est systématiquement "
             "livrée avec une poignée montée et le kit de motorisation dans un colis "
             "accessoire à part ; cette poignée reste nécessaire pour retourner le vantail "
             "et nettoyer la face extérieure du vitrage.")

    f.ajoute("%s — Le volet de toit solaire TRYBA VTS ne se monte que sur la FT84" % PREFIXE,
             PAGES["vts_exclusif"],
             "Le volet de toit solaire TRYBA VTS proposé avec la fenêtre de toit TRYBA "
             "FT84 a été développé spécifiquement pour cette fenêtre et ne pourra donc se "
             "monter que sur la TRYBA FT84. Il est proposé uniquement avec une "
             "motorisation solaire SIMU IO et une commande radio compatible Somfy IO. Le "
             "tarif précise par ailleurs que ce volet n'est pas considéré comme étant "
             "occultant à cent pour cent.")

    f.ajoute("%s — Jumelage horizontal et jumelage vertical des châssis" % PREFIXE,
             PAGES["jumeles"],
             "Le tarif de la fenêtre de toit FT84 prévoit deux montages de châssis "
             "jumelés. Le montage horizontal admet une largeur entre tableaux de 140 mm "
             "au minimum et 460 mm au maximum, et requiert une tôle de jumelage qui est "
             "tarifée. Le montage vertical demande une hauteur minimale entre tableaux de "
             "270 mm ; le tarif ne porte aucune tôle ni aucun montant pour ce montage. "
             "Cette absence est signalée telle quelle et ne doit pas être comblée par le "
             "montant du jumelage horizontal.")

    f.ajoute("%s — Vitrages décoratifs indisponibles sur la gamme" % PREFIXE,
             PAGES["vitrage_pv"],
             "Pour des raisons de tenue mécanique des vitrages, les vitrages décoratifs ne "
             "sont pas disponibles pour la gamme de fenêtre de toit TRYBA FT84. Le tarif "
             "ne propose que les compositions thermiques standard, et ne porte aucune "
             "plus-value de vitrage pour cette gamme.")

    f.ajoute("%s — Coloris réservés au régime dimensions sur mesure" % PREFIXE,
             PAGES["couleurs_vts"],
             "Plusieurs coloris de la fenêtre de toit FT84 et du volet de toit solaire "
             "TRYBA VTS ne sont disponibles qu'en dimensions sur mesure. C'est le cas du "
             "RAL 8019 brun gris pour le capotage extérieur de la fenêtre ainsi que pour "
             "le coffre et les coulisses du volet, et du RAL 7047 télégris pour le tablier "
             "à lames aluminium du volet. Le RAL 7043 gris signalisation B pour le "
             "capotage, le coffre et les coulisses, et le RAL 7016 gris anthracite pour le "
             "tablier, restent disponibles dans les deux régimes.")

    f.ajoute("%s — Ventilation manuelle et absence de ventilation permanente" % PREFIXE,
             PAGES["ventilation"],
             "La fenêtre de toit TRYBA FT84 n'est pas pourvue d'un système de ventilation "
             "permanent du type grille d'entrée d'air Mini ESEA ou ISOLA2. Elle est "
             "cependant équipée d'une gâche à multi-étage permettant une ventilation "
             "manuelle, avec trois positions : FT84 verrouillée sans ventilation, "
             "ventilation de nuit à petit débit, et ventilation maximum à grand débit. La "
             "sécurité est garantie en position ventilation de nuit comme en position "
             "ventilation maximum.")

    f.ajoute("%s — Étanchéité non garantie en position de ventilation" % PREFIXE,
             PAGES["ventilation"],
             "Le tarif de la fenêtre de toit FT84 avertit qu'il faut systématiquement "
             "remettre la poignée en position fermée, repérée zéro, pour enclencher les "
             "crochets dans l'une des trois positions des gâches. Il n'y a pas de garantie "
             "d'étanchéité dans les cas de ventilation de nuit et de ventilation maximum. "
             "La position zéro correspond aux crochets sortis, la position 1 aux crochets "
             "rentrés pour une ouverture en projection, la position 2 à l'ouverture en "
             "rotation pour la mise en position de nettoyage.")

    f.ajoute("%s — Le dormant de rénovation ne porte aucun montant au tarif" % PREFIXE,
             PAGES["sur_mesure"],
             "La fiche info produit de la fenêtre de toit FT84 décrit un dormant "
             "rénovation spécial permettant une pose sans dégâts ni travaux intérieurs, en "
             "taille standard ou sur mesure. Le tarif ne porte pour ce dormant ni ligne "
             "dédiée, ni plus-value, ni mention de gratuité. Cette absence est signalée "
             "telle quelle : elle ne doit être interprétée ni comme une inclusion au prix "
             "du châssis nu, ni comme une impossibilité.")

    f.ajoute("%s — Abergement : divergence entre la composition du prix et le poste ardoises" % PREFIXE,
             PAGES["composition_chassis"],
             "Le tarif de la fenêtre de toit FT84 présente deux énoncés qu'il ne concilie "
             "pas. La légende de la page 8 range l'abergement en RAL 7043 ou RAL 8019 "
             "parmi ce que comprend le prix du châssis nu. Les grilles des pages 10 et 11 "
             "portent en revanche un abergement ardoises tarifé séparément, exprimé en "
             "nombre de tôles droite et gauche. Le tarif n'indique pas si l'abergement "
             "inclus couvre uniquement les couvertures en tuile. Les deux énoncés sont "
             "rapportés avec leur page, sans arbitrage.")

    f.ajoute("%s — Pression conventionnelle retenue pour le dimensionnement des vitrages" % PREFIXE,
             PAGES["vitrage_description"],
             "Selon les dimensions et selon les prescriptions du DTU, le tarif de la "
             "fenêtre de toit FT84 a tenu compte d'une pression conventionnelle de "
             "800 Pascals, correspondant au classement VA2. Le cas échéant, des glaces "
             "plus épaisses ont été prévues et tarifées. Tout calcul selon une pression "
             "conventionnelle supérieure à 800 Pascals devra faire l'objet d'une demande "
             "spéciale.")

    return f


# --------------------------------------------------------------------------
# F6 — Transverses (aucun montant)
# --------------------------------------------------------------------------

def construit_transverses():
    f = Fichier("Tarif_FT84_TRANSVERSES.md", "transverses",
                "existence et localisation, sans aucun montant")

    f.ajoute("%s — Vitrage standard sans volet de toit solaire" % PREFIXE,
             PAGES["vitrage_pv"],
             "Sans volet de toit TRYBA VTS, la fenêtre de toit FT84 reçoit en standard un "
             "vitrage thermique Isol'4 de code 383.16TN.613, composé d'un verre intérieur "
             "33/2 de 6 mm Isol'4, d'une lame de 16 mm et d'un verre extérieur de 4 mm "
             "TRYBA SUN+. Son coefficient Ug est de 1,0 W/m².K, son épaisseur totale de "
             "26 mm, sa surface maximale de 2,81 m². Sa transmission lumineuse est de 35 % "
             "et son facteur solaire de 20 %. Ce vitrage est marqué standard et ne porte "
             "aucune plus-value.")

    f.ajoute("%s — Vitrage standard avec volet de toit solaire" % PREFIXE,
             PAGES["vitrage_pv"],
             "Avec volet de toit TRYBA VTS, la fenêtre de toit FT84 reçoit un vitrage "
             "thermique différent, de code 383.16TN.386, composé d'un verre intérieur 33/2 "
             "de 6 mm Isol'4, d'une lame de 16 mm et d'un verre extérieur de 4 mm sans "
             "couche TRYBA SUN+. Son coefficient Ug est de 1,0 W/m².K, son épaisseur "
             "totale de 26 mm, sa surface maximale de 2,81 m². Sa transmission lumineuse "
             "est de 73 % et son facteur solaire de 47 %. La présence du volet change donc "
             "la composition du vitrage.")

    f.ajoute("%s — Aucune plus-value de vitrage au tarif de la gamme" % PREFIXE,
             PAGES["vitrage_pv"],
             "Le tarif de la fenêtre de toit FT84 ne porte aucune plus-value de vitrage. "
             "Les deux compositions proposées, avec et sans volet de toit solaire, sont "
             "l'une et l'autre marquées standard dans la colonne de plus-value vitrage, et "
             "les vitrages décoratifs sont déclarés indisponibles sur la gamme. Une "
             "demande de plus-value de vitrage sur cette gamme n'a donc pas de réponse "
             "tarifaire et relève d'une demande spéciale auprès du service Produits.")

    f.ajoute("%s — Description du vitrage isolant thermique Isol'4" % PREFIXE,
             PAGES["vitrage_description"],
             "Le vitrage isolant thermique Isol'4 est un double vitrage qui a reçu sur une "
             "de ses vitres, la vitre intérieure en standard, une mince couche de sels "
             "métalliques par métallisation, et qui comprend entre les deux vitres du gaz "
             "Argon. L'argon utilisé dans les vitrages isolants est garanti non toxique. "
             "La valeur vitrage comprise dans le prix des châssis correspond en standard à "
             "l'Isol'4 33/2-16G TRYBA SUN+4 pour la gamme FT84, avec écarteurs thermiques "
             "noirs de série.")

    f.ajoute("%s — Taux de TVA et éligibilité CITE par composant" % PREFIXE, PAGES["tva"],
             "L'édition toutes taxes comprises du tarif de la fenêtre de toit FT84 indique "
             "le taux de TVA retenu pour chaque composant. La fenêtre de toit TRYBA FT84 "
             "relève d'une TVA à 5,5 % et est signalée éligible CITE. Le volet de toit "
             "solaire TRYBA VTS relève d'une TVA à 10 % et est signalé non éligible CITE. "
             "La valeur vitrage relève d'une TVA à 5,5 % et est signalée éligible CITE. "
             "Ces taux sont donnés à titre d'information : les montants toutes taxes "
             "comprises sont lus au tarif, jamais recalculés.", edition="TTC")

    f.ajoute("%s — Prix hors éco-participation" % PREFIXE, PAGES["lecture"],
             "Tous les montants du tarif de la fenêtre de toit FT84, dans l'édition hors "
             "taxes comme dans l'édition toutes taxes comprises, sont indiqués hors "
             "éco-participation. Le tarif porte cette mention en pied de chacune de ses "
             "pages. L'éco-participation n'est chiffrée nulle part dans ce document et se "
             "recherche dans le barème en vigueur.")

    f.ajoute("%s — Charte de qualité des vitrages TRYBA" % PREFIXE,
             PAGES["vitrage_description"],
             "Le contrôle qualité des vitrages TRYBA est strict. Les unités de vitrage "
             "contrôlées ne devront en aucun cas comprendre plus de deux défauts tels que "
             "bulles, points ou taches dont le diamètre est supérieur à 0,5 mm. Les "
             "rayures ne seront pas acceptées dès lors que leur nombre est supérieur à "
             "deux et que leur longueur est supérieure à 15 mm. Les traces de ventouses de "
             "manutention ne sont pas considérées comme un défaut. La brochure de test de "
             "la qualité des vitrages TRYBA est disponible auprès des commerciaux.")

    f.ajoute("%s — Interférences lumineuses et implantation près d'un radiateur" % PREFIXE,
             PAGES["vitrage_description"],
             "Dans les vitrages isolants, les multiples réflexions sur les interfaces "
             "verre-air peuvent créer sous certaines incidences des interférences sous "
             "forme de franges diversement colorées. Ce phénomène naturel, purement "
             "optique, n'est pas un défaut du produit et ne compromet en rien ses "
             "qualités. Il ne peut en aucun cas être la cause d'un remplacement du vitrage "
             "ni entraîner l'application des garanties TRYBA. En cas d'implantation "
             "particulière avec un radiateur à proximité d'un vitrage, le tarif demande "
             "30 cm entre vitrage et radiateur.")

    f.ajoute("%s — Bon de commande de la fenêtre de toit et du volet de toit solaire" % PREFIXE,
             PAGES["bon_commande"],
             "Le tarif de la fenêtre de toit FT84 comporte un bon de commande commun à la "
             "fenêtre de toit 84 et au volet de toit solaire TRYBA. Il demande par "
             "position la localisation, la quantité, les dimensions du tableau intérieur "
             "en largeur et en hauteur, la pente en degrés, les informations de jumelage "
             "de châssis avec la largeur entre tableaux et la position de la jumelle, le "
             "coloris intérieur de la FT84, le coloris du coffre et des coulisses, la "
             "présence ou non d'un volet TRYBA VTS et le coloris de son tablier. Le type "
             "de couverture est demandé une fois pour la commande.")

    f.ajoute("%s — Historique des évolutions du tarif" % PREFIXE, PAGES["evolutions"],
             "Le tarif de la fenêtre de toit FT84 consigne ses évolutions successives : "
             "ajout de la mention de prix hors éco-participation, hausse de 2 % applicable "
             "au 1er janvier 2025, hausse de 3 % supplémentaire applicable au 30 juin "
             "2025, hausse de 3 % supplémentaire applicable au 1er mai 2026, puis "
             "intégration directe de l'ensemble des hausses dans les grilles tarifaires "
             "au 28 mai 2026. Les grilles de la présente édition intègrent donc déjà ces "
             "hausses : aucune ne doit être appliquée une seconde fois.")

    f.ajoute("%s — Valeur vitrage : chapitre non migré, à consulter au tarif" % PREFIXE,
             PAGES["lecture"],
             "Le tarif de la fenêtre de toit FT84 mentionne une valeur vitrage, abrégée "
             "Val. Vit., présentée comme une base de calcul des plus-values de vitrage. "
             "Aucun montant de valeur vitrage n'est repris dans ce corpus. Motif : les "
             "montants figurant dans le fichier de données ne se retrouvent pas dans la "
             "grille tarifaire du document de référence, et la gamme ne porte par ailleurs "
             "aucune plus-value de vitrage à laquelle cette base s'appliquerait. Toute "
             "demande portant sur la valeur vitrage doit être adressée au service Produits.")

    return f


# --------------------------------------------------------------------------
# Chapitre gele : Val. Vit. sur mesure
# --------------------------------------------------------------------------

def gele_val_vit(lignes, c_ht, c_ttc):
    n = 0
    for L in lignes:
        h, prod = decoupe_hauteur(L["hauteur"])
        if prod != "VAL_VIT":
            continue
        for cle in c_ht:
            if L["cells"].get(c_ht[cle]) is not None:
                n += 1
            if L["cells"].get(c_ttc[cle]) is not None:
                n += 1
    JOURNAL["postes_geles"].append(
        {"chapitre": "Valeur vitrage (Val. Vit.), regime dimensions sur mesure",
         "montants_geles": n,
         "motif": "les montants ne figurent sur aucune grille du PDF de reference ; "
                  "la seule occurrence PDF est la grille illustrative de la page 8, "
                  "dont les valeurs et la structure de bandes sont perimees ; la gamme "
                  "ne porte aucune plus-value de vitrage a laquelle cette base "
                  "s'appliquerait",
         "action": "aucun chunk chiffre produit ; existence signalee sans montant"})
    return n


# --------------------------------------------------------------------------
# Point d'entree
# --------------------------------------------------------------------------

def main():
    os.makedirs(OUT, exist_ok=True)
    lignes, c_sm_ht, c_sm_ttc, c_st_ht, c_st_ttc = lit_excel()

    JOURNAL["pages_exclues"] = [
        {"page": 5, "motif": "Fiche Info Produit edition 05-2026, deja migree a "
                             "l'identique dans FIP_FT84_05-2026.md"},
        {"page": 8, "motif": "grille illustrative : structure de bandes et valeurs "
                             "perimees, sans rapport avec la grille reelle de la page 10 ; "
                             "seules la regle de lecture et la legende du chassis nu sont "
                             "reprises, aucun montant"},
        {"page": 24, "motif": "dos de couverture ; la couche texte y superpose un titre "
                              "etranger, Tarif Portes de garage juillet 2018"},
    ]

    n_geles = gele_val_vit(lignes, c_sm_ht, c_sm_ttc)

    fichiers = [
        construit_methode(),
        construit_prix_sur_mesure(lignes, c_sm_ht, c_sm_ttc),
        construit_prix_stock(lignes, c_st_ht, c_st_ttc),
        construit_options(lignes),
        construit_faisabilites(),
        construit_transverses(),
    ]

    total = 0
    for f in fichiers:
        p = f.ecrit()
        total += len(f.chunks)
        pire = max(c["mots"] for c in f.chunks)
        print("%-40s %3d chunks   max %3d mots" % (f.nom, len(f.chunks), pire))
    print("%-40s %3d chunks" % ("TOTAL", total))
    print("Montants Val. Vit. geles :", n_geles)

    with open(os.path.join(OUT, "journal_generation_FT84.json"), "w",
              encoding="utf-8") as fh:
        json.dump(JOURNAL, fh, ensure_ascii=False, indent=2)
    for k, v in JOURNAL.items():
        print("  journal.%-28s %d entree(s)" % (k, len(v)))


if __name__ == "__main__":
    main()
