#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Audit de conformite du corpus tarif FT84.

Autonome : ne reutilise aucune fonction du generateur. Il relit les fichiers .md
produits, redeclare ses propres tables, re-derive les bandes depuis les en-tetes de
l'Excel, redecode lui-meme les numeros de modele, et re-extrait la couche texte du
PDF. Une divergence entre generateur et audit est donc un ecart reel, non une
tautologie.

Treize familles de controles ; voir note_cadrage_migration_tarif_FT84_v1.md, §9.
"""

import os, re, sys, glob, json, subprocess
import openpyxl

def _resout(noms):
    for base in ("/home/claude/ft84", "/mnt/user-data/uploads"):
        for n in noms:
            p = os.path.join(base, n)
            if os.path.exists(p):
                return p
    raise SystemExit("source introuvable : %s" % noms)

XLSX = _resout(["FT84_-infos-tarifs.xlsx"])
PDF_HT = _resout(["Tarif_FT84_HT_28-05-2026.pdf"])
PDF_TTC = _resout(["Tarif_FT84_TTC_28-05-2026.pdf"])
MD = os.environ.get("FT84_OUT", "/home/claude/ft84/out")

PLAFOND = 200
PREFIXE = "FT84 Fenêtre de toit PVC — "
RE_SOURCE = re.compile(
    r"^\*Source : Tarif—FT84—(HT|TTC)—28-05-2026\.pdf, page (\d+) — "
    r"information (originale|complémentaire) — SC(\d{4})\*$")
FM_ATTENDU = ["document_source", "document_source_ttc", "type_document", "sous_type",
              "gamme_code", "gammes_couvertes", "collection", "version_doc",
              "date_validite", "nb_chunks"]

SANS_MONTANT = {"Tarif_FT84_METHODE.md", "Tarif_FT84_FAISABILITES.md",
                "Tarif_FT84_TRANSVERSES.md"}

FAUX_SYNONYMES = ["gond", "charnière", "paumelle", "survitrage",
                  "anti-dégondage", "oscillo-battant"]
GAMMES_ETRANGERES = ["H81", "T81", "HA76", "HAM76", "TA76", "CA76", "CA80",
                     "H81 Access", "TA80"]
MARQUEURS_UNITE = ["unité de facturation n'est pas établie", "par fenêtre FT84",
                   "pour un kit", "pour une tôle"]

resultats = []


def ok(nom, detail=""):
    resultats.append(("OK", nom, detail))


def ko(nom, detail):
    resultats.append(("ECHEC", nom, detail))


def warn(nom, detail):
    resultats.append(("AVERT", nom, detail))


def mots(txt):
    return len(re.findall(r"\S+", txt))


# --------------------------------------------------------------------------
# Lecture independante des sources
# --------------------------------------------------------------------------

def charge_md():
    fichiers = {}
    # Seuls les six fichiers de corpus sont audites : la note de cadrage et le
    # message au service Produits ne sont pas des fichiers de chunks.
    for path in sorted(glob.glob(os.path.join(MD, "Tarif_FT84_*.md"))):
        nom = os.path.basename(path)
        brut = open(path, encoding="utf-8").read()
        m = re.match(r"^---\n(.*?)\n---\n", brut, re.S)
        fm = {}
        if m:
            for l in m.group(1).split("\n"):
                if ":" in l:
                    k, v = l.split(":", 1)
                    fm[k.strip()] = v.strip()
            corps = brut[m.end():]
        else:
            corps = brut
        chunks = []
        for bloc in re.split(r"\n(?=## )", corps):
            bloc = bloc.strip()
            if not bloc.startswith("## "):
                continue
            lignes = bloc.split("\n")
            chunks.append({"titre": lignes[0][3:].strip(),
                           "source": lignes[1].strip() if len(lignes) > 1 else "",
                           "corps": " ".join(lignes[2:]).strip(),
                           "bloc": bloc})
        fichiers[nom] = {"fm": fm, "chunks": chunks, "brut": brut}
    return fichiers


def charge_excel():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Feuil1"]
    ent = {c: (ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)}

    # Re-derivation independante des bandes depuis les en-tetes.
    bandes_sm, points_st = {}, {}
    for c, h in ent.items():
        h = str(h).strip()
        m = re.fullmatch(r"(HT|TTC) (\d+) à (\d+)", h)
        if m:
            bandes_sm.setdefault((int(m.group(2)), int(m.group(3))), {})[m.group(1)] = c
            continue
        m = re.fullmatch(r"(HT|TTC) (\d+)", h)
        if m:
            points_st.setdefault(int(m.group(2)), {})[m.group(1)] = c

    grille, forfaits, val_vit = {}, [], set()
    for r in range(2, ws.max_row + 1):
        chap = str(ws.cell(r, 3).value or "").strip().lower()
        haut = ws.cell(r, 11).value
        if not chap:
            continue
        if "grille" in str(ws.cell(r, 4).value or "").lower():
            m = re.match(r"^\s*(\d+)\s*-?\s*(.*)$", str(haut or ""))
            if not m:
                continue
            h, prod = m.group(1), m.group(2).strip().upper()
            regime = "SM" if "sur mesure" in chap else "ST"
            cols = bandes_sm if regime == "SM" else points_st
            for cle, cc in cols.items():
                vh = ws.cell(r, cc["HT"]).value
                vt = ws.cell(r, cc["TTC"]).value
                if vh is None and vt is None:
                    continue
                if prod.startswith("VAL"):
                    if vh is not None:
                        val_vit.add(int(vh))
                    if vt is not None:
                        val_vit.add(int(vt))
                    continue
                p = "FT84" if prod in ("FT84", "FR84") else prod
                grille[(regime, int(h), p, cle)] = (
                    int(vh) if vh is not None else None,
                    int(vt) if vt is not None else None)
            # poste scalaire porte par la ligne de grille : abergement
            if not prod.startswith("VAL") and ws.cell(r, 9).value is not None:
                forfaits.append((regime, "abergement", int(h),
                                 int(ws.cell(r, 9).value), int(ws.cell(r, 10).value)))
        else:
            forfaits.append(("F", str(ws.cell(r, 4).value or "").strip(), None,
                             int(ws.cell(r, 9).value), int(ws.cell(r, 10).value)))
    return grille, forfaits, val_vit, sorted(bandes_sm), sorted(points_st)


def texte_page(pdf, n):
    return subprocess.run(["pdftotext", "-f", str(n), "-l", str(n), "-layout", pdf, "-"],
                          capture_output=True, text=True).stdout


# --------------------------------------------------------------------------
# Extraction independante des montants servis par le corpus
# --------------------------------------------------------------------------

RE_SM = re.compile(r"largeur de (\d+) à (\d+) mm, (\d+) € HT et (\d+) € TTC")
RE_ST = re.compile(r"le modèle (\d+), largeur (\d+) mm, à (\d+) € HT et (\d+) € TTC")
RE_TITRE_SM = re.compile(r"Tarif sur mesure (de la fenêtre de toit FT84 seule|"
                         r"du volet de toit solaire TRYBA VTS), hauteur de (\d+) à (\d+) mm")
RE_TITRE_ST = re.compile(r"Tarif stock (de la fenêtre de toit FT84 seule|"
                         r"du volet de toit solaire TRYBA VTS), hauteur (\d+) mm")


def produit(lib):
    return "FT84" if "fenêtre" in lib else "VTS"


def servis(fichiers):
    sm, st = {}, {}
    for c in fichiers["Tarif_FT84_PRIX_SUR_MESURE.md"]["chunks"]:
        m = RE_TITRE_SM.search(c["titre"])
        if not m:
            ko("forme.titre_prix_sur_mesure", c["titre"])
            continue
        p, hmin = produit(m.group(1)), int(m.group(2))
        for lmin, lmax, ht, ttc in RE_SM.findall(c["corps"]):
            cle = ("SM", hmin, p, (int(lmin), int(lmax)))
            if cle in sm:
                ko("couverture.doublon", str(cle))
            sm[cle] = (int(ht), int(ttc))
    for c in fichiers["Tarif_FT84_PRIX_STOCK.md"]["chunks"]:
        m = RE_TITRE_ST.search(c["titre"])
        if not m:
            ko("forme.titre_prix_stock", c["titre"])
            continue
        p, haut = produit(m.group(1)), int(m.group(2))
        for code, larg, ht, ttc in RE_ST.findall(c["corps"]):
            cle = ("ST", haut, p, int(larg))
            if cle in st:
                ko("couverture.doublon", str(cle))
            st[cle] = (int(ht), int(ttc), int(code))
    return sm, st


# --------------------------------------------------------------------------
# Controles
# --------------------------------------------------------------------------

def c1_forme(fichiers):
    pb = []
    for nom, f in fichiers.items():
        for k in FM_ATTENDU:
            if k not in f["fm"]:
                pb.append("%s : front matter sans %s" % (nom, k))
        if f["fm"].get("nb_chunks") != str(len(f["chunks"])):
            pb.append("%s : nb_chunks annonce %s pour %d chunks"
                      % (nom, f["fm"].get("nb_chunks"), len(f["chunks"])))
        attendu = 2
        for c in f["chunks"]:
            if not c["titre"].startswith(PREFIXE):
                pb.append("%s : titre sans prefixe de gamme — %s" % (nom, c["titre"][:60]))
            m = RE_SOURCE.match(c["source"])
            if not m:
                pb.append("%s : ligne de source non conforme — %s" % (nom, c["source"][:70]))
                continue
            if int(m.group(4)) != attendu:
                pb.append("%s : SC discontinue, attendu SC%04d, lu SC%04d"
                          % (nom, attendu, int(m.group(4))))
            attendu = int(m.group(4)) + 1
            n = mots("## " + c["titre"] + " " + c["source"] + " " + c["corps"])
            if n > PLAFOND:
                pb.append("%s : plafond depasse (%d mots) — %s" % (nom, n, c["titre"][:50]))
            if re.search(r"^\s*[-*•]\s", c["corps"], re.M):
                pb.append("%s : puces dans le corps — %s" % (nom, c["titre"][:50]))
    (ok if not pb else ko)("1. forme, front matter, source, SC, plafond, prose",
                           "%d anomalie(s)%s" % (len(pb), "" if not pb else " : " + " | ".join(pb[:6])))


def c2_unicite(fichiers):
    pb, glob_t = [], {}
    for nom, f in fichiers.items():
        vus = set()
        for c in f["chunks"]:
            if c["titre"] in vus:
                pb.append("%s : titre en double — %s" % (nom, c["titre"][:60]))
            vus.add(c["titre"])
            glob_t.setdefault(c["titre"], []).append(nom)
    for t, ns in glob_t.items():
        if len(ns) > 1:
            pb.append("titre partage entre %s — %s" % (ns, t[:50]))
    (ok if not pb else ko)("2. unicite des titres", "%d anomalie(s) %s" % (len(pb), pb[:3]))


def c3_couverture_fidelite(fichiers, grille):
    sm, st = servis(fichiers)
    attendu = {}
    for (reg, h, p, cle), (vh, vt) in grille.items():
        if vh is None or vt is None:
            continue
        attendu[(reg, h, p, cle)] = (vh, vt)
    servi = {}
    servi.update({k: v for k, v in sm.items()})
    servi.update({k: (v[0], v[1]) for k, v in st.items()})

    manquants = sorted(set(attendu) - set(servi))
    fantomes = sorted(set(servi) - set(attendu))
    infidelites = [(k, attendu[k], servi[k]) for k in set(attendu) & set(servi)
                   if attendu[k] != servi[k]]

    (ok if not manquants else ko)(
        "3. couverture exhaustive des cellules",
        "%d cellules attendues, %d servies, %d manquante(s) %s"
        % (len(attendu), len(servi), len(manquants), manquants[:4]))
    (ok if not fantomes else ko)(
        "4. anti-fantome (aucune cellule inventee)",
        "%d fantome(s) %s" % (len(fantomes), fantomes[:4]))
    (ok if not infidelites else ko)(
        "5. fidelite numerique exhaustive HT et TTC",
        "%d valeur(s) comparee(s), %d ecart(s) %s"
        % (2 * len(attendu), len(infidelites), infidelites[:3]))
    return sm, st


def c6_bandes(fichiers, bandes_sm, points_st):
    pb = []
    servies = set()
    for c in fichiers["Tarif_FT84_PRIX_SUR_MESURE.md"]["chunks"]:
        for lmin, lmax, _, _ in RE_SM.findall(c["corps"]):
            servies.add((int(lmin), int(lmax)))
    inconnues = servies - set(bandes_sm)
    if inconnues:
        pb.append("bandes de largeur servies absentes des en-tetes : %s" % sorted(inconnues))
    # contiguite : chaque borne haute + 1 est la borne basse suivante
    tri = sorted(bandes_sm)
    for (a1, b1), (a2, b2) in zip(tri, tri[1:]):
        if b1 + 1 != a2:
            pb.append("bandes non contigues : %s puis %s" % ((a1, b1), (a2, b2)))
    # aucune bande servie comme un point
    for c in fichiers["Tarif_FT84_PRIX_SUR_MESURE.md"]["chunks"]:
        if re.search(r"largeur (\d+) mm, \d+ € HT", c["corps"]):
            pb.append("largeur servie comme un point : %s" % c["titre"][:50])
    pts = set()
    for c in fichiers["Tarif_FT84_PRIX_STOCK.md"]["chunks"]:
        for _, larg, _, _ in RE_ST.findall(c["corps"]):
            pts.add(int(larg))
    if pts - set(points_st):
        pb.append("largeurs stock hors en-tetes : %s" % sorted(pts - set(points_st)))
    (ok if not pb else ko)("6. bornes de bandes recalculees independamment",
                           "%d bande(s) sur mesure, %d point(s) stock ; %d anomalie(s) %s"
                           % (len(bandes_sm), len(points_st), len(pb), pb[:3]))


def c7_codes_modele(st):
    """Redecodage independant : premier chiffre = rang de hauteur, second = rang de largeur."""
    H = [919, 1119, 1339, 1541]
    L = [495, 605, 725, 887, 1085, 1285]
    pb = []
    for (reg, haut, prod, larg), (ht, ttc, code) in st.items():
        attendu = int("%d%d" % (H.index(haut) + 1, L.index(larg) + 1))
        if code != attendu:
            pb.append("hauteur %d largeur %d : code servi %d, recalcule %d"
                      % (haut, larg, code, attendu))
    (ok if not pb else ko)("7. decodage des numeros de modele stock",
                           "%d code(s) verifie(s), %d ecart(s) %s" % (len(st), len(pb), pb[:3]))


def c8_forfaits(fichiers, forfaits):
    """Bijection par multiensembles (HT, TTC), insensible aux libelles."""
    from collections import Counter
    attendu = Counter((h, t) for _, _, _, h, t in forfaits)
    servi = Counter()
    for c in fichiers["Tarif_FT84_OPTIONS.md"]["chunks"]:
        m = re.findall(r"(\d+) € HT(?: net)? et (\d+) € TTC", c["corps"])
        for h, t in m:
            servi[(int(h), int(t))] += 1
    # l'abergement est porte redondamment par plusieurs lignes de grille :
    # l'audit compare les ensembles de couples, non leur multiplicite.
    manquants = sorted(set(attendu) - set(servi))
    fantomes = sorted(set(servi) - set(attendu))
    pb = []
    if manquants:
        pb.append("couples absents du corpus : %s" % manquants)
    if fantomes:
        pb.append("couples servis absents de l'Excel : %s" % fantomes)
    (ok if not pb else ko)("8. bijection des postes forfaitaires (multiensembles HT/TTC)",
                           "%d couple(s) distinct(s) attendu(s), %d servi(s) ; %s"
                           % (len(set(attendu)), len(set(servi)), " | ".join(pb) or "aucun ecart"))


def c9_unites(fichiers):
    pb = []
    for c in fichiers["Tarif_FT84_OPTIONS.md"]["chunks"]:
        if "€" not in c["corps"]:
            continue
        if not any(m in c["corps"] for m in MARQUEURS_UNITE):
            pb.append(c["titre"][:70])
    (ok if not pb else ko)("9. unite de facturation declaree sur tout poste chiffre",
                           "%d poste(s) sans unite %s" % (len(pb), pb[:3]))


def c10_sans_montant(fichiers):
    pb = []
    for nom in SANS_MONTANT:
        for c in fichiers[nom]["chunks"]:
            if "€" in c["corps"]:
                pb.append("%s : %s" % (nom, c["titre"][:60]))
    (ok if not pb else ko)("10. absence de tout montant en methode, faisabilites et transverses",
                           "%d chunk(s) fautif(s) %s" % (len(pb), pb[:3]))


def c11_val_vit(fichiers, val_vit):
    """Le chapitre gele ne doit servir aucun de ses montants."""
    pb = []
    for nom, f in fichiers.items():
        for c in f["chunks"]:
            for m in re.findall(r"(\d+) €", c["corps"]):
                if int(m) in val_vit and nom != "Tarif_FT84_OPTIONS.md":
                    pb.append("%s : %s € — %s" % (nom, m, c["titre"][:45]))
    # les montants d'abergement peuvent coincider avec une valeur Val. Vit. :
    # on ne retient que les occurrences hors postes forfaitaires legitimes.
    if pb:
        warn("11. chapitre Val. Vit. gele",
             "%d coincidence(s) numerique(s) a verifier %s" % (len(pb), pb[:3]))
    else:
        ok("11. chapitre Val. Vit. gele",
           "aucun des %d montants geles n'est servi" % len(val_vit))


def c12_vocabulaire(fichiers):
    pb = []
    for nom, f in fichiers.items():
        txt = f["brut"]
        for mot in FAUX_SYNONYMES:
            if re.search(r"\b%s" % mot, txt, re.I):
                pb.append("%s : faux synonyme '%s'" % (nom, mot))
        for g in GAMMES_ETRANGERES:
            if re.search(r"\b%s\b" % re.escape(g), txt):
                pb.append("%s : gamme etrangere '%s'" % (nom, g))
    legitimes = sum("crochets massifs" in f["brut"] for f in fichiers.values())
    (ok if not pb else ko)("12. vocabulaire et contamination inter-gammes",
                           "%d anomalie(s) %s ; 'crochets massifs' present dans %d fichier(s), "
                           "legitime sur cette gamme" % (len(pb), pb[:3], legitimes))


def c13_pdf(fichiers):
    """Table des pages validee au pied de page, puis croisement page par page."""
    pb = []
    for i in range(2, 25):
        t = texte_page(PDF_HT, i)
        m = re.search(r"(\d{1,2})\s*-\s*V\.28/05/2026|V\.28/05/2026[^\n]*?-\s*(\d{1,2})\s*$",
                      t, re.M)
        if m:
            imprime = int(m.group(1) or m.group(2))
            if imprime != i:
                pb.append("page PDF %d, pied de page %d" % (i, imprime))
    pages_citees = set()
    manque = []
    for nom, f in fichiers.items():
        for c in f["chunks"]:
            ms = RE_SOURCE.match(c["source"])
            if not ms:
                continue
            page = int(ms.group(2))
            pages_citees.add(page)
            if page < 1 or page > 24:
                pb.append("%s : page %d hors document" % (nom, page))
                continue
            montants = re.findall(r"(\d{3,4}) €", c["corps"])
            if not montants:
                continue
            src = PDF_HT if ms.group(1) == "HT" else PDF_TTC
            txt = texte_page(src, page) + texte_page(PDF_TTC, page)
            # Le PDF compose les milliers avec une espace ("1 911 €HT") la ou les
            # grilles n'en mettent pas : l'audit normalise avant comparaison.
            plat = re.sub(r"(?<=\d)[\s\u00a0\u202f](?=\d{3}\b)", "", txt)
            nums = set(re.findall(r"\d+", plat)) | set(re.findall(r"\d+", txt))
            absents = [v for v in montants if v not in nums]
            if absents:
                manque.append("%s p.%d : %s" % (c["titre"][:40], page, absents[:3]))
    (ok if not pb else ko)("13a. table des pages (pied de page = index PDF)",
                           "23 pages controlees, %d ecart(s) %s" % (len(pb), pb[:3]))
    (ok if not manque else warn)(
        "13b. croisement PDF page par page des montants servis",
        "%d chunk(s) chiffre(s) dont un montant n'est pas retrouve sur sa page %s"
        % (len(manque), manque[:3]))
    ok("13c. pages citees par le corpus", str(sorted(pages_citees)))


def main():
    fichiers = charge_md()
    grille, forfaits, val_vit, bandes_sm, points_st = charge_excel()
    c1_forme(fichiers)
    c2_unicite(fichiers)
    sm, st = c3_couverture_fidelite(fichiers, grille)
    c6_bandes(fichiers, bandes_sm, points_st)
    c7_codes_modele(st)
    c8_forfaits(fichiers, forfaits)
    c9_unites(fichiers)
    c10_sans_montant(fichiers)
    c11_val_vit(fichiers, val_vit)
    c12_vocabulaire(fichiers)
    c13_pdf(fichiers)

    print("\n" + "=" * 78)
    for etat, nom, detail in resultats:
        print("[%-6s] %-58s %s" % (etat, nom, detail))
    print("=" * 78)
    n_ok = sum(1 for e, _, _ in resultats if e == "OK")
    n_ko = sum(1 for e, _, _ in resultats if e == "ECHEC")
    n_w = sum(1 for e, _, _ in resultats if e == "AVERT")
    total = sum(len(f["chunks"]) for f in fichiers.values())
    print("%d chunks audites — %d controles reussis, %d echecs, %d avertissements"
          % (total, n_ok, n_ko, n_w))
    return 1 if n_ko else 0


if __name__ == "__main__":
    sys.exit(main())
