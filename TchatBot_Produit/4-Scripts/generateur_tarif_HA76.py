#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Générateur de chunks Markdown pour le tarif HA76 (Porte d'entrée Aluminium).

Hérite de note_cadrage_migration_tarif_H81_v1.md pour le coeur "modèle"
(règles 1 à 7) et des acquis T81 / CA76 pour ce que H81 ne couvrait pas
(fichier METHODE, exposition des divergences, fichier FAISABILITES).

Huit fichiers produits :
  1. METHODE              cadre de lecture du tarif
  2. PRIX_MODELES         un chunk par modèle x configuration de vantaux réellement tarifée
  3. OPTIONS_MODELES      un chunk par couple option x modèle chiffré
  4. CARACTERISTIQUES     un chunk par modèle (vitrage, Ud, enveloppe dimensionnelle)
  5. COMPAT_EQUIPEMENTS   un chunk par équipement
  6. CATALOGUE_OPTIONS    un chunk par référence d'option hors modèle, forfaitaire
  7. FAISABILITES         faisabilités non tarifaires
  8. TRANSVERSES          orientation sans montant (%, €/m², €/ml, grille gelée)
Plus Message_service_produit_HA76.md.

Transversal : SC continue par fichier depuis SC0002, ligne de source normée,
plafond 200 mots, prose sans puces, journal des colonnes non mappées,
anti-fantôme, fidélité littérale (aucun montant calculé).
"""
import openpyxl
import re
import json
import unicodedata
from collections import OrderedDict, defaultdict

XLSX = "/mnt/user-data/uploads/HA76_-infos-tarifs.xlsx"
PDF_SOURCE = "Tarif—HA76—HT—23-06-2026.pdf"      # nom affiché, em-dashes
PDF_FILENAME = "Tarif_HA76_HT_23-06-2026.pdf"    # nom réel, underscores
PAGEMAP = "/home/claude/pagemap.json"
OUTDIR = "/mnt/user-data/outputs"
PLAFOND = 200

GAMME = "HA76"
DESIGNATION = "Porte d'entrée Aluminium"
PREFIXE = f"{GAMME} {DESIGNATION}"

ALERTS = []
JOURNAL = []

# ------------------------------------------------------------------ utilitaires
def clean(v):
    if v is None:
        return ""
    return str(v).strip()


def fmt_euro(v):
    """2614 -> '2 614' (espace fine insécable). None si non numérique."""
    if v in (None, ""):
        return None
    try:
        n = int(round(float(str(v).replace(" ", "").replace(",", "."))))
    except (ValueError, TypeError):
        return None
    return f"{n:,}".replace(",", "\u202f")


def to_num(v):
    try:
        return float(str(v).replace(" ", "").replace("\u202f", "").replace(",", "."))
    except (ValueError, TypeError):
        return None


def norm_ud(v):
    """'Ud porte vitrée : 1,7 W/m2.K' -> '1,7 W/m².K'."""
    s = clean(v)
    if not s:
        return ""
    s = s.replace("m2", "m²")
    if ":" in s:
        s = s.split(":", 1)[1]
    s = re.sub(r"(\d)\s*W", r"\1 W", s)
    return re.sub(r"\s+", " ", s).strip()


def norm_texte(v):
    """Aplati un champ multi-lignes de l'Excel en prose."""
    s = clean(v).replace("\r", "\n")
    s = re.sub(r"\n+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s.rstrip(".").strip()


def count_words(*parts):
    return len(re.findall(r"\S+", " ".join(p for p in parts if p)))


def source_line(page, sc, nature="originale"):
    mot = "pages" if "à" in str(page) else "page"
    return f"*Source : {PDF_SOURCE}, {mot} {page} — information {nature} — SC{sc:04d}*"


class Fichier:
    """Accumule des chunks, numérote les SC depuis SC0002, contrôle le plafond."""

    def __init__(self, nom, titre_yaml, perimetre):
        self.nom = nom
        self.titre_yaml = titre_yaml
        self.perimetre = perimetre
        self.chunks = []
        self.sc = 2

    def add(self, titre, corps, page, nature="originale"):
        src = source_line(page, self.sc, nature)
        # \s+ engloberait l'espace fine insécable des montants : on l'exclut
        corps = re.sub(r"[^\S\u202f]+", " ", corps).strip()
        n = count_words(titre, src, corps)
        if n > PLAFOND:
            ALERTS.append(f"[{self.nom}] plafond dépassé ({n} mots) : {titre}")
        self.chunks.append((titre, src, corps, n))
        self.sc += 1

    def write(self):
        fm = [
            "---",
            f"document_source: {PDF_FILENAME}",
            "type_document: tarif",
            f"sous_type: {self.nom.lower()}",
            f"gamme_code: {GAMME}",
            f'gamme_nom: "{DESIGNATION}"',
            'collection: "TRYBA ALUMINIUM"',
            "materiau: aluminium",
            'version_doc: "2026.06"',
            "date_validite: 2026-06-23",
            f"nb_chunks: {len(self.chunks)}",
            "audiences: [ADV, commercial]",
            "---",
            "",
        ]
        body = []
        for titre, src, corps, _ in self.chunks:
            body.append(f"## {titre}\n{src}\n\n{corps}\n")
        path = f"{OUTDIR}/Tarif_{GAMME}_{self.nom}.md"
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(fm) + "\n".join(body))
        return path, len(self.chunks)


# ------------------------------------------------------------------ lecture Excel
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Feuil1"]
from openpyxl.utils import column_index_from_string as CI, get_column_letter as CL

HDR = {CL(c): clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)}


def cell(r, letter):
    return clean(ws.cell(r, CI(letter)).value)


PAGES_MODELE = json.load(open(PAGEMAP, encoding="utf-8"))

lignes_modele = [r for r in range(2, ws.max_row + 1) if cell(r, "C")]
lignes_chapitre = [r for r in range(2, ws.max_row + 1)
                   if cell(r, "CF") and not cell(r, "C")]

# groupement par modèle : une ligne Excel est un EMPLACEMENT D'AFFICHAGE,
# pas un axe tarifaire (fait vérifié à l'étape 2).
groupes = OrderedDict()
for r in lignes_modele:
    groupes.setdefault(cell(r, "C"), []).append(r)


def premiere_valeur(rs, letter):
    for r in rs:
        if cell(r, letter):
            return cell(r, letter)
    return ""


# modèles dont le TTC 1V diverge entre leurs propres lignes -> TTC gelé (D1)
TTC_GELE = set()
for m, rs in groupes.items():
    if len({cell(r, "N") for r in rs}) > 1:
        TTC_GELE.add(m)
    else:
        ht, ttc = to_num(cell(rs[0], "L")), to_num(cell(rs[0], "N"))
        if ht and ttc and ttc < ht:
            TTC_GELE.add(m)

# ------------------------------------------------------------------ 1. METHODE
def f_methode():
    f = Fichier("METHODE", "Méthode de lecture",
                "règles de lecture du tarif HA76 et périmètre du corpus")

    f.add(f"{PREFIXE} — Portée du tarif et distinction avec la gamme HAM76",
          "Ce tarif porte exclusivement sur la porte d'entrée aluminium de la gamme HA76, "
          "collection TRYBA Aluminium, dans son édition du 23 juin 2026. Il ne s'applique "
          "pas à la porte d'entrée monobloc aluminium de la gamme HAM76, qui est une gamme "
          "distincte disposant de son propre tarif, de ses propres modèles et de ses propres "
          "prix. Les deux gammes partagent un vocabulaire technique très proche mais leurs "
          "montants ne sont jamais interchangeables. Aucun prix, aucune plus-value et aucune "
          "caractéristique de ce tarif ne doit être rapporté à la gamme HAM76.", 1)

    f.add(f"{PREFIXE} — Unité tarifaire : le modèle, sa ligne et sa référence",
          "Le prix d'une porte HA76 est attaché au modèle individuel. Chaque modèle appartient "
          "à une seule ligne parmi la ligne Vitrée, la ligne Contemporaine, la ligne "
          "Traditionnelle et la ligne Accord, et à une seule référence parmi Crystal, Accord, "
          "Excellence, Passion, Artiste, Mesure, Vitrée, Panneau lisse et Porte du mois "
          "permanente. La ligne et la référence sont des propriétés descriptives du modèle et "
          "non des axes de prix. Le tarif comprend quatre-vingt-douze modèles.", 9)

    f.add(f"{PREFIXE} — Le prix d'un modèle ne dépend pas de la dimension",
          "Le tarif d'un modèle de porte HA76 est forfaitaire : il ne varie pas avec la largeur "
          "ni avec la hauteur commandées. Les limites dimensionnelles figurant au tarif sont des "
          "limites de fabrication et non des paliers de prix. Une demande de prix portant sur une "
          "dimension particulière reçoit donc le tarif du modèle, sous réserve que la dimension "
          "demandée reste dans les limites de fabrication du modèle.", 17)

    f.add(f"{PREFIXE} — Le nombre de vantaux est un préalable à toute réponse de prix",
          "Le tarif distingue deux configurations, un vantail et deux vantaux, avec un prix "
          "distinct pour chacune. Une demande de prix dont la configuration de vantaux n'est pas "
          "précisée ne peut pas recevoir de réponse : la configuration doit être demandée avant "
          "toute restitution de montant. Quinze modèles ne sont tarifés qu'en un vantail et n'ont "
          "aucun prix en deux vantaux.", 9)

    f.add(f"{PREFIXE} — Les prix hors taxes et toutes taxes comprises sont lus, jamais calculés",
          "Les montants hors taxes s'entendent hors éco-participation. Les montants toutes taxes "
          "comprises figurant au tarif ne se déduisent pas des montants hors taxes par application "
          "d'un taux : le rapport entre les deux varie d'une ligne de tarif à l'autre. Chaque "
          "montant doit donc être restitué tel qu'il figure au tarif, et jamais obtenu par calcul, "
          "par conversion ni par interpolation entre deux valeurs.", 1)

    f.add(f"{PREFIXE} — Montants exprimés au mètre carré, au mètre linéaire ou en pourcentage",
          "Certaines plus-values du tarif ne sont pas des forfaits mais des montants au mètre "
          "carré, au mètre linéaire ou des pourcentages appliqués au prix du châssis. C'est le cas "
          "des plus-values de vitrages, des habillages, des bavettes, des couvre-joints, des "
          "meneaux battants facturés à la longueur et de l'offre couleurs. Ces montants ne "
          "figurent pas dans ce corpus car leur usage suppose un calcul. La demande doit être "
          "orientée vers la page du tarif qui les porte.", 3)

    f.add(f"{PREFIXE} — Ce que ce corpus ne contient pas",
          "Ce corpus ne contient ni la grille de tarif des fixes latéraux vitrés et impostes "
          "vitrées, ni les plus-values exprimées au mètre carré, au mètre linéaire ou en "
          "pourcentage, ni les exemples de calcul du tarif. Sur ces points, la réponse consiste à "
          "renvoyer l'utilisateur à la page correspondante du tarif. Toute demande dont la réponse "
          "n'est pas présente en toutes lettres dans un chunk doit recevoir une réponse d'absence, "
          "jamais une estimation.", 3)

    f.add(f"{PREFIXE} — Portée des limites de fabrication indiquées par modèle",
          "Les limites de fabrication portées au tarif pour chaque modèle de la gamme HA76 sont "
          "données pour une menuiserie en aluminium blanc équipée d'une serrure six points. Les "
          "restrictions propres aux autres teintes et aux options de ferrage se lisent aux "
          "chapitres correspondants du tarif et peuvent réduire ces limites. Une limite citée "
          "pour un modèle ne vaut donc pas engagement de faisabilité pour une teinte ou un "
          "ferrage particuliers.", 30)
    return f


# ------------------------------------------------------------------ 2. PRIX
CONFIGS = [("L", "N", "un vantail", "1 vantail"),
           ("M", "O", "deux vantaux", "2 vantaux")]


def f_prix():
    f = Fichier("PRIX_MODELES", "Prix des modèles",
                "prix des modèles de porte HA76 par configuration de vantaux")
    for m, rs in groupes.items():
        r0 = rs[0]
        ligne, ref = cell(r0, "H"), cell(r0, "I")
        page = PAGES_MODELE.get(m) or 30
        for col_ht, col_ttc, libelle, court in CONFIGS:
            ht = fmt_euro(cell(r0, col_ht))
            if ht is None:                       # anti-fantôme
                continue
            ttc = fmt_euro(cell(r0, col_ttc))
            titre = (f"{PREFIXE} — Tarif {m} {court} "
                     f"({ligne.lower()}, référence {ref})")
            if m in TTC_GELE and col_ttc == "N":
                corps = (f"En {libelle}, le modèle {m} de la gamme HA76, porte d'entrée "
                         f"aluminium, {ligne.lower()}, référence {ref}, est proposé au tarif de "
                         f"{ht} € HT hors éco-participation. Le montant toutes taxes comprises "
                         f"n'est pas restitué pour ce modèle : les sources du tarif portent pour "
                         f"lui des valeurs toutes taxes comprises divergentes. Le prix toutes "
                         f"taxes comprises doit être lu directement sur l'édition TTC du tarif.")
            elif ttc:
                corps = (f"En {libelle}, le modèle {m} de la gamme HA76, porte d'entrée "
                         f"aluminium, {ligne.lower()}, référence {ref}, est proposé au tarif de "
                         f"{ht} € HT, soit {ttc} € TTC. Le montant hors taxes s'entend hors "
                         f"éco-participation.")
            else:
                corps = (f"En {libelle}, le modèle {m} de la gamme HA76, porte d'entrée "
                         f"aluminium, {ligne.lower()}, référence {ref}, est proposé au tarif de "
                         f"{ht} € HT hors éco-participation. Le tarif ne porte pas de montant "
                         f"toutes taxes comprises pour cette configuration.")
            f.add(titre, corps, page)
    return f


# ------------------------------------------------------------------ 3. OPTIONS MODELE
FAMILLES = [
    ("AE", "AF", "AG", None, None, "option"),          # option spécifique
    ("AS", "AX", "AY", "AV", "BA", "vitrage"),         # options de vitrage
    ("BB", "BD", "BE", None, None, "vitrage Artiste"),  # ligne Artiste
    ("BG", "BI", "BJ", None, None, "teinte de grille"),
    ("BO", "BR", "BS", "BQ", None, "vitrage analogue"),
    ("BW", "BW", "BX", "BV", None, "panneau analogue"),
]

# libellé restitué depuis l'en-tête de bloc de la page modèle du PDF :
# la colonne BT de l'Excel est vide alors que le montant existe (point 7 du message).
LIBELLE_RESTITUE = {"BW": "Panneau analogue pour fixe latéral"}


def option_retenue(libelle, pv_brut, remarque):
    """Règle 4 adaptée : on retient les plus-values chiffrées, et les
    plus-values nulles explicites, mais jamais les renvois transverses."""
    if pv_brut == "":
        return False, "plus-value vide"
    if "voir page" in remarque.lower():
        return False, "renvoi à une page transverse"
    return True, ""


def f_options():
    f = Fichier("OPTIONS_MODELES", "Options et plus-values par modèle",
                "options et plus-values rattachées à un modèle de porte HA76")
    vus = set()
    exclus = defaultdict(int)
    for m, rs in groupes.items():
        r0 = rs[0]
        ligne, ref = cell(r0, "H"), cell(r0, "I")
        page = PAGES_MODELE.get(m) or 30
        for col_lib, col_ht, col_ttc, col_desc, col_rem, nature in FAMILLES:
            for r in rs:
                if col_lib in LIBELLE_RESTITUE:
                    lib = LIBELLE_RESTITUE[col_lib] if cell(r, col_lib) else ""
                else:
                    lib = norm_texte(cell(r, col_lib))
                if not lib:
                    continue
                pv_brut = cell(r, col_ht)
                rem = cell(r, col_rem) if col_rem else ""
                ok, motif = option_retenue(lib, pv_brut, rem)
                if not ok:
                    exclus[motif] += 1
                    continue
                cle = (m, lib.lower())
                if cle in vus:
                    continue
                vus.add(cle)
                ht, ttc = fmt_euro(pv_brut), fmt_euro(cell(r, col_ttc))
                if lib.strip().lower() == "vitrage ornemental":
                    qual = norm_texte(cell(r, col_desc)) if col_desc else ""
                    if not qual:
                        qual = norm_texte(rem)
                    if re.match(r"(?i)^groupe", qual):
                        lib = f"Vitrage ornemental du {qual[0].lower()}{qual[1:]}"
                lib_court = lib if len(lib) < 70 else lib[:67].rstrip() + "…"
                titre = (f"{PREFIXE} — Option {lib_court} sur {m} "
                         f"({ligne.lower()}, référence {ref})")
                if to_num(pv_brut) == 0:
                    corps = (f"Sur le modèle {m} de la gamme HA76, porte d'entrée aluminium, "
                             f"{ligne.lower()}, référence {ref}, l'option {lib} est proposée sans "
                             f"plus-value : elle est comprise dans le tarif du modèle.")
                elif ttc:
                    corps = (f"Sur le modèle {m} de la gamme HA76, porte d'entrée aluminium, "
                             f"{ligne.lower()}, référence {ref}, l'option {lib} est facturée en "
                             f"plus-value de {ht} € HT, soit {ttc} € TTC.")
                else:
                    corps = (f"Sur le modèle {m} de la gamme HA76, porte d'entrée aluminium, "
                             f"{ligne.lower()}, référence {ref}, l'option {lib} est facturée en "
                             f"plus-value de {ht} € HT. Le tarif ne porte pas de montant toutes "
                             f"taxes comprises pour cette option.")
                if nature == "panneau analogue":
                    corps += (" Cette plus-value s'ajoute au prix de grille des fixes de la "
                              "page 28 du tarif, qui n'est pas repris dans ce corpus.")
                if col_desc:
                    d = norm_texte(cell(r, col_desc))
                    if d and d.lower() not in lib.lower() and count_words(corps, d) < 140:
                        corps += f" Cette option correspond à : {d}."
                f.add(titre, corps, page)
    for motif, n in exclus.items():
        JOURNAL.append(f"options écartées ({motif}) : {n} occurrences")
    return f


# ------------------------------------------------------------------ 4. CARACTERISTIQUES
def f_caracteristiques():
    f = Fichier("CARACTERISTIQUES", "Caractéristiques par modèle",
                "vitrage de base, performance thermique et limites dimensionnelles par modèle")
    for m, rs in groupes.items():
        r0 = rs[0]
        ligne, ref = cell(r0, "H"), cell(r0, "I")
        page = PAGES_MODELE.get(m) or 30
        base = norm_texte(premiere_valeur(rs, "W"))
        dimv = norm_texte(premiere_valeur(rs, "U"))
        ud = norm_ud(premiere_valeur(rs, "T"))
        lm = [to_num(premiere_valeur(rs, c)) for c in ("AK", "AO")]
        lM = [to_num(premiere_valeur(rs, c)) for c in ("AM", "AQ")]
        hm = [to_num(premiere_valeur(rs, c)) for c in ("AL", "AP")]
        hM = [to_num(premiere_valeur(rs, c)) for c in ("AN", "AR")]
        titre = (f"{PREFIXE} — Caractéristiques {m} "
                 f"({ligne.lower()}, référence {ref})")
        corps = (f"Le modèle {m} de la gamme HA76, porte d'entrée aluminium, {ligne.lower()}, "
                 f"référence {ref}, reçoit en modèle de base : {base}." if base else
                 f"Le modèle {m} de la gamme HA76, porte d'entrée aluminium, {ligne.lower()}, "
                 f"référence {ref}.")
        if dimv:
            corps += f" {dimv}."
        if ud:
            corps += f" Sa performance thermique est de {ud}."
        vals = [v for v in lm + lM + hm + hM if v]
        if len(vals) == 8:
            corps += (f" Ses limites de fabrication vont de "
                      f"{int(min(lm))} à {int(max(lM))} mm de largeur hors tout et de "
                      f"{int(min(hm))} à {int(max(hM))} mm de hauteur hors tout, les limites "
                      f"exactes dépendant du profil de dormant retenu. Ces limites sont "
                      f"données pour une menuiserie en aluminium blanc équipée d'une serrure "
                      f"six points.")
        else:
            corps += " Le tarif ne porte pas de limites de fabrication pour ce modèle."
        f.add(titre, corps, page)
    return f


# ------------------------------------------------------------------ 5. EQUIPEMENTS
def f_equipements():
    f = Fichier("COMPAT_EQUIPEMENTS", "Compatibilité des équipements",
                "compatibilité de montage des équipements par modèle")
    eq = OrderedDict()
    for m, rs in groupes.items():
        for r in rs:
            nom = cell(r, "Y")
            if not nom:
                continue
            # le poussoir central est chiffré : il relève des options (règle 6, exception)
            if nom.lower().startswith("poussoir"):
                continue
            eq.setdefault(nom, OrderedDict())[m] = cell(r, "X")
    for nom, d in eq.items():
        oui = [m for m, s in d.items() if s == "check vert"]
        non = [m for m, s in d.items() if s == "croix rouge"]
        nom_aff = nom.rstrip("*").strip()
        if nom.endswith("**"):
            titre = (f"{PREFIXE} — Compatibilité du passe-lettres avec le poussoir central "
                     f"par modèle")
            note = (" Sur ces modèles, le tarif signale une incompatibilité entre le "
                    "passe-lettres et le poussoir central : la pose d'un passe-lettres entraîne "
                    "la suppression des deux rectangles du motif.")
        else:
            titre = f"{PREFIXE} — Compatibilité de l'équipement {nom_aff.lower()} par modèle"
            note = ""
        if nom.endswith("**"):
            corps = (f"Sur la gamme HA76, porte d'entrée aluminium, les modèles "
                     f"{' et '.join(oui)} portent un passe-lettres signalé par un renvoi de note "
                     f"au tarif.{note} La compatibilité du passe-lettres sur les autres modèles "
                     f"de la gamme fait l'objet d'un chunk distinct.")
        elif len(oui) <= 30:
            liste = ", ".join(oui)
            corps = (f"Sur la gamme HA76, porte d'entrée aluminium, l'équipement {nom_aff.lower()} "
                     f"peut équiper les modèles suivants : {liste}. Sur les autres modèles de la "
                     f"gamme, son montage n'est pas réalisable. Cet équipement n'est pas chiffré "
                     f"sur les pages modèles du tarif.{note}")
        else:
            corps = (f"Sur la gamme HA76, porte d'entrée aluminium, l'équipement {nom_aff.lower()} "
                     f"peut équiper {len(oui)} modèles et n'est pas réalisable sur les {len(non)} "
                     f"autres. La liste des modèles compatibles se lit sur les pages modèles du "
                     f"tarif. Cet équipement n'est pas chiffré sur ces pages.{note}")
        f.add(titre, corps, "30 à 121")
    return f


# ------------------------------------------------------------------ 6. CATALOGUE
# chapitres retenus : prix forfaitaires à l'unité. Exclus : €/m², €/ml, exemples.
CHAP_PAGES = {
    "Ferrage": 21,
    "PV vitrages panneaux": 122,
    "Garniture - bequilles et poignées standards": 123,
    "Garniture - béquilles design et rosettes": 124,
    "Garnitures - poussoirs et heutoirs": 125,
    "Options et accessoires": 126,
    "Meneaux complémentaire": 127,
    "Biométrie": 128,
}
CHAP_LIBELLE = {
    "Fixes": "fixes",
    "Ferrage": "ferrage",
    "PV vitrages panneaux": "plus-values de vitrages pour panneaux",
    "Garniture - bequilles et poignées standards": "garnitures, béquilles et poignées de tirage standards",
    "Garniture - béquilles design et rosettes": "garnitures, béquilles design et rosettes seules",
    "Garnitures - poussoirs et heutoirs": "garnitures, poussoirs et heurtoirs",
    "Options et accessoires": "options de vantail et de sécurité",
    "Meneaux complémentaire": "meneaux complémentaires",
    "Biométrie": "biométrie",
}
CHAP_EXCLUS = {
    "Vitrages": "plus-values au mètre carré",
    "Vitrages fixes": "plus-values au mètre carré",
    "Renforts, élargisseurs et tapées": "prix au mètre linéaire",
    "Bavettes": "prix au mètre linéaire",
    "Couvre-joints": "prix au mètre linéaire",
    "Exemple de calculs": "exemples de calcul internement incohérents",
}
# corrections de libellé consignées (règle d'arbitrage CA76 : le PDF fait foi)
CORRECTIONS = []


def f_catalogue():
    f = Fichier("CATALOGUE_OPTIONS", "Options de catalogue",
                "options, accessoires et garnitures tarifés indépendamment du modèle")
    for r in lignes_chapitre:
        ch = cell(r, "CF")
        if ch in CHAP_EXCLUS:
            continue
        if ch == "Fixes":
            tab = cell(r, "CG").strip()
            if tab != "Meneau battant":
                continue
            page = 28
        elif ch in CHAP_PAGES:
            page = CHAP_PAGES[ch]
        else:
            JOURNAL.append(f"chapitre non traité : {ch!r} ligne {r}")
            continue

        ht_brut = cell(r, "CK")
        if ht_brut == "":
            JOURNAL.append(f"ligne catalogue sans prix écartée : {ch} / {cell(r,'CG')} / {cell(r,'CH')}")
            continue

        tab = norm_texte(cell(r, "CG"))
        desig = norm_texte(cell(r, "CH"))
        det = norm_texte(cell(r, "CI"))

        # meneaux complémentaires : seule la plus-value forfaitaire est retenue
        if ch == "Meneaux complémentaire" and not desig.lower().startswith("plus-value"):
            JOURNAL.append("meneau complémentaire au mètre linéaire écarté (règle transverses)")
            continue

        if ch == "Fixes":
            longueur = det
            if longueur == "250":
                CORRECTIONS.append("meneau battant : longueur 250 mm de l'Excel rendue à 350 mm "
                                   "conformément au tableau de la page 28 du PDF")
                longueur = "350"
            libelle = f"meneau battant d'une longueur de {longueur} mm"
            titre = f"{PREFIXE} — Tarif du meneau battant de {longueur} mm"
        else:
            if desig.strip().lower() in ("pv", "prix", "plus-value"):
                desig = ""
            morceaux = [x for x in (tab, desig, det) if x]
            libelle = " — ".join(morceaux) if morceaux else CHAP_LIBELLE.get(ch, ch)
            court = libelle if len(libelle) < 70 else libelle[:67].rstrip() + "…"
            titre = f"{PREFIXE} — Tarif catalogue {CHAP_LIBELLE.get(ch, ch)} : {court}"

        ht, ttc = fmt_euro(ht_brut), fmt_euro(cell(r, "CL"))
        contexte = CHAP_LIBELLE.get(ch, ch)
        if to_num(ht_brut) == 0:
            corps = (f"Sur la gamme HA76, porte d'entrée aluminium, au chapitre {contexte}, la "
                     f"référence {libelle} est proposée sans plus-value, quel que soit le "
                     f"modèle de porte retenu.")
        elif ttc:
            corps = (f"Sur la gamme HA76, porte d'entrée aluminium, au chapitre {contexte}, la "
                     f"référence {libelle} est facturée {ht} € HT, soit {ttc} € TTC.")
        else:
            corps = (f"Sur la gamme HA76, porte d'entrée aluminium, au chapitre {contexte}, la "
                     f"référence {libelle} est facturée {ht} € HT.")
        if to_num(ht_brut) != 0:
            corps += " Ce montant est un forfait, indépendant du modèle de porte retenu."
        f.add(titre, corps, page)

    # information originale du PDF sans équivalent Excel
    f.add(f"{PREFIXE} — Tarif du meneau dormant",
          "Sur la gamme HA76, porte d'entrée aluminium, le meneau dormant de référence AL10111, "
          "désigné MD94 au tarif, ne fait l'objet d'aucune plus-value. Il est réalisable en "
          "configuration en T ou filante. Le meneau en croix n'est pas réalisable.", 127)
    return f


# ------------------------------------------------------------------ 7. FAISABILITES
def f_faisabilites():
    f = Fichier("FAISABILITES", "Faisabilités et restrictions",
                "faisabilités non tarifaires : teintes, limites dimensionnelles, fixes")

    f.add(f"{PREFIXE} — Faisabilité des teintes sur les références entièrement vitrées",
          "Sur la gamme HA76, porte d'entrée aluminium, les références Vitrage TRYBA, Crystal et "
          "Mesure acceptent les quatre familles de teintes du tarif : les teintes RAL des groupes "
          "1 et 2, la sublimation bois, l'anodisation et les laquages anodiques. Ce sont les "
          "seules références de la gamme sur lesquelles l'anodisation et les laquages anodiques "
          "sont réalisables.", 25)

    f.add(f"{PREFIXE} — Faisabilité des teintes sur la référence Panneau lisse",
          "Sur la gamme HA76, porte d'entrée aluminium, la référence Panneau lisse accepte les "
          "teintes RAL des groupes 1 et 2 ainsi que la sublimation bois. L'anodisation et les "
          "laquages anodiques ne sont pas réalisables sur cette référence.", 25)

    for ref in ("Excellence", "Passion", "Accord"):
        f.add(f"{PREFIXE} — Faisabilité des teintes sur la référence {ref}",
              f"Sur la gamme HA76, porte d'entrée aluminium, la référence {ref} accepte les "
              f"teintes RAL des groupes 1 et 2 ainsi que la sublimation bois, à l'exception de la "
              f"teinte CMX. L'anodisation et les laquages anodiques ne sont pas réalisables sur "
              f"cette référence.", 25)

    f.add(f"{PREFIXE} — Limites dimensionnelles hors tout en un vantail",
          "Sur la gamme HA76, porte d'entrée aluminium, une porte à un vantail se réalise de "
          "807 à 1 407 mm de largeur hors tout avec les dormants AL10101, dit L68, et AL10108, "
          "dit LZ108, et de 811 à 1 411 mm de largeur hors tout avec le dormant AL10100, dit L70. "
          "La largeur minimale est déterminée par le rayon d'ouverture et la saillie de la "
          "poignée ; la dimension maximale est déterminée par le poids de 160 kg accepté par les "
          "paumelles.", 17)

    f.add(f"{PREFIXE} — Limites dimensionnelles hors tout en deux vantaux",
          "Sur la gamme HA76, porte d'entrée aluminium, une porte à deux vantaux se réalise de "
          "1 094 à 2 700 mm de largeur hors tout avec les dormants AL10101 et AL10108, et de "
          "1 098 à 2 704 mm de largeur hors tout avec le dormant AL10100. La largeur minimale est "
          "déterminée par le rayon d'ouverture et la saillie de la poignée, le vantail semi-fixe "
          "imposant une largeur minimale de pareclose de 150 mm. La dimension maximale est "
          "déterminée par le poids de 160 kg accepté par les paumelles et par l'inertie du profil "
          "ouvrant semi-fixe AL10105.", 17)

    f.add(f"{PREFIXE} — Réalisation des fixes latéraux vitrés et impostes vitrées",
          "Sur la gamme HA76, porte d'entrée aluminium, les fixes latéraux vitrés et les impostes "
          "vitrées se réalisent à partir de 300 mm, tous dormants confondus. La dimension de "
          "300 mm par 300 mm n'est techniquement pas réalisable. Le prix de ces fixes se lit sur "
          "la table dimensionnelle de la page 28 du tarif ; il n'est pas repris dans ce corpus.", 28)

    f.add(f"{PREFIXE} — Configurations de meneaux réalisables",
          "Sur la gamme HA76, porte d'entrée aluminium, le meneau dormant AL10111, dit MD94, et le "
          "meneau battant AL10106J, dit MB94, ne se réalisent qu'en configuration en T. Le meneau "
          "en croix n'est réalisable dans aucune des deux configurations. La longueur du meneau "
          "battant se calcule à partir de la largeur hors tout, dont on retire 229 mm avec les "
          "dormants AL10101 et AL10108, et 233 mm avec le dormant AL10100.", 28)

    f.add(f"{PREFIXE} — Modèles tarifés en un seul vantail",
          "Sur la gamme HA76, porte d'entrée aluminium, quinze modèles ne sont tarifés qu'en un "
          "vantail et n'ont aucun prix en deux vantaux : Bordeaux, Brest, Grenoble, Lille, Lyon, "
          "Marseille, Montpellier, Nantes, Nice, Nîmes, Reims, Rennes, Strasbourg, Toulon et "
          "Toulouse. Ces quinze modèles relèvent tous de la référence Porte du mois permanente.", 9)
    return f


# ------------------------------------------------------------------ 8. TRANSVERSES
def f_transverses():
    f = Fichier("TRANSVERSES", "Pages transverses",
                "existence et localisation des tarifs transverses, sans montant")

    blocs = [
        ("l'offre couleurs", 23,
         "L'offre couleurs de la gamme HA76, porte d'entrée aluminium, répartit les teintes en "
         "groupes. Le groupe 1 est sans plus-value ; les autres groupes, ainsi que les teintes RAL "
         "granité dites autres, appliquent une plus-value exprimée en pourcentage du prix du "
         "châssis. Ce corpus ne reproduit aucun de ces pourcentages, car leur usage suppose un "
         "calcul. Les taux et la composition de chaque groupe se lisent aux pages 23 et 24 du "
         "tarif."),
        ("la couleur des accessoires", 26,
         "La gamme HA76, porte d'entrée aluminium, distingue les teintes claires des teintes "
         "sombres pour déterminer la couleur des accessoires. Cette définition conditionne le "
         "choix des accessoires mais ne porte aucun montant. Le classement de chaque teinte se lit "
         "à la page 26 du tarif."),
        ("les plus-values de vitrages pour porte entièrement vitrée", 27,
         "Les plus-values de vitrages applicables à une porte HA76 entièrement vitrée sont "
         "exprimées au mètre carré de surface vitrée du châssis. Ce corpus ne reproduit aucun de "
         "ces montants, car leur usage suppose une multiplication par la surface. La grille, avec "
         "les performances thermiques et les codes de vitrage associés, se lit à la page 27 du "
         "tarif."),
        ("les plus-values de vitrages pour fixes", 29,
         "Les plus-values de vitrages applicables aux fixes de la gamme HA76 sont exprimées au "
         "mètre carré de surface vitrée. Ce corpus ne reproduit aucun de ces montants, car leur "
         "usage suppose une multiplication par la surface. La grille se lit à la page 29 du tarif."),
        ("la table de tarif des fixes latéraux vitrés et impostes vitrées", 28,
         "Le tarif des fixes latéraux vitrés et des impostes vitrées de la gamme HA76 se lit sur "
         "une table à double entrée, en largeur et en hauteur. Ce corpus ne reproduit aucune "
         "valeur de cette table : le tarif n'énonce pas comment lire une dimension intermédiaire "
         "entre deux colonnes ou deux lignes de la table, si bien qu'aucune réponse ne peut être "
         "donnée sans risque d'arrondi. La table se lit à la page 28 du tarif."),
        ("les habillages, renforts, élargisseurs et tapées", 129,
         "Les renforts statiques, élargisseurs, compléments d'habillage et tapées tubulaires de la "
         "gamme HA76 sont tarifés au mètre linéaire, coupés sur mesure, avec un montant qui varie "
         "selon la teinte. Ce corpus ne reproduit aucun de ces montants, car leur usage suppose "
         "une multiplication par la longueur. La grille se lit à la page 129 du tarif."),
        ("les bavettes extérieures", 130,
         "Les bavettes extérieures de la gamme HA76 sont tarifées au mètre linéaire, avec un "
         "montant qui varie selon la teinte. Ce corpus ne reproduit aucun de ces montants. La "
         "grille se lit à la page 130 du tarif."),
        ("les couvre-joints", 131,
         "Les couvre-joints intérieurs, extérieurs et spécial rénovation de la gamme HA76 sont "
         "tarifés au mètre linéaire, avec un montant qui varie selon la teinte. Ce corpus ne "
         "reproduit aucun de ces montants. La grille se lit à la page 131 du tarif."),
        ("les meneaux facturés à la longueur", 127,
         "Le meneau battant de la gamme HA76 peut être facturé à la longueur réelle, mesurée en "
         "fond de feuillure, à un prix au mètre linéaire. Ce corpus ne reproduit pas ce montant. "
         "Il se lit à la page 127 du tarif. La plus-value forfaitaire pour deux fixations à angle "
         "droit, elle, figure dans le corpus."),
        ("les exemples de calcul", 132,
         "Le tarif de la gamme HA76 comporte des exemples de calcul complets. Ce corpus ne les "
         "reprend pas : les totaux annoncés ne correspondent pas à la somme des lignes qui les "
         "composent. Ces exemples se lisent à la page 132 du tarif et ne doivent pas servir de "
         "référence de prix."),
    ]
    for nom, page, corps in blocs:
        f.add(f"{PREFIXE} — Existence et localisation des tarifs de {nom}", corps, page)
    return f


# ------------------------------------------------------------------ message produit
def message_produit(stats):
    txt = f"""# Message au service Produits — anomalies relevées sur le tarif HA76

Objet : anomalies de documentation relevées lors de la migration du tarif HA76 vers le
corpus du chatbot ADV.

Bonjour,

Dans le cadre de la préparation du corpus du chatbot ADV, le tarif HA76 du 23 juin 2026
et son fichier Excel associé ont été analysés ligne à ligne. Sept anomalies ont été
relevées. Aucune n'a été corrigée d'autorité : les points portant sur une valeur ont été
gelés dans le corpus, en attente de votre arbitrage. Un seul point, portant sur un libellé
manifestement fautif, a été rectifié et est signalé ci-dessous.

## 1. Prix TTC incohérents sur les quinze modèles Porte du mois — bloquant

Sur les quinze modèles de la référence Porte du mois permanente (Bordeaux, Brest,
Grenoble, Lille, Lyon, Marseille, Montpellier, Nantes, Nice, Nîmes, Reims, Rennes,
Strasbourg, Toulon, Toulouse), le prix HT en un vantail est constant, mais le prix TTC
varie entre les lignes d'un même modèle et se trouve inférieur au prix HT.

Exemples : Bordeaux est à 3 595 € HT, avec un TTC porté tantôt à 3 490 €, tantôt à
3 593 €. Toulouse est à 4 174 € HT, avec les mêmes valeurs TTC de 3 490 € puis 3 593 €.
Soixante occurrences au total.

Le PDF joint étant une édition HT seule, la source de contrôle ne permet pas de trancher.
En conséquence, le corpus ne sert aucun prix TTC pour ces quinze modèles : il donne le
prix HT et renvoie l'ADV à l'édition TTC du tarif. Merci de nous indiquer les valeurs TTC
exactes.

## 2. Longueur de meneau battant divergente entre l'Excel et le PDF

Le tableau des meneaux battants de la page 28 du PDF donne les longueurs 320, 350, 400,
500 mm et suivantes. L'Excel porte 320, puis 250, puis 400 mm pour les mêmes prix de 25,
30 et 33 €. La valeur 250 rompt la progression et contredit le PDF.

S'agissant d'un libellé et non d'un montant, et le PDF se déclarant document de
référence, la valeur a été rectifiée à 350 mm dans le corpus. Merci de confirmer, et de
corriger l'Excel source.

## 3. Sommaire décalé de quatre pages sur toute la seconde moitié du tarif

Le sommaire de la page 3 annonce le chapitre Options et accessoires à la page 118, les
Meneaux complémentaires à la page 123 et les Évolutions du tarif à la page 131. Le
contenu réel et les pieds de page situent ces chapitres aux pages 122, 127 et 135. Le
décalage est de quatre pages et concerne tout ce qui suit les pages modèles. La section
avant, jusqu'à la page 29, est exacte.

Un ADV qui suit le sommaire arrive systématiquement quatre pages trop tôt.

## 4. Page étrangère insérée en fin de document

La page 137 du PDF est un document intitulé « TARIF PORTES DE GARAGE — JUILLET 2018 —
CONCESSIONNAIRES », sans rapport avec la gamme HA76 et antérieur de huit ans. Il est
inséré à la suite du tarif. Merci d'en confirmer le retrait.

## 5. Fiche Info Produit insérée dans le tarif

La page 20 du tarif reproduit la Fiche Info Produit HA76 d'avril 2024. Ce document est
par ailleurs diffusé et migré pour lui-même. Sa présence dans le tarif crée un risque de
double source si son édition évolue d'un côté sans évoluer de l'autre. Elle a été exclue
du périmètre du corpus tarif.

## 6. Exemples de calcul internement incohérents

À la page 132, l'exemple « Porte vitrée » additionne 2 776 € pour le modèle, 116 € pour
le soubassement, 38 € pour le vitrage chinchilla et 64 € pour le meneau battant, soit
2 994 € HT, mais annonce un total de 2 922 € HT. En TTC, la somme des lignes donne
3 801 € pour un total annoncé de 3 797 €. Les autres exemples de la page présentent le
même type d'écart.

Ces exemples ont été exclus du corpus. Ils restent visibles pour l'ADV dans le tarif
papier, où ils peuvent induire un chiffrage faux.

## 7. Libellé d'option manquant sur le panneau analogue

Cinq modèles portent une plus-value de panneau analogue chiffrée dans l'Excel — 1 032 €
sur Blizzard, Lombarde et Sirocco, 2 160 € sur Paris et Paris plein — accompagnée d'une
description technique, mais la colonne de libellé de l'option est vide dans l'Excel. Le
montant existe sans que l'option porte un nom.

La page 95 du PDF, vérifiée visuellement, intitule ce bloc « Panneau analogue pour fixe
latéral » et précise que la plus-value s'ajoute au prix de grille de la page Fixes. Le
libellé a donc été restitué depuis le PDF et les cinq plus-values sont servies. Merci de
confirmer cet intitulé et de renseigner la colonne dans l'Excel source.

Ces plus-values restent d'un usage partiel tant que la grille des fixes n'est pas
intégrée, puisqu'elles s'y ajoutent (voir le point d'arbitrage ci-dessous).

## Point d'arbitrage complémentaire — règle de lecture des fixes

La page 28 indique que la table de tarif des fixes vitrés est « en lecture directe, en
fonction de la dimension L x H du fixe ». Elle n'énonce pas comment traiter une dimension
qui tombe entre deux colonnes ou entre deux lignes de la table — par exemple une largeur
de 650 mm. Le tarif T81 énonce explicitement, lui, qu'une colonne couvre une bande de
dimensions.

En l'absence de règle écrite, les 553 prix de cette table n'ont pas été intégrés au
corpus : les servir supposerait un arrondi implicite, donc un prix potentiellement faux.
Merci de nous préciser la règle de lecture applicable, ce qui permettra d'intégrer ce
chapitre.

## Récapitulatif

| Point | Nature | Traitement dans le corpus |
|---|---|---|
| 1. TTC Porte du mois | valeur | TTC gelé sur 15 modèles |
| 2. Longueur meneau 250 mm | libellé | rectifié à 350 mm d'après le PDF |
| 3. Sommaire décalé | navigation | pieds de page retenus |
| 4. Page portes de garage | corps étranger | exclue |
| 5. FIP insérée | double source | exclue |
| 6. Exemples de calcul | valeur | exclus |
| 7. Libellé panneau analogue | libellé manquant dans l'Excel | restitué depuis le PDF page 95 |
| Fixes | règle de lecture absente | chapitre gelé, 553 prix non servis |

Corpus produit : {stats} chunks répartis en huit fichiers.

Bien cordialement,
"""
    path = f"{OUTDIR}/Message_service_produit_HA76.md"
    open(path, "w", encoding="utf-8").write(txt)
    return path


# ------------------------------------------------------------------ journal colonnes
MAPPEES = set("A B C D E F G H I J K L M N O T U W X Y AE AF AG AS AV AX AY BA BB BD BE "
              "BG BI BJ BO BQ BR BS AK AL AM AN AO AP AQ AR CF CG CH CI CK CL".split())


def journal_colonnes():
    for c in range(1, ws.max_column + 1):
        letter = CL(c)
        if letter in MAPPEES:
            continue
        n = sum(1 for r in range(2, ws.max_row + 1) if cell(r, letter))
        if n:
            JOURNAL.append(f"colonne remplie non mappée : {letter} [{HDR[letter]}] — {n} lignes")


# ------------------------------------------------------------------ main
if __name__ == "__main__":
    fichiers = [f_methode(), f_prix(), f_options(), f_caracteristiques(),
                f_equipements(), f_catalogue(), f_faisabilites(), f_transverses()]
    journal_colonnes()
    total = 0
    print("=== Fichiers produits ===")
    for f in fichiers:
        path, n = f.write()
        total += n
        print(f"  {path.split('/')[-1]:38s} {n:4d} chunks")
    print(f"  {'TOTAL':38s} {total:4d} chunks")
    print(f"\n=== Message service produits ===\n  {message_produit(total).split('/')[-1]}")
    print("\n=== Corrections consignées ===")
    for c in sorted(set(CORRECTIONS)):
        print("  -", c)
    print("\n=== Journal ===")
    for j in JOURNAL:
        print("  -", j)
    print("\n=== Alertes plafond ===")
    for a in ALERTS:
        print("  !", a)
    if not ALERTS:
        print("  aucune")
