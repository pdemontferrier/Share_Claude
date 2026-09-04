#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Générateur de chunks Markdown pour le tarif CA76 / CAG76
(coulissant aluminium et coulissant à galandage, collection TRYBA ALUMINIUM).

Conforme à note_cadrage_migration_tarif_CA76_v1.md (règles C1 à C8), elle-même
dérivée de la note T81 (dimensionnelle) et non de la note H81 (forfaitaire).

Six fichiers :
  F1 METHODE       cotes, lecture par bandes, méthodes de calcul      (règle C3)
  F2 PRIX_CHASSIS  les huit grilles dimensionnelles              (règles C1, C2)
  F3 OPTIONS       plus-values forfaitaires chiffrées                 (règle C4)
  F4 PROPORTION.   plus-values en pourcentage, absentes de l'Excel    (règle C5)
  F5 FAISABILITES  restrictions et impossibilités, sans montant       (règle C6)
  F6 TRANSVERSES   existence et localisation, sans montant            (règle C7)

Extensions par rapport au gabarit T81 :
  1. clause de seuil de croisée renforcée, récupérée de la GÉOMÉTRIE du PDF
     (remplissage de cellule) et absente de toute couche textuelle ;
  2. table des pourcentages alimentée depuis le PDF, l'Excel n'en portant aucun ;
  3. distinction des TROIS états de cellule vide : zéro (sans plus-value),
     impossibilité produit, non-renseigné ;
  4. découpage borné par la page PDF, une grille s'étalant sur deux ou trois
     pages : un chunk ne cite jamais une page où son contenu ne figure pas ;
  5. double préfixe de gamme, le tarif couvrant deux produits (CA76 et CAG76).

Principes hérités : fidélité numérique (toute valeur recopiée de la cellule,
jamais calculée), anti-fantôme, non-invention, SC continue par fichier depuis
SC0002, ligne de source normée, plafond de 200 mots, journal exhaustif.
"""
import re
import sys
from collections import OrderedDict, defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

XLSX = "/mnt/user-data/uploads/CA76_-infos-tarifs.xlsx"
PDF = "/mnt/user-data/uploads/Tarif_CA76_HT_19-06-2026.pdf"
FEUILLE = "Feuil2"
PDF_SOURCE = "Tarif—CA76—HT—19-06-2026.pdf"     # nom affiché, em-dashes
PDF_YAML = "Tarif_CA76_HT_19-06-2026.pdf"       # nom en front matter, underscores
OUTDIR = "/mnt/user-data/outputs"
PLAFOND = 200

GAMME = "CA76"
PREFIXE_CA = "CA76 Coulissant Aluminium — "
PREFIXE_CAG = "CAG76 Coulissant Aluminium à galandage — "

# ====================================================== index de colonnes
# ATTENTION : l'ordre diffère de T81 (chap, tableau, désignation, détails).
# Ici : gamme, clé, chapitre, DÉSIGNATION, DÉTAILS, TABLEAU.
C_GAMME, C_CLE, C_CHAP, C_DES, C_DET, C_TAB = 0, 1, 2, 3, 4, 5
C_MHT, C_MTTC = 6, 7                      # colonnes constantes, non reprises
C_HT, C_TTC, C_PVHT, C_PVTTC = 8, 9, 10, 11
C_HAUTEUR = 12
COLS_HT_L = list(range(13, 72))           # Px L 600..6400 HT
COLS_TTC_L = list(range(72, 131))         # Px L 600..6400 TTC
NCOL = 131

# ====================================================== table des pages
# Établie contre les EN-TÊTES et les PIEDS de page du PDF, relevés page par page.
# Le sommaire général (p. 3) et les sommaires de section (p. 4, 10, 21, 25, 31,
# 42, 46, 51, 57) sont DÉCALÉS DE +1 à partir de la page 36 — la page
# « Plus-values vitrages compositions libres » n'y est pas indexée — et ne sont
# donc PAS utilisés. L'audit revérifie chaque attribution contre le PDF.
PAGES = {
    ("Laquage bloc-baie", "forfait laquage"): 30,
    ("Vitrage", "vitrage altitude"): 33,
    ("Vitrage", "pv vitrage thermique, pho, secu"): 34,
    ("Vitrage ornementaux", "Vitrage ornementaux"): 35,
    ("PV composition libre", "composition libre"): 36,
    ("Remplissage", "pv soubassements"): 37,
    ("Croisillons", "Alu laqué"): 38,
    ("Croisillons", "Alu 10 mm"): 39,
    ("Croisillons", "motif à la grecque"): 39,
    ("Croisillons", "/ champ"): 40,
    ("Croisillons", "gravure 8 à 10 mm"): 40,
    ("Croisillons", "croisillons rapportés en alu"): 41,
    ("Meneaux complémentaires", "vitrage de 28 et 32 mm"): 43,
    ("Meneaux complémentaires", "double meneau"): 43,
    ("Grilles d'entrée d'air", "grille d'air"): 45,
    ("Poignée-croisement possible", "Poignée standard"): 47,
    ("Poignée-croisement impossible", "Poignée en option"): 47,
    ("Poignée-croisement impossible", "Poignée fixe en option"): 48,
    ("Ferrage standard", "pv rabaisser poignée"): 50,
    ("Ferrage standard", "verrou médian"): 50,
    ("Pièces appui, élargisseurs", "Pieces d'appui"): 52,
    ("Pièces appui, élargisseurs", "Profilé complémentaires"): 52,
    ("Elargisseurs, tapées", "u d'assemblage - cache rainure"): 53,
    ("Elargisseurs, tapées", "élargisseur et complément d'habillage"): 53,
    ("Tapées de doublage", "élargisseur et complément d'habillage"): 54,
    ("Bavettes extérieures", "bavettes extérieurs"): 55,
    ("Couvre-joints", "couvre joints int+ext"): 56,
    ("Couvre-joints", "couvre joints spécial réno"): 56,
}

# L'Excel étiquette une même table sur deux chapitres différents : la première
# ligne des tapées de doublage porte le chapitre « Tapées de doublage », les deux
# suivantes le chapitre « Elargisseurs, tapées ». Défaut de saisie relevé à
# l'étape 2. La réattribution se fait sur la DÉSIGNATION et corrige d'un coup le
# chapitre, la page et le libellé.
CHAP_OVERRIDE_DES = {
    ("Elargisseurs, tapées", "Tapées de doublage"): "Tapées de doublage",
}

# ====================================================== grilles (règle C1)
# (chapitre, tableau) -> (préfixe, libellé long, synonyme d'usage, rails)
GRILLES = OrderedDict([
    (("2 vantaux 2 rails", "2V 2 rails"),
     (PREFIXE_CA, "coulissant 2 vantaux sur 2 rails",
      "baie coulissante à deux vantaux", "2 rails")),
    (("3 vantaux 2 rails", "3V 2 rails"),
     (PREFIXE_CA, "coulissant 3 vantaux sur 2 rails",
      "baie coulissante à trois vantaux", "2 rails")),
    (("4 vantaux 2 rails", "4V 2 rails"),
     (PREFIXE_CA, "coulissant 4 vantaux sur 2 rails",
      "baie coulissante à quatre vantaux", "2 rails")),
    (("3 vantaux 3 rails", "3V 3 rails"),
     (PREFIXE_CA, "coulissant 3 vantaux sur 3 rails",
      "baie coulissante à trois vantaux sur trois rails", "3 rails")),
    (("6 vantaux 3 rails", "6V 3 rails"),
     (PREFIXE_CA, "coulissant 6 vantaux sur 3 rails",
      "baie coulissante à six vantaux sur trois rails", "3 rails")),
    (("Galandage 1V", "Galandage 1V"),
     (PREFIXE_CAG, "coulissant à galandage 1 vantail",
      "baie coulissante à galandage à un vantail", "1 rail")),
    (("Galandage 2V", "Galandage 2V"),
     (PREFIXE_CAG, "coulissant à galandage 2 vantaux",
      "baie coulissante à galandage à deux vantaux", "1 ou 2 rails")),
    (("Galandage 4 V", "Galandage 4 V"),
     (PREFIXE_CAG, "coulissant à galandage 4 vantaux",
      "baie coulissante à galandage à quatre vantaux", "1 ou 2 rails")),
])

# Extension 4 : une grille s'étale sur plusieurs pages du PDF. Chaque tranche de
# largeurs a SA page ; le découpage ne franchit jamais une frontière de page,
# pour qu'un chunk ne cite jamais une page où son contenu ne figure pas.
PAGES_GRILLE = {
    "2V 2 rails":    [(11, 1000, 2600), (12, 2700, 4200)],
    "3V 2 rails":    [(13, 1600, 3500), (14, 3600, 5400)],
    "4V 2 rails":    [(15, 2100, 3500), (16, 3600, 5000), (17, 5100, 6400)],
    "3V 3 rails":    [(18, 2100, 3800), (19, 3900, 5500)],
    "6V 3 rails":    [(20, 4200, 6300)],
    "Galandage 1V":  [(22, 600, 1800)],
    "Galandage 2V":  [(23, 1000, 3100)],
    "Galandage 4 V": [(24, 2100, 4300)],
}

# ====================================================== unités (règle C4)
# Relevées PAGE PAR PAGE dans le PDF. L'Excel n'en porte aucune. Une entrée
# absente => unité non établie : le chunk le dit et renvoie à la page.
UNITES = {
    ("Vitrage", "pv vitrage thermique, pho, secu", None):
        "par mètre carré de surface vitrée du châssis, la surface minimale de "
        "facturation étant de 0,5 mètre carré",
    ("Vitrage ornementaux", None, None):
        "par mètre carré de surface vitrée du châssis, la surface minimale de "
        "facturation étant de 0,5 mètre carré",
    ("PV composition libre", None, None):
        "par mètre carré de surface vitrée du châssis",
    ("Remplissage", "pv soubassements", None): "par mètre carré de panneau",
    ("Croisillons", "Alu laqué", None): "par champ, et non au mètre linéaire",
    ("Croisillons", "Alu 10 mm", None): "par champ, et non au mètre linéaire",
    ("Croisillons", "/ champ", None): "par champ, et non au mètre linéaire",
    ("Croisillons", "croisillons rapportés en alu", None):
        "par champ, et non au mètre linéaire",
    ("Croisillons", "gravure 8 à 10 mm", None): "par volume gravé",
    ("Croisillons", "motif à la grecque", None):
        "forfaitaire, par châssis, les quatre angles compris",
    ("Meneaux complémentaires", None, None):
        "par mètre linéaire de longueur réelle, mesurée en fond de feuillure",
    ("Grilles d'entrée d'air", "grille d'air", None):
        "forfaitaire, pour l'ensemble d'une grille, mortaise et vis comprises",
    ("Ferrage standard", "verrou médian", None): "par pièce",
    ("Pièces appui, élargisseurs", "Pieces d'appui", None):
        "par mètre linéaire, posé sur châssis",
    ("Pièces appui, élargisseurs", "Profilé complémentaires", None):
        "par mètre carré, posé sur châssis",
    ("Elargisseurs, tapées", None, None):
        "par mètre linéaire d'habillage coupé sur mesure",
    ("Tapées de doublage", None, None):
        "par mètre linéaire, posé sur châssis",
    ("Bavettes extérieures", None, None):
        "par mètre linéaire d'habillage coupé sur mesure",
    ("Couvre-joints", None, None):
        "par mètre linéaire d'habillage coupé sur mesure",
    ("Laquage bloc-baie", "forfait laquage", None):
        "forfaitaire, par volet roulant",
    # unité non établie sur la page : traitement explicite plutôt que silence
    ("Vitrage", "vitrage altitude", None):
        "forfaitaire, aucune unité n'étant indiquée sur la page",
    ("Poignée-croisement possible", None, None):
        "forfaitaire, aucune unité n'étant indiquée sur la page",
    ("Poignée-croisement impossible", None, None):
        "forfaitaire, aucune unité n'étant indiquée sur la page",
    ("Ferrage standard", "pv rabaisser poignée", None):
        "forfaitaire, aucune unité n'étant indiquée sur la page",
}

# ====================================================== contextes de libellé
CONTEXTE = {
    ("Vitrage", "vitrage altitude"): "vitrage d'altitude",
    ("Vitrage", None): "vitrage",
    ("PV composition libre", None): "vitrage en composition libre",
    ("Remplissage", None): "panneau de soubassement",
    ("Croisillons", "gravure 8 à 10 mm"): "motif Art Déco",
    ("Croisillons", "motif à la grecque"): "croisillon à la grecque",
    ("Croisillons", "croisillons rapportés en alu"): "croisillon rapporté en aluminium",
    ("Croisillons", "Alu 10 mm"): "croisillon en aluminium de 10 mm",
    ("Croisillons", "Alu laqué"): "croisillon intégré en aluminium laqué",
    ("Croisillons", None): "croisillon",
    ("Meneaux complémentaires", "double meneau"): "double meneau complémentaire",
    ("Meneaux complémentaires", None): "meneau complémentaire",
    ("Grilles d'entrée d'air", None): "grille d'entrée d'air",
    ("Poignée-croisement possible", None): "poignée",
    ("Poignée-croisement impossible", "Poignée fixe en option"): "poignée fixe en option",
    ("Poignée-croisement impossible", None): "poignée en option",
    ("Ferrage standard", "verrou médian"): "verrou médian",
    ("Ferrage standard", None): "ferrage Secure+",
    ("Pièces appui, élargisseurs", "Pieces d'appui"): "pièce d'appui",
    ("Pièces appui, élargisseurs", None): "profilé complémentaire",
    ("Elargisseurs, tapées", "u d'assemblage - cache rainure"):
        "U d'assemblage et cache rainure",
    ("Elargisseurs, tapées", None): "élargisseur et complément d'habillage",
    ("Tapées de doublage", None): "tapée de doublage",
    ("Bavettes extérieures", None): "bavette extérieure",
    ("Couvre-joints", "couvre joints spécial réno"): "couvre-joint spécial rénovation",
    ("Couvre-joints", None): "couvre-joint intérieur et extérieur",
    ("Laquage bloc-baie", None): "forfait de laquage du volet roulant",
}

# ====================================================== discriminants (PDF p.38)
# Trois prix distincts sous des colonnes strictement identiques dans l'Excel.
# La finition n'existe que dans la colonne Désignation du tableau du PDF.
# RATTACHEMENT PAR LE MONTANT, jamais par l'ordre des lignes.
# L'Excel fusionnant les profils I18 et I26, le troisième prix vaut chêne d'or
# sur l'un et tons bois sur l'autre : le chunk nomme les deux, la fusion venant
# de la source. Aucune séparation n'est inventée.
DISCRIMINANTS = {
    ("Croisillons", "Alu laqué", "I18+26", 18): "laqué blanc RAL 9016, en 18 ou 26 mm",
    ("Croisillons", "Alu laqué", "I18+26", 31): "laqué RAL, en 18 ou 26 mm",
    ("Croisillons", "Alu laqué", "I18+26", 23): "chêne d'or en 18 mm ou tons bois en 26 mm",
    ("Croisillons", "Alu laqué", "I18+26", 14): "laqué blanc RAL 9016, en 18 ou 26 mm",
    ("Croisillons", "Alu laqué", "I18+26", 28): "laqué RAL, en 18 ou 26 mm",
    ("Croisillons", "Alu laqué", "I18+26", 20): "chêne d'or en 18 mm ou tons bois en 26 mm",
    ("Croisillons", "Alu laqué", "I45", 37): "laqué blanc RAL 9016, en 45 mm",
    ("Croisillons", "Alu laqué", "I45", 50): "laqué RAL, en 45 mm",
    ("Croisillons", "Alu laqué", "I45", 25): "laqué blanc RAL 9016, en 45 mm",
    ("Croisillons", "Alu laqué", "I45", 39): "laqué RAL, en 45 mm",
}

# ====================================================== libellés fautifs
# Règle d'arbitrage : une divergence portant sur une VALEUR est exposée avec
# attribution ; une divergence portant sur un LIBELLÉ ou un CODE DE RÉFÉRENCE,
# dont l'une des deux sources est manifestement fautive, est tranchée en faveur
# du PDF, qui se déclare document de référence à sa page 2 et qui est celui que
# l'ADV a sous les yeux. Chaque correction est consignée au journal.
# Aucune de ces corrections ne touche un montant.
LIBELLES_CORRIGES = {
    "vitrage de 28 et 32 mm": ("vitrage de 28 et 36 mm", 43),
    "Blanc et ral 7017": ("Blanc et RAL 7016 granité", 54),
    "Teinte Std grp 2": ("Teinte standard groupe 1", 54),
    "FR15_DV": ("FR15", 45),
    "ME30_DV": ("ME30_CE", 45),
    "ME30+RA_DV": ("ME30+RA_CE", 45),
    "FR12_DV": ("FR12", 45),
    "ISOLA2-45_DV": ("ISOLA2-45_CE", 45),
    "ISOLA245+RA_DV": ("ISOLA245+RA_CE", 45),
    "ISOLA-HY_DV": ("ISOLA-HY_CE", 45),
    "ISOLA-HY+RA_DV": ("ISOLA-HY+RA_CE", 45),
}


def corriger_libelle(txt):
    t = " ".join(str(txt or "").split())
    # En-tête de colonne recopié dans la désignation : « Prix  Blanc et RAL 7016 Gr ».
    # Le PDF n'a pas de désignation de ce nom, il a une colonne « Prix HT ».
    m = re.match(r"^Prix\s+(.+)$", t)
    if m and not re.match(r"^Prix\s+HT", t):
        JOURNAL.append(f"LIBELLÉ NETTOYÉ, en-tête de colonne recopié dans la "
                       f"désignation : « {t} » → « {m.group(1)} »")
        t = m.group(1)
    if t in LIBELLES_CORRIGES:
        bon, page = LIBELLES_CORRIGES[t]
        JOURNAL.append(f"LIBELLÉ CORRIGÉ d'après le PDF page {page} : "
                       f"« {t} » → « {bon} » (aucun montant touché)")
        return bon
    return t


# ====================================================== divergences exposées
# Règle : une divergence portant sur une VALEUR n'est jamais arbitrée en
# silence. Le chunk sert la valeur de la source primaire, puis expose l'autre
# lecture avec attribution de page, jusqu'à l'arbitrage produit.
DIVERGENCES = {
    ("Poignée-croisement impossible", "Poignée en option", "Delta Flap", "prix"):
        "Le tarif diverge sur ce poste : l'Excel des tarifs le chiffre à ce "
        "montant, alors que la page 47 du tarif range la Delta Flap parmi les "
        "poignées standard et porte la mention « sans PV » pour l'ensemble de la "
        "ligne. L'arbitrage est en cours ; les deux lectures sont rapportées.",
    ("Poignée-croisement possible", "Poignée standard", "Novem", "traversant"):
        "Le tarif diverge sur ce poste : l'Excel des tarifs le chiffre à ce "
        "montant, alors que la page 47 marque le cylindre traversant impossible "
        "sur la Novem et que la page 49 des faisabilités le confirme, toutes "
        "teintes confondues. L'arbitrage est en cours ; les deux lectures sont "
        "rapportées.",
}


# ====================================================== nature de la colonne Détails
# La colonne Détails est hétérogène : elle porte tantôt une référence produit,
# tantôt une épaisseur, un motif, une teinte, une configuration de cylindre, ou
# une simple annotation de saisie. Annoncer indistinctement « la référence X »
# affirme une catégorie fausse dans la majorité des cas — défaut invisible pour
# l'audit, relevé à la relecture d'échantillon. Chaque nature est ici rattachée
# à la page du PDF qui la nomme ; à défaut, la formulation reste neutre.
DETAIL_NATURE = {
    "Remplissage": "l'épaisseur de {} mm",
    "Poignée-croisement possible": "la configuration {}",
    "Poignée-croisement impossible": "la configuration {}",
}
DETAIL_MOTIF = {"MG1 à 7 sauf 2", "MG2", "MG8", "MG9", "MGE2B et MGE3B"}
# Annotations de saisie, sans valeur pour l'ADV : écartées et consignées.
DETAIL_ANNOTATION = {"prix", "prix + forfait", "0.15"}
# Mise en phrase des détails repris dans un titre : l'Excel les abrège, un titre
# auto-porteur ne peut pas dire « Poignée Novem traversant ». Aucune information
# n'est ajoutée, seule l'abréviation est développée d'après la page citée.
DETAIL_LIBELLE = {
    "laquée ral": "en laquage RAL",
    "demi-cylindre": "avec demi-cylindre",
    "demi-cylindre int": "avec demi-cylindre intérieur",
    "traversant": "avec cylindre traversant",
    "cylindre traversant": "avec cylindre traversant",
    "Autre couleur": "en autre couleur",
    "Blanc et ral 7016": "en blanc ou RAL 7016 granité",
    "Teinte Std grp 1": "en teinte standard groupe 1",
    "28": "en épaisseur 28 mm",
    "36": "en épaisseur 36 mm",
}


def dire_detail(x):
    """Développe l'abréviation ; à défaut rend la valeur telle quelle. Les
    valeurs listées avec « et » dans la source sont éclatées pour que
    l'énumération ne produise pas deux « et » consécutifs."""
    t = " ".join(str(x or "").split())
    if t in DETAIL_LIBELLE:
        return [DETAIL_LIBELLE[t]]
    if " et " in t:
        return [p.strip() for p in t.split(" et ") if p.strip()]
    return [t] if t else []


# Type de jonction (PDF p. 38 : « Croisillon T ou Croix » / « Croisillon filant »).
# L'Excel abrège et orthographie « fillant » : libellé de structure, le tarif fait foi.
JONCTIONS = {
    "croisillons t": "en T ou en croix", "croisillon t": "en T ou en croix",
    "croisillons fillant": "filant", "croisillon fillant": "filant",
}

# ====================================================== exclusions de périmètre
CHAP_GRILLE = {k[0] for k in GRILLES}
CHAP_EXCLUS = {
    # totaux additionnés ET périmés : la grille donne 2 873 € et 8 459 € là où
    # la page 58 annonce 2 930 € et 8 890 €. Aucun montant migré.
    "Exemple de calcul": "exemples de calcul périmés (écart vérifié contre les grilles)",
    # 19 valeurs divergentes entre Excel et PDF, décalage systématique d'une
    # ligne. GELÉ tant que l'arbitrage produit n'est pas rendu.
    "Vitrage ornementaux": "divergence Excel/PDF sur 19 valeurs, arbitrage en cours",
}

# Extension 3 : troisième état de cellule vide. Une case blanche vaut ici soit
# une impossibilité produit, soit un non-renseigné. Les impossibilités relevées
# contre le PDF sont consignées et reprises en faisabilité (F5), pas en prix.
IMPOSSIBILITES = {
    ("Poignée-croisement possible", "Poignée standard", "Halo", "traversant"):
        ("page 47", "la poignée Halo n'admet pas le cylindre traversant"),
    ("Remplissage", "pv soubassements",
     "Panneau phonique (Rw = 38 dB) Groupe 1 sans PV", "36"):
        ("page 37", "le panneau phonique n'existe pas en épaisseur 36 mm"),
    ("Croisillons", "gravure 8 à 10 mm",
     "Gravure transparente (ADP) sur vitrage transparent", "MG9"):
        ("page 40", "la gravure transparente ADP n'est pas réalisable sur le motif MG9"),
    ("Croisillons", "gravure 8 à 10 mm",
     "Gravure transparente (ADP) sur vitrage sablé + 160 % PV vitrage pour verre sablé",
     "MG9"):
        ("page 40", "la gravure transparente ADP sur vitrage sablé n'est pas "
                    "réalisable sur le motif MG9"),
}

ALERTS, JOURNAL = [], []


# ====================================================== utilitaires
def clean(v):
    return "" if v is None else " ".join(str(v).replace("\u00a0", " ").split())


def vide(v):
    """Extension 3 : '' et None sont deux encodages de la même vacuité."""
    return v is None or (isinstance(v, str) and not v.strip())


def fmt_euro(v):
    """Recopie la valeur de la cellule, mise en forme. Jamais de calcul."""
    if vide(v):
        return None
    try:
        n = int(round(float(v)))
    except (ValueError, TypeError):
        return None
    return f"{n:,}".replace(",", "\u202f")


def count_words(*parts):
    return len(re.findall(r"\S+", " ".join(p for p in parts if p)))


def sc_id(n):
    return f"SC{n:04d}"


def source_line(page, sc, nature="originale"):
    return f"*Source : {PDF_SOURCE}, page {page} — information {nature} — {sc}*"


def emit(title, source, body):
    n = count_words("##", title, source, body)      # le marqueur compte aussi
    if n > PLAFOND:
        ALERTS.append(f"PLAFOND DÉPASSÉ ({n} mots) : {title[:70]}")
    return f"## {title}\n{source}\n\n{body}\n"


def page_of(chap, tab, des=None):
    return PAGES.get((chap, tab), PAGES.get((chap, None), "?"))


def contexte_of(chap, tab):
    return CONTEXTE.get((chap, tab), CONTEXTE.get((chap, None), chap.lower()))


def unite_of(chap, tab, des=None):
    for k in ((chap, tab, des), (chap, tab, None), (chap, None, None)):
        if k in UNITES:
            return UNITES[k]
    return None


def phrase_unite(chap, tab, des=None):
    """Le montant servi est unitaire : le total revient à l'ADV, jamais au modèle."""
    u = unite_of(chap, tab, des)
    if u is None:
        JOURNAL.append(f"UNITÉ NON ÉTABLIE : {chap} / {tab} / {des} — le chunk "
                       f"renvoie à la page {page_of(chap, tab, des)}")
        return (f" Le tarif exprime ce montant dans une unité de facturation qui "
                f"doit être lue page {page_of(chap, tab, des)} du tarif.")
    if u.startswith("forfaitaire"):
        return f" Ce montant est {u}."
    return (f" Ce montant s'entend {u} : le total s'obtient en le multipliant par "
            f"la quantité concernée, calcul qui revient à l'ADV.")


def enumere(items):
    items = [i for i in items if i]
    if not items:
        return ""
    if len(items) == 1:
        return items[0]
    return ", ".join(items[:-1]) + " et " + items[-1]


def yaml_front(sous_type, nb):
    return (
        "---\n"
        f"document_source: {PDF_YAML}\n"
        "type_document: tarif\n"
        f"sous_type: {sous_type}\n"
        f"gamme_code: {GAMME}\n"
        'gamme_nom: "Coulissant Aluminium"\n'
        'gammes_couvertes: [CA76, CAG76]\n'
        'collection: "TRYBA ALUMINIUM"\n'
        "materiau: aluminium\n"
        'version_doc: "2026.06"\n'
        "date_validite: 2026-06-19\n"
        f"nb_chunks: {nb}\n"
        "audiences: [ADV, commercial]\n"
        "---\n\n"
    )


# ====================================================== chargement
def load_rows():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[FEUILLE]
    raw = list(ws.iter_rows(values_only=True))
    header = list(raw[0]) + [None] * (NCOL - len(raw[0]))
    rows = []
    for i, r in enumerate(raw[1:], start=2):
        r = list(r) + [None] * (NCOL - len(r))
        if all(vide(v) or v in ("HT", "TTC") for v in r):
            continue                       # séparateur : aucune information
        rows.append({"xl": i, "v": r})
    largeurs = {j: int(str(header[j]).split()[2]) for j in COLS_HT_L}
    return header, rows, largeurs


def hors_gamme(rows):
    return [(r["xl"], clean(r["v"][C_GAMME]), clean(r["v"][C_CHAP]),
             clean(r["v"][C_DES]))
            for r in rows
            if clean(r["v"][C_GAMME]) and clean(r["v"][C_GAMME]) != GAMME]


# ====================================================== bandes (fait p. 6)
def echelle_L(rows, chap, tab, largeurs):
    ech = set()
    for r in rows:
        v = r["v"]
        if clean(v[C_CHAP]) == chap and clean(v[C_TAB]) == tab:
            for j in COLS_HT_L:
                if not vide(v[j]):
                    ech.add(largeurs[j])
    return sorted(ech)


def bande_basse(echelle, val):
    """Bande couverte : (précédente + 1)..val. None si première : aucun plancher
    n'est inventé, le chunk écrira « jusqu'à N mm » (règle C2)."""
    i = echelle.index(val)
    return None if i == 0 else echelle[i - 1] + 1


def dire_bande(bas, haut):
    return f"jusqu'à {haut} mm" if bas is None else f"de {bas} à {haut} mm"


# ====================================================== extension 1 : croisée
BEIGE = (0.997, 0.92, 0.828)


def carte_croisee_renforcee():
    """Le marquage « croisée renforcée obligatoire automatique sans plus-value »
    est un attribut PAR CELLULE porté par le seul remplissage graphique : il
    n'existe ni dans l'Excel, ni dans la couche texte du PDF. On le récupère par
    les coordonnées, en ancrant sur les en-têtes de largeur et la colonne des
    hauteurs. Aucune valeur n'est inventée : si l'ancrage échoue, la cellule est
    consignée et la clause n'est pas produite pour la page concernée."""
    try:
        import pdfplumber
    except ImportError:
        JOURNAL.append("CROISÉE RENFORCÉE : pdfplumber indisponible, clause non "
                       "produite (aucune invention)")
        return {}
    carte = defaultdict(set)
    rejets = 0
    with pdfplumber.open(PDF) as doc:
        for tab, tranches in PAGES_GRILLE.items():
            for pno, _, _ in tranches:
                p = doc.pages[pno - 1]
                mots = p.extract_words()
                cand = [w for w in mots if re.fullmatch(r"\d{3,4}", w["text"])]
                if not cand:
                    continue
                ytop = min(w["top"] for w in cand)
                cols = sorted({(round((w["x0"] + w["x1"]) / 2, 1), int(w["text"]))
                               for w in cand
                               if w["top"] < ytop + 6 and int(w["text"]) % 100 == 0})
                xleft = min(w["x0"] for w in cand)
                ligs = sorted({(round((w["top"] + w["bottom"]) / 2, 1), int(w["text"]))
                               for w in cand
                               if w["x0"] < xleft + 25 and w["top"] > ytop + 6
                               and 900 <= int(w["text"]) <= 2600
                               and int(w["text"]) % 100 == 0})
                if not (cols and ligs):
                    continue
                for rect in p.rects:
                    if rect.get("non_stroking_color") != BEIGE or rect["height"] >= 20:
                        continue
                    cx = (rect["x0"] + rect["x1"]) / 2
                    cy = (rect["top"] + rect["bottom"]) / 2
                    c = min(cols, key=lambda t: abs(t[0] - cx))
                    l = min(ligs, key=lambda t: abs(t[0] - cy))
                    if abs(c[0] - cx) > 18 or abs(l[0] - cy) > 10:
                        rejets += 1          # pastille de légende, ou hors table
                        continue
                    carte[(tab, l[1])].add(c[1])
    JOURNAL.append(f"CROISÉE RENFORCÉE : {sum(len(v) for v in carte.values())} "
                   f"cellules récupérées de la géométrie du PDF, {rejets} "
                   f"rectangles hors table écartés (pastilles de légende)")
    return carte


def clause_croisee_ligne(carte, tab, hauteur, largeurs_ligne, ech_L):
    """Le marquage est un effet de SEUIL : pour une bande de hauteur donnée, il
    couvre toutes les largeurs tarifées au-delà d'un seuil, jusqu'au bout de la
    ligne. La forme est vérifiée contre les largeurs RÉELLEMENT TARIFÉES de la
    ligne — et non contre l'échelle entière de la grille, les dernières lignes
    étant tronquées. Si la forme n'est pas vérifiée, les largeurs sont énumérées
    au lieu d'être résumées par un seuil : aucune extrapolation.
    Retourne (clause, largeurs_marquées)."""
    marquees = sorted(carte.get((tab, hauteur), set()))
    if not marquees:
        return "", []
    seuil = min(marquees)
    attendu = [L for L in largeurs_ligne if L >= seuil]
    if marquees == attendu:
        bas = bande_basse(ech_L, seuil)
        return (f" À partir de {bas if bas is not None else seuil} mm de largeur, "
                f"la croisée renforcée est obligatoire et automatique, sans "
                f"plus-value."), marquees
    JOURNAL.append(f"CROISÉE RENFORCÉE non contiguë ({tab}, hauteur {hauteur}) : "
                   f"largeurs énumérées au lieu du seuil "
                   f"({len(marquees)} marquées, {len(attendu)} attendues)")
    return (" La croisée renforcée est obligatoire et automatique, sans plus-value, "
            "en largeur " + enumere([f"{L} mm" for L in marquees]) + "."), marquees


# ====================================================== F2 : prix des châssis
def lignes_grille(rows, chap, tab, largeurs):
    out = []
    for r in rows:
        v = r["v"]
        if clean(v[C_CHAP]) != chap or clean(v[C_TAB]) != tab:
            continue
        cells = []
        for j, k in zip(COLS_HT_L, COLS_TTC_L):
            if vide(v[j]) and vide(v[k]):
                continue
            if vide(v[j]) or vide(v[k]):
                JOURNAL.append(f"HT/TTC désaligné ligne Excel {r['xl']} "
                               f"({chap}/{tab}, largeur {largeurs[j]})")
            cells.append((largeurs[j], v[j], v[k]))
        if cells:
            out.append((v[C_HAUTEUR], cells, r["xl"]))
    return out


def page_de_largeur(tab, L):
    for pno, lo, hi in PAGES_GRILLE[tab]:
        if lo <= L <= hi:
            return pno
    return "?"


def gen_f2(rows, largeurs, carte):
    chunks, sc = [], 2
    for (chap, tab), (prefixe, libelle, synonyme, rails) in GRILLES.items():
        ech_L = echelle_L(rows, chap, tab, largeurs)
        lignes = lignes_grille(rows, chap, tab, largeurs)
        ech_H = sorted({h for h, _, _ in lignes})
        for h, cells, xl in lignes:
            idx = [ech_L.index(L) for L, _, _ in cells]
            if idx != list(range(idx[0], idx[0] + len(idx))):
                JOURNAL.append(f"largeurs non contiguës ligne Excel {xl} "
                               f"({chap}/{tab}) — bandes à vérifier")
            h_bas = bande_basse(ech_H, h)
            # la clause de croisée est établie UNE FOIS par ligne de hauteur,
            # puis apposée aux seules tranches qui contiennent une largeur marquée
            clause, marquees = clause_croisee_ligne(
                carte, tab, h, [c[0] for c in cells], ech_L)
            items = [f"{dire_bande(bande_basse(ech_L, L), L)} : "
                     f"{fmt_euro(ht)} € HT et {fmt_euro(ttc)} € TTC"
                     for L, ht, ttc in cells]
            i = 0
            while i < len(items):
                # extension 4 : la tranche ne franchit pas une frontière de page
                pno = page_de_largeur(tab, cells[i][0])
                jmax = i + 1
                while (jmax < len(items)
                       and page_de_largeur(tab, cells[jmax][0]) == pno):
                    jmax += 1
                j = jmax
                while j > i:
                    lot = items[i:j]
                    Ls = [c[0] for c in cells[i:j]]
                    l_lo = bande_basse(ech_L, Ls[0])
                    titre_l = (f"largeurs {dire_bande(l_lo, Ls[-1])}"
                               if (i or j < len(items)) else "toutes largeurs tarifées")
                    title = (f"{prefixe}Tarif du {libelle}, hauteur "
                             f"{dire_bande(h_bas, h)}, {titre_l}")
                    src = source_line(pno, sc_id(sc))
                    body = (f"Sur la grille de prix du {libelle}, collection TRYBA "
                            f"ALUMINIUM, aussi appelé {synonyme}, le prix du châssis "
                            f"nu vitrage standard compris, en cote tarif de hauteur "
                            f"{dire_bande(h_bas, h)}, est le suivant selon la cote "
                            f"tarif en largeur : " + " ; ".join(lot) + ".")
                    if any(L in marquees for L in Ls):
                        body += clause
                    body += (" Ce prix s'entend" if len(lot) == 1
                             else " Ces prix s'entendent")
                    body += (" hors éco-participation, pour un châssis sans "
                             "complément.")
                    if count_words("##", title, src, body) <= PLAFOND or j == i + 1:
                        chunks.append(emit(title, src, body))
                        sc += 1
                        i = j
                        break
                    j -= 1
    verifier_titres_uniques("prix", chunks)
    return chunks


# ====================================================== F3 : options
COLLISIONS = {}


def groupes_options(rows):
    """Regroupement strictement iso-prix : mêmes chapitre, tableau, désignation,
    HT et TTC. Seules les variantes de la colonne Détails fusionnent.
    Le scalaire retenu est HT/TTC, à défaut PV HT/PV TTC : les deux colonnes
    coexistent sur une même ligne (meneaux) et ne sont pas interchangeables."""
    g = OrderedDict()
    for r in rows:
        v = r["v"]
        chap, tab, des, det = (clean(v[C_CHAP]), clean(v[C_TAB]),
                               clean(v[C_DES]), clean(v[C_DET]))
        if clean(v[C_GAMME]) != GAMME:
            continue
        if (chap, des) in CHAP_OVERRIDE_DES:
            nouveau = CHAP_OVERRIDE_DES[(chap, des)]
            JOURNAL.append(f"CHAPITRE RÉATTRIBUÉ d'après le PDF : Excel {r['xl']} — "
                           f"« {chap} » → « {nouveau} » pour « {des} »")
            chap = nouveau
        if not chap or chap in CHAP_GRILLE:
            continue
        if chap in CHAP_EXCLUS:
            JOURNAL.append(f"EXCLU du périmètre ({CHAP_EXCLUS[chap]}) : "
                           f"Excel {r['xl']} — {chap} / {des}")
            continue
        ht = v[C_HT] if not vide(v[C_HT]) else v[C_PVHT]
        ttc = v[C_TTC] if not vide(v[C_TTC]) else v[C_PVTTC]
        if vide(ht):
            imp = IMPOSSIBILITES.get((chap, tab, des, det))
            if imp:
                JOURNAL.append(f"IMPOSSIBILITÉ PRODUIT (case vide ≠ gratuité, "
                               f"{imp[0]}) : Excel {r['xl']} — {imp[1]} — reprise "
                               f"en faisabilité, aucun prix généré")
            elif tab == "pv sur grilles prix":
                JOURNAL.append(f"PLUS-VALUE PROPORTIONNELLE sans valeur dans "
                               f"l'Excel : Excel {r['xl']} — {chap} / {des} — "
                               f"pourcentage repris du PDF page 7 (fichier F4)")
            else:
                JOURNAL.append(f"ligne sans montant, non générée (anti-fantôme) : "
                               f"Excel {r['xl']} — {chap} / {tab} / {des}")
            continue
        key = (chap, tab, corriger_libelle(v[C_DES]),
               int(round(float(ht))),
               None if vide(ttc) else int(round(float(ttc))))
        g.setdefault(key, []).append(corriger_libelle(det) if det else det)
    return g


def appliquer_discriminants(groupes):
    """Réinjecte dans la clé le discriminant relevé au PDF, RATTACHÉ PAR LE
    MONTANT. Si le montant ne correspond à aucune entrée relevée, rien n'est
    appliqué et le poste retombe sous la règle d'exclusion."""
    nouveau, applique = OrderedDict(), 0
    for (chap, tab, des, ht, ttc), variantes in groupes.items():
        d = DISCRIMINANTS.get((chap, tab, des, ht))
        if d:
            des = f"{des} {d}"
            applique += 1
        nouveau[(chap, tab, des, ht, ttc)] = variantes
    if applique:
        JOURNAL.append(f"DISCRIMINANT REPRIS DU PDF : {applique} postes complétés "
                       f"par leur finition, rattachement par le montant")
    groupes.clear()
    groupes.update(nouveau)


def exclure_indiscriminables(groupes):
    """Deux prix différents sous des colonnes strictement identiques : le
    discriminant manque. Aucune invention n'étant admise, ces postes ne sont pas
    générés mais consignés."""
    par_cle = defaultdict(list)
    for k, variantes in groupes.items():
        par_cle[(k[0], k[1], k[2], tuple(variantes))].append(k)
    for cle, keys in par_cle.items():
        if len(keys) > 1:
            for k in keys:
                groupes.pop(k, None)
            JOURNAL.append(
                f"DISCRIMINANT MANQUANT — {len(keys)} prix distincts sous des "
                f"colonnes identiques ({cle[0]} / {cle[1]} / {cle[2]}) : postes "
                f"NON générés, discriminant à reprendre du PDF page "
                f"{page_of(cle[0], cle[1])}")


def indexer_collisions(groupes):
    COLLISIONS.clear()
    compte = defaultdict(int)
    for (chap, tab, des, ht, ttc) in groupes:
        compte[(chap, tab, des)] += 1
    COLLISIONS.update({k: n for k, n in compte.items() if n > 1})


def jonction_of(variantes):
    v = [JONCTIONS.get(" ".join(str(x).lower().split())) for x in variantes if x]
    return v[0] if v and len(set(v)) == 1 and v[0] else None


def phrase_variantes(chap, variantes):
    v = [x for x in variantes if x and x not in DETAIL_ANNOTATION]
    ecartes = [x for x in variantes if x in DETAIL_ANNOTATION]
    for x in set(ecartes):
        JOURNAL.append(f"DÉTAIL ÉCARTÉ, annotation de saisie sans valeur pour "
                       f"l'ADV : « {x} » ({chap})")
    if not v:
        return ""
    if chap in DETAIL_NATURE:
        gabarit = DETAIL_NATURE[chap]
        dits = [gabarit.format(x) for x in v]
        return (f" Ce montant vaut pour {enumere(dits)}." if len(v) == 1
                else f" Ce montant vaut pour {enumere(dits)}.")
    if all(x in DETAIL_MOTIF for x in v):
        motifs = [m for x in v for m in dire_detail(x)]
        return f" Ce montant vaut pour {enumere(['le motif ' + m for m in motifs])}."
    if len(v) == 1:
        return f" Le tarif rattache ce montant à la mention « {v[0]} »."
    return (" Le tarif rattache ce montant aux mentions "
            + enumere([f"« {x} »" for x in v]) + ".")


def libelle_poste(chap, tab, des, variantes):
    ctx = contexte_of(chap, tab)
    if des:
        base = des
    elif len(variantes) == 1 and variantes[0]:
        base = variantes[0]
    else:
        base = tab
    base = " ".join(str(base).replace("\n", " ").split())
    if base and base.lower() in ctx.lower():
        base = ""
    j = jonction_of(variantes) if chap == "Croisillons" else None
    if j:
        return " ".join(f"{ctx} {base}, {j}".split())
    suffixe = ""
    if des and (chap, tab, des) in COLLISIONS:
        var = [d for v in variantes if v for d in dire_detail(v)]
        if var:
            suffixe = " " + enumere(var)
    return " ".join(f"{ctx} {base}{suffixe}".split())


def verifier_titres_uniques(nom, chunks):
    vus = set()
    for c in chunks:
        t = c.split("\n", 1)[0]
        if t in vus:
            ALERTS.append(f"TITRE NON DISCRIMINANT ({nom}) : {t[:80]}")
        vus.add(t)


def desambiguiser(groupes):
    """Garde-fou générique : deux postes de prix différents ne peuvent pas porter
    le même titre. Si deux libellés finaux coïncident malgré tout — homonymie
    entre deux tableaux d'un même chapitre — le libellé du tableau est ajouté.
    Aucun discriminant n'est inventé : celui du tableau vient de la source."""
    vus = defaultdict(list)
    for k, variantes in groupes.items():
        vus[libelle_poste(k[0], k[1], k[2], variantes)].append(k)
    suffixes = {}
    for lib, keys in vus.items():
        if len(keys) > 1:
            for k in keys:
                suffixes[k] = f" ({k[1]})"
            JOURNAL.append(f"TITRE DÉSAMBIGUÏSÉ par le libellé de tableau : "
                           f"{len(keys)} postes portaient « {lib} »")
    return suffixes


def gen_f3(rows):
    chunks, sc = [], 2
    groupes = groupes_options(rows)
    appliquer_discriminants(groupes)
    exclure_indiscriminables(groupes)
    indexer_collisions(groupes)
    suffixes = desambiguiser(groupes)
    for (chap, tab, des, ht, ttc), variantes in groupes.items():
        libelle = libelle_poste(chap, tab, des, variantes) + \
            suffixes.get((chap, tab, des, ht, ttc), "")
        title = f"{PREFIXE_CA}{libelle[0].upper() + libelle[1:]}, plus-value tarif"
        src = source_line(page_of(chap, tab, des), sc_id(sc))
        if ht == 0:
            body = (f"Dans le tarif du coulissant aluminium CA76, le poste "
                    f"« {libelle} » ne donne lieu à aucune plus-value : le tarif "
                    f"le chiffre à 0 € HT et 0 € TTC, il est donc compris sans "
                    f"supplément.")
        else:
            ttc_txt = (f", soit {fmt_euro(ttc)} € TTC" if ttc is not None else "")
            body = (f"Dans le tarif du coulissant aluminium CA76, le poste "
                    f"« {libelle} » est chiffré en plus-value à {fmt_euro(ht)} € HT"
                    f"{ttc_txt}.")
            if ttc is None:
                body += " Le tarif ne porte pas de valeur TTC pour ce poste."
                JOURNAL.append(f"TTC absent, exposé dans le chunk : {chap}/{tab}/{des}")
            body += phrase_unite(chap, tab, des)
        if not jonction_of(variantes) and not (des and (chap, tab, des) in COLLISIONS):
            body += phrase_variantes(chap, variantes)
        if ht != 0:
            body += " Cette plus-value s'entend hors éco-participation."
        for det_brut in variantes:
            d = DIVERGENCES.get((chap, tab, des, det_brut))
            if d:
                body += " " + d
                JOURNAL.append(f"DIVERGENCE EXPOSÉE (non arbitrée) : {chap} / "
                               f"{des} / {det_brut} — les deux lectures sont "
                               f"rapportées dans le chunk")
                break
        chunks.append(emit(title, src, body))
        sc += 1
    verifier_titres_uniques("options", chunks)
    return chunks


# ====================================================== F4 : proportionnelles
# Extension 2 : AUCUN pourcentage ne figure dans l'Excel, qui n'en garde que la
# coquille (deux lignes de doublage sans valeur, la ligne « Dormant à ailettes »
# étant même absente). Les taux sont relevés page par page dans le PDF.
# Le taux est transcrit littéralement ; son APPLICATION revient à l'ADV, jamais
# au modèle — extension du précédent T4 sur le montant unitaire.
PROPORTIONNELLES = [
    (7, "Plus-value du dormant à ailettes sur les grilles de prix", "+ 3 %",
     "au dormant à ailettes, en 40 mm comme en 70 mm",
     "Elle s'applique sur le prix lu dans les grilles de prix des coulissants à "
     "2 rails. Elle n'est pas applicable aux coulissants à 3 rails ni au "
     "galandage, sur lesquels le tarif indique que les dormants à ailette sont "
     "impossibles."),
    (7, "Plus-value du doublage 120 et 140 mm sur les grilles de prix", "+ 5 %",
     "au dormant large avec doublage de 120 ou de 140 mm",
     "Elle s'applique sur le prix lu dans les grilles de prix des coulissants à "
     "2 rails. Elle n'est pas applicable aux coulissants à 3 rails ni au "
     "galandage, sur lesquels le tarif indique que le doublage est impossible."),
    (7, "Plus-value du doublage 160 mm sur les grilles de prix", "+ 10 %",
     "au dormant large avec doublage de 160 mm",
     "Elle s'applique sur le prix lu dans les grilles de prix des coulissants à "
     "2 rails. Elle n'est pas applicable aux coulissants à 3 rails ni au "
     "galandage, sur lesquels le tarif indique que le doublage est impossible."),
    (26, "Plus-value du groupe de couleurs 2", "+ 15 %",
     "aux teintes du groupe 2, qui rassemble l'anodisé nature, les laquages "
     "imitation anodisation, une série de RAL granité mat et les teintes Futura",
     "Le groupe 1 ne porte, lui, aucune plus-value."),
    (27, "Plus-value du groupe sublimation, tons bois", "+ 25 %",
     "aux teintes du groupe sublimation, tons bois exclusifs chêne d'or, chêne "
     "foncé, chêne mat, chêne irlandais et acajou veiné",
     "Le tarif signale que ces tons bois sublimés sont des teintes approchantes "
     "des décors PVC."),
    (27, "Plus-value des autres RAL granité", "+ 25 %",
     "aux RAL granité mat situés hors des groupes 1 et 2",
     "Le tarif précise que ces teintes, ainsi que les RAL lisses mats, sont "
     "servies sous réserve de faisabilité et qu'une demande doit impérativement "
     "être faite à l'usine. Les RAL lisses brillants et les autres teintes Futura "
     "ne sont pas disponibles."),
    (27, "Plus-value de l'anodisation champagne", "+ 25 %",
     "à la finition anodisé champagne",
     "Le tarif impose de regrouper le chantier sur une seule commande, habillages "
     "compris, pour respecter le bain d'anodisation."),
    (40, "Plus-value de la gravure Art Déco sur vitrage sablé", "+ 160 %",
     "à la gravure transparente ADP réalisée sur un vitrage sablé",
     "Ce taux porte sur la plus-value du vitrage sablé, et non sur le prix du "
     "châssis. Il s'ajoute au prix de la gravure elle-même."),
]


def gen_f4():
    chunks, sc = [], 2
    for page, titre, taux, cible, complement in PROPORTIONNELLES:
        title = f"{PREFIXE_CA}{titre}"
        src = source_line(page, sc_id(sc))
        body = (f"Dans le tarif du coulissant aluminium CA76, une plus-value de "
                f"{taux} s'applique {cible}. {complement} Ce taux est exprimé en "
                f"pourcentage : le tarif n'en donne pas la contrepartie en euros, "
                f"et son application au prix concerné revient à l'ADV. Cette "
                f"plus-value s'entend hors éco-participation.")
        chunks.append(emit(title, src, body))
        sc += 1
    verifier_titres_uniques("proportionnelles", chunks)
    return chunks


# ====================================================== F1, F5, F6
# Rédigés à partir des pages lues et vérifiées. Aucun montant en F5 ni F6
# (règles C6 et C7) ; les valeurs citées en F1 sont des cotes, jamais des prix.
F1_BLOCS = [
    (PREFIXE_CA, 6, "Distinction entre cote tarif et cote de fabrication",
     "Le tarif du coulissant aluminium CA76 distingue deux jeux de cotes. Les "
     "cotes de tarif, notées L exposant T et H exposant T, servent de référence "
     "au chiffrage des devis et désignent le châssis nu. Les cotes de "
     "fabrication, notées L et H, servent de référence à la commande et à la "
     "fabrication, compléments compris. Toutes les cotes sont exprimées en "
     "millimètres. Un prix de grille se lit toujours sur une cote de tarif : "
     "utiliser une cote de fabrication à sa place conduit à lire le prix d'un "
     "autre châssis. Avant tout chiffrage, il faut donc établir laquelle des deux "
     "cotes est en main."),
    (PREFIXE_CA, 6, "Lecture des grilles de prix par bandes de dimensions",
     "Les grilles de prix du coulissant aluminium CA76 ne donnent pas un prix "
     "pour une dimension exacte mais pour une bande de dimensions. Le tarif "
     "l'énonce explicitement : une colonne intitulée 1000 couvre les largeurs de "
     "901 à 1000 millimètres, une ligne intitulée 600 couvre les hauteurs de 501 "
     "à 600 millimètres. Le prix se lit à l'intersection de la bande de largeur "
     "et de la bande de hauteur qui contiennent la cote de tarif du châssis. "
     "Aucun calcul ni aucune interpolation n'est à faire entre deux valeurs de la "
     "grille. Les prix indiqués valent pour un châssis sans complément, vitrage "
     "standard compris."),
    (PREFIXE_CA, 7, "Profils de dormant et passage à la cote de fabrication",
     "Le tarif du coulissant aluminium CA76 référence plusieurs profils de "
     "dormant. Le dormant neuf sans ailette existe en 2 rails sous la référence "
     "AU10100 et en 3 rails sous la référence AU10101. Le dormant à ailette "
     "existe en 40 millimètres sous la référence AU10103 et en 70 millimètres "
     "sous la référence AU10111. Le dormant large associe la référence AU10108 au "
     "couvre-joint CJ22 pour un doublage de 120 millimètres, la référence AU10108 "
     "au couvre-joint CJ02 pour 140 millimètres, et la référence AU10109 au "
     "couvre-joint CJ02 pour 160 millimètres. La cote de fabrication en largeur "
     "correspond à la largeur du tableau fini augmentée de 30 millimètres."),
    (PREFIXE_CA, 8, "Cotes utiles et clair de vitre intérieur",
     "Sur le coulissant aluminium CA76, les cotes utiles donnent les dimensions "
     "de clair de vitre intérieur hors joint, valables tous dormants confondus. "
     "La cote A, entre ouvrant et clair de jour intérieur, est de 111 "
     "millimètres. La cote B est de 40 millimètres, la cote C de 152 millimètres "
     "et la cote D de 101 millimètres. Ces cotes servent au calcul du clair de "
     "vitre et non au chiffrage : le prix se lit sur la cote de tarif."),
    (PREFIXE_CA, 11, "Principe de la croisée renforcée sur les grilles de prix",
     "Sur le coulissant aluminium CA76, certaines combinaisons de hauteur et de "
     "largeur imposent une croisée renforcée. Le tarif la signale directement "
     "dans ses grilles de prix, sur les cellules concernées, et précise qu'elle "
     "est obligatoire et automatique et qu'elle ne donne lieu à aucune "
     "plus-value. Elle concerne les grandes hauteurs, à partir d'un seuil de "
     "largeur propre à chaque bande de hauteur. Sur le coulissant à galandage "
     "CAG76, la croisée standard est systématique et la croisée renforcée n'a "
     "pas lieu d'être."),
    (PREFIXE_CA, 11, "Poids maximal par vantail",
     "Le tarif du coulissant aluminium CA76 fixe une règle de ferrage commune à "
     "toutes ses grilles de prix : le poids maximal par vantail est de 200 "
     "kilogrammes. Le tarif renvoie aux pages des limites de fabrication du "
     "battant par épaisseur de vitrage pour vérifier cette limite sur une "
     "configuration donnée. Cette règle vaut également pour le coulissant à "
     "galandage CAG76."),
    (PREFIXE_CA, 37, "Cotes de référence des panneaux de soubassement",
     "Sur le coulissant aluminium CA76, les dimensions des panneaux de "
     "soubassement se calculent en coulissant comme en frappe. La hauteur h est "
     "la hauteur de soubassement, mesurée de l'extérieur du dormant à l'axe de la "
     "traverse. La largeur l vaut la largeur L pour un châssis à un vantail, L "
     "divisé par deux pour deux vantaux, L divisé par trois pour trois vantaux et "
     "L divisé par quatre pour quatre vantaux. Le soubassement standard a une "
     "hauteur de 350 millimètres."),
    (PREFIXE_CA, 37, "Méthode de calcul du prix d'un soubassement",
     "Sur le coulissant aluminium CA76, le prix d'un soubassement est une "
     "plus-value à appliquer sur la valeur du vitrage. Le tarif décrit la "
     "méthode : partir du prix du châssis complet vitrage compris, chercher la "
     "valeur vitrage correspondant à la taille du remplissage, calculer la "
     "plus-value du panneau, l'ajouter au châssis, puis ajouter le prix de la "
     "traverse et de ses fixations. Cette suite d'opérations revient à l'ADV. "
     "Le tarif renvoie pour cette recherche à une grille de prix des fixes qui "
     "ne figure pas dans le tarif CA76, et annonce des soubassements précalculés "
     "au bas des grilles de prix qui n'y figurent pas non plus : ces deux "
     "renvois sont sans objet dans ce document."),
    (PREFIXE_CA, 38, "Méthode de calcul du prix des croisillons",
     "Sur le coulissant aluminium CA76, le prix des croisillons, incorporés comme "
     "rapportés, se calcule au nombre de champs et non au mètre linéaire. Un "
     "champ est une surface de vitrage délimitée par les croisillons, par le "
     "dormant ou par l'ouvrant. Le tarif distingue deux types de champs : ceux "
     "dont les croisillons présentent une jonction en T ou en croix, et ceux dont "
     "les croisillons sont filants, c'est-à-dire sans jonction. Le comptage des "
     "champs revient à l'ADV."),
    (PREFIXE_CA, 54, "Méthode de calcul du prix des tapées de doublage",
     "Sur le coulissant aluminium CA76, les tapées sur les montants verticaux "
     "sont filantes et les tapées sur les traverses horizontales viennent buter "
     "contre les tapées montantes. Les dimensions de fabrication du châssis font "
     "office de cote de référence pour le calcul du prix des tapées, dont la "
     "longueur facturée vaut la largeur augmentée de deux fois la hauteur. Ce "
     "calcul revient à l'ADV : le tarif donne un prix au mètre linéaire, pas un "
     "prix par châssis."),
    (PREFIXE_CA, 58, "Composition d'un prix de coulissant",
     "Un prix de coulissant aluminium CA76 se compose du prix du châssis nu, lu "
     "dans la grille correspondant au nombre de vantaux et de rails, auquel "
     "s'ajoutent d'une part les plus-values exprimées en pourcentage sur ce prix, "
     "pour le dormant et pour la teinte, d'autre part les plus-values chiffrées "
     "en euros des compléments retenus : vitrage, remplissage, croisillons, "
     "ferrage, grille d'entrée d'air et accessoires de pose. Chaque plus-value "
     "est chiffrée séparément dans le tarif et l'addition revient à l'ADV. Les "
     "prix s'entendent hors éco-participation."),
    (PREFIXE_CA, 3, "Vocabulaire et distinction entre CA76, CAG76 et CA80 New",
     "CA76 désigne le coulissant aluminium de la collection TRYBA ALUMINIUM et "
     "CAG76 sa déclinaison à galandage, dans laquelle le vantail se range dans "
     "une réserve maçonnée. Les deux produits partagent un même tarif mais ont "
     "des grilles de prix, des faisabilités et des règles de croisée distinctes. "
     "Ni l'un ni l'autre ne doit être confondu avec le CA80 New, gamme voisine de "
     "coulissant aluminium dont les prix, les cotes et les options sont propres. "
     "Sur cette gamme, la crémone est le terme exact : le CA76 est fermé par une "
     "crémone Secure+ à crochets inox."),
]

F5_BLOCS = [
    (PREFIXE_CA, 9, "Compositions de châssis réalisables sur le coulissant CA76",
     "Le coulissant aluminium CA76 admet, en 2 rails, les compositions à deux "
     "vantaux, à trois vantaux et à quatre vantaux, et, en 3 rails, les "
     "compositions à trois vantaux et à six vantaux. Le tarif indique que toute "
     "autre composition est irréalisable. Chacune de ces cinq compositions "
     "dispose de sa propre grille de prix, et une combinaison de dimensions "
     "absente de la grille n'est pas tarifée."),
    (PREFIXE_CAG, 9, "Compositions de châssis réalisables sur le galandage CAG76",
     "Le coulissant à galandage CAG76 admet, en 1 rail, les compositions à un "
     "vantail et à deux vantaux, et, en 2 rails, les compositions à deux vantaux "
     "et à quatre vantaux. Le tarif indique que toute autre composition est "
     "irréalisable. Trois grilles de prix couvrent ces compositions. Le tarif "
     "présente toutefois une divergence sur ce point : la grille du galandage à "
     "quatre vantaux se déclare valable sur un ou deux rails, alors que la page "
     "des faisabilités ne montre cette composition qu'en deux rails. Les deux "
     "énoncés sont rapportés tels quels, sans arbitrage."),
    (PREFIXE_CA, 18, "Restrictions propres aux coulissants à 3 rails",
     "Sur le coulissant aluminium CA76 en 3 rails, qu'il s'agisse de la "
     "composition à trois vantaux ou de celle à six vantaux, le tarif signale "
     "que le coffre de volet roulant, les dormants à ailette et le doublage sont "
     "impossibles. Les plus-values de dormant qui figurent sur les grilles à 2 "
     "rails ne sont donc pas applicables à ces deux grilles, et les tableaux "
     "correspondants n'y sont d'ailleurs pas imprimés."),
    (PREFIXE_CAG, 23, "Restriction du coffre de volet roulant sur le galandage",
     "Sur le coulissant à galandage CAG76, le tarif signale que le coffre de "
     "volet roulant est impossible sur un galandage à deux rails. La restriction "
     "est portée sur les grilles du galandage à deux vantaux et du galandage à "
     "quatre vantaux."),
    (PREFIXE_CA, 32, "Ordre de composition obligatoire d'un triple vitrage",
     "Sur le coulissant aluminium CA76, toute association de types de vitrage en "
     "triple vitrage doit respecter un ordre de priorité. Le vitrage solaire, "
     "TRYBASUN ou STOPSOL, se place côté extérieur. Le vitrage ornemental ou "
     "muni de croisillons Art Déco se place au milieu. Le vitrage à couche Isol'3 "
     "se place côté intérieur et côté extérieur. Le vitrage feuilleté se place "
     "côté intérieur, et également côté extérieur en cas d'allège. La position du "
     "vitrage phonique est sans importance. Il est impératif d'avoir un vitrage à "
     "couche thermique côté intérieur et côté extérieur."),
    (PREFIXE_CA, 32, "Ordre de composition obligatoire d'un double vitrage",
     "Sur le coulissant aluminium CA76, toute association de types de vitrage en "
     "double vitrage doit respecter un ordre de priorité. Le vitrage solaire, "
     "TRYBASUN ou STOPSOL, se place côté extérieur. Le vitrage ornemental ou muni "
     "de croisillons Art Déco se place côté extérieur. Le vitrage à couche Isol'3 "
     "se place côté intérieur. Le vitrage feuilleté se place côté intérieur, et "
     "également côté extérieur en cas d'allège. Il est impératif d'avoir au moins "
     "un vitrage à couche thermique."),
    (PREFIXE_CA, 32, "Limites dimensionnelles des vitrages",
     "Sur le coulissant aluminium CA76, le rapport maximal entre la largeur et la "
     "hauteur d'un verre de 4 millimètres est de six. Les vitrages dont les "
     "dimensions sont inférieures à 190 par 350 millimètres ne peuvent pas être "
     "fabriqués en TPS et reçoivent des écarteurs traditionnels en inox noir."),
    (PREFIXE_CA, 32, "Exclusions de la certification Cekal",
     "Sur le coulissant aluminium CA76, tous les vitrages sont titulaires de la "
     "certification Cekal, à l'exception des vitrages Art Déco, des vitrages "
     "munis de croisillons laiton ou couleur, du vitrage Cathédrale et des "
     "vitrages de petites dimensions, c'est-à-dire inférieurs à 350 par 350 "
     "millimètres avec un écarteur de 16 millimètres, ou inférieurs à 410 par 410 "
     "millimètres avec un écarteur de 20 millimètres."),
    (PREFIXE_CA, 33, "Conditions du vitrage d'altitude",
     "Sur le coulissant aluminium CA76, le vitrage d'altitude s'impose à partir "
     "d'une altitude de mille mètres et n'est disponible qu'en double vitrage sur "
     "cette gamme. Toute fabrication destinée à un site situé au-delà de cette "
     "altitude doit faire l'objet d'une saisie spéciale. Le tarif pose deux "
     "conditions impératives : un écarteur de 12 millimètres sans croisillons, ou "
     "un écarteur de 14 millimètres avec croisillons. Le tarif précise que les "
     "déformations optiques dues aux différences de pression ne peuvent pas "
     "motiver un remplacement du vitrage."),
    (PREFIXE_CA, 33, "Contraintes de transport et de pose du vitrage",
     "Les châssis du coulissant aluminium CA76 sont livrés en cassette, sanglés "
     "et protégés par des cales en polystyrène, sauf pour un ensemble dont une "
     "dimension dépasse 2,20 mètres : le service expédition doit alors être "
     "contacté. À partir d'une dimension de vitrage supérieure à 2000 par 2000 "
     "millimètres ou d'un poids supérieur à 80 kilogrammes, le vitrage est livré "
     "non posé et les cales de vitrage ne sont pas livrées. En deçà de ces "
     "limites, le vitrage peut être livré posé ou non, sur demande."),
    (PREFIXE_CA, 35, "Vitrages ornementaux indisponibles en triple vitrage",
     "Sur le coulissant aluminium CA76, les verres Opale, Delta, Gothique et "
     "Mastercarré ne sont pas disponibles en triple vitrage, pour des raisons "
     "techniques. Les verres Chinchilla, Cathédrale, Granité, Dépoli, Sablé et "
     "Stopsol sont disponibles en double comme en triple vitrage."),
    (PREFIXE_CA, 37, "Limites dimensionnelles des panneaux de soubassement",
     "Sur le coulissant aluminium CA76, un panneau de soubassement mesure au "
     "minimum 195 par 195 millimètres et au maximum 3000 par 1500 millimètres. La "
     "hauteur minimale d'un soubassement est de 250 millimètres. Le panneau "
     "standard se compose d'un isolant thermique et de deux faces en tôle laquée "
     "blanc. Les panneaux moulurés ne sont pas disponibles sur le CA76, à cause du "
     "croisement des vantaux."),
    (PREFIXE_CA, 37, "Épaisseur du panneau phonique",
     "Sur le coulissant aluminium CA76, le panneau de soubassement phonique "
     "n'existe qu'en épaisseur 28 millimètres. Le tarif ne le chiffre pas en 36 "
     "millimètres, épaisseur dans laquelle il n'est pas réalisable, alors que le "
     "panneau standard existe dans les deux épaisseurs."),
    (PREFIXE_CA, 38, "Conditions de garantie des croisillons intégrés",
     "Sur le coulissant aluminium CA76, les croisillons doivent toujours être "
     "incorporés en laissant au moins 2 millimètres entre le croisillon et le "
     "vitrage pour bénéficier de la garantie TRYBA. Le certificat Cekal reste "
     "applicable avec un espace d'un millimètre, mais cette exécution entraîne "
     "des bruits de contact entre croisillons et vitres et des taches de contact "
     "sous l'effet des pressions : elle n'est pas couverte par la garantie."),
    (PREFIXE_CA, 38, "Écarteurs et position des croisillons intégrés",
     "Sur le coulissant aluminium CA76, un châssis muni de croisillons intégrés "
     "reçoit des écarteurs de vitrage traditionnels en inox noir, ou en TPS selon "
     "les impératifs techniques. En triple vitrage, les croisillons se situent "
     "entre le verre intermédiaire et le verre extérieur. Pour les croisillons en "
     "finition laquée, seule la teinte RAL est reprise, et non la finition "
     "granitée. Si la finition anodisée nature a été retenue pour le châssis, les "
     "croisillons sont laqués dans une teinte équivalente, le RAL 9006."),
    (PREFIXE_CA, 39, "Dimensions minimales du croisillon à la grecque",
     "Sur le coulissant aluminium CA76, la dimension minimale de vitrage pour "
     "recevoir un croisillon à la grecque est de 400 par 400 millimètres en un "
     "vantail et de 250 par 400 millimètres en deux vantaux. Un croisillon "
     "intermédiaire est ajouté à partir d'une cote Y supérieure à 1700 "
     "millimètres. Les cotes X et Y à retenir dépendent de la taille du vitrage "
     "et figurent dans le tableau de la page des croisillons à la grecque."),
    (PREFIXE_CA, 40, "Conditions de réalisation des motifs Art Déco",
     "Sur le coulissant aluminium CA76, la gravure Art Déco n'est réalisable que "
     "sur un vitrage de 6 millimètres. Aucune gravure n'est possible sur un "
     "vitrage ornemental ni sur un vitrage d'altitude. Il est impossible de "
     "combiner un vitrage Isol'3 associé à un vitrage TRYBASUN avec un motif Art "
     "Déco. La gravure transparente ADP n'est pas réalisable en largeur de 18 "
     "millimètres, ni sur le motif MG9. La taille maximale d'un vitrage gravé est "
     "de 1600 par 2500 millimètres."),
    (PREFIXE_CA, 41, "Restriction d'épaisseur des croisillons rapportés",
     "Sur le coulissant aluminium CA76, les croisillons rapportés ne sont "
     "possibles qu'avec un vitrage de 28 millimètres. Ils ne peuvent donc pas "
     "équiper un châssis en triple vitrage de 36 millimètres, y compris le triple "
     "vitrage standard de la gamme."),
    (PREFIXE_CA, 30, "Faisabilité des teintes du bloc-baie en face intérieure",
     "Sur le coulissant aluminium CA76 associé à un coffre Chrono One 200 ou 230, "
     "la face intérieure du coffre PVC n'accepte que trois traitements : le RAL "
     "9010 brillant et le RAL 9016 granité, qui donnent un coffre blanc, et un "
     "laquage liquide dans un RAL granité. Toutes les autres finitions, laquages "
     "nature, champagne et bronze, anodisations, tons bois exclusifs, Futura "
     "sablés et Rouille, sont impossibles en face intérieure."),
    (PREFIXE_CA, 30, "Faisabilité des teintes du bloc-baie en face extérieure",
     "Sur le coulissant aluminium CA76 associé à un coffre Chrono One 200 ou 230, "
     "la face extérieure se décline sur la cornière, le lambrequin, les coulisses "
     "et la lame finale en aluminium. Les RAL brillants et granités, les laquages "
     "nature, champagne et bronze ainsi que les Futura sablés et le Rouille y sont "
     "réalisables. Les anodisations sont impossibles sur les quatre éléments. Le "
     "chêne d'or est impossible sur cornière et lambrequin mais donne un CD sur "
     "coulisses et lame finale, et le chêne irlandais y donne un laqué RAL 1019. "
     "Le chêne foncé, le chêne mat et l'acajou veiné sont impossibles partout."),
    (PREFIXE_CA, 30, "Restrictions du laquage du volet roulant",
     "Sur le coulissant aluminium CA76, il est impossible de proposer un volet "
     "roulant laqué en laquage champagne et en laquage bronze. Le laquage deux "
     "faces n'est réalisable qu'avec des manœuvres par moteur filaire ou par "
     "moteur radio sans manœuvre de secours."),
    (PREFIXE_CA, 44, "Règles d'implantation des grilles d'entrée d'air",
     "Sur le coulissant aluminium CA76, l'implantation de la grille d'entrée d'air "
     "suit deux règles. Si la menuiserie est équipée d'un coffre de volet roulant, "
     "la grille est proposée posée sur la trappe de visite du caisson. Sinon, "
     "l'entrée d'air est placée sur la traverse haute du vantail principal. À la "
     "pose, il faut s'assurer que l'entrée d'air est parfaitement centrée sur la "
     "mortaise, orifice de passage d'air."),
    (PREFIXE_CA, 44, "Limites d'utilisation des grilles d'entrée d'air",
     "Sur le coulissant aluminium CA76, la pose d'une grille d'entrée d'air "
     "centrée sur le clair de jour du vitrage, en conservant des gardes latérales "
     "de 20 millimètres, suppose une largeur hors tout du battant suffisante. Le "
     "tarif fixe cette largeur à 408 millimètres pour les grilles Mini ESEA 30 et "
     "Mini ESEA 22, et à 525 millimètres pour les grilles ISOLA 45 et ISOLA HY. "
     "Les grilles Mini ESEA et Isola2 sont toutes deux possibles sur un coffre de "
     "volet roulant."),
    (PREFIXE_CA, 47, "Cas où le croisement des vantaux est impossible",
     "Sur le coulissant aluminium CA76, le croisement des vantaux est impossible "
     "dans trois cas : lorsqu'une poignée autre que la Halo équipe le semi-fixe, "
     "lorsqu'un cylindre traversant est associé à une poignée intérieure et "
     "extérieure, et lorsqu'un demi-cylindre équipe le semi-fixe. La poignée "
     "standard, elle, est valable pour le vantail principal comme pour le "
     "semi-fixe."),
    (PREFIXE_CA, 48, "Combinaisons de poignées autorisées avec cylindre traversant",
     "Sur le coulissant aluminium CA76, seules certaines combinaisons de poignée "
     "intérieure et de poignée extérieure sont autorisées avec un cylindre "
     "traversant. Le tarif admet la Toulon en intérieur avec la Toulon en "
     "extérieur, la Klass en intérieur avec la Klass en extérieur, la Kort en "
     "intérieur avec la Kort en extérieur, et la Delta Flap en intérieur avec la "
     "Delta Flap en extérieur. Aucune autre combinaison n'est autorisée."),
    (PREFIXE_CA, 49, "Faisabilités de la poignée Novem et Novemde",
     "Sur le coulissant aluminium CA76, la poignée Novem et sa variante Novemde "
     "sont disponibles en blanc RAL 9016 brillant, en noir RAL 9005 brillant, en "
     "titane laquage brillant et en RAL granité, mais pas en titane anodisation. "
     "Dans ces quatre teintes, elle admet la configuration en poignée intérieure "
     "seule, la poignée intérieure avec demi-cylindre et la poignée fixe non "
     "manœuvrante. Elle n'admet pas la double poignée avec cylindre traversant."),
    (PREFIXE_CA, 49, "Faisabilités de la poignée Toulon",
     "Sur le coulissant aluminium CA76, la poignée Toulon est disponible en blanc "
     "RAL 9016 brillant, en noir RAL 9005 brillant, en titane anodisation et en "
     "RAL granité, mais pas en titane laquage brillant. Dans ces quatre teintes, "
     "elle admet la poignée intérieure seule, la poignée intérieure avec "
     "demi-cylindre et la double poignée avec cylindre traversant. Elle n'existe "
     "pas en poignée fixe non manœuvrante."),
    (PREFIXE_CA, 49, "Faisabilités de la poignée Klass",
     "Sur le coulissant aluminium CA76, la poignée Klass est disponible en blanc "
     "RAL 9016 brillant, en noir RAL 9005 brillant, en titane laquage brillant et "
     "en RAL granité, mais pas en titane anodisation. Dans ces quatre teintes, "
     "elle admet la poignée intérieure seule et la poignée intérieure avec "
     "demi-cylindre. En double poignée avec cylindre traversant, elle n'est admise "
     "qu'avec une Kort en extérieur. Elle n'existe pas en poignée fixe non "
     "manœuvrante."),
    (PREFIXE_CA, 49, "Faisabilités des poignées Kort et Delta Flap",
     "Sur le coulissant aluminium CA76, les poignées Kort et Delta Flap sont "
     "disponibles en blanc RAL 9016 brillant, en noir RAL 9005 brillant, en titane "
     "laquage brillant et en RAL granité, mais pas en titane anodisation. Dans ces "
     "quatre teintes, elles admettent les quatre configurations : poignée "
     "intérieure seule, poignée intérieure avec demi-cylindre, double poignée avec "
     "cylindre traversant et poignée fixe non manœuvrante."),
    (PREFIXE_CA, 49, "Faisabilités de la poignée Halo",
     "Sur le coulissant aluminium CA76, la poignée Halo est disponible en blanc "
     "RAL 9016 brillant, en noir RAL 9005 brillant, en titane laquage brillant et "
     "en RAL granité, mais pas en titane anodisation. Elle admet la poignée "
     "intérieure seule et la poignée intérieure avec demi-cylindre. Elle n'admet "
     "pas la double poignée avec cylindre traversant. La poignée fixe non "
     "manœuvrante est annoncée à venir et n'est pas disponible à ce jour. La Halo "
     "est la seule poignée admise sur le semi-fixe lorsque les vantaux se croisent."),
    (PREFIXE_CA, 50, "Caractéristiques du ferrage standard Secure+",
     "Le coulissant aluminium CA76 est fermé par une crémone Secure+ à crochets "
     "inox, manœuvrée par rotation à 90 degrés, avec un entraxe de 750 "
     "millimètres entre les crochets. Le crochet basculant lutte contre l'effet "
     "bilame. Les crochets assurent une résistance à l'arrachement vingt fois "
     "supérieure à la norme NF EN 13126-19. Le dispositif anti-crochetage "
     "interdit une effraction simple par action sur les crochets, sous brevet "
     "numéro 03 12282, et l'anti-fausse manœuvre empêche toute collision entre le "
     "crochet et la gâche à la fermeture."),
    (PREFIXE_CA, 50, "Hauteur de poignée selon les dimensions du vantail en fenêtre",
     "Sur le coulissant aluminium CA76, la hauteur de poignée dépend de la hauteur "
     "du vantail. En fenêtre, un vantail de 440 à 509 millimètres reçoit une "
     "poignée à 240 millimètres et une crémone à un point de fermeture ; de 510 à "
     "909, une poignée à 320 et un point de fermeture ; de 910 à 1009, une poignée "
     "à 320 et deux points ; de 1010 à 1109, une poignée à 420 ; de 1110 à 1209, "
     "une poignée à 520 ; de 1210 à 1309, une poignée à 620 ; de 1310 à 1559, une "
     "poignée à 720, ces quatre derniers cas en deux points de fermeture."),
    (PREFIXE_CA, 50, "Hauteur de poignée et points de fermeture en porte-fenêtre",
     "Sur le coulissant aluminium CA76 en porte-fenêtre, un vantail de 1660 à 2009 "
     "millimètres reçoit une poignée à 1070 millimètres et une crémone à trois "
     "points courts, ramenée à deux points si la hauteur de poignée est hors "
     "standard. Un vantail de 2010 à 2800 millimètres reçoit également une poignée "
     "à 1070 millimètres et une crémone à trois points, ramenée à deux points si "
     "la hauteur de poignée est hors standard ; une version à cinq points est "
     "possible en option, mais elle exclut alors toute hauteur de poignée hors "
     "standard. Le cylindre n'est possible qu'à partir d'une hauteur de battant de "
     "910 millimètres."),
]

F6_BLOCS = [
    (PREFIXE_CA, 26, "Existence et localisation de l'offre couleurs du groupe 1",
     "Le coulissant aluminium CA76 propose un premier groupe de couleurs sans "
     "plus-value, qui rassemble des teintes monocolores, une gamme bicolore "
     "granité mat et un laquage nature lisse mat en imitation anodisation. La "
     "page de l'offre couleurs du tarif donne la liste complète de ces teintes "
     "avec leur code RAL et leur finition, ainsi que les associations "
     "intérieur-extérieur admises en bicolore. Le tarif précise que les teintes "
     "ne sont pas contractuelles et que les laquages imitation anodisation ne "
     "sont pas équivalents à l'anodisation."),
    (PREFIXE_CA, 26, "Existence et localisation de l'offre couleurs du groupe 2",
     "Le coulissant aluminium CA76 propose un second groupe de couleurs, qui "
     "porte une plus-value en pourcentage. Il rassemble l'anodisé nature, les "
     "laquages champagne et bronze en imitation anodisation, une série de RAL en "
     "granité mat, les teintes Futura sablées et une offre bicolore incluant les "
     "tons bois sublimés. La page de l'offre couleurs du tarif donne la liste "
     "complète de ces teintes et les associations intérieur-extérieur admises."),
    (PREFIXE_CA, 27, "Existence et localisation du groupe sublimation tons bois",
     "Le coulissant aluminium CA76 propose un groupe sublimation, qui porte une "
     "plus-value en pourcentage et rassemble cinq tons bois exclusifs : chêne "
     "d'or, chêne foncé, chêne mat, chêne irlandais et acajou veiné. La page "
     "correspondante du tarif donne les associations monocolores et bicolores "
     "admises avec les teintes RAL granité et les laquages aluminium. Le tarif "
     "signale que ces tons bois sublimés sont des teintes approchantes des décors "
     "PVC."),
    (PREFIXE_CA, 27, "Existence et localisation de la liste des autres RAL granité",
     "Au-delà des groupes 1 et 2, le coulissant aluminium CA76 peut recevoir une "
     "large liste d'autres RAL en granité mat, en monocolore comme en bicolore, "
     "avec une plus-value en pourcentage. La page correspondante du tarif énumère "
     "ces références RAL une à une, des beiges et jaunes aux gris, bruns, bleus et "
     "verts. Ces teintes, ainsi que les RAL lisses mats, sont servies sous réserve "
     "de faisabilité et supposent une demande à l'usine."),
    (PREFIXE_CA, 27, "Existence et localisation de l'offre anodisation champagne",
     "Le coulissant aluminium CA76 peut recevoir une finition anodisé champagne, "
     "en monocolore ou en bicolore avec un blanc signalisation granité mat en "
     "intérieur, avec une plus-value en pourcentage. La page correspondante du "
     "tarif précise les associations admises et impose de regrouper le chantier "
     "sur une seule commande, habillages compris, pour respecter le bain "
     "d'anodisation."),
    (PREFIXE_CA, 28, "Existence et localisation des couleurs de poignée et de cache-cylindre",
     "Sur le coulissant aluminium CA76, la teinte de la poignée, du cache-cylindre "
     "et de la grille de ventilation découle de la teinte du dormant et de "
     "l'ouvrant, et non d'un choix indépendant. La page des couleurs des "
     "accessoires du tarif donne, groupe de couleurs par groupe de couleurs et "
     "teinte par teinte, la correspondance à appliquer. Cette correspondance ne "
     "porte pas de plus-value propre : c'est le groupe de couleurs du châssis qui "
     "en porte une."),
    (PREFIXE_CA, 29, "Existence et localisation des règles de couleur des joints",
     "Sur le coulissant aluminium CA76, les joints sont noirs quelles que soient "
     "les teintes intérieure et extérieure du châssis, et les embouts de finition "
     "sont toujours noirs. La page des couleurs des accessoires et des joints du "
     "tarif précise ces règles et le positionnement des joints en coupe. Pour le "
     "groupe sublimation en bicolore, la teinte des accessoires suit celle de la "
     "face intérieure."),
    (PREFIXE_CA, 39, "Existence et localisation des cotes des croisillons à la grecque",
     "Sur le coulissant aluminium CA76, le motif à la grecque suit des cotes X et "
     "Y qui dépendent du nombre de vantaux et de la taille du vitrage. La page des "
     "croisillons à la grecque du tarif donne un tableau à double entrée, en un "
     "vantail et en deux vantaux, croisant des plages de largeur et de hauteur de "
     "vitrage pour fixer le couple de cotes à retenir. Ces cotes doivent être lues "
     "directement sur cette page."),
    (PREFIXE_CA, 40, "Existence et localisation du catalogue des motifs Art Déco",
     "Le coulissant aluminium CA76 peut recevoir des motifs Art Déco gravés, "
     "identifiés par les codes MG1 à MG9 dans le logiciel Syscon et le référentiel "
     "LOOK, auxquels s'ajoutent les motifs en étoile sans cintre MGE2B et MGE3B. "
     "La page des croisillons Art Déco du tarif présente chaque motif, sa largeur "
     "de gravure et sa dimension minimale par volume. Les motifs sont à dimensions "
     "variables, à l'exception du MG8 dont la diagonale de losange est fixe."),
    (PREFIXE_CA, 59, "Existence et localisation de l'historique des évolutions du tarif",
     "Le tarif du coulissant aluminium CA76 tient une page d'évolutions qui "
     "consigne les modifications successives, la page concernée et la date "
     "d'application. La version en vigueur porte une hausse générale appliquée à "
     "l'ensemble du tarif au premier mai 2026. Cette page doit être consultée pour "
     "vérifier si une valeur a évolué depuis la dernière édition consultée."),
]


def gen_statique(blocs, nom):
    chunks, sc = [], 2
    for prefixe, page, titre, corps in blocs:
        chunks.append(emit(f"{prefixe}{titre}",
                           source_line(page, sc_id(sc), "complémentaire"), corps))
        sc += 1
    verifier_titres_uniques(nom, chunks)
    return chunks


# ====================================================== journal des colonnes
def journal_colonnes(header, rows):
    mappees = {C_GAMME, C_CLE, C_CHAP, C_DES, C_DET, C_TAB,
               C_HT, C_TTC, C_PVHT, C_PVTTC, C_HAUTEUR}
    mappees |= set(COLS_HT_L) | set(COLS_TTC_L)
    out = []
    for j in range(NCOL):
        remplies = sum(1 for r in rows if not vide(r["v"][j]))
        lettre = get_column_letter(j + 1)
        if j in (C_MHT, C_MTTC):
            out.append(f"  {lettre} ({header[j]}) : {remplies} lignes — colonne "
                       f"constante, sans information, non reprise en chunk")
        elif j == C_CLE:
            out.append(f"  {lettre} ({header[j]}) : {remplies} lignes — identifiant "
                       f"de ligne, utilisé pour la traçabilité, non repris en chunk")
        elif j not in mappees and remplies:
            out.append(f"  {lettre} ({header[j]}) : {remplies} lignes — NON MAPPÉE")
        if remplies == 0:
            out.append(f"  {lettre} ({header[j]}) : colonne VIDE")
    return out


# ====================================================== écriture
def write_file(fname, sous_type, chunks):
    path = f"{OUTDIR}/{fname}"
    with open(path, "w", encoding="utf-8") as f:
        f.write(yaml_front(sous_type, len(chunks)))
        f.write("\n".join(chunks))
    return path, len(chunks)


def main():
    header, rows, largeurs = load_rows()
    print(f"Lignes lues (séparateurs exclus) : {len(rows)}")

    for xl, g, chap, des in hors_gamme(rows):
        JOURNAL.append(f"ligne hors gamme EXCLUE : Excel {xl} — gamme {g} — "
                       f"{chap} / {des}")

    carte = carte_croisee_renforcee()

    results = [
        write_file("Tarif_CA76_METHODE.md", "methode",
                   gen_statique(F1_BLOCS, "methode")),
        write_file("Tarif_CA76_PRIX_CHASSIS.md", "prix",
                   gen_f2(rows, largeurs, carte)),
        write_file("Tarif_CA76_OPTIONS.md", "options", gen_f3(rows)),
        write_file("Tarif_CA76_PLUS_VALUES_PROPORTIONNELLES.md", "proportionnelles",
                   gen_f4()),
        write_file("Tarif_CA76_FAISABILITES.md", "faisabilites",
                   gen_statique(F5_BLOCS, "faisabilites")),
        write_file("Tarif_CA76_TRANSVERSES.md", "transverses",
                   gen_statique(F6_BLOCS, "transverses")),
    ]

    print("\n=== Fichiers générés ===")
    total = 0
    for path, n in results:
        print(f"  {path.split('/')[-1]:42s} : {n:4d} chunks")
        total += n
    print(f"  {'TOTAL':42s} : {total:4d} chunks")

    print("\n=== Alertes ===")
    print("  Aucune." if not ALERTS else "\n".join("  " + a for a in ALERTS))

    print("\n=== Journal ===")
    for l in JOURNAL:
        print("  " + l)

    print("\n=== Colonnes ===")
    for l in journal_colonnes(header, rows):
        print(l)
    return 0


if __name__ == "__main__":
    sys.exit(main())
