#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Générateur de chunks Markdown pour le tarif T81 (fenêtre et porte-fenêtre PVC).
Conforme à note_cadrage_migration_tarif_T81_v1.md (règles T1 à T7).

Six fichiers :
  F1 METHODE          cotes de tarif, lecture des grilles, vocabulaire (règle T3)
  F2 PRIX_CHASSIS     grilles dimensionnelles                     (règles T1, T2)
  F3 OPTIONS          plus-values forfaitaires                    (règle T4)
  F4 CHASSIS_SPECIAUX cintres, formes, angles, CVR, habillage     (règle T5)
  F5 FAISABILITES     restrictions et compatibilités, sans montant(règle T6)
  F6 TRANSVERSES      orientation sans montant                    (règle T7)

Principes hérités de H81 : fidélité numérique (toute valeur recopiée de la
cellule, jamais calculée), anti-fantôme, SC continue par fichier depuis SC0002,
ligne de source normée, plafond de 200 mots, journal des colonnes non mappées.
"""
import re
import sys
import unicodedata
from collections import OrderedDict, defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

XLSX = "/mnt/user-data/uploads/T81_-infos-tarifs.xlsx"
FEUILLE = "Feuil1"
PDF_SOURCE = "Tarif—T81—HT—09-07-2026.pdf"      # nom affiché dans la ligne de source
PDF_YAML = "Tarif_T81_HT_09-07-2026.pdf"        # nom dans le front matter
OUTDIR = "/mnt/user-data/outputs"
PLAFOND = 200
GAMME = "T81"
DESIGNATION = "Fenêtre PVC"
PREFIXE = f"{GAMME} {DESIGNATION} — "

# ============================================================ index de colonnes
C_GAMME, C_CLE, C_CHAP, C_TAB, C_DES, C_DET = 0, 1, 2, 3, 4, 5
C_HT, C_TTC = 8, 9
PV_VANTAIL = [("Fixe", 10, 13), ("1V", 11, 14), ("2V", 12, 15)]
PV_FORME = [("arc de cercle", 16, 20), ("plein cintre", 17, 21),
            ("gothique", 18, 22), ("anse de panier", 19, 23)]
C_HAUTEUR = 28
COLS_HT_L = list(range(29, 62))     # AD..BJ : Px L xxx HT
COLS_TTC_L = list(range(62, 95))    # BK..CQ : Px L xxx TTC

# ============================================================ table des pages
# Établie contre les en-têtes de page du PDF (chaque page porte son titre) et le
# sommaire général de la page 3. Les sommaires de section (p. 23, 36, 45, 59, 67,
# 81) portent une pagination périmée et ne sont PAS utilisés. L'audit revérifie
# chaque attribution en cherchant le montant sur la page citée.
PAGES = {
    ("Cote tarif", "fenetre à frappe"): 6,
    ("Cote tarif", "coulissant"): 8,
    ("1 OF", "1V fr"): 11,
    ("1 OF", "1V grande hauteur"): 11,
    ("2 OF", "2V fr"): 12,
    ("2 OF", "2V grande hauteur"): 12,
    ("Châssis fixe", "Fixe"): 13,
    ("Châssis à soufflet", "SN"): 14,
    ("Châssis à soufflet", "SN poignée latérale"): 14,
    ("Châssis à soufflet", "SA"): 14,
    ("Coulissant", "coulissant"): 15,
    ("Offre couleurs", "laquage accessoires"): 18,
    ("Vitrage", "PV vitrage thermique, pho et secu"): 26,
    ("Vitrage", "PV vitrage ornementaux"): 27,
    ("Vitrage", "PV composition libre"): 28,
    ("Remplissage", "panneaux"): 29,
    ("Remplissage", "moulures"): 30,
    ("Croisillons", "Alu laqué"): 31,
    ("Croisillons", "Alu 10 mm"): 32,
    ("Croisillons", "motif à la grecque"): 32,
    ("Croisillons", "gravure 10 à 18 mm"): 33,
    ("Croisillons", "gravure 8 à 10 mm"): 33,
    ("Croisillons", "PVC rapporté 1F"): 34,
    ("Croisillons", "PVC rapporté 2F"): 34,
    ("Croisillons", "Croisillons viennois"): 34,
    ("Poignées", "en option"): 37,
    ("Crémones", "prix et finition"): 38,
    ("Meneaux", "menau battant"): 40,
    ("Grilles d'air", "grilles d'air"): 43,
    ("Chatières", "chatière"): 44,
    ("Judas optiques", "Judas"): 44,
    ("Ferrage", "pv options"): 47,
    ("Ferrage", "hauteur poignée"): 47,
    ("Ferrage SA", "pv tringle"): 50,
    ("Ferrage SA", "pv guide"): 50,
    ("Ferrage SA", "pv flexible"): 50,
    ("Ferrage SA", "pv compas"): 51,
    ("Ferrage SA", "commandes spéciales pour SA"): 51,
    ("Ferrage coulissant", "f et pf coulissantes équipées d'une serrure"): 52,
    ("Ferrage coulissant", "2 coulissant à translation"): 52,
    ("Ferrage PF", "poignée de tirage"): 54,
    ("Ferrage PF", "serrures sb"): 54,
    ("Ferrage PF", "serrures cc"): 54,
    ("Ferrage RC2", "PV RC2 fe et pf"): 55,
    ("Ferrage RC2", "PV rabaisser poignée"): 55,
    ("Pieces d'appui", "pieces d'appui neuf"): 60,
    ("Pieces d'appui", "pieces d'appui dormant neuf"): 60,
    ("Pieces d'appui", "pièces d'appui dormant réno"): 60,
    ("Elargisseurs", "sans armature"): 62,
    ("Elargisseurs", "avec armature"): 62,
    ("Elargisseurs", "panneaux élargisseurs"): 62,
    ("Profilés", "profilés de finition"): 63,
    ("Profilés", "tapée et compensateur de feuillure"): 63,
    ("Profilés", "couvre-joints"): 63,
    ("Tapée de doublage", "cotes de référence tarif"): 64,
    ("Accouplements statiques", "profil d'acc droit sans volet roulant"): 65,
    ("Accouplements statiques", "accouplement d'angle"): 65,
    ("Accouplements statiques", "accessoires pour volet traditionnel"): 65,
    ("Seuils", "seuils"): 66,
    ("Angles", "tarification"): 70,
    ("Angles", "acbat"): 70,
    ("Angles", "meneau/traverse"): 70,
    ("Cintres", "principe de tarification"): 73,
    ("Cintres", "oscillo-battant"): 73,
    ("Cintres", "meneau / traverse"): 73,
    ("CVR", "plus-values"): 75,
    ("Forme sur accessoires", "pv pour cintrage"): 76,
    ("Forme sur accessoires", "pv pour découpe des accessoires"): 76,
    ("Châssis spé - Croisillons", "alu laqué blanc"): 77,
    ("Châssis spé - Croisillons", "croisillons"): 77,
    ("Habillage PVC blanc sur vitrage", "habillage pvc blanc"): 78,
    ("Habillage PVC blanc sur dormant", "prix"): 79,
}

# ============================================================ grilles (règle T1)
# (chapitre, tableau) -> (libellé long, synonyme d'usage, continuité en hauteur)
GRILLES = OrderedDict([
    (("1 OF", "1V fr"),
     ("châssis à 1 ouvrant à la française", "fenêtre à un vantail", None)),
    (("1 OF", "1V grande hauteur"),
     ("châssis à 1 ouvrant à la française en grande hauteur",
      "fenêtre à un vantail de grande hauteur", ("1 OF", "1V fr"))),
    (("2 OF", "2V fr"),
     ("châssis à 2 ouvrants égaux à la française", "fenêtre à deux vantaux", None)),
    (("2 OF", "2V grande hauteur"),
     ("châssis à 2 ouvrants égaux à la française en grande hauteur",
      "fenêtre à deux vantaux de grande hauteur", ("2 OF", "2V fr"))),
    (("Châssis fixe", "Fixe"), ("châssis fixe", "fenêtre fixe", None)),
    (("Châssis à soufflet", "SN"),
     ("châssis à soufflet normal (SN)", "soufflet normal", None)),
    (("Châssis à soufflet", "SN poignée latérale"),
     ("châssis à soufflet normal à poignée latérale (SN)",
      "soufflet normal à poignée latérale", None)),
    (("Châssis à soufflet", "SA"),
     ("châssis à soufflet d'aération avec ferme-imposte (SA)",
      "soufflet d'aération", None)),
    (("Coulissant", "coulissant"),
     ("coulissant à translation", "fenêtre coulissante", None)),
])

# grille d'habillage : traitée dans F4, discriminée par l'ordre des lignes
HABILLAGE = ("Habillage PVC blanc sur vitrage", "habillage pvc blanc")
HABILLAGE_FORMES = ["cintré", "rectangulaire"]   # ligne 1 = cintré (légende p. 78)

# ============================================================ libellés F3
CONTEXTE = {
    ("Cote tarif", None): "recoupe d'ailette",
    ("Offre couleurs", None): "laquage",
    ("Vitrage", None): "vitrage",
    ("Remplissage", "panneaux"): "panneau de remplissage",
    ("Remplissage", "moulures"): "moulure de remplissage",
    ("Croisillons", None): "croisillon",
    ("Poignées", None): "poignée",
    ("Crémones", None): "crémone à l'ancienne, finition",
    ("Meneaux", "menau battant"): "meneau battant complémentaire",
    ("Meneaux", None): "meneau complémentaire",
    ("Grilles d'air", None): "grille d'entrée d'air",
    ("Chatières", None): "chatière",
    ("Judas optiques", None): "judas optique",
    ("Ferrage", None): "ferrage",
    ("Ferrage SA", None): "ferrage du soufflet d'aération",
    ("Ferrage coulissant", None): "ferrage du coulissant à translation",
    ("Ferrage PF", None): "ferrage de porte-fenêtre",
    ("Ferrage RC2", "PV RC2 fe et pf"): "plus-value du ferrage RC2 sur fenêtre et porte-fenêtre",
    ("Ferrage RC2", "PV rabaisser poignée"): "plus-value pour rabaisser la poignée en ferrage RC2",
    ("Ferrage RC2", None): "ferrage RC2",
    ("Pieces d'appui", None): "pièce d'appui",
    ("Elargisseurs", None): "élargisseur",
    ("Profilés", None): "profilé complémentaire",
    ("Tapée de doublage", None): "tapée de doublage",
    ("Accouplements statiques", None): "accouplement statique",
    ("Seuils", None): "seuil",
    ("Angles", None): "châssis en angle",
    ("CVR", None): "adaptation sur CVR",
    ("Châssis spé - Croisillons", None): "croisillon de châssis spécial",
    ("Habillage PVC blanc sur dormant", None): "habillage PVC blanc sur dormant",
    ("Cintres", None): "cintrage",
    ("Forme sur accessoires", None): "découpe d'accessoire",
}

# ============================================================ unités (règle T4)
# Le tarif exprime la plupart des plus-values dans une UNITÉ DE FACTURATION que
# l'Excel ne porte pas : elle n'existe que dans le PDF. Table établie page par
# page. Une entrée absente => unité non établie : le chunk le dit et renvoie à
# la page, plutôt que de servir un montant nu.
UNITES = {
    # unité par (chapitre, tableau, désignation), à défaut (chapitre, tableau),
    # à défaut (chapitre, None). Chaque entrée est relevée sur la page du PDF.
    ("Vitrage", None, None): "par mètre carré de surface vitrée du châssis",
    ("Remplissage", "panneaux", None): "par mètre carré",
    ("Remplissage", "moulures", None): "par pièce",
    ("Croisillons", None, None): "par champ",
    ("Meneaux", None, None): "par mètre linéaire",
    ("Offre couleurs", "laquage accessoires", "Plus-value laquage accessoires"):
        "par mètre linéaire",
    ("Offre couleurs", "laquage accessoires", "Forfait laquage volet roulant"):
        "par volet",
    ("Ferrage", "pv options", None): "par pièce",
    ("Ferrage", "hauteur poignée", None):
        "forfaitaire, aucune unité n'étant indiquée sur la page",
    ("Ferrage SA", "pv tringle", None): "par mètre linéaire",
    ("Ferrage SA", "pv guide", None): "par sachet de dix pièces",
    ("Ferrage SA", "pv flexible", None): "par pièce",
    ("Ferrage SA", "pv compas", None): "par pièce",
    ("Ferrage SA", "commandes spéciales pour SA", "Commandes spéciale SA F25"):
        "par pièce",
    ("Ferrage SA", "commandes spéciales pour SA",
     "Cmd spéciale SA ferme imposte : 1 compas"): "par ensemble",
    ("Ferrage SA", "commandes spéciales pour SA",
     "Cmd spéciale SA ferme imposte : 1 compas + 1 verrouillage"): "par ensemble",
    ("Ferrage SA", "commandes spéciales pour SA",
     "Cmd spéciale SA ferme imposte : 1 compas + 2 verrouillages"): "par ensemble",
    ("Ferrage coulissant", "2 coulissant à translation", None): "par unité",
    ("Ferrage PF", "serrures sb", None): "par ensemble",
    ("Ferrage PF", "serrures cc", None): "par ensemble",
    ("Ferrage RC2", "PV RC2 fe et pf", None): "par battant",
    ("Ferrage RC2", "PV rabaisser poignée", None):
        "forfaitaire, aucune unité n'étant indiquée sur la page",
    ("Pieces d'appui", None, None): "par mètre linéaire",
    ("Elargisseurs", "sans armature", None): "par mètre linéaire",
    ("Elargisseurs", "avec armature", None): "par mètre linéaire",
    ("Elargisseurs", "panneaux élargisseurs", None): "par mètre carré",
    ("Profilés", "profilés de finition", None): "par mètre linéaire",
    ("Profilés", "couvre-joints", None): "par mètre linéaire",
    ("Profilés", "tapée et compensateur de feuillure", "TP126 - Blanc"):
        "par mètre linéaire",
    ("Profilés", "tapée et compensateur de feuillure", "TP126 - Plaxage"):
        "par mètre linéaire",
    ("Profilés", "tapée et compensateur de feuillure", "5334 - Blanc"): "par montant",
    ("Profilés", "tapée et compensateur de feuillure", "53347 - Plaxage"): "par montant",
    ("Tapée de doublage", None, None): "par mètre linéaire",
    ("Accouplements statiques", None, None): "par mètre linéaire",
    ("Seuils", "seuils", "5120SN"): "par pièce",
    ("Seuils", "seuils", "AK10123"): "par pièce",
    ("Seuils", "seuils", "AK10123-RAS1"): "par pièce",
    ("Seuils", "seuils", "AK10123-RAS2"): "par pièce",
    ("Seuils", "seuils", "5263"): "par mètre linéaire",
    ("Cote tarif", None, None): "par mètre linéaire",
    ("Ferrage SA", "commandes spéciales pour SA", "Commandes spéciale SA CEFI"):
        "par ensemble",
    ("Ferrage coulissant", "f et pf coulissantes équipées d'une serrure", None):
        "par ensemble",
    ("Ferrage PF", "poignée de tirage", None): "par ensemble",
    ("Angles", "tarification", None): "par châssis, vitrage non compris",
    ("Angles", "acbat", None): "par paumelle",
    ("Angles", "meneau/traverse", None): "par fixation",
    ("Cintres", "principe de tarification", None): "par châssis, vitrage non compris",
    ("Cintres", "oscillo-battant", None): "par châssis",
    ("Cintres", "meneau / traverse", None): "par fixation",
    ("CVR", "plus-values", "Blanc"): "par mètre linéaire",
    ("CVR", "plus-values", "Paxage 1 face"): "par mètre linéaire",
    ("CVR", "plus-values", "Paxage 2 face"): "par mètre linéaire",
    ("CVR", "plus-values", "Blanc panneau"): "par mètre carré",
    ("CVR", "plus-values", "Paxage 1 face panneau"): "par mètre carré",
    ("CVR", "plus-values", "Paxage 2 face panneau"): "par mètre carré",
    ("CVR", "plus-values", "Forfait de coupe"): "forfaitaire pour l'ensemble",
    ("Forme sur accessoires", "pv pour cintrage", None): "par pièce",
    ("Forme sur accessoires", "pv pour découpe des accessoires", None): "par angle",
    ("Châssis spé - Croisillons", "alu laqué blanc", "Champs"): "par champ",
    ("Châssis spé - Croisillons", "alu laqué blanc", None): "par unité",
    ("Châssis spé - Croisillons", "croisillons", None): "par fixation",
    ("Habillage PVC blanc sur dormant", "prix", "Plaque PVC épaisseur 3 mm"):
        "par mètre carré",
    ("Habillage PVC blanc sur dormant", "prix", None): "par unité",
    ("Poignées", None, None): "forfaitaire, aucune unité n'étant indiquée sur la page",
    ("Crémones", None, None): "forfaitaire, aucune unité n'étant indiquée sur la page",
    ("Grilles d'air", None, None): "forfaitaire, aucune unité n'étant indiquée sur la page",
    ("Chatières", None, None): "forfaitaire, aucune unité n'étant indiquée sur la page",
    ("Judas optiques", None, None): "forfaitaire, aucune unité n'étant indiquée sur la page",
}

# Discriminants absents de l'Excel, relevés dans la colonne Désignation du
# tableau du PDF et rattachés par le montant, non par l'ordre des lignes.
DISCRIMINANTS = {
    ("Croisillons", "Alu laqué", "I18+26", 18): "laqué blanc RAL 9016, 18 ou 26 mm",
    ("Croisillons", "Alu laqué", "I18+26", 31): "laqué RAL, 18 ou 26 mm",
    ("Croisillons", "Alu laqué", "I18+26", 22): "chêne d'or 18 mm ou tons bois 26 mm",
    ("Croisillons", "Alu laqué", "I18+26", 13): "laqué blanc RAL 9016, 18 ou 26 mm",
    ("Croisillons", "Alu laqué", "I18+26", 28): "laqué RAL, 18 ou 26 mm",
    ("Croisillons", "Alu laqué", "I18+26", 21): "chêne d'or 18 mm ou tons bois 26 mm",
    ("Croisillons", "Alu laqué", "I45", 36): "laqué blanc RAL 9016, 45 mm",
    ("Croisillons", "Alu laqué", "I45", 49): "laqué RAL, 45 mm",
    ("Croisillons", "Alu laqué", "I45", 24): "laqué blanc RAL 9016, 45 mm",
    ("Croisillons", "Alu laqué", "I45", 38): "laqué RAL, 45 mm",
    ("Croisillons", "PVC rapporté 2F", "NS28/2 NS28/CRM13", 28): "blanc",
    ("Croisillons", "PVC rapporté 2F", "NS28/2 NS28/CRM13", 21): "blanc",
    ("Croisillons", "PVC rapporté 2F", "NS28/2 NS28/CRM13", 39): "décor 2 faces",
    ("Croisillons", "PVC rapporté 2F", "NS28/2 NS28/CRM13", 33): "décor 2 faces",
}

CHAP_GRILLE = {k[0] for k in GRILLES} | {HABILLAGE[0]}
CHAP_SPECIAUX = {"Cintres", "Forme sur accessoires", "Angles", "CVR",
                 "Châssis spé - Croisillons", "Habillage PVC blanc sur dormant"}
CHAP_EXCLUS = {"Exemple de calcul"}   # arbitrage : totaux additionnés, hors corpus

ALERTS, JOURNAL = [], []

# ============================================================ utilitaires
def clean(v):
    return "" if v is None else str(v).strip()


def fmt_euro(v):
    """Recopie la valeur de la cellule, mise en forme. Jamais de calcul."""
    if v in (None, ""):
        return None
    try:
        n = int(round(float(v)))
    except (ValueError, TypeError):
        return None
    return f"{n:,}".replace(",", "\u202f")


def count_words(*parts):
    return len(re.findall(r"\S+", " ".join(parts)))


def sc_id(n):
    return f"SC{n:04d}"


def source_line(page, sc, nature="originale"):
    return f"*Source : {PDF_SOURCE}, page {page} — information {nature} — {sc}*"


def emit(title, source, body):
    n = count_words("##", title, source, body)   # le marqueur compte aussi
    if n > PLAFOND:
        ALERTS.append(f"PLAFOND DÉPASSÉ ({n} mots) : {title[:70]}")
    return f"## {title}\n{source}\n\n{body}\n"


def page_of(chap, tab):
    return PAGES.get((chap, tab), PAGES.get((chap, None), "?"))


def contexte_of(chap, tab):
    return CONTEXTE.get((chap, tab), CONTEXTE.get((chap, None), chap.lower()))


def unite_of(chap, tab, des=None):
    """Unité de facturation relevée sur la page, ou None si non établie."""
    for k in ((chap, tab, des), (chap, tab, None), (chap, None, None)):
        if k in UNITES:
            return UNITES[k]
    return None


def phrase_unite(chap, tab, des=None):
    """Le montant est unitaire : le total revient à l'ADV, jamais au modèle."""
    u = unite_of(chap, tab, des)
    if u is None:
        JOURNAL.append(f"UNITÉ NON ÉTABLIE : {chap} / {tab} / {des} — le chunk "
                       f"renvoie à la page {page_of(chap, tab)}")
        return (f" Le tarif exprime ce montant dans une unité de facturation qui "
                f"doit être lue page {page_of(chap, tab)} du tarif.")
    if u.startswith("forfaitaire"):
        return f" Ce montant est {u}."
    return (f" Ce montant s'entend {u} : le total s'obtient en le multipliant par "
            f"la quantité concernée, calcul qui revient à l'ADV.")


def phrase_unite_pluriel(chap, tab, des=None):
    return (phrase_unite(chap, tab, des)
            .replace(" Ce montant s'entend ", " Ces montants s'entendent ")
            .replace(" Ce montant est ", " Ces montants sont ")
            .replace("en le multipliant", "en les multipliant"))


def enumere(items):
    """['a','b','c'] -> 'a, b et c'. Aucune invention, simple mise en phrase."""
    items = [i for i in items if i]
    if not items:
        return ""
    if len(items) == 1:
        return items[0]
    return ", ".join(items[:-1]) + " et " + items[-1]


def yaml_front(sous_type, nb):
    return (
        "---\n"
        f"document_source: {PDF_YAML}\n"
        "type_document: tarif\n"
        f"sous_type: {sous_type}\n"
        f"gamme_code: {GAMME}\n"
        f'gamme_nom: "{DESIGNATION}"\n'
        'collection: "TRYBA PVC"\n'
        "materiau: PVC\n"
        'version_doc: "2026.07"\n'
        "date_validite: 2026-07-09\n"
        f"nb_chunks: {nb}\n"
        "audiences: [ADV, commercial]\n"
        "---\n\n"
    )


# ============================================================ chargement
def load_rows():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[FEUILLE]
    raw = list(ws.iter_rows(values_only=True))
    header = list(raw[0]) + [None] * (95 - len(raw[0]))
    rows = []
    for i, r in enumerate(raw[1:], start=2):
        r = list(r) + [None] * (95 - len(r))
        if all(v is None for v in r):
            continue
        rows.append({"xl": i, "v": r})
    largeurs = {}
    for j in COLS_HT_L:
        largeurs[j] = int(str(header[j]).split()[2])
    return header, rows, largeurs


def hors_gamme(rows):
    """Lignes dont la colonne gamme n'est pas T81 : exclues, consignées."""
    out = []
    for r in rows:
        g = clean(r["v"][C_GAMME])
        if g and g != GAMME:
            out.append((r["xl"], g, clean(r["v"][C_CHAP]), clean(r["v"][C_DES])))
    return out


# ============================================================ bandes (fait p. 10)
def echelle(rows, chap, tab, largeurs):
    """Échelle des largeurs tarifées de la grille = en-tête du tableau du PDF."""
    ech = set()
    for r in rows:
        v = r["v"]
        if clean(v[C_CHAP]) == chap and clean(v[C_TAB]) == tab:
            for j in COLS_HT_L:
                if v[j] is not None:
                    ech.add(largeurs[j])
    return sorted(ech)


def bande(valeurs, val):
    """Bande couverte par une cote : (précédente + 1) .. val. None si première."""
    i = valeurs.index(val)
    return None if i == 0 else valeurs[i - 1] + 1


def dire_bande(bas, haut, unite="mm"):
    if bas is None:
        return f"jusqu'à {haut} {unite}"
    return f"de {bas} à {haut} {unite}"


# ============================================================ F2 : prix châssis
def lignes_grille(rows, chap, tab, largeurs):
    """[(hauteur, [(largeur, ht, ttc), ...]), ...] dans l'ordre du fichier."""
    out = []
    for r in rows:
        v = r["v"]
        if clean(v[C_CHAP]) != chap or clean(v[C_TAB]) != tab:
            continue
        cells = []
        for j, k in zip(COLS_HT_L, COLS_TTC_L):
            if v[j] is None and v[k] is None:
                continue
            if v[j] is None or v[k] is None:
                JOURNAL.append(f"HT/TTC désaligné ligne Excel {r['xl']} "
                               f"({chap}/{tab}, largeur {largeurs[j]})")
            cells.append((largeurs[j], v[j], v[k]))
        if cells:
            out.append((v[C_HAUTEUR], cells, r["xl"]))
    return out


def gen_prix_grille(chap, tab, libelle, synonyme, continu, rows, largeurs,
                    sc_start=2, forme=None):
    """Règle T1/T2. Découpage piloté par le plafond, jamais par une constante."""
    ech_L = echelle(rows, chap, tab, largeurs)
    lignes = lignes_grille(rows, chap, tab, largeurs)
    hauteurs = [h for h, _, _ in lignes]
    if forme:                                  # habillage : 2 séries par hauteur
        hauteurs = sorted(set(hauteurs))
    ech_H = sorted(set(hauteurs))
    if continu:
        ech_H = sorted(set(ech_H) | set(h for h, _, _ in
                                        lignes_grille(rows, *continu, largeurs)))

    chunks, sc = [], sc_start
    for h, cells, xl in lignes:
        # contiguïté : un trou en milieu de ligne fausserait les bandes
        idx = [ech_L.index(L) for L, _, _ in cells]
        if idx != list(range(idx[0], idx[0] + len(idx))):
            JOURNAL.append(f"largeurs non contiguës ligne Excel {xl} ({chap}/{tab})")

        h_bas = bande(ech_H, h)
        items = []
        for L, ht, ttc in cells:
            l_bas = bande(ech_L, L)
            items.append(f"en largeur {dire_bande(l_bas, L)}, "
                         f"{fmt_euro(ht)} € HT et {fmt_euro(ttc)} € TTC")

        # empaquetage glouton sous le plafond
        i = 0
        while i < len(items):
            j = len(items)
            while j > i:
                lot = items[i:j]
                l_lo = bande(ech_L, cells[i][0])
                titre_l = (f"largeurs {dire_bande(l_lo, cells[j - 1][0])}"
                           if (i or j < len(items)) else "toutes largeurs tarifées")
                nom = libelle if not forme else f"{libelle} {forme}"
                title = (f"{PREFIXE}Tarif {nom}, hauteur "
                         f"{dire_bande(h_bas, h)}, {titre_l}")
                src = source_line(page_of(chap, tab), sc_id(sc))
                body = (f"Sur la grille de prix du {nom} de la fenêtre PVC T81, "
                        f"aussi appelé {synonyme}, pour une cote tarif en hauteur "
                        f"{dire_bande(h_bas, h)}, le tarif est le suivant : "
                        + " ; ".join(lot) + ". Ces prix s'entendent hors "
                        "éco-participation et valent pour un châssis sans complément.")
                if count_words("##", title, src, body) <= PLAFOND or j == i + 1:
                    chunks.append(emit(title, src, body))
                    sc += 1
                    i = j
                    break
                j -= 1
    return chunks, sc


def gen_f2(rows, largeurs):
    chunks, sc = [], 2
    for (chap, tab), (lib, syn, cont) in GRILLES.items():
        c, sc = gen_prix_grille(chap, tab, lib, syn, cont, rows, largeurs, sc)
        chunks += c
    return chunks


# ============================================================ F3 : options
def groupes_options(rows):
    """Regroupement strictement iso-prix : même chapitre, tableau, désignation,
    HT et TTC. Seules les variantes de la colonne Détails sont fusionnées."""
    g = OrderedDict()
    for r in rows:
        v = r["v"]
        chap, tab = clean(v[C_CHAP]), clean(v[C_TAB])
        if (not chap or chap in CHAP_GRILLE or chap in CHAP_SPECIAUX
                or chap in CHAP_EXCLUS or clean(v[C_GAMME]) != GAMME):
            continue
        if v[C_HT] is None:
            JOURNAL.append(f"ligne sans montant, non générée (anti-fantôme) : "
                           f"Excel {r['xl']} — {chap} / {tab} / {clean(v[C_DES])}")
            continue
        key = (chap, tab, clean(v[C_DES]), v[C_HT], v[C_TTC])
        g.setdefault(key, []).append(clean(v[C_DET]))
    return g


# Type de jonction des croisillons. Le tarif écrit « Croisillon T ou Croix » et
# « Croisillon filant » (page 31) ; l'Excel abrège et orthographie « fillant ».
# C'est un libellé de structure, pas une valeur : le tarif fait foi.
JONCTIONS = {
    "croisillons t": "en T ou en croix", "croisillon t": "en T ou en croix",
    "croisillons fillant": "filant", "croisillon fillant": "filant",
}


def jonction_of(variantes):
    v = [JONCTIONS.get(" ".join(str(x).lower().split())) for x in variantes if x]
    return v[0] if v and len(set(v)) == 1 and v[0] else None


def phrase_variantes(variantes):
    v = [x.replace("\n", " ") for x in variantes if x]
    if not v:
        return ""
    if len(v) == 1:
        return f" Ce montant correspond à la référence « {v[0]} »."
    return f" Ce montant s'applique aux références {enumere(['« %s »' % x for x in v])}."


COLLISIONS = {}


def appliquer_discriminants(groupes):
    """Réinjecte, dans la clé du groupe, le discriminant relevé au PDF. Le
    rattachement se fait par le montant : si le montant ne correspond à aucune
    entrée relevée, rien n'est appliqué et le poste retombe sous la règle
    d'exclusion."""
    nouveau, applique = OrderedDict(), 0
    for (chap, tab, des, ht, ttc), variantes in groupes.items():
        d = DISCRIMINANTS.get((chap, tab, " ".join(str(des).split()),
                               int(round(float(ht)))))
        if d:
            des = f"{des} {d}"
            applique += 1
        nouveau[(chap, tab, des, ht, ttc)] = variantes
    if applique:
        JOURNAL.append(f"DISCRIMINANT REPRIS DU PDF : {applique} postes complétés "
                       f"par la finition, rattachement par le montant")
    groupes.clear()
    groupes.update(nouveau)


def exclure_indiscriminables(groupes, chapitre_page):
    """Deux prix différents sous des colonnes strictement identiques (chapitre,
    tableau, désignation, détails) : le discriminant manque dans l'Excel. Aucune
    invention n'étant admise, ces postes ne sont pas générés mais consignés."""
    par_cle = defaultdict(list)
    for k, variantes in groupes.items():
        chap, tab, des, ht, ttc = k
        par_cle[(chap, tab, des, tuple(variantes))].append(k)
    retires = []
    for cle, keys in par_cle.items():
        if len(keys) > 1:
            for k in keys:
                groupes.pop(k, None)
                retires.append(k)
            JOURNAL.append(
                f"DISCRIMINANT MANQUANT — {len(keys)} prix distincts sous des "
                f"colonnes identiques ({cle[0]} / {cle[1]} / {cle[2]} / "
                f"{'+'.join(x for x in cle[3] if x)}) : postes NON générés, "
                f"discriminant à reprendre du PDF page {chapitre_page(cle[0], cle[1])}")
    return retires


def indexer_collisions(groupes):
    """(chap, tab, désignation) portant plusieurs prix : le titre doit alors
    intégrer la variante de la colonne Détails pour rester discriminant."""
    COLLISIONS.clear()
    compte = defaultdict(int)
    for (chap, tab, des, ht, ttc) in groupes:
        compte[(chap, tab, des)] += 1
    COLLISIONS.update({k: v for k, v in compte.items() if v > 1})


def libelle_poste(chap, tab, des, variantes):
    """Libellé auto-discriminant. Quand la désignation est vide, le discriminant
    est repris de la colonne Détails, à défaut du tableau : deux postes de prix
    différents ne peuvent pas porter le même titre."""
    ctx = contexte_of(chap, tab)
    if des:
        base = des
    elif len(variantes) == 1 and variantes[0]:
        base = variantes[0]
    else:
        base = tab
    base = " ".join(str(base).replace("\n", " ").split())
    if base and base.lower() in ctx.lower():
        base = ""
    suffixe = ""
    j = jonction_of(variantes) if chap == "Croisillons" else None
    if j:
        return " ".join(f"{ctx} {base}, {j}".split())
    if des and (chap, tab, des) in COLLISIONS:
        var = [" ".join(str(v).replace("\n", " ").split()) for v in variantes if v]
        if var:
            suffixe = " " + enumere(var)
    return " ".join(f"{ctx} {base}{suffixe}".split())


def verifier_titres_uniques(nom, chunks):
    vus = {}
    for c in chunks:
        t = c.split("\n", 1)[0]
        if t in vus:
            ALERTS.append(f"TITRE NON DISCRIMINANT ({nom}) : {t[:80]}")
        vus[t] = 1


def gen_f3(rows):
    chunks, sc = [], 2
    groupes = groupes_options(rows)
    appliquer_discriminants(groupes)
    exclure_indiscriminables(groupes, page_of)
    indexer_collisions(groupes)
    for (chap, tab, des, ht, ttc), variantes in groupes.items():
        libelle = libelle_poste(chap, tab, des, variantes)
        title = f"{PREFIXE}{libelle[0].upper() + libelle[1:]}, plus-value tarif"
        src = source_line(page_of(chap, tab), sc_id(sc))
        nul = float(ht) == 0
        if nul:
            body = (f"Dans le tarif de la fenêtre PVC T81, le poste « {libelle} » "
                    f"ne donne lieu à aucune plus-value : le tarif le chiffre à "
                    f"0 € HT et 0 € TTC, il est donc compris sans supplément.")
        else:
            body = (f"Dans le tarif de la fenêtre PVC T81, le poste « {libelle} » "
                    f"est chiffré en plus-value à {fmt_euro(ht)} € HT, soit "
                    f"{fmt_euro(ttc)} € TTC.")
            body += phrase_unite(chap, tab, des)
        if not (chap == "Croisillons" and jonction_of(variantes)) \
                and not (des and (chap, tab, des) in COLLISIONS):
            body += phrase_variantes(variantes)
        if not nul:
            body += " Cette plus-value s'entend hors éco-participation."
        chunks.append(emit(title, src, body))
        sc += 1
    verifier_titres_uniques("options", chunks)
    return chunks


# ============================================================ F4 : châssis spéciaux
def gen_f4(rows, largeurs):
    chunks, sc = [], 2

    # 4a. plus-values de cintre par type d'ouvrant (colonnes K..P)
    for r in rows:
        v = r["v"]
        if clean(v[C_CHAP]) != "Cintres" or clean(v[C_TAB]) != "principe de tarification":
            continue
        forme = clean(v[C_DES])
        parts, absents, sans_ttc = [], [], []
        for lab, jh, jt in PV_VANTAIL:
            ht, ttc = v[jh], v[jt]
            nom = {"Fixe": "un châssis fixe", "1V": "un châssis à un vantail",
                   "2V": "un châssis à deux vantaux"}[lab]
            if ht is None:
                absents.append(nom)
                continue
            if ttc is None:
                JOURNAL.append(f"TTC absent : Cintres / {forme} / {lab} "
                               f"(Excel {r['xl']}) — exposé dans le chunk")
                sans_ttc.append(nom)
                parts.append(f"{fmt_euro(ht)} € HT sur {nom}")
            else:
                parts.append(f"{fmt_euro(ht)} € HT, soit {fmt_euro(ttc)} € TTC, sur {nom}")
        title = f"{PREFIXE}Plus-value de la forme {forme.lower()} selon le type d'ouvrant"
        src = source_line(page_of("Cintres", "principe de tarification"), sc_id(sc))
        body = (f"Dans le tarif de la fenêtre PVC T81, la forme {forme.lower()} "
                f"est chiffrée en plus-value à " + enumere(parts) + ".")
        if sans_ttc:
            body += (" Le tarif ne porte pas de valeur TTC pour "
                     + enumere(sans_ttc) + ".")
        if absents:
            body += (" Le tarif ne porte pas de plus-value de cette forme pour "
                     + enumere(absents) + ".")
        body += phrase_unite_pluriel("Cintres", "principe de tarification")
        body += (" Le tarif ajoute par ailleurs une plus-value sur le vitrage, "
                 "exprimée en pourcentage et à lire page "
                 f"{page_of('Cintres', 'principe de tarification')} du tarif.")
        body += " Cette plus-value s'entend hors éco-participation."
        chunks.append(emit(title, src, body))
        sc += 1

    # 4b. plus-values de cintrage sur accessoires (colonnes Q..X)
    for r in rows:
        v = r["v"]
        if clean(v[C_CHAP]) != "Forme sur accessoires" or clean(v[C_TAB]) != "pv pour cintrage":
            continue
        acc = clean(v[C_DES])
        parts = []
        for lab, jh, jt in PV_FORME:
            if v[jh] is None:
                continue
            parts.append(f"{fmt_euro(v[jh])} € HT, soit {fmt_euro(v[jt])} € TTC, "
                         f"pour une forme {lab}")
        title = f"{PREFIXE}Plus-value de cintrage de l'accessoire {acc.lower()}"
        src = source_line(page_of("Forme sur accessoires", "pv pour cintrage"), sc_id(sc))
        body = (f"Dans le tarif de la fenêtre PVC T81, le cintrage de l'accessoire "
                f"« {acc} » est chiffré en plus-value à " + enumere(parts) + ".")
        body += phrase_unite_pluriel("Forme sur accessoires", "pv pour cintrage")
        body += " Cette plus-value s'entend hors éco-participation."
        chunks.append(emit(title, src, body))
        sc += 1

    # 4c. postes scalaires des chapitres châssis spéciaux
    g = OrderedDict()
    for r in rows:
        v = r["v"]
        chap, tab = clean(v[C_CHAP]), clean(v[C_TAB])
        if chap not in CHAP_SPECIAUX or v[C_HT] is None or clean(v[C_GAMME]) != GAMME:
            continue
        g.setdefault((chap, tab, clean(v[C_DES]), v[C_HT], v[C_TTC]), []).append(clean(v[C_DET]))
    exclure_indiscriminables(g, page_of)
    indexer_collisions(g)
    for (chap, tab, des, ht, ttc), variantes in g.items():
        libelle = libelle_poste(chap, tab, des, variantes)
        title = f"{PREFIXE}{libelle[0].upper() + libelle[1:]}, plus-value tarif"
        src = source_line(page_of(chap, tab), sc_id(sc))
        body = (f"Dans le tarif de la fenêtre PVC T81, le poste « {libelle} » "
                f"est chiffré en plus-value à {fmt_euro(ht)} € HT, soit "
                f"{fmt_euro(ttc)} € TTC.")
        body += phrase_unite(chap, tab, des)
        body += phrase_variantes(variantes)
        body += " Cette plus-value s'entend hors éco-participation."
        chunks.append(emit(title, src, body))
        sc += 1
    verifier_titres_uniques("speciaux", chunks)

    # 4d. grille dimensionnelle de l'habillage PVC blanc sur vitrage
    chap, tab = HABILLAGE
    ech_L = echelle(rows, chap, tab, largeurs)
    lignes = lignes_grille(rows, chap, tab, largeurs)
    ech_H = sorted(set(h for h, _, _ in lignes))
    vus = defaultdict(int)
    for h, cells, xl in lignes:
        forme = HABILLAGE_FORMES[min(vus[h], len(HABILLAGE_FORMES) - 1)]
        vus[h] += 1
        h_bas = bande(ech_H, h)
        items = [f"en largeur {dire_bande(bande(ech_L, L), L)}, "
                 f"{fmt_euro(ht)} € HT et {fmt_euro(ttc)} € TTC"
                 for L, ht, ttc in cells]
        i = 0
        while i < len(items):
            j = len(items)
            while j > i:
                lot = items[i:j]
                titre_l = f"largeurs {dire_bande(bande(ech_L, cells[i][0]), cells[j - 1][0])}"
                title = (f"{PREFIXE}Habillage PVC blanc {forme} sur vitrage, hauteur "
                         f"{dire_bande(h_bas, h)}, {titre_l}")
                src = source_line(page_of(chap, tab), sc_id(sc))
                body = (f"Dans le tarif de la fenêtre PVC T81, l'habillage PVC blanc "
                        f"{forme} posé sur vitrage est chiffré, pour une cote tarif "
                        f"en hauteur {dire_bande(h_bas, h)}, comme suit : "
                        + " ; ".join(lot) + ". Cette plus-value s'entend hors "
                        "éco-participation.")
                if count_words("##", title, src, body) <= PLAFOND or j == i + 1:
                    chunks.append(emit(title, src, body))
                    sc += 1
                    i = j
                    break
                j -= 1
    return chunks


# ============================================================ F1, F5, F6
# Chunks rédigés à partir du PDF (pages lues et vérifiées), sans montant pour
# F5 et F6 (règles T6 et T7). Les valeurs citées en F1 sont des cotes, pas des prix.
F1_BLOCS = [
    (6, "Distinction entre cote tarif et cote de fabrication",
     "Le tarif de la fenêtre PVC T81 distingue deux jeux de cotes. Les cotes de "
     "tarif, notées LT et HT, servent de référence au chiffrage des devis. Les "
     "cotes de fabrication, notées L et H, servent de référence à la commande et "
     "à la fabrication des châssis. Toutes les cotes sont exprimées en "
     "millimètres. Un prix lu dans une grille se lit toujours sur une cote de "
     "tarif : utiliser une cote de fabrication à sa place conduit à lire un prix "
     "qui n'est pas celui du châssis commandé. Avant tout chiffrage, il faut "
     "donc établir laquelle des deux cotes est en main."),
    (8, "Passage de la cote de fabrication à la cote de tarif selon le dormant",
     "Sur la fenêtre PVC T81, la relation entre cote de fabrication et cote de "
     "tarif dépend du dormant retenu. Pour les dormants neufs sans ailette, la "
     "cote de tarif est égale à la cote de fabrication. Pour le dormant à "
     "ailette 5114, également désigné LZ102, la cote de tarif vaut la cote de "
     "fabrication augmentée de 35 millimètres en largeur et de 35 millimètres en "
     "hauteur. Les schémas de correspondance figurent aux pages des cotes de "
     "tarif du tarif."),
    (10, "Lecture des grilles de prix par bandes de dimensions",
     "Les grilles de prix de la fenêtre PVC T81 ne donnent pas un prix pour une "
     "dimension exacte mais pour une bande de dimensions. Une colonne intitulée "
     "1000 couvre les largeurs de 901 à 1000 millimètres ; une ligne intitulée "
     "600 couvre les hauteurs de 501 à 600 millimètres. Le prix se lit donc à "
     "l'intersection de la bande de largeur et de la bande de hauteur qui "
     "contiennent la cote de tarif du châssis. Aucun calcul ni aucune "
     "interpolation n'est à faire entre deux valeurs de la grille. Les prix "
     "indiqués sont ceux d'un châssis sans complément."),
    (48, "Les grilles de prix valent abaque des limites de fabrication",
     "Le tarif de la fenêtre PVC T81 précise que les grilles de prix des "
     "différents types de châssis définissent les limites de fabrication des "
     "châssis et peuvent à ce titre être utilisées comme abaques. L'étendue "
     "d'une grille indique donc jusqu'où le châssis se fabrique, et une "
     "combinaison de dimensions absente de la grille n'est pas tarifée. Les "
     "dimensions minimales de fabrication, mesurées à l'extérieur du battant, "
     "sont de 300 millimètres en hauteur et 300 millimètres en largeur pour le "
     "châssis fixe, et de 250 millimètres en hauteur et 400 millimètres en "
     "largeur pour le soufflet d'aération."),
    (10, "Composition d'un prix : le châssis puis ses compléments",
     "Un prix de fenêtre PVC T81 se compose du prix du châssis nu, lu dans la "
     "grille du type d'ouverture concerné, auquel s'ajoutent les plus-values des "
     "compléments retenus : vitrage, remplissage, croisillons, ferrage, "
     "accessoires de pose, teinte et formes particulières. Chaque plus-value est "
     "chiffrée séparément dans le tarif. Les prix s'entendent hors "
     "éco-participation. Le tarif présente des exemples de calcul complets à sa "
     "rubrique dédiée, où le détail de l'addition peut être vérifié."),
    (3, "Vocabulaire des types d'ouverture et abréviations du tarif",
     "Le tarif de la fenêtre PVC T81 emploie des abréviations constantes. OF "
     "désigne l'ouvrant à la française, OB l'oscillo-battant, SN le soufflet "
     "normal, SA le soufflet d'aération avec ferme-imposte, PF la porte-fenêtre, "
     "CVR le coffre de volet roulant. Un châssis à 1 OF est une fenêtre à un "
     "vantail, un châssis à 2 OF une fenêtre à deux vantaux égaux. Sur cette "
     "gamme, la crémone à l'ancienne est le nom d'une option décorative et non "
     "un synonyme de ferrure de verrouillage. Le croisillon est aussi appelé "
     "petit-bois."),
]

F5_BLOCS = [
    (46, "Compatibilité des options avec les ferrages R20 et RC2",
     "Sur la fenêtre PVC T81, le tarif distingue le ferrage R20 et le ferrage "
     "RC2. Sont compatibles avec les deux ferrages l'ouvrant à la française, "
     "l'oscillo-battant, l'oscillo-battant inversé, le soufflet à poignée sur "
     "montant, le soufflet à poignée sur traverse haute, ainsi que la "
     "configuration menuiserie intérieure avec poignée centrée. Ne sont "
     "compatibles qu'avec le ferrage R20, et non avec le ferrage RC2, "
     "l'entrebâilleur sur ouvrant à la française, la poignée centrée, la crémone "
     "à l'ancienne, les serrures SB et CC, le seuil, le châssis fixe, le "
     "soufflet d'aération, le coulissant à translation et les châssis spéciaux."),
    (46, "Restrictions de montage signalées avec le ferrage",
     "Le tarif de la fenêtre PVC T81 signale plusieurs restrictions de montage. "
     "L'oscillo-battant inversé n'est pas compatible avec les châssis spéciaux. "
     "L'entrebâilleur sur ouvrant à la française n'est compatible ni avec le "
     "seuil ni avec le châssis cintré. L'ouvrant à la française est disponible en "
     "ferrage RC2 avec un ferrage symétrique. En configuration SB ou CC, "
     "l'utilisation de l'ouvrant 5404, également désigné Z97, est obligatoire."),
    (38, "Limites d'utilisation de la crémone à l'ancienne selon la configuration",
     "Sur la fenêtre PVC T81, la crémone à l'ancienne est soumise à des limites "
     "d'utilisation qui dépendent de la hauteur mesurée à l'extérieur du battant "
     "et de la configuration du châssis. À hauteur 2000, elle est admise en "
     "fenêtre oscillo-battante et en porte-fenêtre, la béquille étant en option "
     "en fenêtre à ouvrant à la française. À hauteur 2400, elle est admise en "
     "ouvrant à la française, en fenêtre comme en porte-fenêtre, mais pas en "
     "oscillo-battant. À hauteur 2500, elle n'est admise dans aucune "
     "configuration. Le tarif recommande la béquille dans les cas signalés."),
]

F6_BLOCS = [
    (17, "Existence et localisation des tarifs de l'offre couleurs",
     "La fenêtre PVC T81 se décline selon une offre de couleurs organisée en "
     "trois groupes tarifaires. Le premier groupe, monocolore blanc teinté dans "
     "la masse, ne porte aucune plus-value. Les deux autres groupes rassemblent "
     "les plaxages une face et deux faces et portent une plus-value exprimée en "
     "pourcentage du prix du châssis. Les taux applicables, la liste des teintes "
     "de chaque groupe et les conditions de faisabilité des plaxages hors "
     "catalogue figurent aux pages de l'offre couleurs du tarif. Le taux "
     "applicable doit être lu directement sur ces pages."),
    (19, "Existence et localisation des tarifs de teinte des accessoires",
     "Sur la fenêtre PVC T81, la poignée, les paumelles et la grille de "
     "ventilation suivent une offre de couleurs propre, organisée elle aussi en "
     "groupes tarifaires dont le premier est sans plus-value et les suivants "
     "portent une plus-value en pourcentage. La page des couleurs des "
     "accessoires du tarif donne, pour chaque teinte de châssis, la teinte "
     "d'accessoire correspondante, ainsi qu'une table de correspondance entre "
     "les codes de plaxage et les teintes RAL approchantes pour les accessoires "
     "en aluminium. Le taux applicable doit être lu directement sur cette page."),
    (20, "Existence et localisation des règles de couleur des joints",
     "La couleur des joints de la fenêtre PVC T81 dépend de la couleur du "
     "châssis et du type de châssis, vitré ou à panneau. La page des couleurs "
     "des joints du tarif donne la correspondance à appliquer. Cette "
     "correspondance ne porte pas de plus-value ; elle conditionne l'aspect du "
     "châssis et doit être vérifiée sur cette page du tarif au moment de la "
     "commande."),
    (21, "Existence et localisation des tarifs de laquage du bloc-baie",
     "La fenêtre PVC T81 peut être associée à un bloc-baie dont le coffre reçoit "
     "une teinte. Le tarif consacre des pages dédiées aux teintes disponibles "
     "sur les coffres et à la faisabilité des associations entre teinte de "
     "châssis et teinte de coffre, selon le type de coffre, neuf ou rénovation. "
     "La faisabilité et la plus-value applicables doivent être lues directement "
     "sur ces pages du tarif."),
]


def gen_statique(blocs):
    chunks, sc = [], 2
    for page, titre, corps in blocs:
        chunks.append(emit(f"{PREFIXE}{titre}",
                           source_line(page, sc_id(sc), "complémentaire"), corps))
        sc += 1
    return chunks


# ============================================================ journal colonnes
def journal_colonnes(header, rows):
    mappees = set([C_GAMME, C_CLE, C_CHAP, C_TAB, C_DES, C_DET, C_HT, C_TTC, C_HAUTEUR])
    mappees |= {j for _, j, k in PV_VANTAIL for j in (j, k)}
    mappees |= {j for _, j, k in PV_FORME for j in (j, k)}
    mappees |= set(COLS_HT_L) | set(COLS_TTC_L)
    out = []
    for j in range(95):
        remplies = sum(1 for r in rows if r["v"][j] is not None)
        letter = get_column_letter(j + 1)
        if j not in mappees and remplies:
            out.append(f"  {letter} ({header[j]}) : {remplies} lignes — NON MAPPÉE")
        if j in (6, 7) and remplies:
            out.append(f"  {letter} ({header[j]}) : {remplies} lignes — mention "
                       f"de colonne, non reprise en chunk")
        if remplies == 0:
            out.append(f"  {letter} ({header[j]}) : colonne VIDE")
    return out


# ============================================================ écriture
def write_file(fname, sous_type, chunks):
    path = f"{OUTDIR}/{fname}"
    with open(path, "w", encoding="utf-8") as f:
        f.write(yaml_front(sous_type, len(chunks)))
        f.write("\n".join(chunks))
    return path, len(chunks)


def main():
    header, rows, largeurs = load_rows()
    print(f"Lignes lues : {len(rows)}")
    hg = hors_gamme(rows)
    if hg:
        for xl, g, chap, des in hg:
            JOURNAL.append(f"ligne hors gamme EXCLUE (arbitrage en attente) : "
                           f"Excel {xl} — gamme {g} — {chap} / {des}")

    results = [
        write_file("Tarif_T81_METHODE.md", "methode", gen_statique(F1_BLOCS)),
        write_file("Tarif_T81_PRIX_CHASSIS.md", "prix", gen_f2(rows, largeurs)),
        write_file("Tarif_T81_OPTIONS.md", "options", gen_f3(rows)),
        write_file("Tarif_T81_CHASSIS_SPECIAUX.md", "chassis_speciaux",
                   gen_f4(rows, largeurs)),
        write_file("Tarif_T81_FAISABILITES.md", "faisabilites", gen_statique(F5_BLOCS)),
        write_file("Tarif_T81_TRANSVERSES.md", "transverses", gen_statique(F6_BLOCS)),
    ]

    print("\n=== Fichiers générés ===")
    total = 0
    for path, n in results:
        print(f"  {path.split('/')[-1]:34s} : {n:4d} chunks")
        total += n
    print(f"  {'TOTAL':34s} : {total:4d} chunks")

    print("\n=== Alertes plafond ===")
    print("  Aucune." if not ALERTS else "\n".join("  " + a for a in ALERTS))

    print("\n=== Journal ===")
    for l in JOURNAL:
        print("  " + l)
    print("\n=== Colonnes ===")
    for l in journal_colonnes(header, rows):
        print(l)
    return 0


if __name__ == "__main__":
    sys.exit(main())
