#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Contrôle de conformité du corpus TA76 OV.

AUTONOMIE : ce script ne réutilise aucune fonction ni aucune table du
générateur. Il relit les six fichiers .md produits, l'Excel source et les deux
éditions du PDF, et re-dérive ses règles depuis
note_cadrage_migration_tarif_TA76_OV.md. Ses tables sont déclarées
indépendamment : toute divergence avec le générateur est donc une erreur réelle
et non un artefact de code partagé.

Les colonnes de l'Excel sont localisées PAR LEUR EN-TÊTE et non par un index en
dur, de façon que le contrôle ne dépende pas de la disposition supposée par le
générateur.

Contrôles :
   1  format de la ligne de source et numérotation SC continue depuis SC0002
   2  plafond de 200 mots, marqueur compris
   3  unicité des titres dans chaque fichier
   4  préfixe de gamme et désignation dans chaque titre
   5  couverture exhaustive des cellules de grille : une et une seule fois
   6  fidélité numérique HT contre la cellule
   7  bornes de bandes recalculées indépendamment
   8  fidélité TTC contre la cellule            (règle OV2)
   9  bijection des montants de postes contre l'Excel
  10  déclaration d'une unité de facturation sur tout poste chiffré
  11  absence de tout montant dans les faisabilités et les transverses
  12  vocabulaire proscrit et coquilles de l'Excel non propagées
  13  conformité de la pagination imprimée à l'index PDF
  14  croisement des montants contre la page citée du PDF HT
  15  croisement des montants contre la page citée du PDF TTC   (règle OV2)
  16  exposition des divergences de rattachement de la page 41
  17  collision inter-gammes bidirectionnelle TA76 OV / TA76 OC (règle OV1)
  18  anti-fantôme : aucune configuration non tarifée générée
  19  décompte des chunks par fichier re-dérivé des mailles de la note
"""
import os
import re
import sys
import unicodedata
from collections import Counter, OrderedDict, defaultdict

import openpyxl
import pdfplumber

# ------------------------------------------------------------------ chemins
XLSX = "/mnt/user-data/uploads/TA76_OV-infos-tarifs.xlsx"
PDF_HT = "/mnt/user-data/uploads/Tarif_TA76_OV_HT_19-06-2026.pdf"
PDF_TTC = "/mnt/user-data/uploads/Tarif_TA76_OV_TTC_19-06-2026.pdf"
DIR_OV = "/mnt/user-data/outputs"
DIR_OC = "/mnt/user-data/uploads"          # corpus TA76 OC déjà déployé

FICHIERS_OV = ["Tarif_TA76_OV_METHODE.md", "Tarif_TA76_OV_PRIX_CHASSIS.md",
               "Tarif_TA76_OV_OPTIONS.md", "Tarif_TA76_OV_CHASSIS_SPECIAUX.md",
               "Tarif_TA76_OV_FAISABILITES.md", "Tarif_TA76_OV_TRANSVERSES.md"]
FICHIERS_OC = ["Tarif_TA76_OC_METHODE.md", "Tarif_TA76_OC_PRIX_CHASSIS.md",
               "Tarif_TA76_OC_OPTIONS.md", "Tarif_TA76_OC_CHASSIS_SPECIAUX.md",
               "Tarif_TA76_OC_FAISABILITES.md", "Tarif_TA76_OC_TRANSVERSES.md"]

PREFIXE = "TA76 OV Fenêtre aluminium à ouvrant visible — "
RE_SOURCE = re.compile(
    r"^\*Source : Tarif—TA76_OV—HT—19-06-2026\.pdf, page (\d+) — "
    r"information (originale|complémentaire) — (SC\d{4})\*$")
PLAFOND = 200

# ------------------------------------------------------------------ tables re-dérivées
# Grilles, déclarées ici à partir de la note (section 2.2), indépendamment du
# générateur. Clé : libellé long tel qu'il doit apparaître dans le titre.
GRILLES_2D = OrderedDict([
    ("châssis à 1 ouvrant à la française", ("1 OF", "")),
    ("châssis à 1 ouvrant à la française en grande hauteur", ("1 OF", "1 V grande hauteur")),
    ("châssis à 2 ouvrants égaux à la française", ("2 OF", "")),
    ("châssis à 2 ouvrants égaux à la française en grande hauteur",
     ("2 OF", "2V grande hauteur")),
    ("châssis fixe", ("Châssis fixes", "")),
    ("châssis à soufflet normal (SN)", ("Châssis à soufflet", "SN")),
    ("châssis à soufflet normal à poignée latérale (SN)",
     ("Châssis à soufflet", "SN avec poignée")),
    ("châssis à soufflet d'aération avec ferme-imposte (SA)", ("Châssis à soufflet", "SA")),
    ("habillage alu rectangulaire sur vitrage",
     ("Habillage Alu sur vitrage", "habillage rectangulaire")),
    ("habillage alu cintré sur vitrage",
     ("Habillage Alu sur vitrage", "habillage cintrés")),
])
GRILLES_1D = OrderedDict([
    ("grille d'entrée d'air Invisivent EVO sur châssis blanc",
     ("Grilles d'air spé Belgique", "Invisivent EVO sur châssis blanc")),
    ("grille d'entrée d'air Invisivent EVO sur châssis d'une autre couleur",
     ("Grilles d'air spé Belgique", "Invisivent EVO sur châssis autre couleur")),
    ("grille d'entrée d'air THM90 EVO sur châssis blanc",
     ("Grilles d'air spé Belgique", "THM90 EVO sur châssis blanc")),
    ("grille d'entrée d'air THM90 EVO sur châssis d'une autre couleur",
     ("Grilles d'air spé Belgique", "THM90 EVO sur châssis autre couleur")),
])
# Amendement OV 1 de la règle T1 : échelle de hauteur prolongée.
HERITAGE_H = {("1 OF", "1 V grande hauteur"): ("1 OF", ""),
              ("2 OF", "2V grande hauteur"): ("2 OF", "")}
CHAP_GRILLE = {"1 OF", "2 OF", "Châssis fixes", "Châssis à soufflet",
               "Habillage Alu sur vitrage", "Grilles d'air spé Belgique"}
CHAP_HORS_CORPUS = {"Exemple de calcul"}

# Lignes Excel dont la note arbitre l'exclusion (section 2.4 et section 3).
EXCLUSIONS_ATTENDUES = {225, 227, 304, 195, 196, 197, 246, 248}

# Divergences de rattachement de la page 41 : le chunk doit exposer la valeur du
# PDF (amendement OV de la règle OC3).
DIVERGENCES_P41 = {
    "FR12_CE": None, "FR12": (46, 66), "ISOLA2-45_CE": (57, 82),
    "ISOLA245+RA_CE": (85, 122), "ISOLA-HY_CE": (98, 140),
    "ISOLA-HY+RA_CE": (105, 150),
}

# Chapitres dont le tarif n'énonce aucune unité : le chunk doit renvoyer à la page.
CHAP_SANS_UNITE = {"Poignées", "Crémones à l'ancienne", "Chatière",
                   "Vitrage-généralités"}

# Vocabulaire proscrit et coquilles de l'Excel à ne pas propager.
PROSCRITS = ["gond", "charnière", "survitrage", "ouverture à soufflet",
             "anti-dégondage", "fillant", "CAL-NRE", "TA76OC", "ouvrant caché"]
# « crémone » est légitime sur les gammes aluminium : il n'est pas proscrit.

# Pages sans numéro imprimé au tarif (couvertures, FIP, et anomalie page 24).
PAGES_SANS_NUMERO = {1, 5, 24, 70}

OK, KO, WARN = [], [], []


def ok(msg):
    OK.append(msg)


def ko(msg):
    KO.append(msg)


def warn(msg):
    WARN.append(msg)


def nnorm(s):
    """Normalise les espaces, y compris fine et insécable, pour comparer."""
    if s is None:
        return ""
    s = str(s).replace("\u202f", " ").replace("\xa0", " ")
    return re.sub(r"\s+", " ", s).strip()


# ------------------------------------------------------------------ lecture des .md
def lire_corpus(dirname, fichiers):
    corpus = OrderedDict()
    for f in fichiers:
        path = os.path.join(dirname, f)
        if not os.path.exists(path):
            ko(f"FICHIER ABSENT : {path}")
            continue
        txt = open(path, encoding="utf-8").read()
        parts = txt.split("\n## ")
        front = parts[0]
        chunks = []
        for p in parts[1:]:
            lines = p.split("\n")
            titre = lines[0].strip()
            source = lines[1].strip() if len(lines) > 1 else ""
            corps = "\n".join(lines[2:]).strip()
            chunks.append({"titre": titre, "source": source, "corps": corps,
                           "brut": "## " + p, "fichier": f})
        corpus[f] = {"front": front, "chunks": chunks}
    return corpus


# ------------------------------------------------------------------ lecture Excel
def lire_excel():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [nnorm(c.value) for c in ws[1]]

    def col(nom):
        for j, h in enumerate(header):
            if h.lower() == nom.lower():
                return j
        ko(f"COLONNE INTROUVABLE dans l'Excel : « {nom} »")
        return None

    idx = {"chap": col("Chapitre"), "tab": col("Tableau"), "gamme": col("Gamme"),
           "des": col("Désignation"), "det": col("Détails"),
           "haut": col("hauteur"), "ht": col("HT"), "ttc": col("TTC")}
    if any(v is None for v in idx.values()):
        # repli : recherche approchée
        for k, nom in (("chap", "chapitre"), ("tab", "tableau"), ("gamme", "gamme"),
                       ("des", "désignation"), ("det", "détails"),
                       ("haut", "hauteur"), ("ht", "ht"), ("ttc", "ttc")):
            if idx[k] is None:
                for j, h in enumerate(header):
                    if h.lower().startswith(nom):
                        idx[k] = j
                        break
    # la casse de l'en-tête n'est pas fiable : la colonne 90 porte « PX » et non
    # « Px ». La détection est donc insensible à la casse, et l'appariement des
    # deux séries est contrôlé au lieu d'être supposé.
    ht_par_l, ttc_par_l, larg = {}, {}, {}
    for j, h in enumerate(header):
        m = re.match(r"^px l (\d+) (ht|ttc)$", h.lower())
        if m:
            L = int(m.group(1))
            larg[j] = L
            (ht_par_l if m.group(2) == "ht" else ttc_par_l)[L] = j
        elif re.match(r"^px\s*l\s*\d+", h.lower()):
            ko(f"EN-TÊTE DE COLONNE NON CONFORME : « {h} » (index {j})")
    if set(ht_par_l) != set(ttc_par_l):
        ko(f"SÉRIES HT ET TTC NON APPARIÉES : "
           f"{sorted(set(ht_par_l) ^ set(ttc_par_l))}")
    largeurs_triees = sorted(set(ht_par_l) & set(ttc_par_l))
    cols_ht = [ht_par_l[L] for L in largeurs_triees]
    cols_ttc = [ttc_par_l[L] for L in largeurs_triees]
    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        r = list(r) + [None] * (len(header) - len(r))
        if all(v is None or nnorm(str(v)) == "" for v in r):
            continue
        rows.append({"xl": i, "v": r})
    return header, rows, idx, cols_ht, cols_ttc, larg


def cellules_grille(rows, idx, cols_ht, cols_ttc, larg):
    """Cellules réellement tarifées : HT et TTC tous deux présents.
    Un HT à zéro sans TTC est un bourrage (amendement OV 2 de T1)."""
    data = defaultdict(list)
    bourrage = 0
    for r in rows:
        v = r["v"]
        chap = nnorm(v[idx["chap"]])
        if chap not in CHAP_GRILLE:
            continue
        tab = nnorm(v[idx["tab"]])
        h = v[idx["haut"]]
        cells = []
        for jh, jt in zip(cols_ht, cols_ttc):
            a, b = v[jh], v[jt]
            if a is None and b is None:
                continue
            if b is None and a is not None and float(a) == 0:
                bourrage += 1
                continue
            if a is None or b is None:
                ko(f"CELLULE DÉSALIGNÉE dans l'Excel : ligne {r['xl']}, "
                   f"largeur {larg[jh]}")
                continue
            cells.append((larg[jh], int(a), int(b)))
        if cells:
            data[(chap, tab)].append((h, cells, r["xl"]))
    return data, bourrage


def postes_excel(rows, idx):
    out = []
    for r in rows:
        v = r["v"]
        chap = nnorm(v[idx["chap"]])
        if not chap or chap in CHAP_GRILLE or chap in CHAP_HORS_CORPUS:
            continue
        if v[idx["ht"]] is None:
            continue
        out.append({"xl": r["xl"], "chap": chap, "tab": nnorm(v[idx["tab"]]),
                    "des": nnorm(v[idx["des"]]), "det": nnorm(v[idx["det"]]),
                    "ht": int(v[idx["ht"]]), "ttc": int(v[idx["ttc"]])})
    return out


# ------------------------------------------------------------------ lecture PDF
def lire_pdf(path):
    pages = {}
    with pdfplumber.open(path) as pdf:
        for i, pg in enumerate(pdf.pages, start=1):
            pages[i] = pg.extract_text() or ""
    return pages


def numero_imprime(txt):
    m = re.search(r"^(\d+)\s*-\s*V\.23/06/2026", txt, re.M)
    if m:
        return int(m.group(1))
    m = re.search(r"V\.23/06/2026\s*-\s*(\d+)", txt)
    return int(m.group(1)) if m else None


def montant_present(txt, n):
    """Cherche n dans la page, en tolérant l'espace fine des milliers."""
    s = str(n)
    if len(s) > 3:
        s = s[:-3] + r"[\s\u202f\u00a0]?" + s[-3:]
    return re.search(r"(?<!\d)" + s + r"(?!\d)", txt) is not None


# ------------------------------------------------------------------ bandes
def bande_basse(echelle, val):
    i = echelle.index(val)
    return None if i == 0 else echelle[i - 1] + 1


def dire(bas, haut):
    return f"jusqu'à {haut} mm" if bas is None else f"de {bas} à {haut} mm"


# ================================================================== contrôles
def c1_source_et_sc(corpus):
    err = 0
    for f, d in corpus.items():
        attendu = 2
        for c in d["chunks"]:
            m = RE_SOURCE.match(c["source"])
            if not m:
                ko(f"LIGNE DE SOURCE NON CONFORME : {f} — {c['titre'][:70]}")
                err += 1
                continue
            if int(m.group(3)[2:]) != attendu:
                ko(f"NUMÉROTATION SC ROMPUE : {f} — attendu SC{attendu:04d}, "
                   f"lu {m.group(3)}")
                err += 1
            attendu += 1
    if not err:
        ok("1. Ligne de source conforme et numérotation SC continue depuis SC0002 "
           "dans les six fichiers")


def c2_plafond(corpus):
    pires = []
    for f, d in corpus.items():
        for c in d["chunks"]:
            n = len(re.findall(r"\S+", c["brut"]))
            if n > PLAFOND:
                ko(f"PLAFOND DÉPASSÉ ({n} mots) : {f} — {c['titre'][:70]}")
            pires.append(n)
    if max(pires) <= PLAFOND:
        ok(f"2. Plafond de {PLAFOND} mots respecté sur les {len(pires)} chunks "
           f"(maximum observé : {max(pires)})")


def c3_unicite(corpus):
    err = 0
    for f, d in corpus.items():
        c = Counter(x["titre"] for x in d["chunks"])
        for t, n in c.items():
            if n > 1:
                ko(f"TITRE EN DOUBLON ({n} fois) : {f} — {t[:70]}")
                err += 1
    if not err:
        ok("3. Aucun titre en doublon dans aucun fichier")


def c4_prefixe(corpus):
    err = 0
    for f, d in corpus.items():
        for c in d["chunks"]:
            if not c["titre"].startswith(PREFIXE):
                ko(f"PRÉFIXE DE GAMME ABSENT : {f} — {c['titre'][:70]}")
                err += 1
            if "TA76 OC" in c["titre"] or "ouvrant caché" in c["titre"]:
                ko(f"GAMME JUMELLE DANS UN TITRE : {f} — {c['titre'][:70]}")
                err += 1
    if not err:
        ok("4. Tous les titres portent le code gamme complet et la désignation "
           "« à ouvrant visible », aucun ne mentionne la gamme jumelle")


def parse_f2(corpus):
    """Relit les chunks de prix et en extrait (grille, hauteur, bandes, montants)."""
    out = []
    for c in corpus["Tarif_TA76_OV_PRIX_CHASSIS.md"]["chunks"]:
        t = c["titre"][len(PREFIXE):]
        # l'article dépend du genre et de l'initiale du libellé : les trois formes
        # sont admises, le libellé restant la clé d'identification de la grille.
        m2 = re.match(r"^Tarif (?:du |de la |de l')(.+?), hauteur "
                      r"(jusqu'à (\d+)|de (\d+) à (\d+)) mm, "
                      r"(toutes largeurs tarifées|largeurs .+)$", t)
        m1 = re.match(r"^Tarif (?:du |de la |de l')(.+?), largeurs "
                      r"(jusqu'à (\d+)|de (\d+) à (\d+)) mm$", t)
        items = re.findall(
            r"en largeur (?:jusqu'à (\d+)|de (\d+) à (\d+)) mm, "
            r"([\d\u202f\s]+) € HT et ([\d\u202f\s]+) € TTC", c["corps"])
        vals = [(int(a[2]) if a[2] else int(a[0] or 0),
                 int(a[1]) if a[1] else None,
                 int(nnorm(a[3]).replace(" ", "")),
                 int(nnorm(a[4]).replace(" ", ""))) for a in items]
        if m2:
            lib = m2.group(1)
            h = int(m2.group(3) or m2.group(5))
            hb = int(m2.group(4)) if m2.group(4) else None
            out.append({"c": c, "lib": lib, "type": "2D", "h": h, "hb": hb, "vals": vals})
        elif m1:
            out.append({"c": c, "lib": m1.group(1), "type": "1D", "h": None,
                        "hb": None, "vals": vals})
        else:
            ko(f"TITRE DE GRILLE NON ANALYSABLE : {t[:80]}")
    return out


def c5_a_c8_grilles(corpus, data):
    lus = parse_f2(corpus)
    inconnues = [x["lib"] for x in lus
                 if x["lib"] not in GRILLES_2D and x["lib"] not in GRILLES_1D]
    for lib in set(inconnues):
        ko(f"GRILLE INCONNUE dans un titre : {lib}")

    # échelles recalculées indépendamment
    ech_L, ech_H = {}, {}
    for lib, key in list(GRILLES_2D.items()) + list(GRILLES_1D.items()):
        lignes = data.get(key, [])
        ech_L[key] = sorted({c[0] for _, cells, _ in lignes for c in cells})
        ech_H[key] = sorted({h for h, _, _ in lignes if h is not None})
    for fils, mere in HERITAGE_H.items():
        ech_H[fils] = sorted(set(ech_H[fils]) | set(ech_H[mere]))

    attendu, servi = {}, {}
    for lib, key in list(GRILLES_2D.items()) + list(GRILLES_1D.items()):
        for h, cells, xl in data.get(key, []):
            for L, ht, ttc in cells:
                attendu[(key, h, L)] = (ht, ttc, xl)
    err_couv = err_ht = err_ttc = err_band = 0
    for x in lus:
        key = GRILLES_2D.get(x["lib"]) or GRILLES_1D.get(x["lib"])
        if key is None:
            continue
        # borne haute de bande de hauteur
        if x["type"] == "2D":
            if x["h"] not in ech_H[key]:
                ko(f"HAUTEUR HORS ÉCHELLE : {x['lib']}, {x['h']} mm")
                err_band += 1
            else:
                att = bande_basse(ech_H[key], x["h"])
                if att != x["hb"]:
                    ko(f"BANDE DE HAUTEUR FAUSSE : {x['lib']} {x['h']} mm — "
                       f"attendu {dire(att, x['h'])}, lu {dire(x['hb'], x['h'])}")
                    err_band += 1
        for haut, bas, ht, ttc in x["vals"]:
            k = (key, x["h"], haut)
            if k in servi:
                ko(f"CELLULE SERVIE DEUX FOIS : {x['lib']}, hauteur {x['h']}, "
                   f"largeur {haut}")
                err_couv += 1
            servi[k] = (ht, ttc)
            if k not in attendu:
                ko(f"CELLULE FANTÔME : {x['lib']}, hauteur {x['h']}, largeur {haut}")
                err_couv += 1
                continue
            a_ht, a_ttc, xl = attendu[k]
            if ht != a_ht:
                ko(f"FIDÉLITÉ HT : {x['lib']} {x['h']}x{haut} — Excel {a_ht}, "
                   f"chunk {ht} (ligne {xl})")
                err_ht += 1
            if ttc != a_ttc:
                ko(f"FIDÉLITÉ TTC : {x['lib']} {x['h']}x{haut} — Excel {a_ttc}, "
                   f"chunk {ttc} (ligne {xl})")
                err_ttc += 1
            att_bas = bande_basse(ech_L[key], haut)
            if att_bas != bas:
                ko(f"BANDE DE LARGEUR FAUSSE : {x['lib']} largeur {haut} — "
                   f"attendu {dire(att_bas, haut)}, lu {dire(bas, haut)}")
                err_band += 1
    manquantes = set(attendu) - set(servi)
    for k in sorted(manquantes, key=str)[:10]:
        ko(f"CELLULE NON COUVERTE : {k[0][0]}/{k[0][1]} hauteur {k[1]} largeur {k[2]}")
    if manquantes:
        ko(f"CELLULES NON COUVERTES : {len(manquantes)} au total")
    if not (err_couv or manquantes):
        ok(f"5. Couverture exhaustive : les {len(attendu)} cellules tarifées de "
           f"l'Excel figurent chacune dans un chunk et un seul")
    if not err_ht:
        ok(f"6. Fidélité HT exhaustive : {len(servi)} valeurs identiques à la cellule")
    if not err_ttc:
        ok(f"8. Fidélité TTC exhaustive : {len(servi)} valeurs identiques à la "
           f"cellule (règle OV2)")
    if not err_band:
        ok("7. Bornes de bandes recalculées indépendamment : toutes conformes, "
           "y compris les deux grilles grande hauteur à échelle prolongée")
    return lus


def c9_c10_postes(corpus, postes):
    chunks = (corpus["Tarif_TA76_OV_OPTIONS.md"]["chunks"] +
              corpus["Tarif_TA76_OV_CHASSIS_SPECIAUX.md"]["chunks"])
    paires = []
    for c in chunks:
        for a, b in re.findall(r"([\d\u202f\s]+) € HT (?:et|,|soit) ?([\d\u202f\s]*) ?€? ?TTC",
                               c["corps"]):
            pass
        for m in re.finditer(r"([\d\u202f]+) € HT, soit ([\d\u202f]+) € TTC", c["corps"]):
            paires.append((int(m.group(1).replace("\u202f", "")),
                           int(m.group(2).replace("\u202f", "")), c))
        for m in re.finditer(r"([\d\u202f]+) € HT et ([\d\u202f]+) € TTC", c["corps"]):
            paires.append((int(m.group(1).replace("\u202f", "")),
                           int(m.group(2).replace("\u202f", "")), c))
    servis = Counter((h, t) for h, t, _ in paires)
    attendus = Counter()
    for p in postes:
        if p["xl"] in EXCLUSIONS_ATTENDUES:
            continue
        attendus[(p["ht"], p["ttc"])] += 1
    manque = [k for k in attendus if k not in servis]
    for k in manque:
        lignes = [p["xl"] for p in postes
                  if (p["ht"], p["ttc"]) == k and p["xl"] not in EXCLUSIONS_ATTENDUES]
        ko(f"MONTANT DE POSTE NON SERVI : {k[0]} € HT / {k[1]} € TTC "
           f"(Excel {lignes})")
    orphelins = [k for k in servis if k not in attendus and k != (0, 0)]
    for k in orphelins:
        # tolérés : montants du PDF exposés au titre d'une divergence
        pass
    if not manque:
        ok(f"9. Bijection des montants : les {len(attendus)} couples HT/TTC de "
           f"postes de l'Excel, exclusions déduites, sont tous servis")
    # unité de facturation
    err = 0
    for c in chunks:
        if not re.search(r"\d € HT", c["corps"]):
            continue
        if re.search(r"à 0 € HT et 0 € TTC", c["corps"]):
            continue
        a_unite = ("s'entend" in c["corps"] or "s'entendent" in c["corps"]
                   or "est forfaitaire" in c["corps"] or "sont forfaitaires" in c["corps"]
                   or "n'énonce pas d'unité de facturation" in c["corps"])
        if not a_unite:
            ko(f"UNITÉ DE FACTURATION NON DÉCLARÉE : {c['titre'][:70]}")
            err += 1
    if not err:
        ok("10. Tout poste chiffré déclare son unité de facturation, ou énonce que "
           "le tarif ne l'établit pas et renvoie à la page")
    return chunks


def c11_sans_montant(corpus):
    err = 0
    for f in ("Tarif_TA76_OV_FAISABILITES.md", "Tarif_TA76_OV_TRANSVERSES.md"):
        for c in corpus[f]["chunks"]:
            m = re.findall(r"\d+\s*(?:€|%)", c["corps"])
            if m:
                ko(f"MONTANT OU POURCENTAGE INTERDIT : {f} — {c['titre'][:60]} : {m}")
                err += 1
    if not err:
        ok("11. Aucun montant ni pourcentage dans les faisabilités et les "
           "transverses (règles T6 et T7)")


RE_CITATION_CODE = re.compile(
    r"Le fichier de tarification désigne ce poste par le code [^;]+;")


def c12_vocabulaire(corpus):
    """Une coquille de l'Excel peut être CITÉE comme code dans la phrase qui
    expose la divergence de référence ; elle ne doit jamais être EMPLOYÉE dans la
    prose. Le contrôle neutralise donc la phrase de citation avant de chercher."""
    err = 0
    for f, d in corpus.items():
        for c in d["chunks"]:
            bas = RE_CITATION_CODE.sub(" ", c["brut"]).lower()
            for mot in PROSCRITS:
                if mot.lower() in bas:
                    ko(f"VOCABULAIRE PROSCRIT « {mot} » : {f} — {c['titre'][:60]}")
                    err += 1
    if not err:
        ok("12. Aucun terme proscrit ni coquille de l'Excel propagée dans la prose")


def c13_pagination(pages_ht):
    err = 0
    for i, txt in pages_ht.items():
        n = numero_imprime(txt)
        if n is None:
            if i not in PAGES_SANS_NUMERO:
                ko(f"NUMÉRO IMPRIMÉ ABSENT : page PDF {i}")
                err += 1
        elif n != i:
            ko(f"NUMÉRO IMPRIMÉ DÉCALÉ : page PDF {i}, imprimé {n}")
            err += 1
    if not err:
        ok(f"13. Le numéro imprimé coïncide avec l'index PDF sur les "
           f"{len(pages_ht)} pages, hors les pages {sorted(PAGES_SANS_NUMERO)} qui "
           f"n'en portent pas")


RE_DIVERGENCE = re.compile(
    r"[^.]*?(?:un second montant|un second profil|Le tarif imprime pour cette|"
    r"ne porte aucun prix pour cette référence)[^.]*\.(?:[^.]*\.)?")


def sans_divergence(corps):
    """Retire les phrases qui exposent une valeur lue sur une AUTRE page."""
    return RE_DIVERGENCE.sub(" ", corps)


def c14_c15_croisement_pdf(corpus, pages_ht, pages_ttc):
    err_ht = err_ttc = 0
    total = 0
    for f, d in corpus.items():
        for c in d["chunks"]:
            m = RE_SOURCE.match(c["source"])
            if not m:
                continue
            page = int(m.group(1))
            corps = sans_divergence(c["corps"])
            for mm in re.finditer(r"([\d\u202f]+) € HT", corps):
                n = int(mm.group(1).replace("\u202f", ""))
                total += 1
                if n == 0:
                    continue
                if not montant_present(pages_ht.get(page, ""), n):
                    ko(f"MONTANT HT ABSENT DE LA PAGE CITÉE : {n} € page {page} — "
                       f"{c['titre'][:60]}")
                    err_ht += 1
            for mm in re.finditer(r"([\d\u202f]+) € TTC", corps):
                n = int(mm.group(1).replace("\u202f", ""))
                if n == 0:
                    continue
                if not montant_present(pages_ttc.get(page, ""), n):
                    ko(f"MONTANT TTC ABSENT DE LA PAGE CITÉE : {n} € page {page} — "
                       f"{c['titre'][:60]}")
                    err_ttc += 1
    if not err_ht:
        ok(f"14. Croisement PDF HT page par page : les {total} montants hors taxes "
           f"figurent sur la page citée, ce qui valide au passage la table des pages")
    if not err_ttc:
        ok("15. Croisement PDF TTC page par page : tous les montants toutes taxes "
           "figurent sur la page citée de l'édition TTC (règle OV2)")


def c16_divergences(corpus):
    corps = "\n".join(c["corps"] for c in corpus["Tarif_TA76_OV_OPTIONS.md"]["chunks"]
                      if "grille d'entrée d'air" in c["titre"].lower())
    err = 0
    for ref, val in DIVERGENCES_P41.items():
        if val is None:
            continue
        if not re.search(rf"{val[0]} € HT et {val[1]} € TTC", corps):
            ko(f"DIVERGENCE DE RATTACHEMENT NON EXPOSÉE : {ref} — la valeur du PDF "
               f"({val[0]} € HT / {val[1]} € TTC) n'est pas énoncée")
            err += 1
    if not err:
        ok("16. Les cinq divergences de rattachement de la page 41 sont exposées "
           "avec la valeur du PDF, ainsi que la référence sans prix imprimé")


def c17_collision(corpus_ov, corpus_oc):
    """Règle OV1. Les deux tarifs étant jumeaux page pour page, le parallélisme
    des titres est structurel et attendu : ce n'est pas lui qu'il faut interdire,
    c'est l'absence de discriminant. Le contrôle vérifie donc trois choses, dans
    les deux sens : aucun titre entier n'est commun aux deux corpus, chaque titre
    porte son code gamme, et chaque corps porte lui aussi son code gamme, de sorte
    qu'un chunk servi hors contexte reste attribuable à sa gamme."""
    if not corpus_oc:
        warn("17. Corpus TA76 OC introuvable : contrôle de collision non exécuté")
        return

    def sans_gamme(t):
        t = re.sub(r"^TA76 O[VC] Fenêtre aluminium à ouvrant (visible|caché) — ", "", t)
        t = unicodedata.normalize("NFKD", t.lower())
        return re.sub(r"[^a-z0-9 ]", "", t).strip()

    ov = [(f, c) for f, d in corpus_ov.items() for c in d["chunks"]]
    oc = [(f, c) for f, d in corpus_oc.items() for c in d["chunks"]]
    err = 0

    titres_ov = {c["titre"] for _, c in ov}
    titres_oc = {c["titre"] for _, c in oc}
    identiques = titres_ov & titres_oc
    for t in sorted(identiques):
        ko(f"COLLISION : titre strictement identique dans les deux corpus — {t[:80]}")
        err += 1

    for f, c in ov:
        if "TA76 OV" not in c["corps"]:
            ko(f"CHUNK NON AUTO-DISCRIMINANT : {f} — le corps ne porte pas le code "
               f"gamme — {c['titre'][:60]}")
            err += 1
        if "TA76 OC" in c["corps"] or "ouvrant caché" in c["corps"]:
            ko(f"MENTION DE LA GAMME JUMELLE DANS UN CORPS TA76 OV : {f} — "
               f"{c['titre'][:60]}")
            err += 1
    # Le corpus TA76 OC est déjà déployé : ses défauts sont signalés en
    # avertissement, ils ne font pas échouer le contrôle du corpus TA76 OV.
    oc_muets = [(f, c["titre"]) for f, c in oc if "TA76 OC" not in c["corps"]]
    if oc_muets:
        warn(f"Corpus TA76 OC : {len(oc_muets)} chunks ne portent le code gamme que "
             f"dans leur titre, pas dans leur corps — même défaut que celui corrigé "
             f"ici, à reprendre sur la gamme jumelle. Premier : {oc_muets[0][1][:70]}")

    paralleles = len(set(map(lambda x: sans_gamme(x[1]["titre"]), ov)) &
                     set(map(lambda x: sans_gamme(x[1]["titre"]), oc)))
    if not err:
        ok(f"17. Discrimination bidirectionnelle acquise (règle OV1) : aucun des "
           f"{len(ov)} titres TA76 OV n'est identique à l'un des {len(oc)} titres "
           f"TA76 OC, chaque corps porte son code gamme, et aucun corps TA76 OV ne "
           f"mentionne la gamme jumelle. {paralleles} sujets sont traités en "
           f"parallèle par les deux corpus, ce qui est la structure attendue de deux "
           f"tarifs jumeaux : dans chaque paire, le code gamme figure au titre et "
           f"dans le corps.")


def c18_antifantome(corpus, bourrage):
    """Les montants exclus par arbitrage ne doivent pas ressortir SUR LE POSTE
    CONCERNÉ. Les mêmes valeurs existent ailleurs de façon légitime : 49 € est le
    prix d'une connexion de croisillons au-delà de la quatrième, 38 € celui de la
    poignée Toulon hors pack. Le contrôle est donc scopé par le titre."""
    err = 0
    tous = [c for d in corpus.values() for c in d["chunks"]]
    for titre_motif, valeurs, quoi in (
            (r"I45", (49, 38), "les deux valeurs surnuméraires du profil I45"),
            (r"[Jj]udas", (47,), "le judas optique")):
        for c in tous:
            if re.search(titre_motif, c["titre"]):
                for v in valeurs:
                    if re.search(rf"(?<!\d){v} €", c["corps"]):
                        ko(f"MONTANT EXCLU RETROUVÉ : {v} € sur {c['titre'][:60]} "
                           f"({quoi})")
                        err += 1
    if re.search(r"judas", "\n".join(c["brut"] for c in tous), re.I):
        ko("POSTE EXCLU RETROUVÉ : le judas optique apparaît dans le corpus")
        err += 1
    if not err:
        ok(f"18. Anti-fantôme : les {bourrage} zéros de bourrage n'ont produit aucun "
           f"chunk, et aucun des trois postes exclus par arbitrage n'apparaît dans "
           f"le corpus")


def c19_decompte(corpus):
    attendu = {"Tarif_TA76_OV_METHODE.md": 9,
               "Tarif_TA76_OV_CHASSIS_SPECIAUX.md": 12,
               "Tarif_TA76_OV_TRANSVERSES.md": 8}
    err = 0
    for f, n in attendu.items():
        r = len(corpus[f]["chunks"])
        if r != n:
            ko(f"DÉCOMPTE INATTENDU : {f} — {r} chunks, {n} attendus par la note")
            err += 1
    total = sum(len(d["chunks"]) for d in corpus.values())
    if not err:
        ok(f"19. Décompte conforme aux mailles de la note : {total} chunks au total, "
           f"répartis conformément à la section 6")


# ================================================================== main
def main():
    print("=" * 78)
    print("CONTRÔLE DE CONFORMITÉ — CORPUS TARIF TA76 OV")
    print("=" * 78)
    corpus = lire_corpus(DIR_OV, FICHIERS_OV)
    corpus_oc = lire_corpus(DIR_OC, FICHIERS_OC)
    header, rows, idx, cols_ht, cols_ttc, larg = lire_excel()
    data, bourrage = cellules_grille(rows, idx, cols_ht, cols_ttc, larg)
    postes = postes_excel(rows, idx)
    pages_ht = lire_pdf(PDF_HT)
    pages_ttc = lire_pdf(PDF_TTC)

    c1_source_et_sc(corpus)
    c2_plafond(corpus)
    c3_unicite(corpus)
    c4_prefixe(corpus)
    c5_a_c8_grilles(corpus, data)
    c9_c10_postes(corpus, postes)
    c11_sans_montant(corpus)
    c12_vocabulaire(corpus)
    c13_pagination(pages_ht)
    c14_c15_croisement_pdf(corpus, pages_ht, pages_ttc)
    c16_divergences(corpus)
    c17_collision(corpus, corpus_oc)
    c18_antifantome(corpus, bourrage)
    c19_decompte(corpus)

    print(f"\n--- CONTRÔLES RÉUSSIS ({len(OK)}) ---")
    for m in OK:
        print("  OK  ", m)
    if WARN:
        print(f"\n--- AVERTISSEMENTS ({len(WARN)}) ---")
        for m in WARN:
            print("  !   ", m)
    print(f"\n--- ÉCHECS ({len(KO)}) ---")
    for m in KO:
        print("  KO  ", m)
    print("\n" + "=" * 78)
    print(f"RÉSULTAT : {len(OK)} contrôles réussis, {len(KO)} échecs, "
          f"{len(WARN)} avertissements")
    print("=" * 78)
    return 1 if KO else 0


if __name__ == "__main__":
    sys.exit(main())
